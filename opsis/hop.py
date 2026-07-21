"""
Parser impor data HOP (Hari Operasi) dari workbook konfirmasi stok batu bara /
BBM Sulawesi. Dipakai bersama oleh management command `import_hop` dan form
upload di dashboard OPSIS (opsis.views.hop_import).

Struktur sheet mentah 'Batubara' / 'BBM':
    baris 1  : judul (diabaikan)
    baris 2  : header -> NO | SISTEM | PEMBANGKIT | ASET | DMN (MW) | <tanggal...>
    baris 3+ : satu baris per pembangkit; kolom F.. berisi nilai HOP per tanggal.

Catatan:
- Kolom SISTEM adalah merged-cell (hanya baris pertama tiap grup terisi) ->
  di-forward-fill.
- Di bawah tabel utama ada baris sisa dengan NO kosong -> di-skip (hanya baris
  dengan NO numerik yang diproses).
"""
import datetime

# (nama_sheet, kode_kategori) yang diproses dari workbook
SHEETS = [('Batubara', 'batubara'), ('BBM', 'bbm')]

# Indeks kolom (1-based) sesuai layout header baris 2
COL_NO, COL_SISTEM, COL_PEMBANGKIT, COL_ASET, COL_DMN = 1, 2, 3, 4, 5
COL_TANGGAL_MULAI = 6  # kolom tanggal pertama


def _to_float(v):
    if v is None or v == '':
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _as_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None


def parse_sheet(ws, kategori):
    """
    Parse satu worksheet -> list dict pembangkit:
        {nama, kategori, sistem, aset, dmn_mw, urutan, snaps: {date: hop}}
    """
    # Header baris 2: petakan indeks kolom -> tanggal
    header = [c.value for c in ws[2]]
    date_cols = {}  # idx(1-based) -> date
    for idx, val in enumerate(header, start=1):
        if idx < COL_TANGGAL_MULAI:
            continue
        d = _as_date(val)
        if d:
            date_cols[idx] = d

    pembangkit = []
    sistem_terakhir = ''
    urutan = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        no = row[COL_NO - 1] if len(row) >= COL_NO else None
        if not isinstance(no, (int, float)):
            continue  # baris sisa / subtotal tanpa NO -> lewati
        nama = row[COL_PEMBANGKIT - 1] if len(row) >= COL_PEMBANGKIT else None
        if not nama or not str(nama).strip():
            continue
        nama = str(nama).strip()

        sistem = row[COL_SISTEM - 1] if len(row) >= COL_SISTEM else None
        if sistem and str(sistem).strip():
            sistem_terakhir = str(sistem).strip()
        sistem = sistem_terakhir

        aset = row[COL_ASET - 1] if len(row) >= COL_ASET else None
        aset = str(aset).strip() if aset else ''
        dmn = _to_float(row[COL_DMN - 1]) if len(row) >= COL_DMN else None

        snaps = {}
        for idx, d in date_cols.items():
            if idx - 1 < len(row):
                hop = _to_float(row[idx - 1])
                if hop is not None:
                    snaps[d] = hop

        urutan += 1
        pembangkit.append({
            'nama': nama, 'kategori': kategori, 'sistem': sistem,
            'aset': aset, 'dmn_mw': dmn, 'urutan': urutan, 'snaps': snaps,
        })
    return pembangkit


def parse_workbook(path_or_file):
    """
    Baca workbook, kembalikan list dict pembangkit gabungan (batu bara + BBM).
    Sheet yang tidak ada dilewati (mis. file hanya berisi salah satu kategori).
    """
    import openpyxl
    wb = openpyxl.load_workbook(path_or_file, data_only=True, read_only=True)
    hasil = []
    for sheet_name, kategori in SHEETS:
        if sheet_name in wb.sheetnames:
            hasil.extend(parse_sheet(wb[sheet_name], kategori))
    wb.close()
    return hasil


def simpan(pembangkit_list):
    """
    Upsert hasil parse ke database. Kembalikan ringkasan dict.
    Impor bersifat aditif untuk snapshot (tanggal lama tidak dihapus), dan
    memperbarui metadata pembangkit (sistem/aset/dmn/urutan).
    """
    from .models import HopPembangkit, HopSnapshot

    n_pb, n_snap, tanggal_max = 0, 0, None
    for d in pembangkit_list:
        pb, _ = HopPembangkit.objects.update_or_create(
            nama=d['nama'], kategori=d['kategori'],
            defaults={
                'sistem': d['sistem'], 'aset': d['aset'],
                'dmn_mw': d['dmn_mw'], 'urutan': d['urutan'], 'aktif': True,
            },
        )
        n_pb += 1
        for tanggal, hop in d['snaps'].items():
            HopSnapshot.objects.update_or_create(
                pembangkit=pb, tanggal=tanggal, defaults={'hop': hop},
            )
            n_snap += 1
            if tanggal_max is None or tanggal > tanggal_max:
                tanggal_max = tanggal
    return {'pembangkit': n_pb, 'snapshot': n_snap, 'tanggal_terakhir': tanggal_max}


def impor_dari_file(path_or_file):
    """Parse + simpan sekaligus."""
    return simpan(parse_workbook(path_or_file))
