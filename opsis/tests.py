"""
Tes prediksi beban berbasis spreadsheet (opsis/prakiraan.py) dan endpoint
penerima kurvanya (api/views.py::prakiraan_beban_endpoint).

Jalankan: python manage.py test opsis
"""
import datetime
import json

from django.contrib.auth.models import User
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone

from .models import (KelompokPeta, KolomEWS, ModePemeliharaan, PantauanKit,
                     Pembangkit, PrakiraanBeban, SnapFreq, SnapFreqRT, SnapLive,
                     TitikEWS)
from . import freq_history, hop_map, mssql, prakiraan, prediksi, sumber_data, views
from auditlog.models import AuditLog

API_KEY = 'kunci-tes-prakiraan'


def _waktu_lokal(tanggal, menit):
    """datetime tz-aware pada `menit` sejak 00:00 waktu lokal."""
    tz = timezone.get_current_timezone()
    return timezone.make_aware(
        datetime.datetime.combine(tanggal, datetime.time.min), tz
    ) + datetime.timedelta(minutes=menit)


class PrakiraanBebanTest(TestCase):
    """Kurva spreadsheet dibaca benar dan dibandingkan ke realisasi SnapLive."""

    @classmethod
    def setUpTestData(cls):
        cls.hari_ini = timezone.localdate()
        cls.besok = cls.hari_ini + datetime.timedelta(days=1)
        cls.kit_a = Pembangkit.objects.create(kode='KITA', nama='PLTU A', kode_kit='KITA')
        cls.kit_b = Pembangkit.objects.create(kode='KITB', nama='PLTU B', kode_kit='KITB')

    def _isi_kurva(self, tanggal, nilai_per_menit):
        for menit, mw in nilai_per_menit.items():
            PrakiraanBeban.objects.create(tanggal=tanggal, menit=menit, mw=mw)

    def _isi_realisasi(self, tanggal, nilai_per_menit):
        """Realisasi dibagi dua pembangkit — total sistem = jumlah keduanya."""
        for menit, mw in nilai_per_menit.items():
            waktu = _waktu_lokal(tanggal, menit)
            SnapLive.objects.create(pembangkit=self.kit_a, waktu=waktu, mw=mw * 0.6)
            SnapLive.objects.create(pembangkit=self.kit_b, waktu=waktu, mw=mw * 0.4)

    def test_kurva_hari_ini_dan_puncak(self):
        self._isi_kurva(self.hari_ini, {0: 800.0, 720: 900.0, 1110: 1000.0})
        self._isi_realisasi(self.hari_ini, {720: 890.0, 1110: 1010.0})

        hasil = prakiraan.predict_beban_hari_ini()

        self.assertEqual(hasil['source'], 'sheet')
        self.assertEqual([p['minute'] for p in hasil['forecast']], [0, 720, 1110])
        self.assertEqual(hasil['prediksi_puncak_siang'], 900.0)
        self.assertEqual(hasil['prediksi_puncak_malam'], 1000.0)
        self.assertEqual(hasil['realisasi_puncak_siang'], 890.0)
        self.assertEqual(hasil['realisasi_puncak_malam'], 1010.0)
        # 'actual' menjumlahkan semua pembangkit per menit, bukan satu baris per KIT
        self.assertEqual(len(hasil['actual']), 2)

    def test_realisasi_none_bila_titik_belum_terlewati(self):
        self._isi_kurva(self.hari_ini, {720: 900.0})
        hasil = prakiraan.predict_beban_hari_ini()
        self.assertEqual(hasil['prediksi_puncak_siang'], 900.0)
        self.assertIsNone(hasil['realisasi_puncak_siang'])

    def test_realisasi_toleransi_2_menit_hanya_ke_belakang(self):
        self._isi_kurva(self.hari_ini, {720: 900.0})
        # SnapLive telat 2 menit (12:02) tidak boleh dipakai untuk titik 12:00;
        # yang 11:58 boleh (toleransi ke belakang).
        self._isi_realisasi(self.hari_ini, {718: 880.0, 722: 999.0})
        hasil = prakiraan.predict_beban_hari_ini()
        self.assertEqual(hasil['realisasi_puncak_siang'], 880.0)

    def test_hari_lain_tidak_ikut_terbawa(self):
        self._isi_kurva(self.hari_ini, {720: 900.0})
        self._isi_kurva(self.besok, {720: 950.0, 1110: 1100.0})

        hasil = prakiraan.predict_beban_hari_ini()
        self.assertEqual(len(hasil['forecast']), 1)

        besok = prakiraan.predict_besok_puncak()
        self.assertEqual(besok['source'], 'sheet')
        self.assertEqual(besok['prediksi_puncak_siang_besok'], 950.0)
        self.assertEqual(besok['prediksi_puncak_malam_besok'], 1100.0)

    def test_besok_belum_diunggah(self):
        besok = prakiraan.predict_besok_puncak()
        self.assertEqual(besok['source'], 'no_sheet')
        self.assertIsNone(besok['prediksi_puncak_siang_besok'])

    def test_akurasi_dihitung_dari_pasangan_yang_punya_realisasi(self):
        # Dua titik punya realisasi (selisih 10 MW dari 1000 -> MAPE 1%),
        # satu titik belum ada realisasinya -> harus dilewati, bukan dianggap 0.
        self._isi_kurva(self.hari_ini, {600: 1010.0, 660: 990.0, 1110: 1200.0})
        self._isi_realisasi(self.hari_ini, {600: 1000.0, 660: 1000.0})

        a = prakiraan.evaluate_accuracy(days=7)
        self.assertEqual(a['n'], 2)
        self.assertAlmostEqual(a['mae'], 10.0, places=2)
        self.assertAlmostEqual(a['rmse'], 10.0, places=2)
        self.assertAlmostEqual(a['mape_percent'], 1.0, places=2)
        self.assertAlmostEqual(a['akurasi_percent'], 99.0, places=2)

    def test_akurasi_kosong_bila_belum_ada_prakiraan(self):
        a = prakiraan.evaluate_accuracy(days=7)
        self.assertEqual(a['n'], 0)
        self.assertIsNone(a['akurasi_percent'])
        self.assertEqual(a['period_days'], 7)

    def test_akurasi_abaikan_hari_di_luar_rentang(self):
        lama = self.hari_ini - datetime.timedelta(days=30)
        self._isi_kurva(lama, {600: 1010.0})
        self._isi_realisasi(lama, {600: 1000.0})
        self.assertEqual(prakiraan.evaluate_accuracy(days=7)['n'], 0)


class SumberPrediksiTest(TestCase):
    """Switch OPSIS_FORECAST_SOURCE memilih implementasi yang benar."""

    @override_settings(OPSIS_FORECAST_SOURCE='sheet')
    def test_default_sheet(self):
        self.assertEqual(prediksi.sumber_aktif(), 'sheet')
        self.assertEqual(prediksi.predict_besok_puncak()['source'], 'no_sheet')

    @override_settings(OPSIS_FORECAST_SOURCE='ML')
    def test_ml_case_insensitive(self):
        self.assertEqual(prediksi.sumber_aktif(), 'ml')

    @override_settings(OPSIS_FORECAST_SOURCE='entah-apa')
    def test_nilai_tak_dikenal_jatuh_ke_sheet(self):
        self.assertEqual(prediksi.sumber_aktif(), 'sheet')


@override_settings(API_KEY=API_KEY)
class PrakiraanBebanEndpointTest(TestCase):
    """POST /api/v1/prakiraan-beban/ — jalur n8n dari Google Sheets."""

    def setUp(self):
        self.url = reverse('api:prakiraan_beban')
        self.hari_ini = timezone.localdate()

    def _post(self, payload, key=API_KEY):
        headers = {'X-API-Key': key} if key else {}
        return self.client.post(self.url, data=json.dumps(payload),
                                content_type='application/json', headers=headers)

    def test_tolak_tanpa_api_key(self):
        self.assertEqual(self._post({'data': []}, key=None).status_code, 401)

    def test_tolak_api_key_salah(self):
        self.assertEqual(self._post({'data': []}, key='salah').status_code, 403)

    def test_terima_jam_hh_mm_dan_menit(self):
        r = self._post({
            'tanggal': self.hari_ini.isoformat(),
            'data': [
                {'jam': '00:00', 'mw': 800},
                {'jam': '18:30:00', 'mw': '1.010,5'.replace('.', '')},  # koma desimal
                {'menit': 720, 'mw': 900.25},
            ],
        })
        self.assertEqual(r.status_code, 200)
        body = r.json()
        self.assertEqual(body['titik_ditulis'], 3)
        self.assertEqual(body['dilewati'], 0)
        self.assertEqual(
            list(PrakiraanBeban.objects.filter(tanggal=self.hari_ini)
                 .order_by('menit').values_list('menit', flat=True)),
            [0, 720, 1110])
        self.assertEqual(
            PrakiraanBeban.objects.get(tanggal=self.hari_ini, menit=1110).mw, 1010.5)

    def test_idempoten_dan_memperbarui_nilai(self):
        payload = {'tanggal': self.hari_ini.isoformat(),
                   'data': [{'menit': 720, 'mw': 900}]}
        self._post(payload)
        payload['data'][0]['mw'] = 950
        self._post(payload)
        self.assertEqual(PrakiraanBeban.objects.count(), 1)
        self.assertEqual(PrakiraanBeban.objects.get().mw, 950.0)

    def test_baris_boleh_bawa_tanggal_sendiri(self):
        besok = self.hari_ini + datetime.timedelta(days=1)
        r = self._post({
            'tanggal': self.hari_ini.isoformat(),
            'data': [
                {'menit': 720, 'mw': 900},
                {'tanggal': besok.isoformat(), 'menit': 720, 'mw': 950},
            ],
        })
        self.assertEqual(r.json()['titik_ditulis'], 2)
        self.assertEqual(sorted(r.json()['tanggal_tertulis']),
                         sorted([self.hari_ini.isoformat(), besok.isoformat()]))

    def test_sel_kosong_dilewati_tanpa_error(self):
        r = self._post({'data': [{'menit': 0, 'mw': ''}, {'menit': 30, 'mw': None},
                                 {'menit': 60, 'mw': 810}]})
        body = r.json()
        self.assertEqual(body['titik_ditulis'], 1)
        self.assertEqual(body['dilewati'], 2)
        self.assertEqual(body['errors'], [])

    def test_nilai_tidak_valid_dilaporkan(self):
        r = self._post({'data': [{'menit': 1500, 'mw': 800},
                                 {'jam': 'pagi', 'mw': 800},
                                 {'menit': 0, 'mw': 'abc'}]})
        body = r.json()
        self.assertEqual(body['titik_ditulis'], 0)
        self.assertEqual(len(body['errors']), 3)
        self.assertEqual(PrakiraanBeban.objects.count(), 0)

    def test_replace_menghapus_titik_yang_tidak_dikirim(self):
        self._post({'tanggal': self.hari_ini.isoformat(),
                    'data': [{'menit': 0, 'mw': 800}, {'menit': 30, 'mw': 810}]})
        # Tanpa replace: titik lama tetap ada.
        self._post({'tanggal': self.hari_ini.isoformat(),
                    'data': [{'menit': 0, 'mw': 805}]})
        self.assertEqual(PrakiraanBeban.objects.count(), 2)
        # Dengan replace: titik 30 ikut terhapus.
        self._post({'tanggal': self.hari_ini.isoformat(), 'replace': True,
                    'data': [{'menit': 0, 'mw': 806}]})
        self.assertEqual(
            list(PrakiraanBeban.objects.values_list('menit', flat=True)), [0])

    def test_replace_tidak_menyentuh_tanggal_lain(self):
        besok = self.hari_ini + datetime.timedelta(days=1)
        PrakiraanBeban.objects.create(tanggal=besok, menit=720, mw=950)
        self._post({'tanggal': self.hari_ini.isoformat(), 'replace': True,
                    'data': [{'menit': 0, 'mw': 800}]})
        self.assertTrue(PrakiraanBeban.objects.filter(tanggal=besok, menit=720).exists())

    def test_tolak_data_bukan_array(self):
        self.assertEqual(self._post({'data': {'menit': 0}}).status_code, 400)

    def test_tolak_payload_kelewat_besar(self):
        r = self._post({'data': [{'menit': 0, 'mw': 1}] * 2001})
        self.assertEqual(r.status_code, 400)

    def test_payload_apa_adanya_dari_node_code_n8n(self):
        """
        Bentuk payload persis seperti yang dihasilkan node Code di
        docs/n8n_prakiraan_beban*.workflow.json: `menit` dan `mw` sudah
        dinormalkan di sisi n8n (sel Jam/MW di .xlsx bisa berupa angka, Date,
        atau teks dengan pemisah ribuan), jadi yang masuk ke sini angka murni.
        """
        besok = self.hari_ini + datetime.timedelta(days=1)
        r = self._post({
            'sumber': 'roh-sulbagsel',
            'data': [
                {'menit': 0, 'mw': 812.5, 'tanggal': self.hari_ini.isoformat()},
                {'menit': 720, 'mw': 1024, 'tanggal': self.hari_ini.isoformat()},
                {'menit': 1110, 'mw': 1187.3, 'tanggal': self.hari_ini.isoformat()},
                {'menit': 720, 'mw': 950, 'tanggal': besok.isoformat()},
                {'menit': 1410, 'mw': 890, 'tanggal': self.hari_ini.isoformat()},
            ],
        })
        body = r.json()
        self.assertEqual(body['titik_ditulis'], 5)
        self.assertEqual(body['errors'], [])
        self.assertEqual(
            PrakiraanBeban.objects.get(tanggal=self.hari_ini, menit=0).mw, 812.5)
        self.assertEqual(
            PrakiraanBeban.objects.get(tanggal=self.hari_ini, menit=1110).mw, 1187.3)
        self.assertEqual(
            PrakiraanBeban.objects.get(tanggal=besok, menit=720).mw, 950.0)
        self.assertEqual(
            PrakiraanBeban.objects.get(tanggal=self.hari_ini, menit=0).sumber,
            'roh-sulbagsel')

    def test_get_baca_balik_kurva(self):
        PrakiraanBeban.objects.create(tanggal=self.hari_ini, menit=1110, mw=1000)
        r = self.client.get(self.url, {'tanggal': self.hari_ini.isoformat()},
                            headers={'X-API-Key': API_KEY})
        body = r.json()
        self.assertEqual(body['jumlah'], 1)
        self.assertEqual(body['data'][0]['jam'], '18:30')


class KeteranganSumberDashboardTest(TestCase):
    """Dashboard OPSIS menyebutkan asal angka prediksi, bukan cuma "Prediksi"."""

    def setUp(self):
        # force_login, bukan login() — AxesBackend menolak authenticate()
        # tanpa objek request.
        user = User.objects.create_superuser('adm-tes', 'adm@contoh.id', 'rahasia-tes-123')
        # ForcePasswordChangeMiddleware mengalihkan user baru ke /ganti-password/
        # sebelum sempat sampai ke dashboard.
        profile = getattr(user, 'profile', None)
        if profile is not None:
            profile.force_password_change = False
            profile.save(update_fields=['force_password_change'])
        self.client.force_login(user)

    def test_dashboard_menyebut_dashboard_roh_sulbagsel(self):
        r = self.client.get('/opsis/')
        self.assertEqual(r.status_code, 200)
        isi = r.content.decode()
        self.assertIn('Dashboard ROH Sulbagsel', isi)
        self.assertIn('beban-sumber-prediksi', isi)

    def test_halaman_analitik_menyebut_sumbernya(self):
        r = self.client.get('/opsis/prediksi-beban/')
        self.assertEqual(r.status_code, 200)
        self.assertIn('Dashboard ROH Sulbagsel', r.content.decode())


class PetaPembangkitTest(TestCase):
    """Peta Pembangkit: pin terpasang, tidak bertumpuk, dan tabel memuat semua unit."""

    @classmethod
    def setUpTestData(cls):
        # Tiga pembangkit di lokasi yang sama persis (rumpun Tello) untuk menguji
        # penyebaran pin, satu PLTA yang posisinya hanya ada di PEMBANGKIT_MAP_POS,
        # satu pembangkit dengan koordinat manual, dan satu yang tak dikenal peta.
        data = [
            ('PLTD GE Tello', 'TELLO5', 'PLTD'),
            ('PLTD Tello Biosolar', 'TELLOB', 'PLTD'),
            ('MPP Tello', 'TELLO_SW', 'PLTD'),
            ('PLTA Bakaru', 'BKARU5', 'PLTA'),
            ('Pembangkit Tak Dikenal', 'XYZ', 'LAIN'),
        ]
        for urutan, (nama, kode, jenis) in enumerate(data):
            Pembangkit.objects.create(nama=nama, kode=kode, jenis=jenis, urutan=urutan)
        cls.manual = Pembangkit.objects.create(nama='PLTS Uji Manual', kode='MANUAL',
                                               jenis='PLTS', urutan=9, peta_x=40, peta_y=60)

    def setUp(self):
        user = User.objects.create_superuser('adm-peta', 'peta@contoh.id', 'rahasia-tes-123')
        profile = getattr(user, 'profile', None)
        if profile is not None:
            profile.force_password_change = False
            profile.save(update_fields=['force_password_change'])
        self.client.force_login(user)

    def test_posisi_peta_manual_menang_atas_tabel_bawaan(self):
        self.assertEqual(self.manual.posisi_peta(), (40, 60))
        self.assertEqual(Pembangkit.objects.get(kode='BKARU5').posisi_peta(),
                         hop_map.posisi_pembangkit('PLTA Bakaru'))
        self.assertIsNone(Pembangkit.objects.get(kode='XYZ').posisi_peta())

    def test_halaman_memuat_semua_pembangkit(self):
        r = self.client.get('/opsis/peta/')
        self.assertEqual(r.status_code, 200)
        isi = r.content.decode()
        for p in Pembangkit.objects.all():
            self.assertIn(p.nama, isi)        # semua unit masuk tabel
        # Yang punya koordinat dapat pin; yang tidak, hanya masuk tabel.
        self.assertEqual(len(r.context['pins']), 5)
        self.assertEqual([t['kode'] for t in r.context['tak_tampil']], ['XYZ'])

    def test_pin_serumpun_tidak_saling_menutup(self):
        r = self.client.get('/opsis/peta/')
        pins = r.context['pins']
        for i, a in enumerate(pins):
            for b in pins[i + 1:]:
                self.assertFalse(
                    abs(a['x'] - b['x']) < views.PIN_MIN_DX and abs(a['y'] - b['y']) < views.PIN_MIN_DY,
                    f"pin {a['kode']} dan {b['kode']} masih bertumpuk",
                )

    def test_legenda_hanya_jenis_yang_ada(self):
        r = self.client.get('/opsis/peta/')
        legenda = {l['kode']: l['jumlah'] for l in r.context['legenda']}
        self.assertEqual(legenda, {'PLTD': 3, 'PLTA': 1, 'PLTS': 1, 'LAIN': 1})


class PetaSimpanPosisiTest(TestCase):
    """Endpoint seret-lepas: menyimpan, mengembalikan, dan menolak yang tidak berhak."""

    @classmethod
    def setUpTestData(cls):
        cls.pb = Pembangkit.objects.create(nama='PLTA Bakaru', kode='BKARU5', jenis='PLTA')
        cls.lain = Pembangkit.objects.create(nama='PLTU Barru', kode='BLUSU5', jenis='PLTU')

    def _login(self, role='opsis', superuser=False):
        buat = User.objects.create_superuser if superuser else User.objects.create_user
        user = buat(f'u-{role}-{superuser}', f'{role}@contoh.id', 'rahasia-tes-123')
        profile = getattr(user, 'profile', None)
        if profile is not None:
            profile.role = role
            profile.force_password_change = False
            profile.save(update_fields=['role', 'force_password_change'])
        self.client.force_login(user)
        return user

    def _kirim(self, data):
        return self.client.post('/opsis/peta/simpan/', data=json.dumps(data),
                                content_type='application/json')

    def test_role_opsis_boleh_menyimpan_posisi(self):
        self._login('opsis')
        r = self._kirim({'posisi': [{'pk': self.pb.pk, 'x': 33.333, 'y': 44.444}]})
        self.assertEqual(r.status_code, 200)
        self.assertTrue(r.json()['ok'])
        self.pb.refresh_from_db()
        self.assertEqual((self.pb.peta_x, self.pb.peta_y), (33.33, 44.44))
        self.assertEqual(self.pb.posisi_peta(), (33.33, 44.44))

    def test_hapus_mengembalikan_ke_posisi_bawaan(self):
        Pembangkit.objects.filter(pk=self.pb.pk).update(peta_x=10, peta_y=10)
        self._login('opsis')
        r = self._kirim({'hapus': [self.pb.pk]})
        self.assertEqual(r.status_code, 200)
        self.pb.refresh_from_db()
        self.assertIsNone(self.pb.peta_x)
        self.assertEqual(self.pb.posisi_peta(), hop_map.posisi_pembangkit('PLTA Bakaru'))

    def test_viewer_ditolak(self):
        self._login('viewer')
        r = self._kirim({'posisi': [{'pk': self.pb.pk, 'x': 10, 'y': 10}]})
        self.assertEqual(r.status_code, 403)
        self.pb.refresh_from_db()
        self.assertIsNone(self.pb.peta_x)

    def test_posisi_di_luar_peta_ditolak_tanpa_menyimpan_sebagian(self):
        self._login('opsis')
        r = self._kirim({'posisi': [{'pk': self.pb.pk, 'x': 20, 'y': 20},
                                    {'pk': self.lain.pk, 'x': 120, 'y': 20}]})
        self.assertEqual(r.status_code, 400)
        self.pb.refresh_from_db()
        self.lain.refresh_from_db()
        self.assertIsNone(self.pb.peta_x)      # tidak ada yang tersimpan setengah
        self.assertIsNone(self.lain.peta_x)

    def test_pk_tidak_dikenal_ditolak(self):
        self._login('opsis')
        r = self._kirim({'posisi': [{'pk': 999999, 'x': 20, 'y': 20}]})
        self.assertEqual(r.status_code, 400)

    def test_pin_manual_tidak_digeser_penyebaran_otomatis(self):
        # Ditaruh persis di atas pin otomatis PLTU Barru → yang otomatis yang mengalah.
        bawaan = hop_map.posisi_pembangkit('PLTU Barru')
        Pembangkit.objects.filter(pk=self.pb.pk).update(peta_x=bawaan[0], peta_y=bawaan[1])
        self._login('opsis', superuser=True)
        r = self.client.get('/opsis/peta/')
        pins = {p['kode']: p for p in r.context['pins']}
        self.assertEqual((pins['BKARU5']['x'], pins['BKARU5']['y']), bawaan)
        self.assertNotEqual((pins['BLUSU5']['x'], pins['BLUSU5']['y']), bawaan)


class ModePemeliharaanTest(TestCase):
    """Sakelar pemeliharaan menutup /opsis/* dan membukanya lagi tanpa deploy."""

    def setUp(self):
        ModePemeliharaan._cache = {'obj': None, 'ts': 0.0}   # cache antar-tes tidak boleh bocor
        self.user = self._buat('staf-opsis', role='opsis')
        self.client.force_login(self.user)

    def tearDown(self):
        ModePemeliharaan._cache = {'obj': None, 'ts': 0.0}

    def _buat(self, nama, role='opsis', superuser=False):
        buat = User.objects.create_superuser if superuser else User.objects.create_user
        user = buat(nama, f'{nama}@contoh.id', 'rahasia-tes-123')
        profile = getattr(user, 'profile', None)
        if profile is not None:
            profile.role = role
            profile.force_password_change = False
            profile.save(update_fields=['role', 'force_password_change'])
        return user

    def _nyalakan(self, **kwargs):
        mode = ModePemeliharaan.ambil()
        mode.aktif = True
        for k, v in kwargs.items():
            setattr(mode, k, v)
        mode.save()
        return mode

    def test_nonaktif_opsis_tetap_terbuka(self):
        self.assertFalse(ModePemeliharaan.ambil().aktif)     # bawaan: tidak memelihara
        self.assertEqual(self.client.get('/opsis/').status_code, 200)

    def test_aktif_semua_rute_opsis_jadi_halaman_pemeliharaan(self):
        self._nyalakan(judul='OPSIS Dipelihara')
        for url in ('/opsis/', '/opsis/peta/', '/opsis/up2d/', '/opsis/hop/dashboard/'):
            r = self.client.get(url)
            self.assertEqual(r.status_code, 503, url)
            self.assertTemplateUsed(r, 'opsis/pemeliharaan.html')
            self.assertContains(r, 'OPSIS Dipelihara', status_code=503)
            self.assertEqual(r['Retry-After'], '1800')

    def test_endpoint_api_dijawab_json_bukan_html(self):
        self._nyalakan()
        r = self.client.get('/opsis/api/live/')
        self.assertEqual(r.status_code, 503)
        self.assertEqual(r['Content-Type'], 'application/json')
        self.assertTrue(r.json()['pemeliharaan'])

    def test_rute_di_luar_opsis_tidak_terpengaruh(self):
        self._nyalakan()
        # Role opsis dibatasi ke /opsis/ oleh OpsisAccessMiddleware, jadi diuji
        # dengan superuser: yang penting rute non-OPSIS tidak ikut 503.
        self.client.force_login(self._buat('adm-lain', superuser=True))
        r = self.client.get('/device-mon/')
        self.assertNotEqual(r.status_code, 503)

    def test_superuser_menembus_bila_diizinkan(self):
        self._nyalakan(boleh_superuser=True)
        self.client.force_login(self._buat('adm-tembus', superuser=True))
        r = self.client.get('/opsis/')
        self.assertEqual(r.status_code, 200)
        self.assertTrue(r.context['request'].opsis_pemeliharaan)
        self.assertContains(r, 'Mode pemeliharaan aktif')      # pita penanda

    def test_superuser_ikut_ditutup_bila_tidak_diizinkan(self):
        self._nyalakan(boleh_superuser=False)
        self.client.force_login(self._buat('adm-tutup', superuser=True))
        self.assertEqual(self.client.get('/opsis/').status_code, 503)

    def test_dimatikan_lagi_opsis_langsung_terbuka(self):
        mode = self._nyalakan()
        self.assertEqual(self.client.get('/opsis/').status_code, 503)
        mode.aktif = False
        mode.save()
        self.assertEqual(self.client.get('/opsis/').status_code, 200)

    def test_tombol_keluar_memakai_post_bukan_tautan_get(self):
        # LogoutView Django hanya menerima POST: tautan GET ke /logout/ menghasilkan
        # 405, jadi halaman pemeliharaan harus memakai form seperti template lain.
        self._nyalakan()
        isi = self.client.get('/opsis/').content.decode()
        self.assertIn('<form method="post" action="/logout/"', isi)
        self.assertNotIn('href="/logout/"', isi)
        self.assertEqual(self.client.get('/logout/').status_code, 405)

        keluar = self.client.post('/logout/')
        self.assertEqual(keluar.status_code, 302)
        self.assertIsNone(self.client.session.get('_auth_user_id'))

    def test_baris_pengaturan_selalu_tunggal(self):
        ModePemeliharaan.ambil()
        ModePemeliharaan(aktif=True, judul='Baris kedua').save()
        self.assertEqual(ModePemeliharaan.objects.count(), 1)
        self.assertEqual(ModePemeliharaan.objects.get().judul, 'Baris kedua')


class PetaSembunyikanIkonTest(TestCase):
    """Ikon bisa dihilangkan dari peta tanpa menghilangkan pembangkitnya dari tabel."""

    @classmethod
    def setUpTestData(cls):
        # PLTU Barru punya posisi bawaan di hop_map, jadi mengosongkan koordinat
        # saja tidak cukup untuk menghilangkan ikonnya — itulah yang diuji di sini.
        cls.pb = Pembangkit.objects.create(nama='PLTU Barru', kode='BLUSU5', jenis='PLTU')

    def setUp(self):
        user = User.objects.create_superuser('adm-sembunyi', 'sb@contoh.id', 'rahasia-tes-123')
        profile = getattr(user, 'profile', None)
        if profile is not None:
            profile.force_password_change = False
            profile.save(update_fields=['force_password_change'])
        self.client.force_login(user)

    def _kirim(self, data):
        return self.client.post('/opsis/peta/simpan/', data=json.dumps(data),
                                content_type='application/json')

    def test_bawaan_semua_pembangkit_tampil(self):
        self.assertTrue(self.pb.tampil_di_peta)
        r = self.client.get('/opsis/peta/')
        self.assertEqual([p['kode'] for p in r.context['pins']], ['BLUSU5'])

    def test_sembunyikan_menghilangkan_ikon_tapi_tetap_di_tabel(self):
        r = self._kirim({'sembunyi': [self.pb.pk]})
        self.assertEqual(r.status_code, 200)
        self.pb.refresh_from_db()
        self.assertFalse(self.pb.tampil_di_peta)

        r = self.client.get('/opsis/peta/')
        self.assertEqual(r.context['pins'], [])                       # ikon hilang dari peta
        self.assertEqual([t['kode'] for t in r.context['tak_tampil']], ['BLUSU5'])
        self.assertContains(r, 'PLTU Barru')                          # tetap ada di tabel daya

    def test_mengosongkan_koordinat_tidak_menyembunyikan_ikon(self):
        Pembangkit.objects.filter(pk=self.pb.pk).update(peta_x=50, peta_y=50)
        self._kirim({'hapus': [self.pb.pk]})
        self.pb.refresh_from_db()
        self.assertIsNone(self.pb.peta_x)
        self.assertTrue(self.pb.tampil_di_peta)
        self.assertEqual(len(self.client.get('/opsis/peta/').context['pins']), 1)

    def test_menaruh_kembali_di_peta_menampilkan_lagi(self):
        self._kirim({'sembunyi': [self.pb.pk]})
        self._kirim({'posisi': [{'pk': self.pb.pk, 'x': 30, 'y': 40}]})
        self.pb.refresh_from_db()
        self.assertTrue(self.pb.tampil_di_peta)
        self.assertEqual((self.pb.peta_x, self.pb.peta_y), (30.0, 40.0))
        self.assertEqual(len(self.client.get('/opsis/peta/').context['pins']), 1)

    def test_sembunyi_tidak_menghapus_koordinat_lama(self):
        Pembangkit.objects.filter(pk=self.pb.pk).update(peta_x=12.5, peta_y=67.5)
        self._kirim({'sembunyi': [self.pb.pk]})
        self.pb.refresh_from_db()
        self.assertEqual((self.pb.peta_x, self.pb.peta_y), (12.5, 67.5))


class KelompokPetaTest(TestCase):
    """Ikon kelompok: satu ikon mewakili beberapa pembangkit sekaligus."""

    @classmethod
    def setUpTestData(cls):
        cls.a = Pembangkit.objects.create(nama='PLTD GE Tello', kode='TELLO5', jenis='PLTD', urutan=1)
        cls.b = Pembangkit.objects.create(nama='MPP Tello', kode='TELLO_SW', jenis='PLTD', urutan=2)
        cls.luar = Pembangkit.objects.create(nama='PLTA Bakaru', kode='BKARU5', jenis='PLTA', urutan=3)

    def setUp(self):
        user = User.objects.create_superuser('adm-grup', 'grup@contoh.id', 'rahasia-tes-123')
        profile = getattr(user, 'profile', None)
        if profile is not None:
            profile.force_password_change = False
            profile.save(update_fields=['force_password_change'])
        self.client.force_login(user)

    def _kirim(self, data):
        return self.client.post('/opsis/peta/simpan/', data=json.dumps(data),
                                content_type='application/json')

    def _buat(self, **ubah):
        isi = {'id': None, 'nama': 'Kompleks Tello', 'keterangan': 'Kawasan Tello',
               'jenis': 'PLTD', 'x': 13.3, 'y': 88.5, 'anggota': [self.a.pk, self.b.pk]}
        isi.update(ubah)
        return self._kirim({'kelompok': [isi]})

    def test_membuat_kelompok_menyerap_ikon_anggotanya(self):
        r = self._buat()
        self.assertEqual(r.status_code, 200, r.content)
        k = KelompokPeta.objects.get()
        self.assertEqual(k.nama, 'Kompleks Tello')
        self.assertEqual({a.pk for a in k.anggota.all()}, {self.a.pk, self.b.pk})

        ctx = self.client.get('/opsis/peta/').context
        # Anggota tidak lagi digambar sendiri; dayanya sudah terhitung di kelompok.
        self.assertEqual([p['kode'] for p in ctx['pins']], ['BKARU5'])
        self.assertEqual({t['kode'] for t in ctx['dalam_kelompok']}, {'TELLO5', 'TELLO_SW'})
        self.assertEqual(len(ctx['kelompok']), 1)
        self.assertEqual([a['nama'] for a in ctx['kelompok'][0]['anggota']],
                         ['PLTD GE Tello', 'MPP Tello'])

    def test_anggota_tidak_masuk_daftar_yang_bisa_diseret(self):
        # Kalau ikut masuk tak_tampil, anggota bisa diseret jadi ikon kedua dan
        # dayanya terhitung dua kali di peta.
        self._buat()
        ctx = self.client.get('/opsis/peta/').context
        self.assertNotIn('TELLO5', [t['kode'] for t in ctx['tak_tampil']])

    def test_semua_pembangkit_tetap_ada_di_tabel(self):
        self._buat()
        isi = self.client.get('/opsis/peta/').content.decode()
        for nama in ('PLTD GE Tello', 'MPP Tello', 'PLTA Bakaru'):
            self.assertIn(nama, isi)

    def test_memperbarui_kelompok_yang_sudah_ada(self):
        self._buat()
        k = KelompokPeta.objects.get()
        r = self._kirim({'kelompok': [{'id': k.pk, 'nama': 'Tello', 'keterangan': '',
                                       'jenis': 'PLTG', 'x': 20, 'y': 30,
                                       'anggota': [self.a.pk]}]})
        self.assertEqual(r.status_code, 200)
        k.refresh_from_db()
        self.assertEqual((k.nama, k.jenis, k.peta_x, k.peta_y), ('Tello', 'PLTG', 20.0, 30.0))
        self.assertEqual([a.pk for a in k.anggota.all()], [self.a.pk])
        # Yang dikeluarkan dari kelompok kembali punya ikon sendiri
        ctx = self.client.get('/opsis/peta/').context
        self.assertIn('TELLO_SW', [p['kode'] for p in ctx['pins']])

    def test_menghapus_kelompok_mengembalikan_ikon_anggotanya(self):
        self._buat()
        k = KelompokPeta.objects.get()
        r = self._kirim({'kelompok_hapus': [k.pk]})
        self.assertEqual(r.status_code, 200)
        self.assertEqual(KelompokPeta.objects.count(), 0)
        ctx = self.client.get('/opsis/peta/').context
        self.assertEqual({p['kode'] for p in ctx['pins']}, {'TELLO5', 'TELLO_SW', 'BKARU5'})

    def test_kelompok_disembunyikan_anggotanya_kembali_sendiri(self):
        self._buat()
        KelompokPeta.objects.update(tampil_di_peta=False)
        ctx = self.client.get('/opsis/peta/').context
        self.assertEqual(ctx['kelompok'], [])
        self.assertEqual({p['kode'] for p in ctx['pins']}, {'TELLO5', 'TELLO_SW', 'BKARU5'})

    def test_nama_kosong_ditolak(self):
        r = self._buat(nama='   ')
        self.assertEqual(r.status_code, 400)
        self.assertEqual(KelompokPeta.objects.count(), 0)

    def test_jenis_ikon_asing_ditolak(self):
        self.assertEqual(self._buat(jenis='PLTX').status_code, 400)
        self.assertEqual(KelompokPeta.objects.count(), 0)

    def test_posisi_di_luar_peta_ditolak(self):
        self.assertEqual(self._buat(x=140).status_code, 400)
        self.assertEqual(KelompokPeta.objects.count(), 0)

    def test_anggota_asing_ditolak(self):
        self.assertEqual(self._buat(anggota=[999999]).status_code, 400)
        self.assertEqual(KelompokPeta.objects.count(), 0)

    def test_id_kelompok_asing_ditolak(self):
        self.assertEqual(self._buat(id=999999).status_code, 400)

    def test_kelompok_tidak_bergeser_penyebaran_otomatis(self):
        # Kelompok selalu 'manual': posisinya harus tetap persis seperti disimpan.
        self._buat(x=15.11, y=80.6)          # persis di posisi bawaan PLTU Barru
        Pembangkit.objects.create(nama='PLTU Barru', kode='BLUSU5', jenis='PLTU', urutan=4)
        ctx = self.client.get('/opsis/peta/').context
        k = ctx['kelompok'][0]
        self.assertEqual((k['x'], k['y']), (15.11, 80.6))
        barru = next(p for p in ctx['pins'] if p['kode'] == 'BLUSU5')
        self.assertNotEqual((barru['x'], barru['y']), (15.11, 80.6))


class TitikEWSAmbangTest(TestCase):
    """Margin dan status titik EWS dihitung di model, bukan di JavaScript."""

    @classmethod
    def setUpTestData(cls):
        cls.kolom = KolomEWS.objects.create(nama='Tegangan')

    def _titik(self, **kwargs):
        data = dict(kolom=self.kolom, nama='GI Uji', skema='UVLS', besaran='v',
                    nominal=150, setting=135, arah='bawah',
                    sumber_tabel='dbo.UJI', sumber_kolom_nilai='VALUE',
                    sumber_kolom_kunci='ANALOG', sumber_nilai_kunci='UJI')
        data.update(kwargs)
        return TitikEWS.objects.create(**data)

    def test_margin_ambang_bawah_relatif_nominal(self):
        t = self._titik()
        self.assertAlmostEqual(t.margin(150), 0.1)
        self.assertAlmostEqual(t.margin(135), 0.0)
        self.assertAlmostEqual(t.margin(130), -1 / 30)

    def test_margin_ambang_atas_berlawanan_arah(self):
        t = self._titik(skema='OVTS', setting=165, arah='atas')
        self.assertAlmostEqual(t.margin(150), 0.1)
        self.assertAlmostEqual(t.margin(170), -1 / 30)

    def test_margin_frekuensi_dalam_hz_absolut(self):
        t = self._titik(besaran='f', nominal=50, setting=49.2, arah='bawah')
        self.assertAlmostEqual(t.margin(50.0), 0.8)
        self.assertAlmostEqual(t.margin(49.0), -0.2)

    def test_margin_arus_relatif_setting(self):
        t = self._titik(besaran='i', nominal=None, setting=200, arah='atas')
        self.assertAlmostEqual(t.margin(100), 0.5)
        self.assertAlmostEqual(t.margin(220), -0.1)

    def test_status_lima_keadaan(self):
        t = self._titik()
        self.assertEqual(t.status(150), 'good')
        self.assertEqual(t.status(139), 'warning')     # margin 2,67% < ambang 3%
        self.assertEqual(t.status(135), 'critical')    # tepat di setting = sudah bekerja
        self.assertEqual(t.status(130), 'critical')
        self.assertEqual(t.status(None), 'plan')       # nilai belum terbaca

    def test_status_rencana_dan_belum_termonitor_selalu_plan(self):
        self.assertEqual(self._titik(status_skema='rencana').status(150), 'plan')
        self.assertEqual(self._titik(sumber_tabel='').status(150), 'plan')

    def test_status_unknown_bila_setting_belum_diisi(self):
        t = self._titik(setting=None)
        self.assertEqual(t.status(150), 'unknown')
        self.assertIsNone(t.margin(150))

    def test_ambang_waspada_bisa_ditimpa_per_titik(self):
        self.assertEqual(self._titik().ambang(), 0.03)
        self.assertEqual(self._titik(ambang_waspada=0.10).ambang(), 0.10)
        # dengan ambang 10%, 139 kV (margin 2,67%) tetap waspada, 150 kV jadi good
        t = self._titik(ambang_waspada=0.10)
        self.assertEqual(t.status(146), 'warning')
        self.assertEqual(t.status(150), 'good')

    def test_satuan_bawaan_mengikuti_besaran(self):
        self.assertEqual(self._titik().satuan_tampil, 'kV')
        self.assertEqual(self._titik(besaran='f').satuan_tampil, 'Hz')
        self.assertEqual(self._titik(besaran='i').satuan_tampil, 'A')
        self.assertEqual(self._titik(satuan='MVAr').satuan_tampil, 'MVAr')

    def test_spesifikasi_sumber_none_bila_belum_diarahkan(self):
        self.assertIsNone(self._titik(sumber_tabel='').spesifikasi_sumber())
        spec = self._titik().spesifikasi_sumber()
        self.assertEqual(spec['tabel'], 'dbo.UJI')
        self.assertEqual(spec['kolom_kunci'], 'ANALOG')
        self.assertEqual(spec['faktor'], 1.0)


class GetNilaiEWSTest(TestCase):
    """
    Pengelompokan query dan penjagaan identifier di mssql.get_nilai_ews().
    Koneksi MSSQL diganti palsu supaya tes tidak butuh historian.
    """

    def setUp(self):
        self.dijalankan = []          # (sql, params) tiap execute()
        self.baris = {}               # nilai balikan per nilai kunci
        uji = self

        class KursorPalsu:
            def __init__(self):
                self._hasil = []

            def execute(self, sql, params=None):
                uji.dijalankan.append((' '.join(sql.split()), list(params or [])))
                if params:
                    self._hasil = [(k, uji.baris[k]) for k in params if k in uji.baris]
                else:
                    self._hasil = [(next(iter(uji.baris.values())),)] if uji.baris else []

            def fetchall(self):
                return self._hasil

            def fetchone(self):
                return self._hasil[0] if self._hasil else None

        class KoneksiPalsu:
            def cursor(self):
                return KursorPalsu()

            def close(self):
                pass

        self._asli = mssql._get_connection
        mssql._get_connection = lambda: KoneksiPalsu()
        self.addCleanup(lambda: setattr(mssql, '_get_connection', self._asli))

    def _spec(self, pk, **kwargs):
        spec = {'pk': pk, 'tabel': 'dbo.RT', 'kolom_nilai': 'VALUE',
                'kolom_kunci': 'ANALOG', 'nilai_kunci': f'TAG{pk}', 'faktor': 1.0}
        spec.update(kwargs)
        return spec

    @override_settings(MSSQL_HOST='127.0.0.1,1433')
    def test_satu_query_untuk_banyak_titik_di_tabel_sama(self):
        self.baris = {'TAG1': 10.0, 'TAG2': 20.0, 'TAG3': 30.0}
        hasil = mssql.get_nilai_ews([self._spec(1), self._spec(2), self._spec(3)])
        self.assertEqual(hasil, {1: 10.0, 2: 20.0, 3: 30.0})
        self.assertEqual(len(self.dijalankan), 1)
        sql, params = self.dijalankan[0]
        self.assertIn('IN (?, ?, ?)', sql)
        self.assertEqual(sorted(params), ['TAG1', 'TAG2', 'TAG3'])

    @override_settings(MSSQL_HOST='127.0.0.1,1433')
    def test_titik_yang_menunjuk_kunci_sama_ikut_terisi(self):
        """Semua skema frekuensi satu sistem membaca satu titik ukur yang sama."""
        self.baris = {'FREQ': 50.02}
        hasil = mssql.get_nilai_ews([
            self._spec(1, nilai_kunci='FREQ'),
            self._spec(2, nilai_kunci='FREQ'),
        ])
        self.assertEqual(hasil, {1: 50.02, 2: 50.02})
        self.assertEqual(len(self.dijalankan), 1)

    @override_settings(MSSQL_HOST='127.0.0.1,1433')
    def test_tabel_berbeda_dipisah_jadi_query_sendiri(self):
        self.baris = {'TAG1': 1.0, 'TAG2': 2.0}
        mssql.get_nilai_ews([self._spec(1), self._spec(2, tabel='dbo.LAIN')])
        self.assertEqual(len(self.dijalankan), 2)

    @override_settings(MSSQL_HOST='127.0.0.1,1433')
    def test_faktor_skala_diterapkan(self):
        self.baris = {'TAG1': 150000.0}
        hasil = mssql.get_nilai_ews([self._spec(1, faktor=0.001)])
        self.assertEqual(hasil[1], 150.0)

    @override_settings(MSSQL_HOST='127.0.0.1,1433')
    def test_kolom_kunci_kosong_membaca_baris_pertama(self):
        self.baris = {'X': 49.98}
        hasil = mssql.get_nilai_ews([self._spec(1, kolom_kunci='', nilai_kunci='')])
        self.assertEqual(hasil[1], 49.98)
        sql, params = self.dijalankan[0]
        self.assertIn('SELECT TOP 1', sql)
        self.assertEqual(params, [])

    @override_settings(MSSQL_HOST='127.0.0.1,1433')
    def test_identifier_berbahaya_tidak_pernah_masuk_query(self):
        self.baris = {'TAG4': 7.0}
        hasil = mssql.get_nilai_ews([
            self._spec(1, tabel='dbo.X; DROP TABLE Y'),
            self._spec(2, kolom_nilai='VALUE; DELETE FROM Z'),
            self._spec(3, kolom_kunci='ANALOG OR 1=1'),
            self._spec(4),
        ])
        self.assertEqual(hasil, {1: None, 2: None, 3: None, 4: 7.0})
        semua_sql = ' '.join(sql for sql, _ in self.dijalankan)
        for jahat in ('DROP TABLE', 'DELETE FROM', 'OR 1=1'):
            self.assertNotIn(jahat, semua_sql)

    @override_settings(MSSQL_HOST='127.0.0.1,1433')
    def test_nilai_kunci_lewat_bind_parameter(self):
        self.baris = {}
        mssql.get_nilai_ews([self._spec(1, nilai_kunci="x' OR '1'='1")])
        sql, params = self.dijalankan[0]
        self.assertNotIn("OR '1'='1", sql)
        self.assertEqual(params, ["x' OR '1'='1"])

    @override_settings(MSSQL_HOST='')
    def test_tanpa_mssql_host_kembalikan_none_tanpa_query(self):
        hasil = mssql.get_nilai_ews([self._spec(1)])
        self.assertEqual(hasil, {1: None})
        self.assertEqual(self.dijalankan, [])

    @override_settings(MSSQL_HOST='127.0.0.1,1433')
    def test_koneksi_gagal_tidak_melempar_exception(self):
        def gagal():
            raise ConnectionError('historian mati')
        mssql._get_connection = gagal
        self.assertEqual(mssql.get_nilai_ews([self._spec(1)]), {1: None})


class HalamanEWSTest(TestCase):
    """Halaman & endpoint EWS tetap tampil meski MSSQL tidak tersedia."""

    @classmethod
    def setUpTestData(cls):
        cls.user = User.objects.create_superuser('ews-user', 'ews@contoh.id', 'rahasia-tes-123')
        # ForcePasswordChangeMiddleware mengalihkan user baru ke /ganti-password/
        # sebelum sempat sampai ke halaman EWS (lihat KeteranganSumberDashboardTest).
        profile = getattr(cls.user, 'profile', None)
        if profile is not None:
            profile.force_password_change = False
            profile.save(update_fields=['force_password_change'])
        cls.kolom = KolomEWS.objects.create(nama='Parameter Tegangan', urutan=1)
        cls.kolom_mati = KolomEWS.objects.create(nama='Kolom Nonaktif', aktif=False)
        cls.titik = TitikEWS.objects.create(
            kolom=cls.kolom, nama='GI Uji', skema='UVLS', sistem='Sulbagsel',
            besaran='v', nominal=150, setting=135, arah='bawah')
        cls.titik_mati = TitikEWS.objects.create(
            kolom=cls.kolom, nama='GI Nonaktif', skema='UVLS', aktif=False)
        cls.titik_kolom_mati = TitikEWS.objects.create(
            kolom=cls.kolom_mati, nama='GI Kolom Mati', skema='UVLS')

    def setUp(self):
        self.client.force_login(self.user)

    @override_settings(MSSQL_HOST='')
    def test_halaman_tampil_tanpa_mssql(self):
        resp = self.client.get(reverse('opsis_ews'))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'EWS Defense Scheme')
        self.assertContains(resp, 'Parameter Tegangan')

    @override_settings(MSSQL_HOST='')
    def test_kolom_dan_titik_nonaktif_tidak_ikut(self):
        resp = self.client.get(reverse('opsis_ews'))
        self.assertEqual(resp.context['titik_count'], 1)
        self.assertNotContains(resp, 'GI Nonaktif')
        self.assertNotContains(resp, 'Kolom Nonaktif')

    @override_settings(MSSQL_HOST='')
    def test_kolom_tanpa_titik_tidak_digambar(self):
        KolomEWS.objects.create(nama='Kolom Kosong')
        resp = self.client.get(reverse('opsis_ews'))
        self.assertNotContains(resp, 'Kolom Kosong')

    @override_settings(MSSQL_HOST='')
    def test_api_kembalikan_status_plan_saat_mssql_kosong(self):
        resp = self.client.get(reverse('opsis_api_ews'))
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.content)
        self.assertTrue(data['terputus'])
        self.assertEqual(data['data'][str(self.titik.pk)],
                         {'nilai': None, 'status': 'plan', 'margin': None})

    def test_butuh_login(self):
        self.client.logout()
        for nama in ('opsis_ews', 'opsis_api_ews'):
            resp = self.client.get(reverse(nama))
            self.assertEqual(resp.status_code, 302)
            self.assertIn('/login', resp['Location'])


class EWSSuntingKartuTest(TestCase):
    """
    Sunting ambang setting langsung dari kartu EWS: siapa yang boleh, apa yang
    divalidasi, dan apa yang TIDAK boleh ikut berubah dari sini.
    """

    @classmethod
    def setUpTestData(cls):
        cls.kolom = KolomEWS.objects.create(nama='Tegangan')
        cls.titik = TitikEWS.objects.create(
            kolom=cls.kolom, nama='GI Uji', skema='UVLS', sistem='Sulbagsel',
            besaran='v', nominal=150, setting=135, arah='bawah', time_delay='1 s',
            sumber_tabel='dbo.RT', sumber_kolom_nilai='VALUE',
            sumber_kolom_kunci='ANALOG', sumber_nilai_kunci='TAG')

    def _user(self, nama, role=None, super_=False):
        if super_:
            u = User.objects.create_superuser(nama, f'{nama}@contoh.id', 'rahasia-tes-123')
        else:
            u = User.objects.create_user(nama, f'{nama}@contoh.id', 'rahasia-tes-123')
        profile = getattr(u, 'profile', None)
        if profile is not None:
            profile.force_password_change = False
            if role:
                profile.role = role
            profile.save()
        return u

    def _simpan(self, muatan):
        return self.client.post(reverse('opsis_ews_simpan'),
                                data=json.dumps(muatan), content_type='application/json')

    # ---- izin --------------------------------------------------------------

    def test_teknisi_boleh_menyimpan(self):
        self.client.force_login(self._user('tek', role='technician'))
        r = self._simpan({'pk': self.titik.pk, 'setting': '133.5'})
        self.assertEqual(r.status_code, 200)
        self.assertTrue(json.loads(r.content)['ok'])
        self.titik.refresh_from_db()
        self.assertEqual(self.titik.setting, 133.5)

    def test_superuser_boleh_menyimpan(self):
        self.client.force_login(self._user('adm', super_=True))
        r = self._simpan({'pk': self.titik.pk, 'setting': '134'})
        self.assertEqual(r.status_code, 200)

    def test_viewer_ditolak(self):
        self.client.force_login(self._user('lihat', role='viewer'))
        r = self._simpan({'pk': self.titik.pk, 'setting': '100'})
        self.assertEqual(r.status_code, 403)
        self.titik.refresh_from_db()
        self.assertEqual(self.titik.setting, 135)      # tidak berubah

    def test_anonim_dialihkan_ke_login(self):
        r = self._simpan({'pk': self.titik.pk, 'setting': '100'})
        self.assertEqual(r.status_code, 302)
        self.assertIn('/login', r['Location'])

    def test_tombol_pensil_hanya_untuk_yang_berwenang(self):
        self.client.force_login(self._user('tek2', role='technician'))
        self.assertTrue(self.client.get(reverse('opsis_ews')).context['bisa_edit'])
        self.client.force_login(self._user('lihat2', role='viewer'))
        resp = self.client.get(reverse('opsis_ews'))
        self.assertFalse(resp.context['bisa_edit'])
        # Nama 'ewsCsrf' selalu ada di dalam JS dan base template punya token
        # sendiri untuk form logout, jadi yang dicek elemen formulirnya.
        self.assertNotContains(resp, 'id="ewsCsrf"')
        self.assertContains(resp, 'id="ews-bisa-edit"')

    # ---- batas kewenangan --------------------------------------------------

    def test_pemetaan_mssql_tidak_bisa_diubah_dari_kartu(self):
        """Field sumber_* hanya lewat site admin — payload nakal harus diabaikan."""
        self.client.force_login(self._user('tek3', role='technician'))
        r = self._simpan({
            'pk': self.titik.pk, 'setting': '130',
            'sumber_tabel': 'dbo.JAHAT', 'sumber_kolom_kunci': 'X',
            'sumber_nilai_kunci': 'Y', 'faktor_skala': 99,
        })
        self.assertEqual(r.status_code, 200)
        self.titik.refresh_from_db()
        self.assertEqual(self.titik.setting, 130)          # yang boleh, berubah
        self.assertEqual(self.titik.sumber_tabel, 'dbo.RT')  # yang tidak boleh, utuh
        self.assertEqual(self.titik.sumber_kolom_kunci, 'ANALOG')
        self.assertEqual(self.titik.sumber_nilai_kunci, 'TAG')
        self.assertEqual(self.titik.faktor_skala, 1.0)

    # ---- validasi ----------------------------------------------------------

    def test_metode_selain_post_ditolak(self):
        self.client.force_login(self._user('tek4', role='technician'))
        self.assertEqual(self.client.get(reverse('opsis_ews_simpan')).status_code, 405)

    def test_json_rusak_ditolak(self):
        self.client.force_login(self._user('tek5', role='technician'))
        r = self.client.post(reverse('opsis_ews_simpan'), data='bukan json',
                             content_type='application/json')
        self.assertEqual(r.status_code, 400)

    def test_titik_tidak_ada_ditolak(self):
        self.client.force_login(self._user('tek6', role='technician'))
        self.assertEqual(self._simpan({'pk': 999999, 'setting': '1'}).status_code, 404)

    def test_nilai_bukan_angka_ditolak(self):
        self.client.force_login(self._user('tek7', role='technician'))
        r = self._simpan({'pk': self.titik.pk, 'setting': 'seratus'})
        self.assertEqual(r.status_code, 400)
        self.assertIn('angka', json.loads(r.content)['error'])
        self.titik.refresh_from_db()
        self.assertEqual(self.titik.setting, 135)

    def test_arah_tidak_dikenal_ditolak(self):
        self.client.force_login(self._user('tek8', role='technician'))
        self.assertEqual(self._simpan({'pk': self.titik.pk, 'arah': 'miring'}).status_code, 400)

    def test_ambang_waspada_harus_positif(self):
        self.client.force_login(self._user('tek9', role='technician'))
        self.assertEqual(self._simpan({'pk': self.titik.pk, 'ambang_waspada': '0'}).status_code, 400)
        self.assertEqual(self._simpan({'pk': self.titik.pk, 'ambang_waspada': '-1'}).status_code, 400)

    def test_koma_desimal_diterima(self):
        """Operator lapangan terbiasa mengetik 133,5 — jangan ditolak."""
        self.client.force_login(self._user('tek10', role='technician'))
        r = self._simpan({'pk': self.titik.pk, 'setting': '133,5'})
        self.assertEqual(r.status_code, 200)
        self.titik.refresh_from_db()
        self.assertEqual(self.titik.setting, 133.5)

    def test_setting_boleh_dikosongkan(self):
        self.client.force_login(self._user('tek11', role='technician'))
        r = self._simpan({'pk': self.titik.pk, 'setting': ''})
        self.assertEqual(r.status_code, 200)
        self.titik.refresh_from_db()
        self.assertIsNone(self.titik.setting)

    def test_time_delay_kepanjangan_ditolak(self):
        self.client.force_login(self._user('tek12', role='technician'))
        self.assertEqual(
            self._simpan({'pk': self.titik.pk, 'time_delay': 'x' * 31}).status_code, 400)

    # ---- jejak audit -------------------------------------------------------

    def test_perubahan_tercatat_di_auditlog(self):
        self.client.force_login(self._user('tek13', role='technician'))
        self._simpan({'pk': self.titik.pk, 'setting': '130', 'arah': 'atas'})
        entri = AuditLog.objects.filter(model_name='TitikEWS').order_by('-id').first()
        self.assertIsNotNone(entri)
        self.assertEqual(entri.action, AuditLog.UPDATE)
        self.assertIn('135.0', entri.detail)     # nilai lama ikut tercatat
        self.assertIn('130.0', entri.detail)     # dan nilai barunya
        self.assertIn('arah', entri.detail)

    def test_tanpa_perubahan_tidak_menulis_auditlog(self):
        self.client.force_login(self._user('tek14', role='technician'))
        awal = AuditLog.objects.filter(model_name='TitikEWS').count()
        r = self._simpan({'pk': self.titik.pk, 'setting': '135'})   # sama seperti sekarang
        self.assertEqual(r.status_code, 200)
        self.assertFalse(json.loads(r.content)['berubah'])
        self.assertEqual(AuditLog.objects.filter(model_name='TitikEWS').count(), awal)


class FreqHistoryGabunganTest(TestCase):
    """
    Riwayat frekuensi dari dua sumber: SYS_FREQ_HIS (historian) + SnapFreqRT
    (rekaman FASOP sendiri). Historian jadi acuan, PostgreSQL menambal lubang.
    """

    def setUp(self):
        self.t0 = datetime.datetime(2026, 8, 24, 15, 0, 0)
        self.t1 = datetime.datetime(2026, 8, 24, 15, 0, 10)
        self._asli = mssql.get_freq_range
        self.addCleanup(lambda: setattr(mssql, 'get_freq_range', self._asli))

    def _historian(self, pasangan):
        """Palsukan SYS_FREQ_HIS: [(detik_ke, hz)] relatif t0, naive seperti MSSQL."""
        data = [(self.t0 + datetime.timedelta(seconds=d), hz) for d, hz in pasangan]
        mssql.get_freq_range = lambda a, b: list(data)

    def _postgres(self, pasangan):
        """Isi SnapFreqRT — aware (USE_TZ), sebagaimana collect_freq_rt menyimpannya."""
        for d, hz in pasangan:
            SnapFreqRT.objects.create(
                waktu=timezone.make_aware(self.t0 + datetime.timedelta(seconds=d)), hz=hz)

    def test_historian_saja(self):
        self._historian([(0, 50.0), (1, 50.1)])
        deret, info = freq_history.ambil_range_detail(self.t0, self.t1)
        self.assertEqual(deret, [(self.t0, 50.0),
                                 (self.t0 + datetime.timedelta(seconds=1), 50.1)])
        self.assertEqual(info['sumber'], 'historian')
        self.assertEqual((info['historian'], info['postgres']), (2, 0))

    def test_postgres_saja_saat_historian_mati(self):
        """Kasus 24 Agustus 2026: SYS_FREQ_HIS berhenti, rekaman FASOP jalan."""
        self._historian([])
        self._postgres([(0, 49.9), (1, 49.8)])
        deret, info = freq_history.ambil_range_detail(self.t0, self.t1)
        self.assertEqual([h for _, h in deret], [49.9, 49.8])
        self.assertEqual(info['sumber'], 'postgres')
        self.assertEqual((info['historian'], info['postgres']), (0, 2))

    def test_postgres_menambal_lubang_historian(self):
        """Historian punya detik 0 dan 3; detik 1-2 ditambal dari PostgreSQL."""
        self._historian([(0, 50.0), (3, 50.3)])
        self._postgres([(1, 49.1), (2, 49.2)])
        deret, info = freq_history.ambil_range_detail(self.t0, self.t1)
        self.assertEqual([h for _, h in deret], [50.0, 49.1, 49.2, 50.3])
        self.assertEqual(info['sumber'], 'gabungan')
        self.assertEqual((info['historian'], info['postgres']), (2, 2))

    def test_historian_menang_saat_detik_yang_sama_ada_di_dua_sumber(self):
        """Historian sumber asli — nilainya tidak boleh ditimpa rekaman FASOP."""
        self._historian([(0, 50.0), (1, 50.1)])
        self._postgres([(0, 11.1), (1, 22.2), (2, 49.5)])
        deret, info = freq_history.ambil_range_detail(self.t0, self.t1)
        self.assertEqual([h for _, h in deret], [50.0, 50.1, 49.5])
        self.assertEqual((info['historian'], info['postgres']), (2, 1))

    def test_hasil_selalu_terurut_waktu(self):
        self._historian([(4, 50.4), (0, 50.0)])
        self._postgres([(2, 49.2)])
        deret, _ = freq_history.ambil_range_detail(self.t0, self.t1)
        self.assertEqual([t for t, _ in deret], sorted(t for t, _ in deret))
        self.assertEqual([h for _, h in deret], [50.0, 49.2, 50.4])

    def test_waktu_postgres_disamakan_ke_naive_lokal(self):
        """
        MSSQL mengembalikan naive waktu lokal, PostgreSQL aware (UTC). Tanpa
        penyamaan, kedua deret meleset sejauh offset zona waktu saat digabung.
        """
        self._historian([])
        self._postgres([(0, 49.9)])
        deret, _ = freq_history.ambil_range_detail(self.t0, self.t1)
        waktu = deret[0][0]
        self.assertTrue(timezone.is_naive(waktu))
        self.assertEqual(waktu, self.t0)

    def test_di_luar_rentang_tidak_ikut(self):
        self._historian([])
        self._postgres([(-30, 48.0), (0, 49.9), (300, 51.0)])
        deret, _ = freq_history.ambil_range_detail(self.t0, self.t1)
        self.assertEqual([h for _, h in deret], [49.9])

    def test_dua_sumber_kosong(self):
        self._historian([])
        deret, info = freq_history.ambil_range_detail(self.t0, self.t1)
        self.assertEqual(deret, [])
        self.assertEqual(info['sumber'], 'kosong')

    def test_historian_meledak_tetap_pakai_postgres(self):
        """Analisis tidak boleh ikut mati kalau MSSQL melempar exception."""
        def meledak(a, b):
            raise RuntimeError('historian mati')
        mssql.get_freq_range = meledak
        self._postgres([(0, 49.9)])
        deret, info = freq_history.ambil_range_detail(self.t0, self.t1)
        self.assertEqual([h for _, h in deret], [49.9])
        self.assertEqual(info['sumber'], 'postgres')

    def test_ambil_range_bentuknya_sama_dengan_getter_lama(self):
        """Dipakai sebagai pengganti langsung mssql.get_freq_range."""
        self._historian([(0, 50.0)])
        self.assertEqual(freq_history.ambil_range(self.t0, self.t1),
                         [(self.t0, 50.0)])

    def test_keterangan_menyebut_jumlah_tambalan(self):
        self._historian([(0, 50.0)])
        self._postgres([(1, 49.1)])
        _, info = freq_history.ambil_range_detail(self.t0, self.t1)
        teks = freq_history.keterangan(info)
        self.assertIn('Historian SCADA', teks)
        self.assertIn('1 detik', teks)

    # ---- sumber ketiga: SnapFreq (cermin historian di PostgreSQL) ----------

    def _snapfreq(self, pasangan):
        """Isi opsis.SnapFreq — cermin SYS_FREQ_HIS yang diisi cron collect_freq."""
        for d, hz in pasangan:
            SnapFreq.objects.create(
                waktu=timezone.make_aware(self.t0 + datetime.timedelta(seconds=d)), hz=hz)

    def test_snapfreq_dipakai_saat_mssql_tak_terjangkau(self):
        """Koneksi MSSQL putus, tapi cerminnya di PostgreSQL masih terbaca."""
        def meledak(a, b):
            raise ConnectionError('MSSQL tak terjangkau')
        mssql.get_freq_range = meledak
        self._snapfreq([(0, 50.0), (1, 50.1)])
        deret, info = freq_history.ambil_range_detail(self.t0, self.t1)
        self.assertEqual([h for _, h in deret], [50.0, 50.1])
        self.assertEqual(info['sumber'], 'snapfreq')
        self.assertEqual((info['historian'], info['snapfreq'], info['postgres']), (0, 2, 0))

    def test_urutan_prioritas_tiga_sumber(self):
        """historian > snapfreq > snapfreqrt; yang lebih dulu tidak ditimpa."""
        self._historian([(0, 50.0)])
        self._snapfreq([(0, 11.1), (1, 50.1)])          # detik 0 kalah dari historian
        self._postgres([(0, 22.2), (1, 33.3), (2, 49.2)])  # detik 0,1 kalah
        deret, info = freq_history.ambil_range_detail(self.t0, self.t1)
        self.assertEqual([h for _, h in deret], [50.0, 50.1, 49.2])
        self.assertEqual((info['historian'], info['snapfreq'], info['postgres']), (1, 1, 1))
        self.assertEqual(info['sumber'], 'gabungan')

    def test_snapfreq_tidak_menolong_saat_historian_berhenti_diisi(self):
        """
        Kasus 24 Agustus 2026: SYS_FREQ_HIS berhenti, jadi cerminnya (SnapFreq)
        ikut kosong untuk rentang itu. Hanya SnapFreqRT yang punya isinya —
        inilah alasan sumber ketiga tidak bisa menggantikan yang kedua.
        """
        self._historian([])                # historian berhenti diisi
        # SnapFreq juga tidak punya apa-apa di rentang ini
        self._postgres([(0, 49.9), (1, 49.8)])
        deret, info = freq_history.ambil_range_detail(self.t0, self.t1)
        self.assertEqual([h for _, h in deret], [49.9, 49.8])
        self.assertEqual(info['snapfreq'], 0)
        self.assertEqual(info['sumber'], 'postgres')


class ResponGetterSumberTest(TestCase):
    """Respon Pembangkit harus memakai getter gabungan, bukan MSSQL langsung."""

    def test_getter_respon_memakai_freq_history(self):
        get_freq, _ = views._respon_getters()
        self.assertIs(get_freq, freq_history.ambil_range)


class NamaSheetExcelTest(TestCase):
    """Nama sheet Excel: karakter terlarang, batas 31 huruf, dan tabrakan."""

    def test_karakter_terlarang_diganti(self):
        nama = views._nama_sheet('PLTU A/B [1]:2*3?4', set())
        for c in r'[]:*?/\\':
            self.assertNotIn(c, nama)

    def test_dipotong_31_karakter(self):
        nama = views._nama_sheet('P' * 60, set())
        self.assertEqual(len(nama), 31)

    def test_nama_kembar_diberi_akhiran(self):
        terpakai = set()
        a = views._nama_sheet('PLTA POSO', terpakai)
        b = views._nama_sheet('PLTA POSO', terpakai)
        self.assertEqual(a, 'PLTA POSO')
        self.assertNotEqual(a, b)
        self.assertIn('(2)', b)

    def test_nama_panjang_kembar_tetap_muat_31(self):
        """Dua nama yang sama-sama terpotong di 31 huruf tidak boleh bertabrakan."""
        terpakai = set()
        a = views._nama_sheet('PEMBANGKIT DENGAN NAMA SANGAT PANJANG SEKALI 1', terpakai)
        b = views._nama_sheet('PEMBANGKIT DENGAN NAMA SANGAT PANJANG SEKALI 2', terpakai)
        self.assertNotEqual(a, b)
        self.assertLessEqual(len(a), 31)
        self.assertLessEqual(len(b), 31)

    def test_nama_kosong_tetap_dapat_judul(self):
        self.assertTrue(views._nama_sheet('', set()))


class ExportBebanPembangkitTest(TestCase):
    """Ekspor Excel beban semua pembangkit — satu sheet per pembangkit."""

    @classmethod
    def setUpTestData(cls):
        cls.user = User.objects.create_superuser('exp', 'exp@contoh.id', 'rahasia-tes-123')
        profile = getattr(cls.user, 'profile', None)
        if profile is not None:
            profile.force_password_change = False
            profile.save(update_fields=['force_password_change'])

        cls.a = Pembangkit.objects.create(kode='KITA', nama='PLTA Alpha')
        cls.b = Pembangkit.objects.create(kode='KITB', nama='PLTU Beta')
        cls.mati = Pembangkit.objects.create(kode='KITC', nama='PLTD Nonaktif', aktif=False)

        cls.tanggal = datetime.date(2026, 7, 10)
        tz = timezone.get_current_timezone()
        for menit, (mw_a, mw_b) in enumerate([(10.0, 20.0), (12.0, 22.0), (14.0, 24.0)]):
            w = timezone.make_aware(
                datetime.datetime.combine(cls.tanggal, datetime.time(0, menit)), tz)
            SnapLive.objects.create(pembangkit=cls.a, waktu=w, mw=mw_a, mvar=1.0, frekuensi=50.0)
            SnapLive.objects.create(pembangkit=cls.b, waktu=w, mw=mw_b, mvar=2.0, frekuensi=50.0)

    def setUp(self):
        self.client.force_login(self.user)

    def _unduh(self, query=''):
        import io as _io
        import openpyxl
        resp = self.client.get(reverse('opsis_export_beban_pembangkit') + query)
        self.assertEqual(resp.status_code, 200)
        return resp, openpyxl.load_workbook(_io.BytesIO(resp.content))

    def test_berkas_xlsx_dengan_nama_yang_menyebut_tanggal(self):
        resp, _ = self._unduh('?tanggal=2026-07-10')
        self.assertIn('spreadsheetml.sheet', resp['Content-Type'])
        self.assertIn('BebanPembangkit_2026-07-10.xlsx', resp['Content-Disposition'])

    def test_satu_sheet_per_pembangkit_aktif(self):
        _, wb = self._unduh('?tanggal=2026-07-10')
        self.assertIn('PLTA Alpha', wb.sheetnames)
        self.assertIn('PLTU Beta', wb.sheetnames)
        self.assertNotIn('PLTD Nonaktif', wb.sheetnames)   # nonaktif tidak ikut

    def test_ringkasan_di_depan_keterangan_di_belakang(self):
        _, wb = self._unduh('?tanggal=2026-07-10')
        self.assertEqual(wb.sheetnames[0], 'Ringkasan Sistem')
        self.assertEqual(wb.sheetnames[-1], 'Keterangan')

    def test_isi_sheet_pembangkit(self):
        _, wb = self._unduh('?tanggal=2026-07-10')
        ws = wb['PLTA Alpha']
        baris = list(ws.iter_rows(values_only=True))
        self.assertEqual(baris[0], ('No', 'Tanggal', 'Waktu', 'MW', 'MVAR', 'Frekuensi (Hz)'))
        self.assertEqual([b[3] for b in baris[1:4]], [10.0, 12.0, 14.0])
        self.assertEqual(baris[1][1], '2026-07-10')
        # ringkasan statistik di kaki sheet
        ekor = [b[2] for b in baris if b and b[2] in ('Rata-rata', 'Minimum', 'Maksimum')]
        self.assertEqual(ekor, ['Rata-rata', 'Minimum', 'Maksimum'])

    def test_ringkasan_menjumlah_semua_pembangkit(self):
        _, wb = self._unduh('?tanggal=2026-07-10')
        baris = list(wb['Ringkasan Sistem'].iter_rows(values_only=True))
        self.assertEqual([b[3] for b in baris[1:4]], [30.0, 34.0, 38.0])   # 10+20, 12+22, 14+24
        self.assertEqual([b[4] for b in baris[1:4]], [50.0, 50.0, 50.0])   # Hz ikut terbawa

    def test_pembangkit_tanpa_data_tetap_punya_sheet(self):
        kosong = Pembangkit.objects.create(kode='KITD', nama='PLTS Kosong')
        _, wb = self._unduh('?tanggal=2026-07-10')
        self.assertIn('PLTS Kosong', wb.sheetnames)
        isi = list(wb['PLTS Kosong'].iter_rows(values_only=True))
        self.assertIn('Tidak ada data pada rentang ini', [c for b in isi for c in b])

    def test_rentang_beberapa_hari(self):
        resp, wb = self._unduh('?mulai=2026-07-08&selesai=2026-07-10')
        self.assertIn('BebanPembangkit_2026-07-08_sd_2026-07-10.xlsx',
                      resp['Content-Disposition'])
        self.assertEqual(len(list(wb['PLTA Alpha'].iter_rows(values_only=True))) - 1, 3 + 4)

    def test_hari_lain_tidak_ikut_terbawa(self):
        tz = timezone.get_current_timezone()
        SnapLive.objects.create(
            pembangkit=self.a,
            waktu=timezone.make_aware(
                datetime.datetime.combine(datetime.date(2026, 7, 11), datetime.time(0, 0)), tz),
            mw=99.0)
        _, wb = self._unduh('?tanggal=2026-07-10')
        nilai = [b[3] for b in wb['PLTA Alpha'].iter_rows(values_only=True)][1:4]
        self.assertNotIn(99.0, nilai)

    def test_rentang_terbalik_ditolak(self):
        resp = self.client.get(reverse('opsis_export_beban_pembangkit')
                               + '?mulai=2026-07-10&selesai=2026-07-08')
        self.assertEqual(resp.status_code, 302)

    def test_rentang_terlalu_panjang_ditolak(self):
        resp = self.client.get(reverse('opsis_export_beban_pembangkit')
                               + '?mulai=2026-06-01&selesai=2026-07-10')
        self.assertEqual(resp.status_code, 302)

    def test_tanggal_ngawur_jatuh_ke_hari_ini(self):
        resp, _ = self._unduh('?tanggal=bukan-tanggal')
        self.assertIn(str(timezone.localdate()), resp['Content-Disposition'])

    def test_butuh_login(self):
        self.client.logout()
        resp = self.client.get(reverse('opsis_export_beban_pembangkit'))
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login', resp['Location'])


class CekArmadaKitTest(TestCase):
    """Diagnosa armada KIT harus tetap jalan (dan diam) saat MSSQL tidak ada."""

    @override_settings(MSSQL_HOST='')
    def test_jalan_tanpa_mssql(self):
        from io import StringIO
        from django.core.management import call_command
        keluaran = StringIO()
        call_command('cek_armada_kit', stdout=keluaran, stderr=StringIO())
        teks = keluaran.getvalue()
        self.assertIn('KIT_REALTIME', teks)
        self.assertIn('RESPON_PLANTS', teks)

    @override_settings(MSSQL_HOST='')
    def test_pembangkit_tanpa_padanan_dilaporkan(self):
        from io import StringIO
        from django.core.management import call_command
        Pembangkit.objects.create(kode='ZZZTEST', nama='PLTU Uji Armada')
        keluaran = StringIO()
        call_command('cek_armada_kit', stdout=keluaran, stderr=StringIO())
        self.assertIn('ZZZTEST', keluaran.getvalue())


class PetaSumberDataTest(TestCase):
    """Peta sumber data: klasifikasi kesegaran & urutan lapisan."""

    @classmethod
    def setUpTestData(cls):
        cls.user = User.objects.create_superuser('sd', 'sd@contoh.id', 'rahasia-tes-123')
        profile = getattr(cls.user, 'profile', None)
        if profile is not None:
            profile.force_password_change = False
            profile.save(update_fields=['force_password_change'])

    def setUp(self):
        self.client.force_login(self.user)

    @override_settings(MSSQL_HOST='')
    def test_halaman_tampil_tanpa_mssql(self):
        resp = self.client.get(reverse('opsis_sumber_data'))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Peta Sumber Data')
        self.assertContains(resp, 'dbo.KIT_REALTIME')

    @override_settings(MSSQL_HOST='')
    def test_lapisan_terurut_agar_regroup_tidak_pecah(self):
        """{% regroup %} hanya menggabungkan yang berurutan — urutan wajib rapi."""
        from itertools import groupby
        hasil = sumber_data.periksa_semua(dengan_mssql=False)
        urut = [k for k, _ in groupby(b['lapis'] for b in hasil)]
        self.assertEqual(urut, sorted(set(urut), key=lambda x: sumber_data.LAPIS_URUT[x]))
        self.assertEqual(len(urut), len(set(urut)))     # tiap lapis muncul sekali

    def test_status_segar_telat_mati(self):
        sekarang = timezone.now()
        self.assertEqual(sumber_data._status_dari_waktu(sekarang)[0], 'segar')
        self.assertEqual(
            sumber_data._status_dari_waktu(sekarang - datetime.timedelta(hours=2))[0], 'telat')
        self.assertEqual(
            sumber_data._status_dari_waktu(sekarang - datetime.timedelta(days=3))[0], 'mati')
        self.assertEqual(sumber_data._status_dari_waktu(None)[0], 'kosong')

    @override_settings(MSSQL_HOST='')
    def test_setiap_entri_punya_kolom_wajib(self):
        for b in sumber_data.periksa_semua(dengan_mssql=False):
            for kunci in ('fitur', 'lapis', 'sumber', 'diisi', 'periksa'):
                self.assertIn(kunci, b, f'{b.get("sumber")} kehilangan {kunci}')

    def test_postgres_kosong_dilaporkan_kosong(self):
        hasil = {b['sumber']: b['periksa'] for b in sumber_data.periksa_semua(dengan_mssql=False)}
        self.assertEqual(hasil['opsis.SnapLive']['status'], 'kosong')

    def test_postgres_terisi_dilaporkan_segar(self):
        kit = Pembangkit.objects.create(kode='SDX', nama='PLTU Sumber Data')
        SnapLive.objects.create(pembangkit=kit, waktu=timezone.now(), mw=1.0)
        hasil = {b['sumber']: b['periksa'] for b in sumber_data.periksa_semua(dengan_mssql=False)}
        self.assertEqual(hasil['opsis.SnapLive']['status'], 'segar')

    def test_butuh_login(self):
        self.client.logout()
        resp = self.client.get(reverse('opsis_sumber_data'))
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login', resp['Location'])


@override_settings(MSSQL_HOST='')   # jangan menembak historian saat menguji akses
class AksesOpsisAMTest(TestCase):
    """
    Role AM (asisten_manager) boleh membuka SELURUH menu /opsis/* dan
    melakukan aksi tulisnya. Sebelum ini AM tertahan di Respons Pembangkit,
    Logsheet, dan Input HOP.

    Aturan role-nya tinggal di devices.models.UserProfile — tes ini menjaga
    supaya halaman dan properti tidak berbeda pendapat.
    """

    HALAMAN_TERBATAS = ['/opsis/respon/', '/opsis/logsheet/', '/opsis/hop/input/']

    def _buat(self, username, role):
        user = User.objects.create_user(username, f'{username}@contoh.id', 'rahasia-tes-123')
        profil = user.profile
        profil.role = role
        profil.force_password_change = False
        profil.save()
        return user

    def _klien(self, user):
        c = self.client_class()
        c.force_login(user)      # force_login: AxesBackend menolak authenticate() tanpa request
        return c

    def test_am_bisa_buka_halaman_opsis_terbatas(self):
        c = self._klien(self._buat('am-akses', 'asisten_manager'))
        for url in self.HALAMAN_TERBATAS:
            with self.subTest(url=url):
                self.assertEqual(c.get(url).status_code, 200)

    def test_am_bisa_semua_aksi_tulis_opsis(self):
        profil = self._buat('am-tulis', 'asisten_manager').profile
        self.assertTrue(profil.bisa_lihat_opsis)
        self.assertTrue(profil.bisa_tulis_opsis)
        self.assertTrue(profil.bisa_sunting_ews)
        self.assertTrue(profil.can_input_hop)

    def test_opsis_view_tetap_lihat_saja(self):
        profil = self._buat('ov-akses', 'opsis_view').profile
        self.assertTrue(profil.bisa_lihat_opsis)
        self.assertFalse(profil.bisa_tulis_opsis)

    def test_viewer_tetap_tertutup(self):
        user = self._buat('viewer-akses', 'viewer')
        self.assertFalse(user.profile.bisa_lihat_opsis)
        c = self._klien(user)
        for url in self.HALAMAN_TERBATAS:
            with self.subTest(url=url):
                self.assertEqual(c.get(url).status_code, 302)

    def test_teknisi_tetap_bisa_sunting_ews_tapi_bukan_tulis_opsis(self):
        # Sunting setting rele memang milik Teknisi; menambah AM tidak boleh
        # diam-diam memberi Teknisi akses input HOP / penanda data.
        profil = self._buat('tek-akses', 'technician').profile
        self.assertTrue(profil.bisa_sunting_ews)
        self.assertFalse(profil.bisa_tulis_opsis)
        self.assertFalse(profil.bisa_lihat_opsis)

    def test_menu_input_hop_muncul_untuk_am(self):
        c = self._klien(self._buat('am-menu', 'asisten_manager'))
        isi = c.get('/opsis/hop/').content.decode()
        self.assertIn('/opsis/hop/input/', isi)

    def test_menu_input_hop_tersembunyi_untuk_opsis_view(self):
        c = self._klien(self._buat('ov-menu', 'opsis_view'))
        isi = c.get('/opsis/hop/').content.decode()
        self.assertNotIn('/opsis/hop/input/', isi)


@override_settings(MSSQL_HOST='')   # dashboard tidak perlu menembak historian
class PantauanKitTest(TestCase):
    """
    Kartu "KIT Terpilih" di dashboard: total MW pembangkit pilihan + chart
    24 jam-nya. Anggotanya dipilih di site admin, jadi tes ini menjaga supaya
    kartu tidak muncul sebelum dikonfigurasi dan angkanya hanya menjumlahkan
    anggota yang benar.
    """

    def setUp(self):
        self.hari_ini = timezone.localdate()
        self.a = Pembangkit.objects.create(kode='PKA', nama='PLTU A', kode_kit='PKA', urutan=1)
        self.b = Pembangkit.objects.create(kode='PKB', nama='PLTU B', kode_kit='PKB', urutan=2)
        self.luar = Pembangkit.objects.create(kode='PKC', nama='PLTD C', kode_kit='PKC', urutan=3)
        self.url = reverse('opsis_api_beban_kit_terpilih')

    def _user_login(self):
        user = User.objects.create_user('pantau-tes', 'p@contoh.id', 'rahasia-tes-123')
        profil = user.profile
        profil.force_password_change = False
        profil.save()
        self.client.force_login(user)
        return user

    def _snap(self, pembangkit, menit, mw):
        SnapLive.objects.create(pembangkit=pembangkit, mw=mw,
                                waktu=_waktu_lokal(self.hari_ini, menit))

    def _nyalakan(self, *anggota):
        pantauan = PantauanKit.ambil()
        pantauan.aktif = True
        pantauan.save()
        pantauan.anggota.set(anggota)
        return pantauan

    def test_singleton_dan_bawaan_mati(self):
        p1 = PantauanKit.ambil()
        p2 = PantauanKit.ambil()
        self.assertEqual(p1.pk, p2.pk)
        self.assertEqual(PantauanKit.objects.count(), 1)
        self.assertFalse(p1.aktif)
        self.assertFalse(p1.tampil())

    def test_selalu_satu_baris_walau_dipaksa_simpan_baru(self):
        PantauanKit.ambil()
        PantauanKit(nama='Coba Kedua').save()
        self.assertEqual(PantauanKit.objects.count(), 1)
        self.assertEqual(PantauanKit.objects.get().nama, 'Coba Kedua')

    def test_api_menjumlahkan_hanya_anggota(self):
        self._nyalakan(self.a, self.b)
        self._snap(self.a, 600, 100.0)
        self._snap(self.b, 600, 25.5)
        self._snap(self.luar, 600, 999.0)      # bukan anggota -> tidak boleh ikut
        self._snap(self.a, 660, 110.0)

        self._user_login()
        body = self.client.get(self.url).json()

        self.assertTrue(body['aktif'])
        self.assertEqual(body['rows'], [{'minute': 600, 'mw': 125.5},
                                        {'minute': 660, 'mw': 110.0}])
        self.assertEqual(body['terakhir'], 110.0)
        self.assertEqual(body['anggota'], ['PLTU A', 'PLTU B'])

    def test_api_abaikan_hari_lain(self):
        self._nyalakan(self.a)
        kemarin = self.hari_ini - datetime.timedelta(days=1)
        SnapLive.objects.create(pembangkit=self.a, mw=500.0,
                                waktu=_waktu_lokal(kemarin, 600))
        self._user_login()
        self.assertEqual(self.client.get(self.url).json()['rows'], [])

    def test_api_diam_saat_belum_dikonfigurasi(self):
        self._user_login()
        body = self.client.get(self.url).json()
        self.assertFalse(body['aktif'])
        self.assertEqual(body['rows'], [])

    def test_anggota_nonaktif_tidak_ikut(self):
        pantauan = self._nyalakan(self.a, self.b)
        self.b.aktif = False
        self.b.save()
        self.assertEqual([p.kode for p in pantauan.anggota_aktif()], ['PKA'])

        self._snap(self.a, 600, 100.0)
        self._snap(self.b, 600, 25.5)
        self._user_login()
        self.assertEqual(self.client.get(self.url).json()['rows'],
                         [{'minute': 600, 'mw': 100.0}])

    def test_kartu_muncul_di_dashboard_hanya_saat_dinyalakan(self):
        self._user_login()
        self.assertNotIn('id="chart-pantauan"', self.client.get('/opsis/').content.decode())

        self._nyalakan(self.a)
        isi = self.client.get('/opsis/').content.decode()
        self.assertIn('id="chart-pantauan"', isi)
        self.assertIn('id="pantauan-mw"', isi)

    def test_kartu_tersembunyi_bila_aktif_tapi_tanpa_anggota(self):
        pantauan = PantauanKit.ambil()
        pantauan.aktif = True
        pantauan.save()
        self.assertFalse(pantauan.tampil())
        self._user_login()
        self.assertNotIn('id="chart-pantauan"', self.client.get('/opsis/').content.decode())
