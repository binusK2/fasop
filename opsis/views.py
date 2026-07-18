import datetime
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.utils import timezone
from django.conf import settings
from .models import Pembangkit, SnapLive, SnapFreq, Trafo, SnapTrafo
from . import mssql
from . import forecast


def _pembangkit_aktif():
    return list(Pembangkit.objects.filter(aktif=True))


def _trafo_aktif_saja(rows):
    """
    Filter hasil mssql.get_beban_trafo()/get_beban_trafo_ibt() agar hanya
    trafo yang aktif di admin (opsis.Trafo) yang ikut ditampilkan/dihitung.
    Trafo baru yang belum terdaftar otomatis didaftarkan sebagai aktif=True
    supaya tidak hilang dari tampilan sebelum sempat dikonfigurasi.

    MSSQL tidak reachable → mssql.get_beban_trafo()/get_beban_trafo_ibt()
    sudah return [] (lihat mssql.py), jadi tidak ada risiko baris palsu
    ikut auto-registrasi di sini.
    """
    existing = {(t.site, t.bay): t.aktif for t in Trafo.objects.all()}
    result = []
    for r in rows:
        key = (r['site'], r['bay'])
        if key not in existing:
            Trafo.objects.get_or_create(site=r['site'], bay=r['bay'])
            existing[key] = True
        if existing[key]:
            result.append(r)
    return result


@login_required
def dashboard(request):
    pembangkit_list = _pembangkit_aktif()
    # Kelompokkan per jenis untuk tampilan grid
    grouped = {}
    for p in pembangkit_list:
        grouped.setdefault(p.jenis, []).append(p)
    return render(request, 'opsis/dashboard.html', {
        'pembangkit_list': pembangkit_list,
        'grouped':         grouped,
    })


@login_required
def pembangkit_detail(request, pk):
    p = get_object_or_404(Pembangkit, pk=pk, aktif=True)
    pembangkit_list = _pembangkit_aktif()
    return render(request, 'opsis/pembangkit.html', {
        'p':               p,
        'pembangkit_list': pembangkit_list,
    })


@login_required
def up2d(request):
    """Dashboard UP2D — Frekuensi Sistem, Beban Trafo, Beban KTT."""
    pembangkit_list = _pembangkit_aktif()
    return render(request, 'opsis/up2d.html', {
        'pembangkit_list': pembangkit_list,
    })


@login_required
def api_live(request):
    """JSON: nilai live semua pembangkit aktif. Dipanggil setiap 5 detik."""
    pembangkit_list = _pembangkit_aktif()
    result = mssql.get_live_data(pembangkit_list)
    data   = result['data']
    response = {
        'data':              data,
        'frekuensi_sistem':  result.get('frekuensi_sistem'),
        'terputus':          not mssql.is_reachable(),
    }
    return JsonResponse(response)


@login_required
def api_trend(request, pk):
    """JSON: data trend satu pembangkit untuk chart. ?jam=1|6|24"""
    p = get_object_or_404(Pembangkit, pk=pk)
    try:
        jam = int(request.GET.get('jam', 1))
        if jam not in (1, 6, 24):
            jam = 1
    except (ValueError, TypeError):
        jam = 1

    rows = mssql.get_trend_data(p, jam)

    labels    = [r['timestamp'] for r in rows]
    frekuensi = [r['frekuensi'] for r in rows]
    mw        = [r['mw']        for r in rows]
    mvar      = [r['mvar']      for r in rows]

    return JsonResponse({
        'labels':    labels,
        'frekuensi': frekuensi,
        'mw':        mw,
        'mvar':      mvar,
        'jam':       jam,
        'nama':      p.nama,
        'warna':     p.warna,
        'terputus':  not mssql.is_reachable(),
    })


@login_required
def export_frekuensi(request):
    """
    Download rekap frekuensi sistem harian dari PostgreSQL (SnapFreq).
    ?tanggal=YYYY-MM-DD  (default: hari ini) — bisa pilih tanggal histori
    selama data masih ada (retensi SnapFreq 30 hari).
    Format: Excel (.xlsx)
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.db.models import Avg
    from django.http import HttpResponse

    tz_local = timezone.get_current_timezone()

    # Parse tanggal
    tanggal_str = request.GET.get('tanggal', '')
    try:
        tanggal = datetime.date.fromisoformat(tanggal_str)
    except ValueError:
        tanggal = timezone.now().astimezone(tz_local).date()

    # Ambil dari SnapFreq (per detik, retensi 30 hari)
    rows = SnapFreq.objects.filter(waktu__date=tanggal).order_by('waktu')

    # Buat workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Frekuensi {tanggal}'

    # Style
    hdr_fill = PatternFill('solid', fgColor='0F172A')
    hdr_font = Font(bold=True, color='60A5FA', size=10)
    thin     = Side(style='thin', color='1E293B')
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    center   = Alignment(horizontal='center', vertical='center')

    # Header
    headers = ['No', 'Tanggal', 'Waktu', 'Frekuensi (Hz)', 'Status']
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = border
        cell.alignment = center

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18

    # Batas frekuensi normal PLN: 49.5 – 50.5 Hz
    batas_bawah, batas_atas = 49.5, 50.5

    data_hz = []  # (hz, waktu_lokal) — untuk ringkasan min/max beserta jamnya
    for i, row in enumerate(rows, 1):
        waktu_lokal = row.waktu.astimezone(tz_local)
        hz = round(row.hz, 4) if row.hz is not None else None
        if hz is not None:
            data_hz.append((hz, waktu_lokal))
        if hz is None:
            status = '—'
        elif hz < batas_bawah:
            status = '⚠ Rendah'
        elif hz > batas_atas:
            status = '⚠ Tinggi'
        else:
            status = '✓ Normal'

        ws.append([
            i,
            waktu_lokal.strftime('%Y-%m-%d'),
            waktu_lokal.strftime('%H:%M:%S'),
            hz,
            status,
        ])

        # Warna baris abnormal
        if hz is not None and (hz < batas_bawah or hz > batas_atas):
            for col in range(1, 6):
                ws.cell(i + 1, col).fill = PatternFill('solid', fgColor='2D1A1A')

        for col in range(1, 6):
            ws.cell(i + 1, col).border = border
            ws.cell(i + 1, col).alignment = center

    # Summary baris terakhir
    if data_hz:
        hz_vals = [h for h, _ in data_hz]
        min_hz, min_waktu = min(data_hz, key=lambda x: x[0])
        max_hz, max_waktu = max(data_hz, key=lambda x: x[0])
        ws.append([])
        ws.append(['', '', 'Rata-rata', round(sum(hz_vals)/len(hz_vals), 4), ''])
        ws.append(['', '', 'Minimum',   min_hz, f"pukul {min_waktu.strftime('%H:%M:%S')}"])
        ws.append(['', '', 'Maksimum',  max_hz, f"pukul {max_waktu.strftime('%H:%M:%S')}"])
        abnormal = sum(1 for h in hz_vals if h < batas_bawah or h > batas_atas)
        ws.append(['', '', 'Abnormal',  abnormal, f'dari {len(hz_vals)} data'])

    # Response
    filename = f'Frekuensi_{tanggal}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def export_beban(request):
    """
    Download rekap beban kit harian dari PostgreSQL (SnapLive).
    ?tanggal=YYYY-MM-DD  (default: hari ini)
    Format: Excel (.xlsx)
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.db.models import Sum
    from django.http import HttpResponse

    tanggal_str = request.GET.get('tanggal', '')
    try:
        tanggal = datetime.date.fromisoformat(tanggal_str)
    except ValueError:
        tanggal = timezone.now().astimezone(timezone.get_current_timezone()).date()

    tz_local = timezone.get_current_timezone()

    # Beban total per menit (sum semua pembangkit)
    rows = (SnapLive.objects
            .filter(waktu__date=tanggal)
            .values('waktu')
            .annotate(total_mw=Sum('mw'))
            .order_by('waktu'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Beban {tanggal}'

    hdr_fill = PatternFill('solid', fgColor='0F172A')
    hdr_font = Font(bold=True, color='34D399', size=10)
    thin     = Side(style='thin', color='1E293B')
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    center   = Alignment(horizontal='center', vertical='center')

    headers = ['No', 'Tanggal', 'Waktu', 'Total Beban (MW)']
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = border
        cell.alignment = center

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 18

    for i, row in enumerate(rows, 1):
        waktu_lokal = row['waktu'].astimezone(tz_local)
        mw = round(row['total_mw'], 2) if row['total_mw'] is not None else None
        ws.append([i, waktu_lokal.strftime('%Y-%m-%d'), waktu_lokal.strftime('%H:%M'), mw])
        for col in range(1, 5):
            ws.cell(i + 1, col).border = border
            ws.cell(i + 1, col).alignment = center

    if rows:
        mw_vals = [r['total_mw'] for r in rows if r['total_mw'] is not None]
        if mw_vals:
            ws.append([])
            ws.append(['', '', 'Rata-rata', round(sum(mw_vals)/len(mw_vals), 2)])
            ws.append(['', '', 'Minimum',   round(min(mw_vals), 2)])
            ws.append(['', '', 'Maksimum',  round(max(mw_vals), 2)])

    filename = f'BebanKit_{tanggal}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def api_hz(request):
    """Hz terkini — ringan, dipanggil tiap 1 detik dari dashboard."""
    hz = mssql.get_current_hz()
    return JsonResponse({'hz': hz, 'terputus': not mssql.is_reachable()})


@login_required
def api_hz_sultra(request):
    """Hz terkini Sultra dari TRANS_KDNEW5_RT (GI KENDARI NEW / COMMON)."""
    hz = mssql.get_freq_sultra()
    return JsonResponse({'hz': hz, 'terputus': not mssql.is_reachable()})


@login_required
def api_hz_baubau(request):
    """Hz terkini Baubau dari TRANS_BAUBAU5_RT (GI BAUBAU / COMMON)."""
    hz = mssql.get_freq_baubau()
    return JsonResponse({'hz': hz, 'terputus': not mssql.is_reachable()})


@login_required
def api_hz_sulteng(request):
    """Hz terkini Sulteng dari TRANS_TLISE5_RT (GI TALISE 150 / COMMON)."""
    hz = mssql.get_freq_sulteng()
    return JsonResponse({'hz': hz, 'terputus': not mssql.is_reachable()})


@login_required
def api_hz_luwuk(request):
    """Hz terkini Luwuk dari TRANS_LUWUK5_RT (GI LUWUK / COMMON)."""
    hz = mssql.get_freq_luwuk()
    return JsonResponse({'hz': hz, 'terputus': not mssql.is_reachable()})


@login_required
def api_freq(request):
    """
    Chart frekuensi sistem.
    ?mode=hari_ini  → rata-rata per menit sejak 00:00 (untuk chart Frekuensi Hari Ini)
    ?menit=N        → last N menit data per detik (default 10, maks 60)
    """
    if request.GET.get('mode') == 'hari_ini':
        rows = mssql.get_freq_hari_ini()
    else:
        try:
            menit = min(max(int(request.GET.get('menit', 10)), 1), 120)
        except (ValueError, TypeError):
            menit = 10
        rows = mssql.get_freq_trend(menit)
    return JsonResponse({'rows': rows, 'terputus': not mssql.is_reachable()})


@login_required
def api_beban(request):
    """
    Chart beban kit — dua seri terpisah, keduanya membentang PENUH 24 jam
    (00:00-23:30) hari ini: 'actual' (SnapLive, resolusi per menit, hanya
    s.d. data terbaru) dan 'forecast' (model gradient boosting, grid 30
    menit, penuh sehari — supaya bisa dibandingkan visual dgn realisasi)
    — lihat opsis.forecast.predict_beban_hari_ini().
    Prediksi & realisasi puncak siang (12:00)/malam (18:30) dikirim terpisah
    supaya UI bisa menampilkan keduanya sekaligus.
    """
    result = forecast.predict_beban_hari_ini()
    if result['actual'] or result['forecast']:
        return JsonResponse({
            'actual':                  result['actual'],
            'forecast':                result['forecast'],
            'source':                  result['source'],
            'prediksi_puncak_siang':   result['prediksi_puncak_siang'],
            'prediksi_puncak_malam':   result['prediksi_puncak_malam'],
            'realisasi_puncak_siang':  result['realisasi_puncak_siang'],
            'realisasi_puncak_malam':  result['realisasi_puncak_malam'],
        })

    # Fallback: SnapLive benar-benar kosong (belum ada data sama sekali) —
    # pakai MSSQL HIS_MEAS_KIT per 15 menit, tanpa prediksi.
    actual = []
    for r in mssql.get_beban_trend():
        try:
            hh, mm = r['timestamp'].split(':')
            actual.append({'minute': int(hh) * 60 + int(mm), 'mw': r['mw']})
        except (ValueError, AttributeError):
            continue
    return JsonResponse({
        'actual': actual, 'forecast': [], 'source': 'mssql',
        'prediksi_puncak_siang': None, 'prediksi_puncak_malam': None,
        'realisasi_puncak_siang': None, 'realisasi_puncak_malam': None,
        'terputus': not mssql.is_reachable(),
    })


@login_required
def api_ping(request):
    """Cek cepat TCP ke MSSQL host:1433. Selesai dalam maks ~2 detik."""
    if not (request.user.is_superuser or request.user.is_staff):
        return JsonResponse({'error': 'Akses ditolak'}, status=403)

    host = getattr(settings, 'MSSQL_HOST', '') or '(kosong)'
    if host == '(kosong)':
        return JsonResponse({'tcp': 'SKIP', 'info': 'MSSQL_HOST belum diset di .env'})

    reachable, h, port = mssql._tcp_ping(host)
    return JsonResponse({
        'host': h,
        'port': port,
        'tcp':  'BERHASIL' if reachable else 'GAGAL',
        'info': 'Host reachable, lanjut cek /opsis/api/diagnose/' if reachable
                else 'Host tidak reachable — cek IP/hostname, firewall, atau pastikan SQL Server berjalan',
    })


@login_required
def api_history(request, pk):
    """
    JSON: historis MW/MVAR dari PostgreSQL (SnapLive) untuk chart trend.
    ?jam=1|6|24|168  (default 24; 168 = 7 hari)
    Data berasal dari collect_live management command, bukan MSSQL langsung.
    """
    p = get_object_or_404(Pembangkit, pk=pk)
    try:
        jam = int(request.GET.get('jam', 24))
        if jam not in (1, 6, 24, 168):
            jam = 24
    except (ValueError, TypeError):
        jam = 24

    since = timezone.now() - datetime.timedelta(hours=jam)
    snaps = (SnapLive.objects
             .filter(pembangkit=p, waktu__gte=since)
             .order_by('waktu')
             .values('waktu', 'mw', 'mvar', 'frekuensi'))

    return JsonResponse({
        'nama':  p.nama,
        'warna': p.warna,
        'jam':   jam,
        'count': snaps.count(),
        'rows': [
            {
                'timestamp': s['waktu'].astimezone(timezone.get_current_timezone()).strftime('%H:%M'),
                'mw':        s['mw'],
                'mvar':      s['mvar'],
                'frekuensi': s['frekuensi'],
            }
            for s in snaps
        ]
    })


@login_required
def api_diagnose(request):
    """
    Endpoint diagnostik — cek koneksi MSSQL dan query data.
    Buka: /opsis/api/diagnose/
    Hanya superuser atau staff yang bisa akses.
    """
    if not (request.user.is_superuser or request.user.is_staff):
        return JsonResponse({'error': 'Akses ditolak'}, status=403)

    tbl = getattr(settings, 'MSSQL_TABLE', 'dbo.HIS_MEAS_KIT')

    import time as _time

    result = {
        'config': {
            'MSSQL_HOST':  getattr(settings, 'MSSQL_HOST',  '') or '(kosong)',
            'MSSQL_DB':    getattr(settings, 'MSSQL_DB',    '') or '(kosong)',
            'MSSQL_USER':  getattr(settings, 'MSSQL_USER',  '') or '(kosong)',
            'MSSQL_TABLE': tbl,
            'MSSQL_PASS':  '***' if getattr(settings, 'MSSQL_PASS', '') else '(kosong)',
        },
        'timing':     {},
        'koneksi':    None,
        'b1_sample':  [],
        'pembangkit': [],
        'error':      None,
    }

    try:
        t0 = _time.monotonic()
        conn = mssql._get_connection()
        result['timing']['koneksi_ms'] = round((_time.monotonic() - t0) * 1000)
        result['koneksi'] = 'BERHASIL'
        cursor = conn.cursor()

        t0 = _time.monotonic()
        try:
            cursor.execute(
                f"SELECT TOP 20 B1, TIME FROM {tbl} WITH (NOLOCK) ORDER BY TIME DESC"
            )
            rows = cursor.fetchall()
            seen = {}
            for r in rows:
                k = r[0].strip() if r[0] else ''
                if k and k not in seen:
                    seen[k] = str(r[1])
            result['b1_sample'] = [{'B1': k, 'TIME': t} for k, t in seen.items()]
        except Exception as e:
            result['b1_sample'] = f'Error: {e}'
        result['timing']['b1_sample_ms'] = round((_time.monotonic() - t0) * 1000)

        pembangkit_list = _pembangkit_aktif()
        info_map = {p.kode: {'kode': p.kode, 'nama': p.nama} for p in pembangkit_list}

        t0 = _time.monotonic()
        for p in pembangkit_list:
            try:
                cursor.execute(
                    f"""
                    SELECT TOP 1 RTRIM(B1), RTRIM(B3), P, Q, TIME
                    FROM {tbl} WITH (NOLOCK)
                    WHERE B1 LIKE ?
                    ORDER BY TIME DESC
                    """,
                    (p.kode + '%',)
                )
                row = cursor.fetchone()
                if row:
                    info_map[p.kode]['max_time'] = str(row[4])
                    info_map[p.kode]['sample'] = {
                        'B1': row[0], 'B3': row[1], 'P': str(row[2]), 'Q': str(row[3]), 'TIME': str(row[4])
                    }
                else:
                    info_map[p.kode]['max_time'] = 'NULL — tidak ada data (cek apakah kode cocok dengan B1)'
            except Exception as e:
                info_map[p.kode]['error'] = str(e)
        result['timing']['pembangkit_ms'] = round((_time.monotonic() - t0) * 1000)

        result['pembangkit'] = list(info_map.values())
        conn.close()

    except Exception as e:
        result['koneksi'] = 'GAGAL'
        result['error']   = str(e)

    return JsonResponse(result, json_dumps_params={'indent': 2})


@login_required
def rangkuman(request):
    """
    Rangkuman harian sistem: beban puncak, frekuensi, durasi abnormal.
    ?periode=kemarin|minggu|bulan
    """
    from django.db.models import Sum, Avg, Max, Min, Count, Case, When, IntegerField
    from django.db.models.functions import TruncDate
    from collections import defaultdict

    tz_local   = timezone.get_current_timezone()
    today      = timezone.now().astimezone(tz_local).date()
    periode    = request.GET.get('periode', 'kemarin')

    if periode == 'minggu':
        start_date = today - datetime.timedelta(days=7)
        end_date   = today - datetime.timedelta(days=1)
    elif periode == 'bulan':
        start_date = today - datetime.timedelta(days=30)
        end_date   = today - datetime.timedelta(days=1)
    else:
        periode    = 'kemarin'
        start_date = today - datetime.timedelta(days=1)
        end_date   = today - datetime.timedelta(days=1)

    # ── Beban: totalisasi per menit, lalu agregat per hari di Python ─
    per_minute = list(
        SnapLive.objects
        .filter(waktu__date__range=(start_date, end_date))
        .values('waktu')
        .annotate(total_mw=Sum('mw'))
        .order_by('waktu')
    )
    daily_mw = defaultdict(list)
    for r in per_minute:
        if r['total_mw'] is None:
            continue
        day = r['waktu'].astimezone(tz_local).date()
        daily_mw[day].append((r['total_mw'], r['waktu'].astimezone(tz_local)))

    # ── Frekuensi: agregat per hari via ORM (efisien) ────────────────
    freq_daily = (
        SnapFreq.objects
        .filter(waktu__date__range=(start_date, end_date), hz__isnull=False)
        .annotate(tanggal=TruncDate('waktu', tzinfo=tz_local))
        .values('tanggal')
        .annotate(
            avg_hz=Avg('hz'),
            min_hz=Min('hz'),
            max_hz=Max('hz'),
            total_detik=Count('id'),
            abnormal=Count(Case(
                When(hz__lt=49.5, then=1),
                When(hz__gt=50.5, then=1),
                output_field=IntegerField(),
            )),
        )
        .order_by('tanggal')
    )
    freq_map = {r['tanggal']: r for r in freq_daily}

    # ── Gabungkan per hari ────────────────────────────────────────────
    all_dates = sorted(set(list(daily_mw.keys()) + list(freq_map.keys())))
    ringkasan = []
    for day in all_dates:
        entry = {'tanggal': day}

        records = daily_mw.get(day, [])
        if records:
            mw_vals       = [r[0] for r in records]
            pk_idx        = mw_vals.index(max(mw_vals))
            entry.update({
                'puncak_mw':    round(max(mw_vals), 2),
                'puncak_waktu': records[pk_idx][1].strftime('%H:%M'),
                'avg_mw':       round(sum(mw_vals) / len(mw_vals), 2),
                'min_mw':       round(min(mw_vals), 2),
                'data_menit':   len(records),
            })
        else:
            entry.update({'puncak_mw': None, 'puncak_waktu': None,
                          'avg_mw': None, 'min_mw': None, 'data_menit': 0})

        freq = freq_map.get(day, {})
        total_detik = freq.get('total_detik', 0)
        abnormal    = freq.get('abnormal', 0)
        entry.update({
            'avg_hz':         round(freq['avg_hz'], 3) if freq.get('avg_hz') else None,
            'min_hz':         round(freq['min_hz'], 3) if freq.get('min_hz') else None,
            'max_hz':         round(freq['max_hz'], 3) if freq.get('max_hz') else None,
            'abnormal_detik': abnormal,
            'abnormal_menit': round(abnormal / 60, 1) if abnormal else 0,
            'abnormal_pct':   round(abnormal / total_detik * 100, 2) if total_detik else 0,
        })
        ringkasan.append(entry)

    # ── Ringkasan keseluruhan periode ─────────────────────────────────
    puncak_list  = [r['puncak_mw']  for r in ringkasan if r.get('puncak_mw')]
    avg_mw_list  = [r['avg_mw']     for r in ringkasan if r.get('avg_mw')]
    avg_hz_list  = [r['avg_hz']     for r in ringkasan if r.get('avg_hz')]
    min_hz_list  = [r['min_hz']     for r in ringkasan if r.get('min_hz')]
    total_abnormal = sum(r.get('abnormal_detik', 0) for r in ringkasan)
    summary = {
        'puncak_mw':      max(puncak_list)                              if puncak_list else None,
        'avg_beban':      round(sum(avg_mw_list)/len(avg_mw_list), 2)  if avg_mw_list else None,
        'avg_hz':         round(sum(avg_hz_list)/len(avg_hz_list), 3)  if avg_hz_list else None,
        'min_hz':         min(min_hz_list)                              if min_hz_list else None,
        'total_abnormal': total_abnormal,
        'abnormal_menit': round(total_abnormal / 60, 1),
        'total_hari':     len(ringkasan),
    }

    # ── Beban Trafo snapshot (realtime) ─────────────────────────────
    trafo_rows = _trafo_aktif_saja(mssql.get_beban_trafo())
    _trafo_grouped = {}
    for r in trafo_rows:
        site = r['site'] or 'Unknown'
        _trafo_grouped.setdefault(site, []).append(r)
    trafo_sites = []
    trafo_total_mw = 0.0
    for site in sorted(_trafo_grouped):
        lst   = _trafo_grouped[site]
        total = round(sum(abs(r['p']) for r in lst if r['p'] is not None), 2)
        p_vals = [abs(r['p']) for r in lst if r['p'] is not None]
        trafo_sites.append({
            'site':   site,
            'tlist':  lst,
            'total':  total,
            'avg':    round(sum(p_vals) / len(p_vals), 2) if p_vals else None,
            'max':    round(max(p_vals), 2) if p_vals else None,
            'min':    round(min(p_vals), 2) if p_vals else None,
            'count':  len(lst),
        })
        trafo_total_mw += total
    trafo_total_mw = round(trafo_total_mw, 2)

    pembangkit_list = _pembangkit_aktif()
    return render(request, 'opsis/rangkuman.html', {
        'pembangkit_list':  pembangkit_list,
        'ringkasan':        ringkasan,
        'summary':          summary,
        'periode':          periode,
        'start_date':       start_date,
        'end_date':         end_date,
        'trafo_sites':    trafo_sites,
        'trafo_total_mw': trafo_total_mw,
    })



# ── Beban Trafo ───────────────────────────────────────────────────────────────

@login_required
def beban_trafo(request):
    """Halaman monitoring beban trafo distribusi dari ALL_TRANS_DATA (BAY TRF52%/TRF42%)."""
    rows = _trafo_aktif_saja(mssql.get_beban_trafo())

    # Kelompokkan per GI (site)
    grouped = {}
    for r in rows:
        site = r['site'] or 'Unknown'
        grouped.setdefault(site, []).append(r)

    # Hitung total P per site — abs() agar nilai negatif tidak kurangi total
    site_totals = {
        site: round(sum(abs(r['p']) for r in trafo_list if r['p'] is not None), 2)
        for site, trafo_list in grouped.items()
    }

    total_mw = round(sum(site_totals.values()), 2)

    return render(request, 'opsis/beban_trafo.html', {
        'pembangkit_list': _pembangkit_aktif(),
        'grouped':         grouped,
        'site_totals':     site_totals,
        'total_mw':        total_mw,
        'terputus':        not mssql.is_reachable(),
    })


KTT_NAME_MAP = {
    'IND_ANTAM':  'ANTAM',
    'IND_CERIA':  'CERIA',
    'IND_TNASA':  'TONASA SEMEN',
    'IND_BSOWA':  'BOSOWA SEMEN',
    'IND_SMLTR4': 'HUADI 4',
    'IND_INDOF':  'INDOFOOD',
    'IND_HUADI':  'HUADI 1',
    'IND_HUADI2': 'HUADI 2',
    'IND_HUADI3': 'HUADI 3',
    'IND_SMLTR5': 'HUADI 5',
}


def _split_ktt(rows):
    """Pisahkan IND_TOTAL dari baris konsumen. Return (consumers, total_mw)."""
    consumers = []
    total_mw  = None
    for r in rows:
        if r['analog'].upper() == 'IND_TOTAL':
            total_mw = r['value']
        else:
            r['nama'] = KTT_NAME_MAP.get(r['analog'].upper(), r['analog'])
            consumers.append(r)
    # Fallback: hitung manual jika IND_TOTAL tidak ada di data
    if total_mw is None:
        total_mw = sum(r['value'] for r in consumers if r['value'] is not None)
    return consumers, round(total_mw, 2) if total_mw is not None else 0


@login_required
def beban_ktt(request):
    """Halaman monitoring beban KTT (konsumen tegangan tinggi) dari IND_LOAD."""
    all_rows = mssql.get_beban_ktt()
    rows, total_mw = _split_ktt(all_rows)
    return render(request, 'opsis/beban_ktt.html', {
        'pembangkit_list': _pembangkit_aktif(),
        'rows':            rows,
        'total_mw':        total_mw,
        'jumlah':          len(rows),
        'terputus':        not mssql.is_reachable(),
    })


@login_required
def api_beban_ktt(request):
    """API JSON untuk refresh otomatis halaman beban KTT."""
    all_rows = mssql.get_beban_ktt()
    rows, total_mw = _split_ktt(all_rows)
    return JsonResponse({
        'rows':     rows,
        'total_mw': total_mw,
        'jumlah':   len(rows),
        'terputus': not mssql.is_reachable(),
    })


@login_required
def api_beban_trafo(request):
    """API JSON untuk refresh otomatis halaman beban trafo."""
    rows = _trafo_aktif_saja(mssql.get_beban_trafo())
    grouped = {}
    for r in rows:
        site = r['site'] or 'Unknown'
        grouped.setdefault(site, []).append(r)
    site_totals = {
        site: round(sum(abs(r['p']) for r in lst if r['p'] is not None), 2)
        for site, lst in grouped.items()
    }
    return JsonResponse({
        'rows':        rows,
        'site_totals': site_totals,
        'total_mw':    round(sum(site_totals.values()), 2),
        'terputus':    not mssql.is_reachable(),
    })


def _beban_trafo_chart_data():
    """
    Bangun payload chart 24 jam daya aktif (P) trafo distribusi — satu chart
    per GI (site) berisi satu garis per trafo di GI tersebut, semua trafo di
    site yang sama berbagi sumbu waktu (label) yang sama.

    Sumber: PostgreSQL (SnapTrafo), diisi tiap menit oleh management command
    'collect_trafo'. Tidak ada fallback ke MSSQL histori — ALL_TRANS_DATA
    cuma snapshot realtime, bukan tabel historian seperti HIS_MEAS_KIT, jadi
    PostgreSQL satu-satunya sumber data historis di sini.

    Trafo dibatasi ke BAY TRF52%/TRF42% (distribusi) saja — registry Trafo
    dipakai bersama dengan halaman Beban Trafo IBT (BAY TRF65%/TRF54%), jadi
    tanpa filter ini trafo IBT ikut nongol di sini walau tidak pernah
    kebagian data dari collect_trafo.
    """
    from django.db.models import Q

    tz_local = timezone.get_current_timezone()
    today    = timezone.now().astimezone(tz_local).date()

    trafo_list = list(
        Trafo.objects.filter(aktif=True)
        .filter(Q(bay__istartswith='TRF52') | Q(bay__istartswith='TRF42'))
    )
    snaps = (SnapTrafo.objects
             .filter(trafo__in=trafo_list, waktu__date=today)
             .order_by('waktu')
             .values('trafo_id', 'waktu', 'p'))

    per_trafo = {}
    count = 0
    for s in snaps:
        per_trafo.setdefault(s['trafo_id'], {})[s['waktu']] = s['p']
        count += 1

    by_site = {}
    for t in trafo_list:
        by_site.setdefault(t.site or 'Unknown', []).append(t)

    sites = []
    for site, trafos in sorted(by_site.items()):
        waktu_set = sorted({w for t in trafos for w in per_trafo.get(t.id, {})})
        labels = [w.astimezone(tz_local).strftime('%H:%M') for w in waktu_set]
        series = []
        for t in trafos:
            vals = per_trafo.get(t.id, {})
            series.append({
                'id':  t.id,
                'bay': t.bay,
                'p': [round(vals[w], 2) if vals.get(w) is not None else None for w in waktu_set],
            })
        sites.append({'site': site, 'labels': labels, 'trafos': series})

    return {'sites': sites, 'count': count}


@login_required
def beban_trafo_chart(request):
    """Halaman chart 24 jam beban trafo distribusi — P & Q per trafo, per GI."""
    return render(request, 'opsis/beban_trafo_chart.html', {
        'pembangkit_list': _pembangkit_aktif(),
        'chart_data':      _beban_trafo_chart_data(),
    })


@login_required
def api_beban_trafo_chart(request):
    """API JSON untuk refresh otomatis chart 24 jam beban trafo distribusi."""
    return JsonResponse(_beban_trafo_chart_data())


# ── Beban Trafo IBT ─────────────────────────────────────────────────────────

@login_required
def beban_trafo_ibt(request):
    """Halaman monitoring beban trafo IBT dari ALL_TRANS_DATA (BAY TRF65%/TRF54%)."""
    rows = _trafo_aktif_saja(mssql.get_beban_trafo_ibt())

    grouped = {}
    for r in rows:
        site = r['site'] or 'Unknown'
        grouped.setdefault(site, []).append(r)

    site_totals = {
        site: round(sum(abs(r['p']) for r in trafo_list if r['p'] is not None), 2)
        for site, trafo_list in grouped.items()
    }
    total_mw = round(sum(site_totals.values()), 2)

    return render(request, 'opsis/beban_trafo_ibt.html', {
        'pembangkit_list': _pembangkit_aktif(),
        'grouped':         grouped,
        'site_totals':     site_totals,
        'total_mw':        total_mw,
        'terputus':        not mssql.is_reachable(),
    })


@login_required
def api_beban_trafo_ibt(request):
    """API JSON untuk refresh otomatis halaman/chart beban trafo IBT."""
    rows = _trafo_aktif_saja(mssql.get_beban_trafo_ibt())
    grouped = {}
    for r in rows:
        site = r['site'] or 'Unknown'
        grouped.setdefault(site, []).append(r)
    site_totals = {
        site: round(sum(abs(r['p']) for r in lst if r['p'] is not None), 2)
        for site, lst in grouped.items()
    }
    return JsonResponse({
        'rows':        rows,
        'site_totals': site_totals,
        'total_mw':    round(sum(site_totals.values()), 2),
        'terputus':    not mssql.is_reachable(),
    })


def _beban_trafo_ibt_chart_data():
    """
    Bangun payload chart 24 jam daya aktif (P) trafo IBT — satu chart per GI
    (site) berisi satu garis per trafo di GI tersebut. Sama persis dengan
    _beban_trafo_chart_data() (trafo distribusi), cuma filter BAY-nya beda
    (TRF65%/TRF54%) dan SUMBER DATANYA SAMA (SnapTrafo, diisi 'collect_trafo').

    Nilai P disimpan & ditampilkan APA ADANYA (bisa negatif — arah aliran
    daya lewat IBT dua arah, jadi tanda minus bermakna, TIDAK di-abs()-kan)
    — beda dengan site_totals di beban_trafo_ibt()/api_beban_trafo_ibt()
    (tabel realtime) yang sengaja pakai abs() krn itu total MAGNITUDE beban,
    bukan tren per-trafo.
    """
    from django.db.models import Q

    tz_local = timezone.get_current_timezone()
    today    = timezone.now().astimezone(tz_local).date()

    trafo_list = list(
        Trafo.objects.filter(aktif=True)
        .filter(Q(bay__istartswith='TRF65') | Q(bay__istartswith='TRF54'))
    )
    snaps = (SnapTrafo.objects
             .filter(trafo__in=trafo_list, waktu__date=today)
             .order_by('waktu')
             .values('trafo_id', 'waktu', 'p'))

    per_trafo = {}
    count = 0
    for s in snaps:
        per_trafo.setdefault(s['trafo_id'], {})[s['waktu']] = s['p']
        count += 1

    by_site = {}
    for t in trafo_list:
        by_site.setdefault(t.site or 'Unknown', []).append(t)

    sites = []
    for site, trafos in sorted(by_site.items()):
        waktu_set = sorted({w for t in trafos for w in per_trafo.get(t.id, {})})
        labels = [w.astimezone(tz_local).strftime('%H:%M') for w in waktu_set]
        series = []
        for t in trafos:
            vals = per_trafo.get(t.id, {})
            series.append({
                'id':  t.id,
                'bay': t.bay,
                'p': [round(vals[w], 2) if vals.get(w) is not None else None for w in waktu_set],
            })
        sites.append({'site': site, 'labels': labels, 'trafos': series})

    return {'sites': sites, 'count': count}


@login_required
def beban_trafo_ibt_chart(request):
    """Halaman chart 24 jam beban trafo IBT — P per trafo, per GI, nilai asli (bisa negatif)."""
    return render(request, 'opsis/beban_trafo_ibt_chart.html', {
        'pembangkit_list': _pembangkit_aktif(),
        'chart_data':      _beban_trafo_ibt_chart_data(),
    })


@login_required
def api_beban_trafo_ibt_chart(request):
    """API JSON untuk refresh otomatis chart 24 jam beban trafo IBT."""
    return JsonResponse(_beban_trafo_ibt_chart_data())


# ── Analitik Prediksi Beban (akurasi model + prediksi puncak besok) ─────────

@login_required
def prediksi_beban(request):
    """Halaman analitik model prediksi: akurasi vs realisasi & prediksi puncak besok."""
    return render(request, 'opsis/prediksi_beban.html', {
        'pembangkit_list': _pembangkit_aktif(),
    })


@login_required
def api_prediksi_beban(request):
    """
    API JSON utk halaman Analitik Prediksi Beban:
    - akurasi: evaluasi walk-forward one-step-ahead 7 hari terakhir vs realisasi
      SnapLive (lihat opsis.forecast.evaluate_accuracy()).
    - besok: prediksi puncak siang (12:00) & malam (18:30) besok, direct
      forecast dari anchor asli (lihat opsis.forecast.predict_besok_puncak()).
    """
    akurasi = forecast.evaluate_accuracy(days=7)
    besok = forecast.predict_besok_puncak()
    return JsonResponse({'akurasi': akurasi, 'besok': besok})
