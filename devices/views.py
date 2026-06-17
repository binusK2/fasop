from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator
from fasop.hashids_helper import encode as _hid
from django.contrib.auth.decorators import login_required
from devices.permissions import (
    require_can_delete, require_can_edit, require_can_manage_lokasi,
    can_delete, can_edit, can_manage_lokasi, is_viewer_only
)
from .models import Device, DeviceType, Icon, SiteLocation, DeviceLog, DeviceEvent, Branch
from .forms import DeviceForm, IconForm
from django.db.models import Count, Q
from django.db.models.functions import TruncMonth, Lower, Trim, ExtractYear, ExtractMonth
from maintenance.models import Maintenance
from django.utils.timezone import now
from django.http import HttpResponse, JsonResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date as date_type
from dateutil.relativedelta import relativedelta
import json
from .device_schema import DEVICE_SCHEMA
from auditlog.utils import log_action as _audit


@login_required
def device_list(request):
    SESSION_KEY = f'devlist_filter_{request.user.pk}'

    # Reset: hapus session filter, redirect balik
    if 'reset' in request.GET:
        request.session.pop(SESSION_KEY, None)
        jenis_param = request.GET.get('jenis', '')
        url = '/devices/'
        if jenis_param:
            url += f'?jenis={jenis_param}'
        return redirect(url)

    # Jika form filter di-submit → pakai GET dan simpan ke session
    if '_filter' in request.GET:
        saved = {
            'q':      request.GET.get('q', ''),
            'lokasi': request.GET.get('lokasi', ''),
            'status': request.GET.get('status_operasi', ''),
            'merk':   request.GET.get('merk', ''),
            'branch': request.GET.get('branch', ''),
        }
        request.session[SESSION_KEY] = saved
        request.session.modified = True
    else:
        saved = request.session.get(SESSION_KEY, {})

    jenis_id       = request.GET.get('jenis')
    search         = saved.get('q', '')
    lokasi         = saved.get('lokasi', '')
    status_operasi = saved.get('status', '')
    merk           = saved.get('merk', '')
    branch_id      = saved.get('branch', '')
    sort           = request.GET.get('sort', 'nama')
    direction      = request.GET.get('dir', 'asc')

    # Cek apakah filter jenis yang dipilih adalah VM SCADA
    _is_vm_jenis = (
        jenis_id and
        DeviceType.objects.filter(pk=jenis_id, name__iexact='VM SCADA').exists()
    )
    if _is_vm_jenis:
        # Untuk VM SCADA: tampilkan VM (host terisi), bukan aset fisik
        devices = Device.objects.filter(is_deleted=False, host__isnull=False)
    else:
        # Semua jenis lain: hanya aset fisik (bukan VM)
        devices = Device.objects.filter(is_deleted=False, host__isnull=True)

    if jenis_id:
        devices = devices.filter(jenis_id=jenis_id)

    if search:
        devices = devices.filter(
            Q(nama__icontains=search) |
            Q(ip_address__icontains=search)
        )

    if lokasi:
        devices = devices.filter(lokasi=lokasi)

    if status_operasi:
        devices = devices.filter(status_operasi=status_operasi)

    if merk:
        devices = devices.filter(merk__iexact=merk)

    if branch_id:
        # Filter lokasi yang masuk ke branch ini
        _branch_lokasi = SiteLocation.objects.filter(branch_id=branch_id).values_list('nama', flat=True)
        devices = devices.filter(lokasi__in=_branch_lokasi)

    # Sorting
    SORT_FIELDS = {
        'nama': 'nama',
        'jenis': 'jenis__name',
        'merk': 'merk',
        'ip': 'ip_address',
        'lokasi': 'lokasi',
        'serial': 'serial_number',
    }
    order_field = SORT_FIELDS.get(sort, 'nama')
    if direction == 'desc':
        order_field = '-' + order_field
    devices = devices.select_related('jenis').order_by(order_field)

    paginator  = Paginator(devices, 50)
    page_num   = request.GET.get('page', 1)
    page_obj   = paginator.get_page(page_num)

    lokasi_list = (
        Device.objects
        .filter(is_deleted=False)
        .exclude(lokasi__isnull=True)
        .exclude(lokasi__exact='')
        .exclude(lokasi__iexact='none')
        .annotate(lokasi_clean=Trim('lokasi'))
        .values_list('lokasi_clean', flat=True)
        .distinct()
        .order_by('lokasi_clean')
    )

    # Merk list — jika jenis dipilih, filter merk sesuai jenis itu
    _merk_qs = Device.objects.filter(is_deleted=False, host__isnull=True).exclude(merk__isnull=True).exclude(merk__exact='')
    if jenis_id:
        _merk_qs = _merk_qs.filter(jenis_id=jenis_id)
    merk_list = _merk_qs.values_list('merk', flat=True).distinct().order_by('merk')

    branch_list   = Branch.objects.all()
    filter_active = bool(search or lokasi or status_operasi or merk or branch_id)

    return render(request, 'devices/device_list.html', {
        'devices':          page_obj,
        'page_obj':         page_obj,
        'paginator':        paginator,
        'search':           search,
        'selected_jenis':   jenis_id,
        'lokasi_list':      lokasi_list,
        'selected_lokasi':  lokasi,
        'selected_status':  status_operasi,
        'selected_merk':    merk,
        'merk_list':        merk_list,
        'branch_list':      branch_list,
        'selected_branch':  branch_id,
        'filter_active':    filter_active,
        'current_sort':     sort,
        'current_dir':      direction,
    })


@login_required
@require_can_edit
def device_create(request):
    host_pk = request.GET.get('host') or request.POST.get('host_prefill')
    if request.method == 'POST':
        form = DeviceForm(request.POST, request.FILES)
        if form.is_valid():
            device = form.save(commit=False)
            spesifikasi_raw = request.POST.get('spesifikasi_json', '{}')
            try:
                device.spesifikasi = json.loads(spesifikasi_raw)
            except (json.JSONDecodeError, ValueError):
                device.spesifikasi = {}
            device.created_by = request.user
            device.save()

            # ── Simpan komponen dari form ──────────────────────────
            komponen_json = request.POST.get('komponen_json', '[]')
            try:
                komponen_list = json.loads(komponen_json)
            except (json.JSONDecodeError, ValueError):
                komponen_list = []

            from devices.models_komponen import DeviceComponent
            from devices.views_komponen import _resolve_tipe
            for k in komponen_list:
                if not k.get('nama'):
                    continue
                DeviceComponent.objects.create(
                    device=device,
                    nama=k.get('nama', '').strip(),
                    tipe_komponen=_resolve_tipe(k.get('tipe_komponen')),
                    posisi=k.get('posisi', '').strip(),
                    merk=k.get('merk', '').strip(),
                    serial_number=k.get('serial_number', '').strip(),
                    status=k.get('status', 'terpasang'),
                    keterangan=k.get('keterangan', '').strip(),
                    created_by=request.user,
                )

            # Audit log
            from devices.device_audit import log_create
            log_create(device, request.user)
            _audit(request, 'create', 'devices', 'Peralatan',
                   device.pk, device.nama,
                   f'{device.jenis} | {device.lokasi}')
            return redirect('device_view', pk=device.pk)
    else:
        initial = {}
        if host_pk:
            try:
                host_device = Device.objects.get(pk=host_pk, is_deleted=False)
                initial['host'] = host_device
            except Device.DoesNotExist:
                pass
        form = DeviceForm(initial=initial)
    from devices.views_komponen import get_tipe_grouped_json
    # Tentukan jenis DeviceType untuk "VM SCADA" agar JS bisa pre-pilih
    try:
        vm_jenis = DeviceType.objects.get(name__iexact='VM SCADA')
        vm_jenis_id = vm_jenis.pk
    except DeviceType.DoesNotExist:
        vm_jenis_id = None
    return render(request, 'devices/device_form.html', {
        'form': form,
        'is_edit': False,
        'device_schema_json': json.dumps(DEVICE_SCHEMA),
        'tipe_komponen_json': get_tipe_grouped_json(),
        'prefill_host_pk': host_pk,
        'vm_jenis_id': vm_jenis_id,
    })


@login_required
@require_can_edit
def device_update(request, pk):
    device = get_object_or_404(Device, pk=pk)
    if request.method == 'POST':
        # Snapshot sebelum perubahan
        from devices.device_audit import log_edit, TRACKED_FIELDS
        import copy
        device_before = copy.copy(device)

        form = DeviceForm(request.POST, request.FILES, instance=device)
        if form.is_valid():
            dev = form.save(commit=False)
            spesifikasi_raw = request.POST.get('spesifikasi_json', '{}')
            try:
                dev.spesifikasi = json.loads(spesifikasi_raw)
            except (json.JSONDecodeError, ValueError):
                dev.spesifikasi = {}
            dev.save()

            # ── Simpan komponen baru dari form ────────────────────
            komponen_json = request.POST.get('komponen_json', '[]')
            try:
                komponen_list = json.loads(komponen_json)
            except (json.JSONDecodeError, ValueError):
                komponen_list = []

            from devices.models_komponen import DeviceComponent
            from devices.views_komponen import _resolve_tipe
            for k in komponen_list:
                if not k.get('nama'):
                    continue
                DeviceComponent.objects.create(
                    device=device,
                    nama=k.get('nama', '').strip(),
                    tipe_komponen=_resolve_tipe(k.get('tipe_komponen')),
                    posisi=k.get('posisi', '').strip(),
                    merk=k.get('merk', '').strip(),
                    serial_number=k.get('serial_number', '').strip(),
                    status=k.get('status', 'terpasang'),
                    keterangan=k.get('keterangan', '').strip(),
                    created_by=request.user,
                )

            # Audit log — bandingkan sebelum vs sesudah
            log_edit(device_before, dev, request.user)
            _audit(request, 'update', 'devices', 'Peralatan',
                   dev.pk, dev.nama,
                   f'{dev.jenis} | {dev.lokasi}')
            return redirect('device_view', pk=device.pk)
    else:
        form = DeviceForm(instance=device)
    from devices.views_komponen import get_tipe_grouped_json
    return render(request, 'devices/device_form.html', {
        'form': form,
        'is_edit': True,
        'device': device,
        'device_schema_json': json.dumps(DEVICE_SCHEMA),
        'existing_spesifikasi': json.dumps(device.spesifikasi or {}),
        'tipe_komponen_json': get_tipe_grouped_json(),
    })


@login_required
@require_can_delete
def device_delete(request, pk):
    device = get_object_or_404(Device, pk=pk)
    from devices.device_audit import log_delete
    log_delete(device, request.user)
    _audit(request, 'delete', 'devices', 'Peralatan',
           device.pk, device.nama,
           f'{device.jenis} | {device.lokasi}')
    jenis_id = device.jenis_id
    device.is_deleted = True
    device.deleted_by = request.user
    device.save()
    from django.urls import reverse
    url = reverse('device_list')
    if jenis_id:
        url += f'?jenis={jenis_id}'
    return redirect(url)


@login_required
def dashboard(request):
    # ── Fork ke dashboard operator ───────────────────────────────
    if not request.user.is_superuser:
        try:
            role = request.user.profile.role
        except Exception:
            role = ''
        if role == 'operator':
            return _dashboard_operator(request)

    # ── Dashboard normal (teknisi / AM / superuser) ──────────────
    # VM (host__isnull=False) tidak dihitung sebagai aset fisik
    _asset_qs = Device.objects.filter(is_deleted=False, host__isnull=True)
    total_devices  = _asset_qs.count()
    dev_operasi    = _asset_qs.filter(status_operasi='operasi').count()
    dev_tdk_operasi= _asset_qs.filter(status_operasi='tidak_operasi').count()

    # Hitung per jenis + persen untuk progress bar
    device_by_type_qs = (
        _asset_qs
        .values('jenis__name', 'jenis__id')
        .annotate(total=Count('id'))
        .order_by('jenis__name')
    )

    # tambah pct untuk progress bar
    device_by_type = []
    for d in device_by_type_qs:
        pct = round((d['total'] / total_devices * 100)) if total_devices else 0
        device_by_type.append({**d, 'pct': pct})

    # Status operasi per jenis → untuk pie chart di dashboard
    import json as _json
    _status_qs = (
        _asset_qs
        .values('jenis__name', 'status_operasi')
        .annotate(total=Count('id'))
        .order_by('jenis__name')
    )
    _status_map = {}
    for row in _status_qs:
        jenis = row['jenis__name'] or 'Lainnya'
        if jenis not in _status_map:
            _status_map[jenis] = {'operasi': 0, 'tidak_operasi': 0}
        _status_map[jenis][row['status_operasi']] = row['total']
    device_status_by_type_json = _json.dumps(_status_map)

    # ── Distribusi peralatan per kelompok (Telkom / SCADA / Prosis) ─
    _KELOMPOK_CFG = [
        {
            'key': 'telkom', 'label': 'Telekomunikasi',
            'icon': 'bi-broadcast-pin', 'color': '#3b82f6', 'bg': '#eff6ff',
            'types': {
                'router', 'switch', 'teleproteksi', 'roip', 'voip',
                'multiplexer', 'plc', 'radio', 'server telkom',
                'master clock', 'ht', 'pheriperal telkom', 'peripheral telkom',
                'microwave', 'repeater', 'tower',
            },
        },
        {
            'key': 'scada', 'label': 'SCADA',
            'icon': 'bi-diagram-3', 'color': '#10b981', 'bg': '#f0fdf4',
            'types': {
                'rtu', 'sas', 'ups', 'server scada', 'vm scada', 'ied bcu',
                'clock server', 'serial server', 'router sas', 'switch sas',
                'inverter sas', 'pheriperal scada', 'peripheral scada', 'gps',
            },
        },
        {
            'key': 'prosis', 'label': 'Proteksi Sistem',
            'icon': 'bi-shield-check', 'color': '#f59e0b', 'bg': '#fffbeb',
            'types': {
                'defense scheme', 'rele defense scheme', 'dfr',
                'master trip', 'ufls', 'server prosis',
            },
        },
    ]
    # Batch: total devices per jenis
    _type_agg = {}
    for row in _asset_qs.values('jenis__name').annotate(total=Count('id')):
        jname = (row['jenis__name'] or 'Lainnya').strip()
        _type_agg.setdefault(jname, {'total': 0, 'maintained': 0})
        _type_agg[jname]['total'] = row['total']
    # Batch: devices with at least one Done maintenance per jenis
    for row in (_asset_qs.filter(maintenance__status='Done')
                .values('jenis__name').annotate(maintained=Count('id', distinct=True))):
        jname = (row['jenis__name'] or 'Lainnya').strip()
        if jname in _type_agg:
            _type_agg[jname]['maintained'] = row['maintained']

    kelompok_data = []
    for _kg in _KELOMPOK_CFG:
        _entries = []
        for _jname, _st in _type_agg.items():
            if _jname.lower().strip() in _kg['types']:
                _entries.append({
                    'name': _jname,
                    'total': _st['total'],
                    'maintained': _st['maintained'],
                })
        _entries.sort(key=lambda x: -x['total'])
        kelompok_data.append({
            'key':        _kg['key'],
            'label':      _kg['label'],
            'icon':       _kg['icon'],
            'color':      _kg['color'],
            'bg':         _kg['bg'],
            'total':      sum(e['total'] for e in _entries),
            'maintained': sum(e['maintained'] for e in _entries),
            'types':      _entries,
        })
    kelompok_json = _json.dumps(kelompok_data)

    total_maintenance = Maintenance.objects.count()
    maintenance_open = Maintenance.objects.filter(status='Open').count()
    maintenance_done = Maintenance.objects.filter(status='Done').count()
    belum_ttd = Maintenance.objects.filter(status='Done', signed_by__isnull=True).count()

    # Selalu tampilkan 6 bulan terakhir, isi 0 jika tidak ada data
    today = date_type.today()
    months_6 = [today.replace(day=1) - relativedelta(months=i) for i in range(5, -1, -1)]
    counts_qs = (
        Maintenance.objects
        .annotate(month=TruncMonth('date'))
        .values('month')
        .annotate(total=Count('id'))
    )
    counts_map = {m['month'].date().replace(day=1): m['total'] for m in counts_qs}
    maintenance_by_month = [
        {'month_label': m.strftime('%b %Y'), 'total': counts_map.get(m, 0)}
        for m in months_6
    ]

    # Maintenance open terbaru (5 record) — tangkap status 'Open' maupun blank/null
    recent_open_maintenance = (
        Maintenance.objects
        .exclude(status='Done')
        .select_related('device', 'device__jenis')
        .order_by('-date')[:5]
    )

    # Maintenance per lokasi (top 8)
    maintenance_by_lokasi = (
        Maintenance.objects
        .exclude(device__lokasi__isnull=True)
        .exclude(device__lokasi__exact='')
        .values('device__lokasi')
        .annotate(total=Count('id'), open=Count('id', filter=Q(status='Open')), done=Count('id', filter=Q(status='Done')))
        .order_by('-total')[:8]
    )
    max_lokasi = maintenance_by_lokasi[0]['total'] if maintenance_by_lokasi else 1

    # ── Gangguan stats ───────────────────────────────────────────
    from gangguan.models import Gangguan
    gangguan_total      = Gangguan.objects.count()
    gangguan_open       = Gangguan.objects.filter(status='open').count()
    gangguan_progress   = Gangguan.objects.filter(status='in_progress').count()
    gangguan_resolved   = Gangguan.objects.filter(status='resolved').count()
    gangguan_closed     = Gangguan.objects.filter(status='closed').count()

    # Trend gangguan 6 bulan (berdasarkan tanggal_gangguan)
    gangguan_counts_qs = (
        Gangguan.objects
        .annotate(month=TruncMonth('tanggal_gangguan'))
        .values('month')
        .annotate(total=Count('id'))
    )
    gangguan_counts_map = {g['month'].date().replace(day=1): g['total'] for g in gangguan_counts_qs}
    gangguan_by_month = [
        {'month_label': m.strftime('%b %Y'), 'total': gangguan_counts_map.get(m, 0)}
        for m in months_6
    ]

    # Gangguan terbaru open/in_progress
    recent_gangguan = (
        Gangguan.objects
        .filter(status__in=['open', 'in_progress'])
        .select_related('peralatan', 'created_by')
        .order_by('-tanggal_gangguan')[:5]
    )

    # ── Health Index summary — baca dari HISnapshot, hindari kalkulasi per-device ─
    from health_index.models import HISnapshot
    _snap_this  = {s['device_id']: s['score'] for s in HISnapshot.objects.filter(
        bulan=today.month, tahun=today.year).values('device_id', 'score')}
    _prev       = (today.replace(day=1) - relativedelta(months=1))
    _snap_prev  = {s['device_id']: s['score'] for s in HISnapshot.objects.filter(
        bulan=_prev.month, tahun=_prev.year).values('device_id', 'score')}
    _snap_scores = {**_snap_prev, **_snap_this}  # this month takes precedence

    hi_summary = {'sangat_baik': 0, 'baik': 0, 'cukup': 0, 'buruk': 0, 'kritis': 0}
    hi_buruk_list = []
    for dev in Device.objects.filter(is_deleted=False).select_related('jenis'):
        s = _snap_scores.get(dev.id, 75)
        if s >= 85:   hi_summary['sangat_baik'] += 1
        elif s >= 70: hi_summary['baik'] += 1
        elif s >= 50: hi_summary['cukup'] += 1
        elif s >= 25:
            hi_summary['buruk'] += 1
            hi_buruk_list.append({'device': dev, 'hi': {'score': s}})
        else:
            hi_summary['kritis'] += 1
            hi_buruk_list.append({'device': dev, 'hi': {'score': s}})
    hi_buruk_list.sort(key=lambda x: x['hi']['score'])
    hi_buruk_list = hi_buruk_list[:5]
    hi_summary_json = _json.dumps(hi_summary)

    # ── Common Enemy aktif ────────────────────────────────────────
    from common_enemy.models import CommonEnemy
    ce_open_list = (
        CommonEnemy.objects
        .filter(status__in=['open', 'in_progress'])
        .select_related('peralatan', 'sub_kategori')
        .order_by('tingkat_keparahan', '-tanggal_laporan')[:5]
    )
    ce_open_total = CommonEnemy.objects.filter(status__in=['open', 'in_progress']).count()

    # ── Gangguan breakdown by kategori & severity ─────────────────
    gangguan_by_kategori_qs = (
        Gangguan.objects
        .values('kategori')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    _kat_labels = {
        'perangkat': 'Perangkat/HW', 'jaringan': 'Jaringan',
        'daya': 'Daya/Power', 'software': 'Software',
        'eksternal': 'Eksternal', 'lainnya': 'Lainnya',
    }
    gangguan_kategori_json = _json.dumps([
        {'label': _kat_labels.get(g['kategori'], g['kategori']), 'total': g['total']}
        for g in gangguan_by_kategori_qs
    ])

    gangguan_by_severity_qs = (
        Gangguan.objects
        .values('tingkat_keparahan')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    _sev_order = ['kritis', 'tinggi', 'sedang', 'rendah']
    _sev_labels = {'kritis': 'Kritis', 'tinggi': 'Tinggi', 'sedang': 'Sedang', 'rendah': 'Rendah'}
    _sev_map = {g['tingkat_keparahan']: g['total'] for g in gangguan_by_severity_qs}
    gangguan_severity_json = _json.dumps([
        {'label': _sev_labels[s], 'total': _sev_map.get(s, 0)}
        for s in _sev_order
    ])

    # Gangguan open per severity (untuk card badges)
    gangguan_open_kritis = Gangguan.objects.filter(status__in=['open', 'in_progress'], tingkat_keparahan='kritis').count()

    # ── Maintenance stacked (preventive vs corrective per bulan) ──
    _maint_type_qs = (
        Maintenance.objects
        .annotate(month=TruncMonth('date'))
        .values('month', 'maintenance_type')
        .annotate(total=Count('id'))
    )
    _maint_type_map = {}
    for _m in _maint_type_qs:
        _mo = _m['month'].date().replace(day=1)
        _t = _m['maintenance_type'] or 'Lainnya'
        if _mo not in _maint_type_map:
            _maint_type_map[_mo] = {}
        _maint_type_map[_mo][_t] = _m['total']
    maintenance_stacked_json = _json.dumps({
        'labels': [m.strftime('%b %Y') for m in months_6],
        'preventive': [_maint_type_map.get(m, {}).get('Preventive', 0) for m in months_6],
        'corrective':  [_maint_type_map.get(m, {}).get('Corrective', 0) for m in months_6],
    })

    # ── Device per lokasi top 10 ──────────────────────────────────
    _dev_lokasi_qs = (
        Device.objects
        .filter(is_deleted=False)
        .exclude(lokasi__isnull=True).exclude(lokasi__exact='')
        .values('lokasi')
        .annotate(
            total=Count('id'),
            operasi=Count('id', filter=Q(status_operasi='operasi')),
            tidak_operasi=Count('id', filter=Q(status_operasi='tidak_operasi')),
        )
        .order_by('-total')[:10]
    )
    device_by_lokasi_json = _json.dumps([
        {
            'lokasi': r['lokasi'], 'total': r['total'],
            'operasi': r['operasi'], 'tidak_operasi': r['tidak_operasi'],
        }
        for r in _dev_lokasi_qs
    ])

    # ── Jadwal kunjungan terdekat ─────────────────────────────────
    try:
        from jadwal.models import JadwalKunjungan
        from datetime import date as _date
        today_d = _date.today()
        jadwal_terdekat = (
            JadwalKunjungan.objects
            .exclude(status='done')
            .filter(
                tahun_rencana=today_d.year,
                bulan_rencana__gte=today_d.month
            )
            .order_by('bulan_rencana', 'lokasi')[:5]
        )
        # Batch progress for these ≤5 jadwals
        from django.db.models import Count as _JCount
        from django.db.models.functions import Upper as _JUpper
        from jadwal.models import JADWAL_EXCLUDED_JENIS as _JEX
        _jt_list = list(jadwal_terdekat)
        if _jt_list:
            _jt_lok = {j.lokasi.upper() for j in _jt_list}
            _jt_dev = {r['lu']: r['c'] for r in Device.objects.filter(
                is_deleted=False, host__isnull=True).exclude(jenis__name__in=_JEX)
                .annotate(lu=_JUpper('lokasi')).filter(lu__in=_jt_lok)
                .values('lu').annotate(c=_JCount('id'))}
            _jt_done = {(r['lu'], r['yr'], r['mo']): r['s'] for r in
                Maintenance.objects.filter(maintenance_type='Preventive',
                    device__is_deleted=False, device__host__isnull=True)
                .exclude(device__jenis__name__in=_JEX)
                .annotate(lu=_JUpper('device__lokasi'),
                    yr=ExtractYear('date'), mo=ExtractMonth('date'))
                .filter(lu__in=_jt_lok).values('lu', 'yr', 'mo')
                .annotate(s=_JCount('device_id', distinct=True))}
            jadwal_terdekat = []
            for j in _jt_list:
                lu = j.lokasi.upper()
                total   = _jt_dev.get(lu, 0)
                selesai = _jt_done.get((lu, j.tahun_rencana, j.bulan_rencana), 0)
                pct     = round(selesai / total * 100) if total else 0
                jadwal_terdekat.append({'jadwal': j, 'progress': {
                    'total': total, 'selesai': selesai,
                    'belum': total - selesai, 'pct': pct,
                }})
        else:
            jadwal_terdekat = []
    except Exception:
        jadwal_terdekat = []

    # ── Statistik per Branch ─────────────────────────────────────
    branch_stats = []
    for br in Branch.objects.prefetch_related('lokasi_set'):
        _br_lokasi = list(br.lokasi_set.values_list('nama', flat=True))
        if not _br_lokasi:
            branch_stats.append({
                'branch': br, 'total': 0, 'operasi': 0,
                'pm_done': 0, 'pm_pct': 0, 'by_jenis': [],
            })
            continue
        _br_qs = Device.objects.filter(
            is_deleted=False, host__isnull=True, lokasi__in=_br_lokasi
        )
        _br_total   = _br_qs.count()
        _br_operasi = _br_qs.filter(status_operasi='operasi').count()
        _br_pm_done = Maintenance.objects.filter(
            device__in=_br_qs, maintenance_type='Preventive',
            date__year=today.year, date__month=today.month,
        ).values('device_id').distinct().count()
        _br_pm_pct  = round(_br_pm_done / _br_total * 100) if _br_total else 0

        by_jenis = list(
            _br_qs.values('jenis__name')
            .annotate(count=Count('id'))
            .order_by('-count')
        )

        branch_stats.append({
            'branch':   br,
            'total':    _br_total,
            'operasi':  _br_operasi,
            'pm_done':  _br_pm_done,
            'pm_pct':   _br_pm_pct,
            'by_jenis': by_jenis,
        })
    branch_stats.sort(key=lambda x: x['total'], reverse=True)

    # ── Notifikasi terbaru belum dibaca ───────────────────────────
    try:
        from notifikasi.models import Notifikasi
        notif_terbaru = (
            Notifikasi.objects
            .filter(
                Q(user=request.user) | Q(user__isnull=True),
                is_read=False
            )
            .select_related('device')
            .order_by('-created_at')[:5]
        )
        notif_unread_total = Notifikasi.objects.filter(
            Q(user=request.user) | Q(user__isnull=True),
            is_read=False
        ).count()
    except Exception:
        notif_terbaru      = []
        notif_unread_total = 0

    return render(request, 'devices/dashboard.html', {
        'total_devices':    total_devices,
        'dev_operasi':      dev_operasi,
        'dev_tdk_operasi':  dev_tdk_operasi,
        'device_by_type':              device_by_type,
        'device_status_by_type_json':  device_status_by_type_json,
        'kelompok_data':    kelompok_data,
        'kelompok_json':    kelompok_json,
        'total_maintenance':    total_maintenance,
        'maintenance_open':     maintenance_open,
        'maintenance_done':     maintenance_done,
        'belum_ttd':            belum_ttd,
        'maintenance_by_month': maintenance_by_month,
        'recent_open_maintenance': recent_open_maintenance,
        'maintenance_by_lokasi':   maintenance_by_lokasi,
        'max_lokasi':              max_lokasi,
        # gangguan
        'gangguan_total':    gangguan_total,
        'gangguan_open':     gangguan_open,
        'gangguan_progress': gangguan_progress,
        'gangguan_resolved': gangguan_resolved,
        'gangguan_closed':   gangguan_closed,
        'gangguan_by_month': gangguan_by_month,
        'recent_gangguan':   recent_gangguan,
        # health index
        'hi_summary':        hi_summary,
        'hi_buruk_list':     hi_buruk_list,
        # common enemy
        'ce_open_list':      ce_open_list,
        'ce_open_total':     ce_open_total,
        # branch
        'branch_stats':      branch_stats,
        # jadwal & notifikasi
        'jadwal_terdekat':   jadwal_terdekat,
        'notif_terbaru':     notif_terbaru,
        'notif_unread_total': notif_unread_total,
        # analytics extras
        'hi_summary_json':          hi_summary_json,
        'gangguan_kategori_json':   gangguan_kategori_json,
        'gangguan_severity_json':   gangguan_severity_json,
        'gangguan_open_kritis':     gangguan_open_kritis,
        'maintenance_stacked_json': maintenance_stacked_json,
        'device_by_lokasi_json':    device_by_lokasi_json,
    })


@login_required
def distribusi_jenis(request):
    """Halaman distribusi status operasi per jenis perangkat — semua jenis tampil dengan chart masing-masing."""
    PALETTE = [
        '#3b82f6', '#10b981', '#f59e0b', '#ef4444', '#8b5cf6',
        '#06b6d4', '#ec4899', '#84cc16', '#f97316', '#64748b',
    ]

    jenis_qs = (
        Device.objects
        .filter(is_deleted=False)
        .values('jenis__id', 'jenis__name', 'status_operasi')
        .annotate(jumlah=Count('id'))
        .order_by('jenis__name')
    )

    # Kumpulkan per jenis
    jenis_map = {}
    for row in jenis_qs:
        jid   = row['jenis__id']
        jname = row['jenis__name'] or 'Lainnya'
        if jid not in jenis_map:
            jenis_map[jid] = {'id': jid, 'name': jname, 'operasi': 0, 'tidak_operasi': 0}
        jenis_map[jid][row['status_operasi']] = row['jumlah']

    jenis_data = []
    for i, (jid, d) in enumerate(sorted(jenis_map.items(), key=lambda x: x[1]['name'])):
        total = d['operasi'] + d['tidak_operasi']
        jenis_data.append({
            'id':            d['id'],
            'name':          d['name'],
            'operasi':       d['operasi'],
            'tidak_operasi': d['tidak_operasi'],
            'total':         total,
            'pct_operasi':   round(d['operasi'] / total * 100) if total else 0,
            'color':         PALETTE[i % len(PALETTE)],
        })

    jenis_json = json.dumps([
        {'name': d['name'], 'operasi': d['operasi'], 'tidak_operasi': d['tidak_operasi']}
        for d in jenis_data
    ])

    return render(request, 'devices/distribusi_jenis.html', {
        'jenis_data': jenis_data,
        'jenis_json': jenis_json,
    })


def distribusi_jenis_detail(request, jenis_id):
    """Halaman drill-down distribusi merk & type untuk satu jenis perangkat."""
    jenis_obj = get_object_or_404(DeviceType, pk=jenis_id)

    devices_qs = Device.objects.filter(is_deleted=False, jenis_id=jenis_id)
    total = devices_qs.count()
    operasi_count = devices_qs.filter(status_operasi='operasi').count()
    tidak_count   = devices_qs.filter(status_operasi='tidak_operasi').count()

    # ── Sebaran per Merk ──────────────────────────────────────────────────────
    merk_qs = (
        devices_qs
        .values('merk')
        .annotate(jumlah=Count('id'))
        .order_by('-jumlah')
    )
    merk_data = [
        {
            'label': r['merk'] or '—',
            'jumlah': r['jumlah'],
            'pct': round(r['jumlah'] / total * 100) if total else 0,
        }
        for r in merk_qs
    ]

    # ── Sebaran per Tipe/Model ────────────────────────────────────────────────
    type_qs = (
        devices_qs
        .values('type')
        .annotate(jumlah=Count('id'))
        .order_by('-jumlah')
    )
    type_data = [
        {
            'label': r['type'] or '—',
            'jumlah': r['jumlah'],
            'pct': round(r['jumlah'] / total * 100) if total else 0,
        }
        for r in type_qs
    ]

    # ── Daftar perangkat (tabel bawah) ────────────────────────────────────────
    device_list = devices_qs.order_by('nama')

    # JSON untuk chart
    merk_json = json.dumps([{'label': d['label'], 'jumlah': d['jumlah']} for d in merk_data])
    type_json = json.dumps([{'label': d['label'], 'jumlah': d['jumlah']} for d in type_data])

    PALETTE = [
        '#3b82f6', '#10b981', '#f59e0b', '#8b5cf6', '#ef4444',
        '#06b6d4', '#ec4899', '#84cc16', '#f97316', '#64748b',
    ]

    return render(request, 'devices/distribusi_jenis_detail.html', {
        'jenis':          jenis_obj,
        'total':          total,
        'operasi_count':  operasi_count,
        'tidak_count':    tidak_count,
        'pct_operasi':    round(operasi_count / total * 100) if total else 0,
        'merk_data':      merk_data,
        'type_data':      type_data,
        'device_list':    device_list,
        'merk_json':      merk_json,
        'type_json':      type_json,
        'palette_json':   json.dumps(PALETTE),
    })


def _dashboard_operator(request):
    """Dashboard khusus role Operator — fokus Inservice Inspection."""
    from inspection.models import Inspection, InspectionCatuDaya, InspectionDefenseScheme
    from inspection.models import InspectionMasterTrip, InspectionUFLS
    from datetime import date as date_type, timedelta
    from django.db.models import Count
    from django.db.models.functions import TruncDate, TruncMonth

    today   = date_type.today()
    user    = request.user

    INSPECTABLE = ['Catu Daya', 'RELE DEFENSE SCHEME', 'MASTER TRIP', 'UFLS']

    # ── Filter berdasarkan ULTG user ──────────────────────────────
    ultg         = None
    ultg_lokasi  = None  # None = semua lokasi
    try:
        profile = request.user.profile
        if profile.role == 'operator' and profile.ultg:
            ultg       = profile.ultg
            ultg_lokasi = ultg.get_lokasi_names()
    except Exception:
        pass

    def filter_by_ultg(qs):
        if ultg_lokasi is not None:
            return qs.filter(device__lokasi__in=ultg_lokasi)
        return qs

    def filter_device_by_ultg(qs):
        if ultg_lokasi is not None:
            return qs.filter(lokasi__in=ultg_lokasi)
        return qs

    # ── Statistik harian ─────────────────────────────────────────
    insp_base       = filter_by_ultg(Inspection.objects.all())
    insp_hari_ini   = insp_base.filter(tanggal__date=today).count()
    insp_hari_ini_u = insp_base.filter(tanggal__date=today, operator=user).count()
    insp_bulan_ini  = insp_base.filter(tanggal__year=today.year, tanggal__month=today.month).count()
    insp_total      = insp_base.count()

    # ── Alarm hari ini ────────────────────────────────────────────
    alarm_count = 0
    for insp in insp_base.filter(tanggal__date=today).select_related('device'):
        try:
            if insp.jenis == 'catu_daya':
                d = insp.detail_catu_daya
                if d.kondisi_rectifier == 'alarm': alarm_count += 1
            elif insp.jenis == 'defense_scheme':
                d = insp.detail_defense_scheme
                if d.kondisi_relay == 'faulty' or d.indikator_led == 'faulty': alarm_count += 1
            elif insp.jenis == 'master_trip':
                d = insp.detail_master_trip
                if d.kondisi_relay == 'faulty' or d.indikator_led == 'faulty': alarm_count += 1
            elif insp.jenis == 'ufls':
                d = insp.detail_ufls
                if d.kondisi_relay == 'faulty' or d.indikator_led == 'faulty': alarm_count += 1
        except Exception:
            pass

    # ── Total device per jenis inspectable ───────────────────────
    device_stats = []
    for jenis_name in INSPECTABLE:
        devs = filter_device_by_ultg(
            Device.objects.filter(is_deleted=False, jenis__name=jenis_name)
        )
        total = devs.count()
        if total == 0:
            continue
        inspected_ids = filter_by_ultg(Inspection.objects.filter(
            device__in=devs,
            tanggal__year=today.year, tanggal__month=today.month
        )).values_list('device_id', flat=True).distinct()
        sudah = len(set(inspected_ids))
        belum = total - sudah
        pct   = round(sudah / total * 100) if total else 0
        device_stats.append({
            'jenis': jenis_name, 'total': total,
            'sudah': sudah, 'belum': belum, 'pct': pct,
        })

    # ── Trend inspeksi 7 hari terakhir ───────────────────────────
    trend_7 = []
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        count = insp_base.filter(tanggal__date=d).count()
        trend_7.append({'label': d.strftime('%d %b'), 'count': count, 'is_today': d == today})

    # ── Trend inspeksi per jenis 30 hari ─────────────────────────
    insp_by_jenis = (
        insp_base
        .filter(tanggal__date__gte=today - timedelta(days=30))
        .values('jenis')
        .annotate(total=Count('id'))
        .order_by('-total')
    )

    # ── Inspeksi terbaru ──────────────────────────────────────────
    recent_inspections = (
        insp_base
        .select_related('device', 'device__jenis', 'operator')
        .order_by('-tanggal')[:10]
    )
    # Tambah status kondisi
    recent_list = []
    for insp in recent_inspections:
        status = 'normal'
        try:
            if insp.jenis == 'catu_daya':
                d = insp.detail_catu_daya
                if d.kondisi_rectifier == 'alarm': status = 'alarm'
            elif insp.jenis in ('defense_scheme', 'master_trip', 'ufls'):
                if insp.jenis == 'defense_scheme': d = insp.detail_defense_scheme
                elif insp.jenis == 'master_trip':  d = insp.detail_master_trip
                else:                              d = insp.detail_ufls
                if d.kondisi_relay == 'faulty' or d.indikator_led == 'faulty':
                    status = 'alarm'
        except Exception:
            pass
        insp.status_kondisi = status
        recent_list.append(insp)

    # ── Device belum diinspeksi bulan ini ─────────────────────────
    all_inspectable = filter_device_by_ultg(
        Device.objects.filter(is_deleted=False, jenis__name__in=INSPECTABLE)
    ).select_related('jenis')
    insp_ids_bulan = insp_base.filter(
        tanggal__year=today.year, tanggal__month=today.month
    ).values_list('device_id', flat=True).distinct()
    belum_insp = all_inspectable.exclude(pk__in=insp_ids_bulan).order_by('jenis__name','lokasi','nama')[:15]

    return render(request, 'devices/dashboard_operator.html', {
        'today':              today,
        'insp_hari_ini':      insp_hari_ini,
        'insp_hari_ini_u':    insp_hari_ini_u,
        'insp_bulan_ini':     insp_bulan_ini,
        'insp_total':         insp_total,
        'alarm_count':        alarm_count,
        'device_stats':       device_stats,
        'trend_7':            trend_7,
        'insp_by_jenis':      list(insp_by_jenis),
        'recent_list':        recent_list,
        'belum_insp':         belum_insp,
        'ultg':               ultg,
    })


@login_required
def device_detail(request, pk):
    device = get_object_or_404(Device, pk=pk)
    maintenance_history = Maintenance.objects.filter(device=device).order_by('-date')
    maintenance_total = maintenance_history.count()
    maintenance_done = maintenance_history.filter(status='Done').count()
    maintenance_open = maintenance_history.filter(status='Open').count()

    # Bangun spesifikasi_display: [{label, value, is_sfp}] berdasarkan schema
    spesifikasi_display = []
    if device.spesifikasi and device.jenis:
        jenis_name = device.jenis.name
        schema_fields = DEVICE_SCHEMA.get(jenis_name, [])
        # Buat mapping key -> label dari schema
        label_map = {f["key"]: f["label"] for f in schema_fields}
        for key, value in device.spesifikasi.items():
            if not value and value != 0:
                continue
            label = label_map.get(key, key.replace("_", " ").title())
            if key == "sfp_speeds" and isinstance(value, list):
                # Tampilkan per port
                for i, speed in enumerate(value, 1):
                    if speed:
                        spesifikasi_display.append({
                            "label": f"SFP Port {i}",
                            "value": speed,
                            "is_sfp": True,
                        })
            else:
                spesifikasi_display.append({
                    "label": label,
                    "value": value if not isinstance(value, list) else ", ".join(str(v) for v in value if v),
                    "is_sfp": False,
                })

    # Hitung umur peralatan (dalam tahun) dari tahun_operasi
    umur_peralatan = None
    if device.tahun_operasi:
        umur_peralatan = date_type.today().year - device.tahun_operasi

    # Hitung Health Index
    from health_index.calculator import calculate_hi
    health_index = calculate_hi(device)

    # Audit log
    from devices.models import DeviceLog, DeviceEvent, ItemBongkar
    device_logs     = DeviceLog.objects.filter(device=device).order_by('-created_at')
    last_update_log = device_logs.first()
    device_events   = DeviceEvent.objects.filter(device=device).order_by('-tanggal', '-created_at').select_related('item_bongkar_ref', 'komponen_terkait').prefetch_related('item_bongkar_set', 'item_bongkar_set__branch')

    # Komponen perangkat
    from devices.views_komponen import get_komponen_for_device, _get_tipe_grouped
    komponen_list = get_komponen_for_device(device)
    tipe_grouped = _get_tipe_grouped()

    # Eviden tambahan
    from devices.models import DeviceEviden, KomponenRusak
    eviden_list = DeviceEviden.objects.filter(device=device).order_by('uploaded_at')

    # Daftar komponen rusak
    komponen_rusak_list = KomponenRusak.objects.filter(device=device).order_by('-tanggal_rusak')

    # Item bongkar — untuk dropdown pemasangan kembali
    item_bongkar_list = ItemBongkar.objects.filter(
        device_asal=device, status='di_gudang'
    ).select_related('branch', 'komponen_terkait').order_by('-tanggal_bongkar')

    # Semua item bongkar device (termasuk yang sudah dipasang kembali)
    item_bongkar_all = ItemBongkar.objects.filter(device_asal=device).select_related('branch').order_by('-tanggal_bongkar')

    # VM children (untuk SERVER SCADA)
    vm_children = device.vm_children.filter(is_deleted=False).order_by('nama')

    # Kandidat host server untuk form tambah VM (Master Station / Server SCADA fisik)
    host_candidates = Device.objects.filter(
        Q(jenis__name__icontains='master station') |
        Q(jenis__name__icontains='server scada'),   # backward compat nama lama
        is_deleted=False,
        host__isnull=True,
    ).exclude(pk=device.pk)

    return render(request, 'devices/device_detail.html', {
        'device':             device,
        'maintenance_history': maintenance_history,
        'maintenance_total':  maintenance_total,
        'maintenance_done':   maintenance_done,
        'maintenance_open':   maintenance_open,
        'spesifikasi_display': spesifikasi_display,
        'umur_peralatan':     umur_peralatan,
        'health_index':       health_index,
        'device_logs':        device_logs,
        'last_update_log':    last_update_log,
        'device_events':      device_events,
        'komponen_list':      komponen_list,
        'tipe_grouped':       tipe_grouped,
        'eviden_list':        eviden_list,
        'today_date':         date_type.today().strftime('%Y-%m-%d'),
        'vm_children':          vm_children,
        'host_candidates':      host_candidates,
        'komponen_rusak_list':  komponen_rusak_list,
        'item_bongkar_list':    item_bongkar_list,
        'item_bongkar_all':     item_bongkar_all,
        'all_branches':         Branch.objects.all(),
    })


@login_required
def device_by_type(request, type_id):
    device_type = get_object_or_404(DeviceType, id=type_id)
    if device_type.name.strip().upper() == 'VM SCADA':
        devices = Device.objects.filter(jenis_id=type_id, is_deleted=False, host__isnull=False)
    else:
        devices = Device.objects.filter(jenis_id=type_id, is_deleted=False, host__isnull=True)
    lokasi_list = (
        Device.objects.filter(is_deleted=False)
        .exclude(lokasi__isnull=True).exclude(lokasi__exact='')
        .annotate(lokasi_clean=Trim('lokasi'))
        .values_list('lokasi_clean', flat=True)
        .distinct().order_by('lokasi_clean')
    )
    return render(request, 'devices/device_list.html', {
        'devices': devices,
        'filter_type': device_type,
        'lokasi_list': lokasi_list,
        'selected_jenis': type_id,
    })


@login_required
def lokasi_list(request):
    lokasi_data = (
        Device.objects
        .filter(is_deleted=False)
        .exclude(lokasi__isnull=True)
        .exclude(lokasi__exact='')
        .exclude(lokasi__iexact='none')
        .annotate(lokasi_clean=Trim('lokasi'))
        .values('lokasi_clean')
        .annotate(total_device=Count('id'))
        .order_by('lokasi_clean')
    )
    # Tambahkan info maintenance open per lokasi
    from maintenance.models import Maintenance
    open_by_lokasi = {}
    for m in Maintenance.objects.filter(status='Open').select_related('device'):
        loc = (m.device.lokasi or '').strip()
        open_by_lokasi[loc] = open_by_lokasi.get(loc, 0) + 1

    # Jenis perangkat per lokasi (untuk filter peta)
    jenis_by_lokasi = {}
    for row in (
        Device.objects
        .filter(is_deleted=False)
        .exclude(lokasi__isnull=True).exclude(lokasi__exact='').exclude(lokasi__iexact='none')
        .annotate(lokasi_clean=Trim('lokasi'))
        .values('lokasi_clean', 'jenis__name')
        .distinct()
        .order_by('lokasi_clean', 'jenis__name')
    ):
        loc = row['lokasi_clean']
        j = row['jenis__name'] or 'Lainnya'
        jenis_by_lokasi.setdefault(loc, []).append(j)

    # Ambil koordinat dari SiteLocation
    site_coords = {sl.nama: sl for sl in SiteLocation.objects.all()}

    lokasi_list_final = []
    for row in lokasi_data:
        row = dict(row)
        row['maintenance_open'] = open_by_lokasi.get(row['lokasi_clean'], 0)
        sl = site_coords.get(row['lokasi_clean'])
        row['latitude']  = sl.latitude  if sl and sl.has_coords else None
        row['longitude'] = sl.longitude if sl and sl.has_coords else None
        row['has_coords'] = bool(row['latitude'] and row['longitude'])
        row['jenis_list'] = jenis_by_lokasi.get(row['lokasi_clean'], [])
        lokasi_list_final.append(row)

    device_types = list(DeviceType.objects.order_by('name'))

    return render(request, 'devices/lokasi_list.html', {
        'lokasi_data':  lokasi_list_final,
        'device_types': device_types,
    })


def _link_to_dict(link, site_coords):
    loc_a = (link.device_a.lokasi or '').strip()
    loc_b = (link.device_b.lokasi or '').strip()
    sl_a  = site_coords.get(loc_a)
    sl_b  = site_coords.get(loc_b)
    return {
        'id':    link.pk,
        'label': link.display_label,
        'tipe':  link.tipe,
        'aktif': link.aktif,
        'device_a': {
            'id':    link.device_a.pk,
            'nama':  link.device_a.nama,
            'jenis': link.device_a.jenis.name if link.device_a.jenis else '',
            'lokasi': loc_a,
        },
        'device_b': {
            'id':    link.device_b.pk,
            'nama':  link.device_b.nama,
            'jenis': link.device_b.jenis.name if link.device_b.jenis else '',
            'lokasi': loc_b,
        },
        'lat_a': float(sl_a.latitude)  if sl_a and sl_a.has_coords else None,
        'lng_a': float(sl_a.longitude) if sl_a and sl_a.has_coords else None,
        'lat_b': float(sl_b.latitude)  if sl_b and sl_b.has_coords else None,
        'lng_b': float(sl_b.longitude) if sl_b and sl_b.has_coords else None,
    }


def api_device_links(request):
    from .models import DeviceLink
    sc  = {sl.nama.strip(): sl for sl in SiteLocation.objects.all()}
    qs  = (DeviceLink.objects
           .select_related('device_a', 'device_a__jenis', 'device_b', 'device_b__jenis'))
    out = []
    for link in qs:
        d = _link_to_dict(link, sc)
        if d['lat_a'] and d['lat_b']:
            out.append(d)
    return JsonResponse({'links': out})


@login_required
@require_can_edit
def api_device_link_create(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    from .models import DeviceLink
    try:
        data = json.loads(request.body)
        a    = get_object_or_404(Device, pk=data['device_a_id'])
        b    = get_object_or_404(Device, pk=data['device_b_id'])
        link = DeviceLink.objects.create(
            device_a   = a,
            device_b   = b,
            tipe       = data.get('tipe', 'fiber'),
            label      = data.get('label', '').strip(),
            keterangan = data.get('keterangan', '').strip(),
        )
        sc = {sl.nama.strip(): sl for sl in SiteLocation.objects.all()}
        return JsonResponse({'ok': True, 'link': _link_to_dict(link, sc)})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@login_required
@require_can_edit
def api_device_link_delete(request, pk):
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    from .models import DeviceLink
    get_object_or_404(DeviceLink, pk=pk).delete()
    return JsonResponse({'ok': True})


@login_required
@require_can_edit
def api_device_link_update(request, pk):
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    from .models import DeviceLink
    link = get_object_or_404(DeviceLink, pk=pk)
    try:
        data       = json.loads(request.body)
        link.tipe  = data.get('tipe', link.tipe)
        link.label = data.get('label', link.label).strip()
        link.aktif = data.get('aktif', link.aktif)
        link.save()
        sc = {sl.nama.strip(): sl for sl in SiteLocation.objects.all()}
        return JsonResponse({'ok': True, 'link': _link_to_dict(link, sc)})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


def api_lokasi_devices(request, lokasi_nama):
    """API endpoint: kembalikan daftar device di suatu lokasi sebagai JSON."""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Unauthenticated'}, status=401)
    qs = (
        Device.objects
        .filter(is_deleted=False, lokasi__iexact=lokasi_nama)
        .select_related('jenis')
        .order_by('jenis__name', 'nama')
    )
    jenis_filter = request.GET.get('jenis', '').strip()
    if jenis_filter:
        qs = qs.filter(jenis__name__iexact=jenis_filter)
    data = [
        {
            'id':    d.id,
            'nama':  d.nama,
            'jenis': d.jenis.name if d.jenis else '-',
            'merk':  d.merk or '',
            'type':  d.type or '-',
            'ip':    str(d.ip_address) if d.ip_address else '-',
            'sn':    d.serial_number or '-',
        }
        for d in qs
    ]
    return JsonResponse({'lokasi': lokasi_nama, 'devices': data})


@login_required
@login_required
def layanan_icon(request):
    icons = Icon.objects.all()

    # Filter
    search = request.GET.get('q', '').strip()
    selected_kondisi = request.GET.get('kondisi', '').strip()
    sort_by  = request.GET.get('sort', 'name')
    sort_dir = request.GET.get('dir', 'asc')

    if search:
        icons = icons.filter(
            Q(name__icontains=search) |
            Q(lokasi_layanan__icontains=search) |
            Q(SID1__icontains=search) |
            Q(SID2__icontains=search) |
            Q(keterangan__icontains=search)
        )
    if selected_kondisi:
        icons = icons.filter(kondisi_operasional__icontains=selected_kondisi)

    # Sort
    SORT_FIELDS = {
        'name':     'name',
        'lokasi':   'lokasi_layanan',
        'bandwidth':'bandwidth',
        'sid1':     'SID1',
        'kontrak':  'kontrak',
        'kondisi':  'kondisi_operasional',
    }
    order_field = SORT_FIELDS.get(sort_by, 'name')
    if sort_dir == 'desc':
        order_field = '-' + order_field
    icons = icons.order_by(order_field)

    icons = list(icons)

    # Hitung jumlah gangguan per icon (berdasarkan FK layanan_icon di model Gangguan)
    from gangguan.models import Gangguan
    gangguan_counts = dict(
        Gangguan.objects.filter(layanan_icon__isnull=False)
        .values('layanan_icon_id')
        .annotate(total=Count('id'))
        .values_list('layanan_icon_id', 'total')
    )
    for icon in icons:
        icon.jumlah_gangguan = gangguan_counts.get(icon.id, 0)

    # Summary counters
    operasi_baik = sum(
        1 for i in icons
        if i.kondisi_operasional and 'Baik' in i.kondisi_operasional
    )
    operasi_gangguan = sum(
        1 for i in icons
        if i.kondisi_operasional and (
            'Gangguan' in i.kondisi_operasional or
            'NOK' in i.kondisi_operasional
        )
    )
    operasi_lain = sum(
        1 for i in icons
        if i.kondisi_operasional and
        'Baik' not in i.kondisi_operasional and
        'Gangguan' not in i.kondisi_operasional and
        'NOK' not in i.kondisi_operasional and
        'Tidak Operasi' in i.kondisi_operasional
    )

    # Total tiket gangguan aktif yang terkait layanan ICON+
    total_gangguan_aktif = Gangguan.objects.filter(
        layanan_icon__isnull=False,
        status__in=('open', 'in_progress')
    ).count()

    # Distinct kondisi values for filter dropdown
    kondisi_list = sorted(set(
        i.kondisi_operasional for i in Icon.objects.all()
        if i.kondisi_operasional
    ))

    # Chart: tren gangguan per bulan (Icon+)
    try:
        tahun_icon = int(request.GET.get('tahun', now().year))
    except (ValueError, TypeError):
        tahun_icon = now().year

    monthly_icon = [0] * 12
    for row in (
        Gangguan.objects
        .filter(layanan_icon__isnull=False, tanggal_gangguan__year=tahun_icon)
        .values('tanggal_gangguan__month')
        .annotate(total=Count('id'))
    ):
        monthly_icon[row['tanggal_gangguan__month'] - 1] = row['total']

    top_icon_gangguan = list(
        Gangguan.objects
        .filter(layanan_icon__isnull=False)
        .values('layanan_icon__name', 'layanan_icon_id')
        .annotate(total=Count('id'))
        .order_by('-total')[:5]
    )

    tahun_icon_list = [
        d.year for d in Gangguan.objects
        .filter(layanan_icon__isnull=False)
        .dates('tanggal_gangguan', 'year', order='DESC')
    ]
    if now().year not in tahun_icon_list:
        tahun_icon_list.insert(0, now().year)

    return render(request, 'devices/layanan_icon.html', {
        'icons':                icons,
        'search':               search,
        'selected_kondisi':     selected_kondisi,
        'kondisi_list':         kondisi_list,
        'operasi_baik':         operasi_baik,
        'gangguan':             operasi_gangguan,
        'lainnya':              operasi_lain,
        'sort_by':              sort_by,
        'sort_dir':             sort_dir,
        'total_gangguan_aktif': total_gangguan_aktif,
        'tahun_icon':           tahun_icon,
        'tahun_icon_list':      tahun_icon_list,
        'monthly_icon':         monthly_icon,
        'top_icon_gangguan':    top_icon_gangguan,
    })


# ══════════════════════════════════════════════════════════════
# FIBER OPTIC VIEWS
# ══════════════════════════════════════════════════════════════

@login_required
def fiber_optic_list(request):
    from .models import FiberOptic
    from gangguan.models import Gangguan

    qs = FiberOptic.objects.all()

    search       = request.GET.get('q', '').strip()
    status_f     = request.GET.get('status', '').strip()
    sort_by      = request.GET.get('sort', 'nama')
    sort_dir     = request.GET.get('dir', 'asc')

    if search:
        qs = qs.filter(
            Q(nama__icontains=search) |
            Q(lokasi_a__icontains=search) |
            Q(lokasi_b__icontains=search) |
            Q(keterangan__icontains=search)
        )
    if status_f:
        qs = qs.filter(status=status_f)

    SORT_MAP = {
        'nama': 'nama', 'lokasi_a': 'lokasi_a', 'lokasi_b': 'lokasi_b',
        'panjang': 'panjang_km', 'status': 'status',
    }
    order_field = SORT_MAP.get(sort_by, 'nama')
    if sort_dir == 'desc':
        order_field = '-' + order_field
    qs = qs.order_by(order_field)

    fo_list = list(qs)

    # Hitung jumlah gangguan per segmen FO
    gangguan_counts = dict(
        Gangguan.objects.filter(fiber_optic__isnull=False)
        .values('fiber_optic_id')
        .annotate(total=Count('id'))
        .values_list('fiber_optic_id', 'total')
    )
    for fo in fo_list:
        fo.jumlah_gangguan = gangguan_counts.get(fo.id, 0)

    # Summary
    total_fo        = len(fo_list)
    total_baik      = sum(1 for f in fo_list if f.status == 'baik')
    total_gangguan  = sum(1 for f in fo_list if f.status == 'gangguan')
    total_perbaikan = sum(1 for f in fo_list if f.status == 'dalam_perbaikan')

    total_aktif = Gangguan.objects.filter(
        fiber_optic__isnull=False,
        status__in=('open', 'in_progress')
    ).count()

    # Chart: tren gangguan per bulan (FO)
    try:
        tahun_fo = int(request.GET.get('tahun', now().year))
    except (ValueError, TypeError):
        tahun_fo = now().year

    monthly_fo = [0] * 12
    for row in (
        Gangguan.objects
        .filter(fiber_optic__isnull=False, tanggal_gangguan__year=tahun_fo)
        .values('tanggal_gangguan__month')
        .annotate(total=Count('id'))
    ):
        monthly_fo[row['tanggal_gangguan__month'] - 1] = row['total']

    top_fo_gangguan = list(
        Gangguan.objects
        .filter(fiber_optic__isnull=False)
        .values('fiber_optic__nama', 'fiber_optic_id')
        .annotate(total=Count('id'))
        .order_by('-total')[:5]
    )

    tahun_fo_list = [
        d.year for d in Gangguan.objects
        .filter(fiber_optic__isnull=False)
        .dates('tanggal_gangguan', 'year', order='DESC')
    ]
    if now().year not in tahun_fo_list:
        tahun_fo_list.insert(0, now().year)

    return render(request, 'devices/fiber_optic_list.html', {
        'fo_list':           fo_list,
        'search':            search,
        'status_filter':     status_f,
        'sort_by':           sort_by,
        'sort_dir':          sort_dir,
        'total_fo':          total_fo,
        'total_baik':        total_baik,
        'total_gangguan':    total_gangguan,
        'total_perbaikan':   total_perbaikan,
        'total_aktif':       total_aktif,
        'STATUS_CHOICES':    FiberOptic.STATUS_CHOICES,
        'tahun_fo':          tahun_fo,
        'tahun_fo_list':     tahun_fo_list,
        'monthly_fo':        monthly_fo,
        'top_fo_gangguan':   top_fo_gangguan,
    })


@login_required
@require_can_edit
def fiber_optic_create(request):
    from .models import FiberOptic, FiberOpticCore, SiteLocation
    lokasi_list = list(SiteLocation.objects.values_list('nama', flat=True).order_by('nama'))
    if request.method == 'POST':
        jumlah_core = int(request.POST['jumlah_core']) if request.POST.get('jumlah_core') else None
        fo = FiberOptic(
            nama            = request.POST.get('nama', '').strip(),
            lokasi_a        = request.POST.get('lokasi_a', '').strip(),
            lokasi_b        = request.POST.get('lokasi_b', '').strip(),
            tipe_kabel      = request.POST.get('tipe_kabel') or None,
            tipe_konektor   = request.POST.get('tipe_konektor') or None,
            tipe_konektor_a = request.POST.get('tipe_konektor_a') or None,
            tipe_konektor_b = request.POST.get('tipe_konektor_b') or None,
            jumlah_core     = jumlah_core,
            konfigurasi     = request.POST.get('konfigurasi') or None,
            panjang_km      = request.POST.get('panjang_km') or None,
            tahun_pasang    = int(request.POST['tahun_pasang']) if request.POST.get('tahun_pasang') else None,
            status          = request.POST.get('status', 'baik'),
            keterangan      = request.POST.get('keterangan', '').strip() or None,
            created_by      = request.user,
        )
        if request.FILES.get('foto_site_a'):
            fo.foto_site_a = request.FILES['foto_site_a']
        if request.FILES.get('foto_site_b'):
            fo.foto_site_b = request.FILES['foto_site_b']
        fo.save()
        if jumlah_core:
            for n in range(1, jumlah_core + 1):
                FiberOpticCore.objects.get_or_create(fiber_optic=fo, nomor_core=n)
        return redirect('fiber_optic_detail', pk=fo.pk)
    return render(request, 'devices/fiber_optic_form.html', {
        'is_edit':           False,
        'TIPE_KABEL':        FiberOptic.TIPE_KABEL_CHOICES,
        'TIPE_KONEKTOR':     FiberOptic.TIPE_KONEKTOR_CHOICES,
        'KONFIGURASI':       FiberOptic.KONFIGURASI_CHOICES,
        'STATUS_CHOICES':    FiberOptic.STATUS_CHOICES,
        'lokasi_list':       lokasi_list,
    })


@login_required
@require_can_edit
def fiber_optic_update(request, pk):
    from .models import FiberOptic, FiberOpticCore, SiteLocation
    fo = get_object_or_404(FiberOptic, pk=pk)
    lokasi_list = list(SiteLocation.objects.values_list('nama', flat=True).order_by('nama'))
    if request.method == 'POST':
        jumlah_core_baru = int(request.POST['jumlah_core']) if request.POST.get('jumlah_core') else None
        fo.nama            = request.POST.get('nama', '').strip()
        fo.lokasi_a        = request.POST.get('lokasi_a', '').strip()
        fo.lokasi_b        = request.POST.get('lokasi_b', '').strip()
        fo.tipe_kabel      = request.POST.get('tipe_kabel') or None
        fo.tipe_konektor   = request.POST.get('tipe_konektor') or None
        fo.tipe_konektor_a = request.POST.get('tipe_konektor_a') or None
        fo.tipe_konektor_b = request.POST.get('tipe_konektor_b') or None
        fo.jumlah_core     = jumlah_core_baru
        fo.konfigurasi     = request.POST.get('konfigurasi') or None
        fo.panjang_km      = request.POST.get('panjang_km') or None
        fo.tahun_pasang    = int(request.POST['tahun_pasang']) if request.POST.get('tahun_pasang') else None
        fo.status          = request.POST.get('status', 'baik')
        fo.keterangan      = request.POST.get('keterangan', '').strip() or None
        if request.FILES.get('foto_site_a'):
            fo.foto_site_a = request.FILES['foto_site_a']
        if request.FILES.get('foto_site_b'):
            fo.foto_site_b = request.FILES['foto_site_b']
        fo.save()
        # Tambah core baru jika jumlah_core bertambah
        if jumlah_core_baru:
            existing = set(fo.cores.values_list('nomor_core', flat=True))
            for n in range(1, jumlah_core_baru + 1):
                if n not in existing:
                    FiberOpticCore.objects.create(fiber_optic=fo, nomor_core=n)
        return redirect('fiber_optic_detail', pk=fo.pk)
    return render(request, 'devices/fiber_optic_form.html', {
        'is_edit':        True,
        'fo':             fo,
        'TIPE_KABEL':     FiberOptic.TIPE_KABEL_CHOICES,
        'TIPE_KONEKTOR':  FiberOptic.TIPE_KONEKTOR_CHOICES,
        'KONFIGURASI':    FiberOptic.KONFIGURASI_CHOICES,
        'STATUS_CHOICES': FiberOptic.STATUS_CHOICES,
        'lokasi_list':    lokasi_list,
    })


@login_required
def fiber_optic_detail(request, pk):
    import json as _json
    from .models import FiberOptic, FiberOpticCore
    fo = get_object_or_404(FiberOptic, pk=pk)
    cores = fo.cores.order_by('nomor_core')
    from gangguan.models import Gangguan
    gangguan_list = Gangguan.objects.filter(fiber_optic=fo).order_by('-tanggal_gangguan')

    # Build JSON for OTDR dashboard
    def _f(v):
        return float(v) if v is not None else None

    cores_json = _json.dumps([{
        'pk':              c.pk,
        'nomor_core':      c.nomor_core,
        'core_num_a':      c.core_num_a,
        'core_num_b':      c.core_num_b,
        'fungsi':          c.fungsi or '',
        'fungsi_b':        c.fungsi_b or '',
        'keterangan':      c.keterangan or '',
        'keterangan_b':    c.keterangan_b or '',
        'status':          c.status,
        'status_a':        c.status_a,
        'status_b':        c.status_b,
        'koneksi_a':       c.koneksi_a or '',
        'koneksi_b':       c.koneksi_b or '',
        # OTDR Site A λ1310
        'a_jarak_1310':         _f(c.otdr_jarak_km_1310),
        'a_redaman_1310':       _f(c.otdr_redaman_db_1310),
        'a_dbkm_1310':          _f(c.otdr_redaman_per_km_1310),
        # OTDR Site A λ1550
        'a_jarak_1550':         _f(c.otdr_jarak_km_1550),
        'a_redaman_1550':       _f(c.otdr_redaman_db_1550),
        'a_dbkm_1550':          _f(c.otdr_redaman_per_km_1550),
        'a_tanggal':            str(c.otdr_tanggal) if c.otdr_tanggal else '',
        'a_catatan':            c.otdr_catatan or '',
        # OTDR Site B λ1310
        'b_jarak_1310':         _f(c.otdr_b_jarak_km_1310),
        'b_redaman_1310':       _f(c.otdr_b_redaman_db_1310),
        'b_dbkm_1310':          _f(c.otdr_b_redaman_per_km_1310),
        # OTDR Site B λ1550
        'b_jarak_1550':         _f(c.otdr_b_jarak_km_1550),
        'b_redaman_1550':       _f(c.otdr_b_redaman_db_1550),
        'b_dbkm_1550':          _f(c.otdr_b_redaman_per_km_1550),
        'b_tanggal':            str(c.otdr_b_tanggal) if c.otdr_b_tanggal else '',
        'b_catatan':            c.otdr_b_catatan or '',
    } for c in cores])

    return render(request, 'devices/fiber_optic_detail.html', {
        'fo':            fo,
        'cores':         cores,
        'gangguan_list': gangguan_list,
        'STATUS_CORE':   FiberOpticCore.STATUS_CORE_CHOICES,
        'cores_json':    cores_json,
    })


@login_required
@require_can_edit
def fiber_optic_core_update(request, fo_pk, core_pk):
    """Update satu baris core via POST (AJAX-friendly)."""
    from .models import FiberOpticCore
    core = get_object_or_404(FiberOpticCore, pk=core_pk, fiber_optic_id=fo_pk)
    if request.method == 'POST':
        # nomor core (editable, must stay unique within same FO)
        new_nomor = request.POST.get('nomor_core', '').strip()
        if new_nomor.isdigit():
            new_nomor_int = int(new_nomor)
            if new_nomor_int != core.nomor_core:
                conflict = FiberOpticCore.objects.filter(
                    fiber_optic_id=fo_pk, nomor_core=new_nomor_int
                ).exclude(pk=core_pk).exists()
                if not conflict:
                    core.nomor_core = new_nomor_int

        site = request.POST.get('site', 'a')
        if site == 'a':
            core.fungsi     = request.POST.get('fungsi', '').strip() or None
            core.keterangan = request.POST.get('keterangan', '').strip() or None
            core.status_a   = request.POST.get('status_a', 'spare')
            core.koneksi_a  = request.POST.get('koneksi_a', '').strip() or None
            # nomor core site A
            nc_a = request.POST.get('nomor_core_a', '').strip()
            core.nomor_core_a = int(nc_a) if nc_a.isdigit() else None
            # λ1310 site A
            core.otdr_jarak_km_1310          = request.POST.get('otdr_jarak_km_1310') or None
            core.otdr_redaman_db_1310        = request.POST.get('otdr_redaman_db_1310') or None
            core.otdr_redaman_per_km_1310    = request.POST.get('otdr_redaman_per_km_1310') or None
            # λ1550 site A
            core.otdr_jarak_km_1550          = request.POST.get('otdr_jarak_km_1550') or None
            core.otdr_redaman_db_1550        = request.POST.get('otdr_redaman_db_1550') or None
            core.otdr_redaman_per_km_1550    = request.POST.get('otdr_redaman_per_km_1550') or None
            core.otdr_tanggal                = request.POST.get('otdr_tanggal') or None
            core.otdr_catatan                = request.POST.get('otdr_catatan', '').strip() or None
        else:  # site == 'b'
            core.fungsi_b     = request.POST.get('fungsi_b', '').strip() or None
            core.keterangan_b = request.POST.get('keterangan_b', '').strip() or None
            core.status_b     = request.POST.get('status_b', 'spare')
            core.koneksi_b    = request.POST.get('koneksi_b', '').strip() or None
            # nomor core site B
            nc_b = request.POST.get('nomor_core_b', '').strip()
            core.nomor_core_b = int(nc_b) if nc_b.isdigit() else None
            # λ1310 site B
            core.otdr_b_jarak_km_1310        = request.POST.get('otdr_b_jarak_km_1310') or None
            core.otdr_b_redaman_db_1310      = request.POST.get('otdr_b_redaman_db_1310') or None
            core.otdr_b_redaman_per_km_1310  = request.POST.get('otdr_b_redaman_per_km_1310') or None
            # λ1550 site B
            core.otdr_b_jarak_km_1550        = request.POST.get('otdr_b_jarak_km_1550') or None
            core.otdr_b_redaman_db_1550      = request.POST.get('otdr_b_redaman_db_1550') or None
            core.otdr_b_redaman_per_km_1550  = request.POST.get('otdr_b_redaman_per_km_1550') or None
            core.otdr_b_tanggal              = request.POST.get('otdr_b_tanggal') or None
            core.otdr_b_catatan              = request.POST.get('otdr_b_catatan', '').strip() or None

        # keep legacy status in sync with the site being updated
        core.status = core.status_a if site == 'a' else core.status_b
        core.save()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'ok': True})
    return redirect('fiber_optic_detail', pk=fo_pk)


@login_required
@require_can_delete
def fiber_optic_delete(request, pk):
    from .models import FiberOptic
    fo = get_object_or_404(FiberOptic, pk=pk)
    if request.method == 'POST':
        fo.delete()
    return redirect('fiber_optic_list')


def api_fiber_optic_json(request):
    """API: semua segmen FO sebagai JSON untuk JS autocomplete di form gangguan."""
    from .models import FiberOptic
    data = list(
        FiberOptic.objects.all().order_by('nama').values(
            'id', 'nama', 'lokasi_a', 'lokasi_b',
            'tipe_kabel', 'tipe_konektor', 'jumlah_core',
            'panjang_km', 'status',
        )
    )
    for d in data:
        if d['panjang_km']:
            d['panjang_km'] = str(d['panjang_km'])
    return JsonResponse({'fiber_optic': data})


def fo_public(request, token):
    """Halaman publik fiber optic via QR Code — tidak perlu login."""
    from .models import FiberOptic, FiberOpticCore
    fo = get_object_or_404(FiberOptic, public_token=token)
    cores = fo.cores.order_by('nomor_core')

    # Gangguan aktif terkait segmen ini
    try:
        from gangguan.models import Gangguan
        gangguan_aktif = Gangguan.objects.filter(
            fiber_optic=fo, status__in=['open', 'in_progress']
        ).order_by('-tanggal_gangguan')
    except Exception:
        gangguan_aktif = []

    # Ringkasan status core per site
    total_core       = cores.count()
    aktif_a_count    = cores.filter(status_a='aktif').count()
    spare_a_count    = cores.filter(status_a='spare').count()
    rusak_a_count    = cores.filter(status_a='rusak').count()
    nonaktif_a_count = cores.filter(status_a='tidak_aktif').count()
    aktif_b_count    = cores.filter(status_b='aktif').count()
    spare_b_count    = cores.filter(status_b='spare').count()
    rusak_b_count    = cores.filter(status_b='rusak').count()
    nonaktif_b_count = cores.filter(status_b='tidak_aktif').count()

    return render(request, 'devices/fiber_optic_public.html', {
        'fo':              fo,
        'cores':           cores,
        'gangguan_aktif':  gangguan_aktif,
        'total_core':      total_core,
        'aktif_a_count':   aktif_a_count,
        'spare_a_count':   spare_a_count,
        'rusak_a_count':   rusak_a_count,
        'nonaktif_a_count': nonaktif_a_count,
        'aktif_b_count':   aktif_b_count,
        'spare_b_count':   spare_b_count,
        'rusak_b_count':   rusak_b_count,
        'nonaktif_b_count': nonaktif_b_count,
    })


@login_required
def fo_qr(request, pk):
    """Halaman cetak QR Code fiber optic."""
    from .models import FiberOptic
    fo = get_object_or_404(FiberOptic, pk=pk)
    if not fo.public_token:
        import secrets
        fo.public_token = secrets.token_urlsafe(20)
        fo.save(update_fields=['public_token'])
    public_url = request.build_absolute_uri(f'/fo/public/{fo.public_token}/')
    return render(request, 'devices/fiber_optic_qr.html', {
        'fo':         fo,
        'public_url': public_url,
    })


@login_required
@require_can_edit
def icon_create(request):
    if request.method == 'POST':
        form = IconForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('layanan_icon')
    else:
        form = IconForm()
    return render(request, 'devices/icon_form.html', {
        'form': form,
        'is_edit': False,
        'kondisi_choices': [('Operasi Baik','Operasi Baik'),('Gangguan','Gangguan'),('Tidak Operasi','Tidak Operasi'),('Dalam Pemeliharaan','Dalam Pemeliharaan')],
    })


@login_required
@require_can_edit
def icon_update(request, pk):
    icon = get_object_or_404(Icon, pk=pk)
    if request.method == 'POST':
        form = IconForm(request.POST, request.FILES, instance=icon)
        if form.is_valid():
            form.save()
            return redirect('layanan_icon')
    else:
        form = IconForm(instance=icon)
    return render(request, 'devices/icon_form.html', {
        'form': form,
        'is_edit': True,
        'icon': icon,
        'kondisi_choices': [('Operasi Baik','Operasi Baik'),('Gangguan','Gangguan'),('Tidak Operasi','Tidak Operasi'),('Dalam Pemeliharaan','Dalam Pemeliharaan')],
    })


@login_required
@require_can_delete
def icon_delete(request, pk):
    icon = get_object_or_404(Icon, pk=pk)
    if request.method == 'POST':
        icon.delete()
    return redirect('layanan_icon')


# =====================================================
# QR CODE VIEW — buka halaman print QR
# =====================================================
@login_required
def device_qr(request, pk):
    device = get_object_or_404(Device, pk=pk)
    from django.urls import reverse
    if device.public_token:
        public_url = request.build_absolute_uri(
            reverse('device_public', args=[device.public_token])
        )
    else:
        import secrets
        device.public_token = secrets.token_urlsafe(20)
        device.save(update_fields=['public_token'])
        public_url = request.build_absolute_uri(
            reverse('device_public', args=[device.public_token])
        )
    return render(request, 'devices/device_qr.html', {
        'device':     device,
        'detail_url': public_url,
        'public_url': public_url,
    })


# =====================================================
# EXPORT DEVICES KE EXCEL
# =====================================================
@login_required
def export_devices_excel(request):
    jenis_id = request.GET.get('jenis')
    search = request.GET.get('q') or ''
    lokasi = request.GET.get('lokasi')

    devices = Device.objects.filter(is_deleted=False).select_related('jenis')

    if jenis_id:
        devices = devices.filter(jenis_id=jenis_id)
    if search:
        devices = devices.filter(Q(nama__icontains=search) | Q(ip_address__icontains=search))
    if lokasi:
        devices = devices.filter(lokasi=lokasi)

    from health_index.calculator import calculate_hi
    from devices.models_komponen import DeviceComponent

    # Pre-fetch komponen for all devices in one query
    devices_list = list(devices.prefetch_related('komponen__tipe_komponen'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Peralatan"

    # Styles
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # HI color fills
    hi_fills = {
        'sangat_baik': PatternFill("solid", fgColor="DCFCE7"),   # green ≥85
        'baik':        PatternFill("solid", fgColor="DBEAFE"),   # blue 70-84
        'cukup':       PatternFill("solid", fgColor="FEF9C3"),   # yellow 50-69
        'buruk':       PatternFill("solid", fgColor="FFEDD5"),   # orange 25-49
        'kritis':      PatternFill("solid", fgColor="FEE2E2"),   # red <25
    }
    hi_fonts = {
        'sangat_baik': Font(bold=True, color="166534"),
        'baik':        Font(bold=True, color="1E40AF"),
        'cukup':       Font(bold=True, color="854D0E"),
        'buruk':       Font(bold=True, color="9A3412"),
        'kritis':      Font(bold=True, color="991B1B"),
    }

    def _hi_category_key(score):
        if score is None:
            return None
        if score >= 85: return 'sangat_baik'
        if score >= 70: return 'baik'
        if score >= 50: return 'cukup'
        if score >= 25: return 'buruk'
        return 'kritis'

    # Judul
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = "INVENTORY PERALATAN FASOP UP2B"
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill("solid", fgColor="EFF6FF")
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:K2')
    from datetime import date
    ws['A2'].value = f"Dicetak: {date.today().strftime('%d %B %Y')}"
    ws['A2'].alignment = Alignment(horizontal="center")
    ws['A2'].font = Font(size=10, italic=True, color="64748B")
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 6  # spacer

    # Header — 11 columns
    headers = ['No', 'Nama', 'Jenis', 'Merk', 'Type/Model', 'Serial Number',
               'IP Address', 'Lokasi', 'Firmware', 'Health Index', 'Kategori HI']
    col_widths = [5, 25, 15, 15, 18, 20, 16, 20, 15, 13, 16]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[4].height = 22

    # Data
    alt_fill = PatternFill("solid", fgColor="F8FAFC")
    for row_idx, d in enumerate(devices_list, 1):
        ws_row = row_idx + 4

        # Calculate HI
        try:
            hi_result = calculate_hi(d, save_snapshot=False)
            hi_score = hi_result.get('score')
            hi_label = hi_result.get('kategori', {}).get('label', '-') if hi_result else '-'
        except Exception:
            hi_score = None
            hi_label = '-'

        row_data = [
            row_idx,
            d.nama,
            d.jenis.name if d.jenis else '-',
            d.merk,
            d.type or '-',
            d.serial_number or '-',
            str(d.ip_address),
            d.lokasi,
            d.firmware_version or '-',
            hi_score if hi_score is not None else '-',
            hi_label,
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=ws_row, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = center_align if col_idx in (1, 10, 11) else Alignment(vertical="center")
            if row_idx % 2 == 0 and col_idx not in (10, 11):
                cell.fill = alt_fill

        # Color-code HI cells (col 10 & 11)
        cat_key = _hi_category_key(hi_score)
        if cat_key:
            for col_idx in (10, 11):
                cell = ws.cell(row=ws_row, column=col_idx)
                cell.fill = hi_fills[cat_key]
                cell.font = hi_fonts[cat_key]

        ws.row_dimensions[ws_row].height = 18

    # Freeze panes
    ws.freeze_panes = 'A5'

    # ── Sheet 2: Komponen Perangkat ────────────────────────────
    ws2 = wb.create_sheet("Komponen Perangkat")

    ws2.merge_cells('A1:L1')
    ws2['A1'].value = "KOMPONEN PERANGKAT FASOP UP2B"
    ws2['A1'].font = Font(bold=True, size=13)
    ws2['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws2['A1'].fill = PatternFill("solid", fgColor="F5F3FF")
    ws2.row_dimensions[1].height = 28

    ws2.merge_cells('A2:L2')
    ws2['A2'].value = f"Dicetak: {date.today().strftime('%d %B %Y')}"
    ws2['A2'].alignment = Alignment(horizontal="center")
    ws2['A2'].font = Font(size=10, italic=True, color="64748B")
    ws2.row_dimensions[2].height = 18
    ws2.row_dimensions[3].height = 6  # spacer

    k_headers = ['No', 'Perangkat', 'Lokasi', 'Jenis', 'Nama Komponen',
                 'Tipe', 'Posisi', 'Merk', 'Model/Type', 'Serial Number', 'Status', 'Keterangan']
    k_widths   = [5,    25,          18,       15,     22,
                  18,    15,      14,      18,          20,              12,       30]

    k_header_fill = PatternFill("solid", fgColor="2D1B69")
    k_header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx, (hdr, width) in enumerate(zip(k_headers, k_widths), 1):
        cell = ws2.cell(row=4, column=col_idx, value=hdr)
        cell.font = k_header_font
        cell.fill = k_header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws2.column_dimensions[get_column_letter(col_idx)].width = width
    ws2.row_dimensions[4].height = 22

    k_alt_fill = PatternFill("solid", fgColor="FAF5FF")
    k_row = 5
    k_no  = 1
    status_fills = {
        'Baik':     PatternFill("solid", fgColor="DCFCE7"),
        'Rusak':    PatternFill("solid", fgColor="FEE2E2"),
        'Degraded': PatternFill("solid", fgColor="FEF9C3"),
    }
    status_fonts = {
        'Baik':     Font(bold=True, color="166534"),
        'Rusak':    Font(bold=True, color="991B1B"),
        'Degraded': Font(bold=True, color="854D0E"),
    }

    for d in devices_list:
        komponents = list(d.komponen.all())
        if not komponents:
            continue
        for k in komponents:
            tipe_nama = k.tipe_komponen.nama if k.tipe_komponen else '-'
            row_data2 = [
                k_no,
                d.nama,
                d.lokasi,
                d.jenis.name if d.jenis else '-',
                k.nama,
                tipe_nama,
                k.posisi or '-',
                k.merk or '-',
                k.model or '-',
                k.serial_number or '-',
                k.status or '-',
                k.keterangan or '',
            ]
            use_alt = (k_no % 2 == 0)
            for col_idx, value in enumerate(row_data2, 1):
                cell = ws2.cell(row=k_row, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = center_align if col_idx == 1 else Alignment(vertical="center", wrap_text=(col_idx == 12))
                if use_alt and col_idx != 11:
                    cell.fill = k_alt_fill
            # Color status cell (col 11)
            st = k.status or ''
            if st in status_fills:
                ws2.cell(row=k_row, column=11).fill = status_fills[st]
                ws2.cell(row=k_row, column=11).font = status_fonts[st]
                ws2.cell(row=k_row, column=11).alignment = center_align
            ws2.row_dimensions[k_row].height = 18
            k_row += 1
            k_no  += 1

    ws2.freeze_panes = 'A5'

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="inventory_peralatan_fasop.xlsx"'
    wb.save(response)
    return response


@login_required
def export_icon_excel(request):
    from datetime import date

    icons = Icon.objects.all()

    # Apply same filters as list view
    search = request.GET.get('q', '').strip()
    selected_kondisi = request.GET.get('kondisi', '').strip()
    if search:
        icons = icons.filter(
            Q(name__icontains=search) |
            Q(lokasi_layanan__icontains=search) |
            Q(SID1__icontains=search) |
            Q(SID2__icontains=search) |
            Q(keterangan__icontains=search)
        )
    if selected_kondisi:
        icons = icons.filter(kondisi_operasional__icontains=selected_kondisi)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Layanan ICON+"

    # Styles
    header_fill  = PatternFill("solid", fgColor="0F172A")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )
    alt_fill = PatternFill("solid", fgColor="F8FAFC")

    COLS = 9

    # Judul
    ws.merge_cells(f'A1:{get_column_letter(COLS)}1')
    ws['A1'].value     = "DATA LAYANAN ICON+ — FASOP UP2B"
    ws['A1'].font      = Font(bold=True, size=13)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A1'].fill      = PatternFill("solid", fgColor="EFF6FF")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f'A2:{get_column_letter(COLS)}2')
    ws['A2'].value     = f"Dicetak: {date.today().strftime('%d %B %Y')}"
    ws['A2'].alignment = Alignment(horizontal="center")
    ws['A2'].font      = Font(size=10, italic=True, color="64748B")
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6  # spacer

    # Header row
    headers    = ['No', 'Nama Layanan', 'Lokasi Layanan', 'Bandwidth', 'SID 1', 'SID 2', 'Kontrak', 'Kondisi Operasional', 'Keterangan']
    col_widths = [5,    28,             25,                14,          22,      22,      18,        22,                   35]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[4].height = 22

    # Conditional fill colours for kondisi
    green_fill  = PatternFill("solid", fgColor="DCFCE7")
    red_fill    = PatternFill("solid", fgColor="FEE2E2")
    yellow_fill = PatternFill("solid", fgColor="FEF3C7")

    # Data rows
    for row_idx, d in enumerate(icons, 1):
        ws_row = row_idx + 4
        row_data = [
            row_idx,
            d.name or '-',
            d.lokasi_layanan or '-',
            d.bandwidth or '-',
            d.SID1 or '-',
            d.SID2 or '-',
            d.kontrak or '-',
            d.kondisi_operasional or '-',
            d.keterangan or '-',
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell            = ws.cell(row=ws_row, column=col_idx, value=value)
            cell.border     = thin_border
            cell.alignment  = center_align if col_idx == 1 else left_align
            if row_idx % 2 == 0:
                cell.fill = alt_fill

        # Colour-code kondisi cell (col 8)
        kondisi_val = d.kondisi_operasional or ''
        kondisi_cell = ws.cell(row=ws_row, column=8)
        if 'Baik' in kondisi_val:
            kondisi_cell.fill = green_fill
            kondisi_cell.font = Font(bold=True, color="065F46")
        elif 'Gangguan' in kondisi_val or 'NOK' in kondisi_val:
            kondisi_cell.fill = red_fill
            kondisi_cell.font = Font(bold=True, color="991B1B")
        elif kondisi_val != '-':
            kondisi_cell.fill = yellow_fill
            kondisi_cell.font = Font(bold=True, color="92400E")

        ws.row_dimensions[ws_row].height = 18

    # Freeze header
    ws.freeze_panes = 'A5'

    # Auto-filter
    ws.auto_filter.ref = f'A4:{get_column_letter(COLS)}{4 + icons.count() if hasattr(icons, "count") else 4 + len(list(icons))}'

    # Response
    from django.http import HttpResponse
    import io
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"layanan_icon_{date.today().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

# ── Device Event Views ──────────────────────────────────────────────────────

@login_required
@require_can_edit
def device_event_add(request, pk):
    """Tambah event kejadian fisik peralatan."""
    import json as _json
    device = get_object_or_404(Device, pk=pk)

    if request.method == 'POST':
        from devices.models import DeviceEvent, ItemBongkar
        from devices.models_komponen import DeviceComponent
        from django.utils import timezone as _tz

        tipe                 = request.POST.get('tipe', '').strip()
        tanggal              = request.POST.get('tanggal', '')
        komponen             = request.POST.get('komponen', '').strip()
        komponen_terkait_pk  = request.POST.get('komponen_terkait_id', '') or None
        nilai_lama           = request.POST.get('nilai_lama', '').strip()
        nilai_baru           = request.POST.get('nilai_baru', '').strip()
        lokasi_asal          = request.POST.get('lokasi_asal', '').strip()
        lokasi_tujuan        = request.POST.get('lokasi_tujuan', '').strip()
        catatan              = request.POST.get('catatan', '').strip()
        foto                 = request.FILES.get('foto')
        # getlist karena field name sama dipakai di fieldPenggantian & fieldPenambahan
        # ambil nilai pertama yang tidak kosong
        merk_komponen_baru   = next((v.strip() for v in request.POST.getlist('merk_komponen_baru')   if v.strip()), '')
        tipe_komponen_baru   = next((v.strip() for v in request.POST.getlist('tipe_komponen_baru')   if v.strip()), '')
        serial_komponen_baru = next((v.strip() for v in request.POST.getlist('serial_komponen_baru') if v.strip()), '')
        posisi_komponen_baru = request.POST.get('posisi_komponen_baru', '').strip()
        disimpan_di          = request.POST.get('disimpan_di', '').strip()
        branch_id_event      = request.POST.get('branch_id', '') or None
        pembongkaran_tipe    = request.POST.get('pembongkaran_tipe', 'komponen').strip()
        config_aspek         = request.POST.get('config_aspek', '').strip()
        item_bongkar_id      = request.POST.get('item_bongkar_id', '') or None
        tipe_komponen_id     = request.POST.get('tipe_komponen_id', '') or None
        alasan_penggantian   = request.POST.get('alasan_penggantian', '').strip()
        # Pembongkaran pakai field terpisah (nama berbeda) agar tidak bentrok
        # dengan field penggantian komponen yang ada di form yang sama
        if tipe == 'pembongkaran':
            branch_id_event    = request.POST.get('bongkar_branch_id', '') or None
            disimpan_di        = request.POST.get('bongkar_disimpan_di', '').strip()
            alasan_penggantian = request.POST.get('bongkar_alasan', '').strip()
        komponen_relokasi_raw = request.POST.getlist('komponen_relokasi_ids')

        komponen_terkait_obj = None
        if komponen_terkait_pk:
            komponen_terkait_obj = DeviceComponent.objects.filter(pk=komponen_terkait_pk).first()

        komponen_relokasi_ids = []
        for v in komponen_relokasi_raw:
            try:
                komponen_relokasi_ids.append(int(v))
            except (ValueError, TypeError):
                pass

        if tipe and tanggal:
            event = DeviceEvent(
                device               = device,
                tipe                 = tipe,
                tanggal              = tanggal,
                komponen             = komponen,
                komponen_terkait     = komponen_terkait_obj,
                nilai_lama           = nilai_lama,
                nilai_baru           = nilai_baru,
                lokasi_asal          = lokasi_asal,
                lokasi_tujuan        = lokasi_tujuan,
                catatan              = catatan,
                dilakukan_oleh       = request.user,
                merk_komponen_baru   = merk_komponen_baru,
                tipe_komponen_baru   = tipe_komponen_baru,
                serial_komponen_baru = serial_komponen_baru,
                posisi_komponen_baru = posisi_komponen_baru,
                pembongkaran_tipe    = pembongkaran_tipe if tipe == 'pembongkaran' else '',
                config_aspek         = config_aspek if tipe == 'modifikasi' else '',
                komponen_relokasi_ids = komponen_relokasi_ids,
                alasan_penggantian   = alasan_penggantian if tipe in ('penggantian', 'penggantian_perangkat', 'pembongkaran') else '',
            )
            if foto:
                event.foto = foto
            event.save()

            # ── RELOKASI ──────────────────────────────────────────────
            if tipe == 'relokasi' and lokasi_tujuan:
                old_lokasi = device.lokasi
                device.lokasi = lokasi_tujuan.upper()
                device.save(update_fields=['lokasi'])
                from devices.device_audit import log_edit
                import copy
                d_before = copy.copy(device)
                d_before.lokasi = old_lokasi
                log_edit(d_before, copy.copy(device), request.user)

            # ── PEMBONGKARAN — buat ItemBongkar ───────────────────────
            elif tipe == 'pembongkaran':
                if pembongkaran_tipe == 'perangkat':
                    nama_b  = device.nama
                    merk_b  = device.merk or ''
                    model_b = device.type or ''
                    sn_b    = device.serial_number or ''
                    komp_b  = None
                else:
                    nama_b  = komponen or (komponen_terkait_obj.nama if komponen_terkait_obj else '')
                    merk_b  = komponen_terkait_obj.merk if komponen_terkait_obj else ''
                    model_b = komponen_terkait_obj.model if komponen_terkait_obj else ''
                    sn_b    = komponen_terkait_obj.serial_number if komponen_terkait_obj else ''
                    komp_b  = komponen_terkait_obj

                item_b = ItemBongkar.objects.create(
                    tipe               = pembongkaran_tipe,
                    nama               = nama_b,
                    merk               = merk_b,
                    model_tipe         = model_b,
                    serial_number      = sn_b,
                    device_asal        = device,
                    komponen_terkait   = komp_b,
                    branch_id          = branch_id_event,
                    lokasi_penyimpanan = disimpan_di,
                    tanggal_bongkar    = tanggal,
                    event_bongkar      = event,
                    alasan_penggantian = alasan_penggantian,
                    catatan            = catatan,
                    created_by         = request.user,
                )
                # Jika seluruh perangkat dibongkar, update lokasi device
                # ke nama branch penyimpanan (atau kosongkan jika tidak ada branch)
                if pembongkaran_tipe == 'perangkat':
                    from devices.models import Branch
                    old_lokasi = device.lokasi
                    if branch_id_event:
                        try:
                            br_obj = Branch.objects.get(pk=branch_id_event)
                            new_lokasi = f'Gudang {br_obj.nama}'
                        except Branch.DoesNotExist:
                            new_lokasi = disimpan_di or ''
                    else:
                        new_lokasi = disimpan_di or ''
                    if new_lokasi != old_lokasi:
                        device.lokasi = new_lokasi
                        device.save(update_fields=['lokasi'])
                if komponen_terkait_obj:
                    # Jika alasan rusak/kinerja → tandai rusak; lainnya → tidak_ada
                    new_status = 'rusak' if alasan_penggantian in ('rusak', 'kinerja') else 'tidak_ada'
                    komponen_terkait_obj.status = new_status
                    if alasan_penggantian:
                        label = dict(DeviceEvent.ALASAN_CHOICES).get(alasan_penggantian, alasan_penggantian)
                        komponen_terkait_obj.keterangan = (
                            (komponen_terkait_obj.keterangan + '\n' if komponen_terkait_obj.keterangan else '')
                            + f'Dibongkar ({tanggal}): {label}'
                        )
                        komponen_terkait_obj.save(update_fields=['status', 'keterangan', 'updated_at'])
                    else:
                        komponen_terkait_obj.save(update_fields=['status', 'updated_at'])

            # ── PEMASANGAN KEMBALI — link ke ItemBongkar ──────────────
            elif tipe == 'pemasangan':
                if item_bongkar_id:
                    try:
                        item_b = ItemBongkar.objects.get(pk=item_bongkar_id, device_asal=device)
                        item_b.status = 'dipasang_kembali'
                        item_b.event_pasang = event
                        item_b.save(update_fields=['status', 'event_pasang'])
                        event.item_bongkar_ref = item_b
                        event.save(update_fields=['item_bongkar_ref'])
                        if item_b.komponen_terkait:
                            item_b.komponen_terkait.status = 'terpasang'
                            item_b.komponen_terkait.tanggal_pasang = _tz.localdate()
                            item_b.komponen_terkait.save(update_fields=['status', 'tanggal_pasang', 'updated_at'])
                    except ItemBongkar.DoesNotExist:
                        pass
                elif komponen_terkait_obj:
                    komponen_terkait_obj.status = 'terpasang'
                    komponen_terkait_obj.tanggal_pasang = _tz.localdate()
                    komponen_terkait_obj.save(update_fields=['status', 'tanggal_pasang', 'updated_at'])

            # ── PENAMBAHAN — buat DeviceComponent baru ────────────────
            elif tipe == 'penambahan':
                new_komp = DeviceComponent.objects.create(
                    device         = device,
                    nama           = komponen or merk_komponen_baru or 'Komponen Baru',
                    merk           = merk_komponen_baru,
                    model          = tipe_komponen_baru,
                    serial_number  = serial_komponen_baru,
                    posisi         = posisi_komponen_baru,
                    tipe_komponen_id = int(tipe_komponen_id) if tipe_komponen_id else None,
                    status         = 'terpasang',
                    tanggal_pasang = tanggal,
                    created_by     = request.user,
                    keterangan     = catatan,
                )
                event.komponen_terkait = new_komp
                event.save(update_fields=['komponen_terkait'])

            # ── PENGGANTIAN PERANGKAT — buat Device baru + ItemBongkar ──
            elif tipe == 'penggantian_perangkat':
                nama_baru   = request.POST.get('pp_nama', '').strip() or device.nama
                merk_baru   = request.POST.get('pp_merk', '').strip() or device.merk
                type_baru   = request.POST.get('pp_type', '').strip()
                sn_baru     = request.POST.get('pp_serial', '').strip()
                lokasi_baru = request.POST.get('pp_lokasi', '').strip() or device.lokasi
                branch_simpan = request.POST.get('branch_id', '') or None
                lokasi_simpan = disimpan_di

                # Buat Device baru sebagai pengganti
                new_device = Device.objects.create(
                    nama          = nama_baru,
                    jenis         = device.jenis,
                    merk          = merk_baru,
                    type          = type_baru or device.type,
                    serial_number = sn_baru,
                    lokasi        = lokasi_baru,
                    status_operasi = 'aktif',
                    created_by    = request.user,
                    keterangan    = f'Menggantikan: {device.nama} (terpasang {tanggal})',
                )
                event.perangkat_pengganti = new_device
                event.save(update_fields=['perangkat_pengganti'])

                # Catat perangkat lama ke ItemBongkar
                ItemBongkar.objects.create(
                    tipe               = 'perangkat',
                    nama               = device.nama,
                    merk               = device.merk or '',
                    model_tipe         = device.type or '',
                    serial_number      = device.serial_number or '',
                    device_asal        = device,
                    branch_id          = branch_simpan,
                    lokasi_penyimpanan = lokasi_simpan,
                    tanggal_bongkar    = tanggal,
                    event_bongkar      = event,
                    alasan_penggantian = alasan_penggantian,
                    catatan            = catatan,
                    created_by         = request.user,
                )

                # Update lokasi & status perangkat lama → pindah ke gudang
                from devices.models import Branch as _Branch
                lokasi_lama_baru = lokasi_simpan or ''
                if branch_simpan:
                    try:
                        br = _Branch.objects.get(pk=branch_simpan)
                        lokasi_lama_baru = f'GUDANG {br.nama.upper()}' + (f' — {lokasi_simpan}' if lokasi_simpan else '')
                    except _Branch.DoesNotExist:
                        pass
                elif lokasi_simpan:
                    lokasi_lama_baru = f'GUDANG — {lokasi_simpan}'

                update_fields = ['status_operasi']
                device.status_operasi = 'tidak_operasi'
                if lokasi_lama_baru:
                    device.lokasi = lokasi_lama_baru
                    update_fields.append('lokasi')
                if alasan_penggantian:
                    label = dict(DeviceEvent.ALASAN_CHOICES).get(alasan_penggantian, alasan_penggantian)
                    device.keterangan = (
                        (device.keterangan + '\n' if device.keterangan else '')
                        + f'Diganti ({tanggal}): {label}'
                    )
                    update_fields.append('keterangan')
                device.save(update_fields=update_fields)

                # Audit untuk device baru
                from devices.device_audit import log_create as _log_create
                _log_create(new_device, request.user)

            # ── PENGGANTIAN — buat KomponenRusak + DeviceComponent baru ──
            elif tipe == 'penggantian':
                from devices.models import KomponenRusak
                nama_rusak = komponen or (komponen_terkait_obj.nama if komponen_terkait_obj else '')
                KomponenRusak.objects.create(
                    device             = device,
                    nama_komponen      = nama_rusak,
                    merk               = komponen_terkait_obj.merk if komponen_terkait_obj else '',
                    tipe               = komponen_terkait_obj.model if komponen_terkait_obj else '',
                    komponen_terkait   = komponen_terkait_obj,
                    tanggal_rusak      = tanggal,
                    disimpan_di        = disimpan_di,
                    branch_id          = branch_id_event,
                    alasan_penggantian = alasan_penggantian,
                    keterangan         = catatan,
                    event              = event,
                    created_by         = request.user,
                )

                # Update status komponen lama
                if komponen_terkait_obj:
                    # improvement → diganti (masih bagus, bukan rusak)
                    # rusak/kinerja → rusak
                    # lainnya → diganti
                    new_status = 'rusak' if alasan_penggantian in ('rusak', 'kinerja') else 'diganti'
                    komponen_terkait_obj.status = new_status
                    komponen_terkait_obj.tanggal_ganti = _tz.localdate()
                    label = dict(DeviceEvent.ALASAN_CHOICES).get(alasan_penggantian, '') if alasan_penggantian else ''
                    if label:
                        komponen_terkait_obj.keterangan = (
                            (komponen_terkait_obj.keterangan + '\n' if komponen_terkait_obj.keterangan else '')
                            + f'Diganti ({tanggal}): {label}'
                        )
                        komponen_terkait_obj.save(update_fields=['status', 'tanggal_ganti', 'keterangan', 'updated_at'])
                    else:
                        komponen_terkait_obj.save(update_fields=['status', 'tanggal_ganti', 'updated_at'])

                # Buat DeviceComponent baru untuk komponen pengganti — selalu dibuat
                # Gunakan data baru jika diisi, fallback ke data lama
                nama_komp_baru  = nama_rusak or (komponen_terkait_obj.nama if komponen_terkait_obj else 'Komponen Baru')
                merk_final      = merk_komponen_baru or (komponen_terkait_obj.merk  if komponen_terkait_obj else '')
                model_final     = tipe_komponen_baru  or (komponen_terkait_obj.model if komponen_terkait_obj else '')
                serial_final    = serial_komponen_baru
                posisi_final    = (komponen_terkait_obj.posisi if komponen_terkait_obj else posisi_komponen_baru) or posisi_komponen_baru
                ket_baru        = (
                    f'Menggantikan: {komponen_terkait_obj.merk or ""} {komponen_terkait_obj.model or ""}'
                    + (f' SN:{komponen_terkait_obj.serial_number}' if komponen_terkait_obj and komponen_terkait_obj.serial_number else '')
                ).strip() if komponen_terkait_obj else ''

                new_komp = DeviceComponent.objects.create(
                    device           = device,
                    nama             = nama_komp_baru,
                    tipe_komponen    = komponen_terkait_obj.tipe_komponen if komponen_terkait_obj else None,
                    posisi           = posisi_final,
                    merk             = merk_final,
                    model            = model_final,
                    serial_number    = serial_final,
                    status           = 'terpasang',
                    tanggal_pasang   = tanggal,
                    keterangan       = ket_baru,
                    created_by       = request.user,
                )
                # Link event ke komponen baru
                event.komponen_terkait = new_komp
                event.save(update_fields=['komponen_terkait'])

            _audit(request, 'other', 'devices', 'Event Peralatan',
                   event.pk, f'{tipe.title()} — {device.nama}',
                   f'Komponen: {komponen} | Catatan: {catatan}')

    return redirect('device_view', pk=pk)


@login_required
def kirim_ke_gudang(request, pk):
    """
    Form untuk mengirim KomponenRusak ke gudang spare part.
    GET : tampilkan form pilih Sparepart (atau buat baru) + branch
    POST: buat MutasiSparepart masuk + link ke KomponenRusak
    """
    from devices.models import KomponenRusak
    from gudang.models import Sparepart, MutasiSparepart
    from django.contrib import messages as dj_messages

    kr = get_object_or_404(KomponenRusak, pk=pk)

    if request.method == 'POST':
        mode        = request.POST.get('mode', 'existing')  # 'existing' or 'new'
        jumlah      = int(request.POST.get('jumlah', 1) or 1)
        keperluan   = request.POST.get('keperluan', '').strip() or f'Dari penggantian komponen: {kr.nama_komponen}'
        branch_id   = request.POST.get('branch_id') or None

        if mode == 'new':
            sp = Sparepart.objects.create(
                nama               = request.POST.get('nama_baru', kr.nama_komponen).strip(),
                kategori           = request.POST.get('kategori_baru', 'Komponen Bekas').strip(),
                merk               = request.POST.get('merk_baru', kr.merk).strip(),
                part_number        = request.POST.get('part_number_baru', '').strip(),
                satuan             = 'pcs',
                lokasi_penyimpanan = '',
                branch_id          = branch_id,
                keterangan         = f'Dari komponen rusak: {kr.device.nama} ({kr.tanggal_rusak})',
                created_by         = request.user,
            )
        else:
            sp_id = request.POST.get('sparepart_id', '').strip()
            if not sp_id:
                dj_messages.error(request, 'Pilih spare part tujuan terlebih dahulu, atau gunakan opsi "Buat spare part baru".')
                return redirect('kirim_ke_gudang', pk=kr.pk)
            sp = get_object_or_404(Sparepart, pk=sp_id, is_deleted=False)

        MutasiSparepart.objects.create(
            sparepart             = sp,
            tipe                  = 'masuk',
            jumlah                = jumlah,
            keperluan             = keperluan,
            sumber_komponen_rusak = kr,
            dilakukan_oleh        = request.user,
        )

        # Update branch di KomponenRusak jika belum diset
        if branch_id and not kr.branch_id:
            kr.branch_id = branch_id
            kr.save(update_fields=['branch'])

        dj_messages.success(request, f'Komponen "{kr.nama_komponen}" berhasil dikirim ke gudang spare part.')
        return redirect('device_view', pk=kr.device_id)

    # GET
    sparepart_list_all = Sparepart.objects.filter(is_deleted=False).order_by('kategori', 'nama')
    return render(request, 'devices/kirim_ke_gudang.html', {
        'kr':                kr,
        'sparepart_list':    sparepart_list_all,
        'all_branches':      Branch.objects.all(),
    })


@login_required
@require_can_delete
def device_event_delete(request, pk, event_pk):
    """Hapus event kejadian fisik."""
    if request.method == 'POST':
        from devices.models import DeviceEvent
        event = get_object_or_404(DeviceEvent, pk=event_pk, device_id=pk)
        event.delete()
    return redirect('device_view', pk=pk)


# ── Manajemen Lokasi (Admin only) ─────────────────────────────────────────────

@login_required
@require_can_manage_lokasi
def lokasi_admin(request):

    if request.method == 'POST':
        action = request.POST.get('action')
        pk     = request.POST.get('pk')

        if action == 'add':
            nama = request.POST.get('nama', '').strip().upper()
            lat  = request.POST.get('latitude', '').strip()
            lng  = request.POST.get('longitude', '').strip()
            ket  = request.POST.get('keterangan', '').strip()
            if nama:
                SiteLocation.objects.get_or_create(
                    nama=nama,
                    defaults={
                        'latitude':    float(lat) if lat else None,
                        'longitude':   float(lng) if lng else None,
                        'keterangan':  ket,
                    }
                )

        elif action == 'edit' and pk:
            site = get_object_or_404(SiteLocation, pk=pk)
            nama = request.POST.get('nama', '').strip().upper()
            lat  = request.POST.get('latitude', '').strip()
            lng  = request.POST.get('longitude', '').strip()
            ket  = request.POST.get('keterangan', '').strip()
            if nama:
                site.nama       = nama
                site.latitude   = float(lat) if lat else None
                site.longitude  = float(lng) if lng else None
                site.keterangan = ket
                site.save()

        elif action == 'delete' and pk:
            SiteLocation.objects.filter(pk=pk).delete()

        return redirect('lokasi_admin')

    sites = SiteLocation.objects.all()
    # Hitung berapa device per lokasi
    device_count = {}
    for d in Device.objects.filter(is_deleted=False).values('lokasi'):
        loc = (d['lokasi'] or '').strip().upper()
        device_count[loc] = device_count.get(loc, 0) + 1

    sites_data = []
    for s in sites:
        sites_data.append({
            'site':         s,
            'device_count': device_count.get(s.nama.upper(), 0),
        })

    return render(request, 'devices/lokasi_admin.html', {
        'sites_data': sites_data,
    })


def api_lokasi_list(request):
    """API: kembalikan daftar lokasi sebagai JSON untuk validasi form."""
    from devices.models import SiteLocation
    locs = list(SiteLocation.objects.values_list('nama', flat=True).order_by('nama'))
    return JsonResponse({'lokasi': locs})


# ── Public Device Page (QR Code) ─────────────────────────────────────────────

def device_public(request, token):
    """Halaman publik perangkat via QR Code — tidak perlu login."""
    device = get_object_or_404(Device, public_token=token, is_deleted=False)

    # Maintenance history ringkas (10 terakhir)
    from maintenance.models import Maintenance
    maintenance_history = (
        Maintenance.objects
        .filter(device=device)
        .order_by('-date')[:10]
    )

    # Gangguan — aktif + 5 terakhir selesai
    from gangguan.models import Gangguan
    gangguan_aktif = Gangguan.objects.filter(
        peralatan=device, status__in=['open', 'in_progress']
    ).order_by('-tanggal_gangguan')
    gangguan_selesai = Gangguan.objects.filter(
        peralatan=device, status__in=['resolved', 'closed']
    ).order_by('-tanggal_gangguan')[:5]

    # Health Index
    try:
        from health_index.calculator import calculate_hi
        health_index = calculate_hi(device, save_snapshot=False)
    except Exception:
        health_index = None

    # Umur peralatan
    from datetime import date as date_type
    umur = None
    if device.tahun_operasi:
        umur = date_type.today().year - device.tahun_operasi

    # ── Inspeksi terakhir ─────────────────────────────────────────
    last_inspection  = None
    can_inspect      = False
    inspection_url   = None
    INSPECTABLE = ['Catu Daya', 'RELE DEFENSE SCHEME', 'MASTER TRIP', 'UFLS']

    jenis_name = device.jenis.name if device.jenis else ''
    if jenis_name in INSPECTABLE:
        can_inspect = True
        inspection_url = f'/inspection/form/{device.pk}/'
        try:
            from inspection.models import Inspection
            last_inspection = (
                Inspection.objects
                .filter(device=device)
                .select_related('operator')
                .order_by('-tanggal')
                .first()
            )
        except Exception:
            pass

    return render(request, 'devices/device_public.html', {
        'device':             device,
        'maintenance_history': maintenance_history,
        'gangguan_aktif':     gangguan_aktif,
        'gangguan_selesai':   gangguan_selesai,
        'health_index':       health_index,
        'umur':               umur,
        'now':                now(),
        'last_inspection':    last_inspection,
        'can_inspect':        can_inspect,
        'inspection_url':     inspection_url,
    })


# ── Peta Jaringan ─────────────────────────────────────────────────────────────

@login_required
def peta_jaringan(request):
    """Halaman peta jaringan — marker per site diwarnai berdasarkan HI rata-rata."""
    site_locations = SiteLocation.objects.filter(
        latitude__isnull=False, longitude__isnull=False
    ).order_by('nama')

    all_lokasi = (
        Device.objects
        .filter(is_deleted=False)
        .exclude(lokasi__isnull=True).exclude(lokasi__exact='')
        .values_list('lokasi', flat=True)
        .distinct()
    )
    total_lokasi    = all_lokasi.count()
    total_berkoord  = site_locations.count()
    total_tanpa     = total_lokasi - total_berkoord

    return render(request, 'devices/peta_jaringan.html', {
        'total_lokasi':   total_lokasi,
        'total_berkoord': total_berkoord,
        'total_tanpa':    total_tanpa,
    })


@login_required
def api_peta_jaringan(request):
    """
    API JSON untuk peta jaringan.
    Kembalikan semua site yang punya koordinat beserta:
    - Daftar device + skor HI masing-masing
    - Skor HI rata-rata site (untuk warna marker)
    - Jumlah gangguan aktif dan maintenance open
    """
    from health_index.calculator import calculate_hi, get_kategori
    from gangguan.models import Gangguan
    from maintenance.models import Maintenance

    site_locations = SiteLocation.objects.filter(
        latitude__isnull=False, longitude__isnull=False
    ).order_by('nama')

    gangguan_per_lokasi = {}
    for g in Gangguan.objects.filter(status__in=['open', 'in_progress']).select_related('peralatan'):
        if g.peralatan:
            lok = (g.peralatan.lokasi or '').strip()
            gangguan_per_lokasi[lok] = gangguan_per_lokasi.get(lok, 0) + 1

    maint_per_lokasi = {}
    for m in Maintenance.objects.filter(status='Open').select_related('device'):
        lok = (m.device.lokasi or '').strip()
        maint_per_lokasi[lok] = maint_per_lokasi.get(lok, 0) + 1

    result = []
    for sl in site_locations:
        devices = Device.objects.filter(
            is_deleted=False, lokasi__iexact=sl.nama
        ).select_related('jenis').order_by('jenis__name', 'nama')

        device_list = []
        scores = []
        for dev in devices:
            try:
                hi = calculate_hi(dev, save_snapshot=False)
                score = hi['score']
                kat   = hi['kategori']
            except Exception:
                score = None
                kat   = None

            scores.append(score)
            device_list.append({
                'id':        dev.pk,
                'nama':      dev.nama,
                'jenis':     dev.jenis.name if dev.jenis else '\u2014',
                'merk':      dev.merk or '',
                'type':      dev.type or '',
                'ip':        str(dev.ip_address) if dev.ip_address else '',
                'status':    dev.status_operasi,
                'hi_score':  score,
                'hi_label':  kat['label']  if kat else '\u2014',
                'hi_accent': kat['accent'] if kat else '#94a3b8',
                'hi_icon':   kat['icon']   if kat else 'bi-circle',
                'url':       f'/view/{_hid(dev.pk)}/',
                'hi_url':    f'/health-index/{_hid(dev.pk)}/',
            })

        valid_scores = [s for s in scores if s is not None]
        avg_score = round(sum(valid_scores) / len(valid_scores)) if valid_scores else None
        site_kat  = get_kategori(avg_score) if avg_score is not None else None

        result.append({
            'nama':           sl.nama,
            'lat':            sl.latitude,
            'lng':            sl.longitude,
            'keterangan':     sl.keterangan or '',
            'total_device':   len(device_list),
            'hi_avg':         avg_score,
            'hi_label':       site_kat['label']  if site_kat else '\u2014',
            'hi_accent':      site_kat['accent'] if site_kat else '#94a3b8',
            'hi_bg':          site_kat['bg']     if site_kat else '#f1f5f9',
            'gangguan_aktif': gangguan_per_lokasi.get(sl.nama, 0),
            'maint_open':     maint_per_lokasi.get(sl.nama, 0),
            'devices':        device_list,
        })

    return JsonResponse({'sites': result})


# ─────────────────────────────────────────────────────────────────────
# EVIDEN TAMBAHAN PERANGKAT
# ─────────────────────────────────────────────────────────────────────
@login_required
@require_can_edit
def device_eviden_add(request, pk):
    """Upload satu atau lebih foto eviden untuk perangkat."""
    device = get_object_or_404(Device, pk=pk)
    if request.method == 'POST':
        from devices.models import DeviceEviden
        fotos = request.FILES.getlist('eviden_foto')
        keterangans = request.POST.getlist('eviden_keterangan')
        for i, foto in enumerate(fotos):
            if not foto:
                continue
            ket = keterangans[i] if i < len(keterangans) else ''
            DeviceEviden.objects.create(
                device      = device,
                foto        = foto,
                keterangan  = ket.strip(),
                uploaded_by = request.user,
            )
    return redirect('device_view', pk=pk)


@login_required
@require_can_delete
def device_eviden_delete(request, pk, eviden_pk):
    """Hapus satu eviden."""
    device = get_object_or_404(Device, pk=pk)
    from devices.models import DeviceEviden
    eviden = get_object_or_404(DeviceEviden, pk=eviden_pk, device=device)
    # Hapus file fisik
    if eviden.foto:
        try:
            import os
            if os.path.isfile(eviden.foto.path):
                os.remove(eviden.foto.path)
        except Exception:
            pass
    eviden.delete()
    return redirect('device_view', pk=pk)


@login_required
@require_can_edit
def device_wiring(request, pk):
    """Editor wiring diagram perangkat."""
    device = get_object_or_404(Device, pk=pk, is_deleted=False)
    if request.method == 'POST':
        import json as _json, base64 as _b64
        from django.core.files.base import ContentFile as _CF
        wiring_raw = request.POST.get('wiring_json', '')
        img_data   = request.POST.get('wiring_img', '')
        update_fields = []
        if wiring_raw:
            try:
                device.wiring_json = _json.loads(wiring_raw)
                update_fields.append('wiring_json')
            except Exception:
                pass
        if img_data and img_data.startswith('data:image/png;base64,'):
            try:
                img_bytes = _b64.b64decode(img_data.split(',')[1])
                fname = f'wiring_{device.pk}.png'
                if device.wiring_img:
                    try:
                        import os; os.remove(device.wiring_img.path)
                    except Exception:
                        pass
                device.wiring_img.save(fname, _CF(img_bytes), save=False)
                update_fields.append('wiring_img')
            except Exception:
                pass
        if update_fields:
            device.save(update_fields=update_fields)
        from django.http import JsonResponse as _JR
        return _JR({'ok': True})


@login_required
def global_search(request):
    query = request.GET.get('q', '').strip()
    devices = []
    gangguan_results = []
    maintenance_results = []

    if query:
        from gangguan.models import Gangguan

        devices = list(
            Device.objects.filter(is_deleted=False, host__isnull=True).filter(
                Q(nama__icontains=query) |
                Q(ip_address__icontains=query) |
                Q(merk__icontains=query) |
                Q(serial_number__icontains=query) |
                Q(lokasi__icontains=query)
            ).select_related('jenis').order_by('nama')[:30]
        )

        gangguan_results = list(
            Gangguan.objects.filter(
                Q(nomor_gangguan__icontains=query) |
                Q(site__icontains=query) |
                Q(executive_summary__icontains=query)
            ).order_by('-tanggal_gangguan')[:20]
        )

        maintenance_results = list(
            Maintenance.objects.filter(
                Q(device__nama__icontains=query) |
                Q(device__lokasi__icontains=query)
            ).select_related('device').order_by('-date')[:20]
        )

    total = len(devices) + len(gangguan_results) + len(maintenance_results)

    return render(request, 'devices/global_search.html', {
        'query': query,
        'devices': devices,
        'gangguan_results': gangguan_results,
        'maintenance_results': maintenance_results,
        'total': total,
    })


@login_required
def global_search_api(request):
    """Live search API — max 5 perangkat, dipakai oleh dropdown topbar."""
    from django.urls import reverse as _rev
    query = request.GET.get('q', '').strip()
    if len(query) < 2:
        return JsonResponse({'results': [], 'total': 0, 'query': query})

    qs = (
        Device.objects.filter(is_deleted=False, host__isnull=True)
        .filter(
            Q(nama__icontains=query) |
            Q(ip_address__icontains=query) |
            Q(merk__icontains=query) |
            Q(serial_number__icontains=query) |
            Q(lokasi__icontains=query)
        )
        .select_related('jenis')
        .order_by('nama')
    )
    total = qs.count()
    results = []
    for d in qs[:5]:
        results.append({
            'id':             d.pk,
            'nama':           d.nama,
            'jenis':          d.jenis.name if d.jenis else '',
            'merk':           d.merk or '',
            'type':           d.type or '',
            'ip_address':     d.ip_address or '',
            'lokasi':         d.lokasi or '',
            'status_operasi': d.status_operasi,
            'foto_url':       d.foto.url if d.foto else '',
            'url':            _rev('device_view', args=[d.pk]),
        })
    return JsonResponse({'results': results, 'total': total, 'query': query})
