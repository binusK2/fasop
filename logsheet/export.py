"""
Export logsheet ke Excel dengan mengisi TEMPLATE asli (format identik).

Template = workbook LOGSHEET_MKS milik operator (VBA dibuang), disimpan di
logsheet/xlsx_template/. Untuk tanggal tertentu, nilai per titik diisi ke 48
kolom waktu pada sheet & baris sesuai LogsheetTitik, menggantikan formula latch
dengan angka statis. Hasilnya sama persis dengan logsheet Excel lama.
"""
import os
from collections import defaultdict

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'xlsx_template',
                             'LOGSHEET_MKS_template.xlsx')

# Kolom (1-based) tempat slot 0 (pukul 00:30) berada, per sheet. 48 slot berturut
# ke kanan. Ditambah saat tiap sheet diverifikasi. F = 6.
SHEET_SLOT_COL0 = {
    'KIT_MW':     6,
    'KIT_MVAR':   6,
    'Trans_MW':   6,
    'Trans_MVar': 6,
    'Trans_Amp':  6,
    'Trans_VBus': 6,
    # 'TRAFO': ...  (trafo menyusul)
}
JUMLAH_SLOT = 48


def load_template():
    import openpyxl
    return openpyxl.load_workbook(TEMPLATE_PATH, data_only=False, keep_vba=False)


def build_workbook(tanggal):
    """
    Kembalikan openpyxl Workbook template yang sudah diisi nilai untuk `tanggal`.
    Hanya sheet yang punya SHEET_SLOT_COL0 & titik ber-baris yang diisi.
    """
    from .models import LogsheetTitik, LogsheetNilai

    wb = load_template()
    titik_qs = (LogsheetTitik.objects
                .filter(aktif=True, baris__isnull=False)
                .exclude(sheet='')
                .exclude(baris=0))
    titik_by_id = {t.id: t for t in titik_qs if t.sheet in SHEET_SLOT_COL0}
    if not titik_by_id:
        return wb

    # 1) Bersihkan SEMUA formula latch (yang mereferensi DASHBOARD) di band 48
    #    kolom data tiap sheet yang dikelola — supaya slot tanpa data tampil
    #    kosong, bukan nilai basi. Formula lain (subtotal/label) dibiarkan.
    sheets = {t.sheet for t in titik_by_id.values()}
    for sheet in sheets:
        col0 = SHEET_SLOT_COL0[sheet]
        ws = wb[sheet]
        for row in ws.iter_rows(min_col=col0, max_col=col0 + JUMLAH_SLOT - 1):
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith('=') and 'DASHBOARD' in v:
                    cell.value = None

    # 2) Isi nilai yang tersedia untuk tanggal ini
    nilai_qs = (LogsheetNilai.objects
                .filter(titik_id__in=titik_by_id.keys(), tanggal=tanggal,
                        nilai__isnull=False)
                .values_list('titik_id', 'slot', 'nilai'))
    for titik_id, slot, nilai in nilai_qs:
        if not (0 <= slot < JUMLAH_SLOT):
            continue
        t = titik_by_id[titik_id]
        col0 = SHEET_SLOT_COL0[t.sheet]
        wb[t.sheet].cell(row=t.baris, column=col0 + slot).value = round(nilai, 2)
    return wb


def workbook_to_response(wb, filename):
    """Serialisasi Workbook -> HttpResponse unduhan .xlsx."""
    import io
    from django.http import HttpResponse
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp
