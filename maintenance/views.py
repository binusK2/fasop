from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from devices.permissions import require_can_edit, require_can_delete, is_viewer_only
from .models import Maintenance, MaintenancePLC, MaintenanceRouter, MaintenanceRadio, MaintenanceVoIP, MaintenanceMux, MaintenanceRectifier, MaintenanceTeleproteksi, MaintenanceGenset, MaintenanceRTU, MaintenanceSAS, MaintenanceRoIP, MaintenanceUPS, BeritaAcaraRecord, BeritaAcaraEviden
from .forms import MaintenanceForm, MaintenancePLCForm, MaintenanceRouterForm, MaintenanceRadioForm, MaintenanceVoIPForm, MaintenanceMuxForm, MaintenanceRectifierForm, MaintenanceTeleproteksiForm, MaintenanceGensetForm, MaintenanceRTUForm, MaintenanceSASForm, MaintenanceRoIPForm, MaintenanceUPSForm
from devices.models import Device, DeviceType
from gangguan.models import Gangguan
from inspection.models import InspectionCatuDaya
from django.db.models import Q, Count
from django.db.models.functions import Trim
from django.http import HttpResponse
from io import BytesIO
import openpyxl
import json
from datetime import date as date_cls
from django.utils import timezone as dj_timezone
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date


# ─────────────────────────────────────────────────────────────────────
# MAPPING: nama jenis perangkat → (form class, template)
# Tambahkan jenis baru di sini tanpa ubah view!
# ─────────────────────────────────────────────────────────────────────
DEVICE_FORM_MAP = {
    'PLC':    (MaintenancePLCForm,    'maintenance/plc_form.html'),
    'ROUTER': (MaintenanceRouterForm, 'maintenance/router_form.html'),
    'SWITCH': (MaintenanceRouterForm, 'maintenance/switch_form.html'),
    'RADIO': (MaintenanceRadioForm, 'maintenance/radio_form.html'),
    'VOIP':        (MaintenanceVoIPForm,  'maintenance/voip_form.html'),
    'MULTIPLEXER':  (MaintenanceMuxForm,        'maintenance/mux_form.html'),
    'RECTIFIER':    (MaintenanceRectifierForm,   'maintenance/rectifier_form.html'),
    'CATU DAYA':    (MaintenanceRectifierForm,   'maintenance/rectifier_form.html'),
    'CATUDAYA':     (MaintenanceRectifierForm,   'maintenance/rectifier_form.html'),
    'RECTIFIER & BATTERY': (MaintenanceRectifierForm, 'maintenance/rectifier_form.html'),
    'TELEPROTEKSI':        (MaintenanceTeleproteksiForm, 'maintenance/teleproteksi_form.html'),
    'GENSET':              (MaintenanceGensetForm,       'maintenance/genset_form.html'),
    'RTU':                 (MaintenanceRTUForm,          'maintenance/rtu_form.html'),
    'SAS':                 (MaintenanceSASForm,          'maintenance/sas_form.html'),
    'SERVER SCADA':        (MaintenanceSASForm,          'maintenance/sas_form.html'),
    'GATEWAY SAS':         (MaintenanceSASForm,          'maintenance/sas_form.html'),
    'ROIP':   (MaintenanceRoIPForm, 'maintenance/roip_form.html'),
    'RoIP':   (MaintenanceRoIPForm, 'maintenance/roip_form.html'),
    'UPS':    (MaintenanceUPSForm,  'maintenance/ups_form.html'),
}

DEFAULT_TEMPLATE = 'maintenance/maintenance_form.html'


def _build_sas_context(dform):
    """Build extra context rows untuk template sas_form.html."""
    if dform is None or dform.__class__.__name__ != 'MaintenanceSASForm':
        return {}
    return {
        'spek_items': [
            ('Merk',                 dform['spek_merk']),
            ('Type',                 dform['spek_type']),
            ('CPU',                  dform['spek_cpu']),
            ('RAM',                  dform['spek_ram']),
            ('GPU',                  dform['spek_gpu']),
            ('Storage Memory',       dform['spek_storage']),
            ('Firmware Version',     dform['spek_firmware']),
            ('Configuration Version',dform['spek_config_ver']),
            ('Maintenance IP',       dform['spek_ip']),
        ],
        'peri_ok_alarm_rows': [
            (dform['peri_eth_switch'], 'Ethernet Switch'),
            (dform['peri_gps'],        'GPS'),
            (dform['peri_eth_serial'], 'Ethernet to Serial'),
            (dform['peri_router'],     'Router'),
        ],
        'perf_status_rows': [
            (dform['indikasi_alarm'], 'Indikasi Alarm / Error'),
            (dform['komm_master'],    'Komunikasi ke Master Station'),
            (dform['komm_ied'],       'Komunikasi ke IED'),
            (dform['time_sync'],      'Time Synchronization'),
        ],
    }


def _get_detail_form_config(device):
    """Return (FormClass, template) berdasarkan jenis perangkat."""
    if not device.jenis:
        return None, DEFAULT_TEMPLATE
    key = device.jenis.name.strip().upper()
    form_class, template = DEVICE_FORM_MAP.get(key, (None, DEFAULT_TEMPLATE))
    return form_class, template


# ─────────────────────────────────────────────────────────────────────
# LIST
# ─────────────────────────────────────────────────────────────────────
@login_required
def maintenance_list(request):
    status    = request.GET.get('status') or ''
    lokasi    = request.GET.get('lokasi') or ''
    jenis_id  = request.GET.get('jenis') or ''
    date_from = request.GET.get('date_from') or ''
    date_to   = request.GET.get('date_to') or ''

    maintenances = Maintenance.objects.select_related(
        'device', 'device__jenis'
    ).order_by('-date')

    if status:    maintenances = maintenances.filter(status=status)
    if lokasi:    maintenances = maintenances.filter(device__lokasi__iexact=lokasi)
    if jenis_id:  maintenances = maintenances.filter(device__jenis_id=jenis_id)
    if date_from: maintenances = maintenances.filter(date__gte=date_from)
    if date_to:   maintenances = maintenances.filter(date__lte=date_to)

    lokasi_list = (
        Maintenance.objects.select_related('device')
        .exclude(device__lokasi__isnull=True)
        .exclude(device__lokasi__exact='')
        .exclude(device__lokasi__iexact='none')
        .annotate(lokasi_clean=Trim('device__lokasi'))
        .values_list('lokasi_clean', flat=True)
        .distinct().order_by('lokasi_clean')
    )

    return render(request, 'maintenance/maintenance_list.html', {
        'maintenances':    maintenances,
        'lokasi_list':     lokasi_list,
        'selected_lokasi': lokasi,
        'selected_status': status,
        'jenis_list':      DeviceType.objects.all(),
        'selected_jenis':  jenis_id,
        'date_from':       date_from,
        'date_to':         date_to,
    })


# ─────────────────────────────────────────────────────────────────────
# CREATE  (otomatis pilih form berdasarkan jenis perangkat)
# ─────────────────────────────────────────────────────────────────────
@login_required
def maintenance_create(request, device_id):
    device = get_object_or_404(Device, id=device_id)
    detail_form_class, template = _get_detail_form_config(device)

    if request.method == 'POST':
        mform = MaintenanceForm(request.POST, request.FILES)
        dform = detail_form_class(request.POST) if detail_form_class else None

        if mform.is_valid() and (dform is None or dform.is_valid()):
            maintenance = mform.save(commit=False)
            maintenance.device = device
            maintenance.maintenance_type = 'Preventive'  # FIX: disabled widget tidak dikirim browser
            maintenance.status = 'Open'  # selalu Open saat baru dibuat
            # Simpan nama pelaksana manual (JSON array dari tag-input)
            names_raw = request.POST.get('pelaksana_names', '[]')
            try:
                maintenance.pelaksana_names = json.loads(names_raw)
            except (json.JSONDecodeError, ValueError):
                maintenance.pelaksana_names = []
            maintenance.save()

            if dform:
                detail = dform.save(commit=False)
                detail.maintenance = maintenance
                detail.save()

            return redirect('maintenance_list')
    else:
        mform = MaintenanceForm()
        dform = detail_form_class() if detail_form_class else None

    slot_fields = []
    if dform and dform.__class__.__name__ == 'MaintenanceMuxForm':
        for letter in 'ABCDEFGH':
            l = letter.lower()
            slot_fields.append((letter, dform['slot_' + l + '_modul'], dform['slot_' + l + '_isian']))

    sas_ctx = _build_sas_context(dform)

    return render(request, template, {
        'maintenance_form': mform,
        'detail_form':      dform,
        'device':           device,
        'slot_fields':      slot_fields,
        **sas_ctx,
    })


# ─────────────────────────────────────────────────────────────────────
# UPDATE STATUS
# ─────────────────────────────────────────────────────────────────────
@login_required
def maintenance_update_status(request, pk):
    maintenance = get_object_or_404(Maintenance, pk=pk)
    if request.method == 'POST':
        old_status = maintenance.status
        maintenance.status = request.POST.get('status')
        maintenance.save()
        # Notif ke AM kalau baru selesai (Done) dan perlu TTD
        if old_status != 'Done' and maintenance.status == 'Done':
            try:
                from notifikasi.views import notif_ke_am
                notif_ke_am(
                    tipe   = 'maintenance_ttd',
                    judul  = f'Maintenance selesai — {maintenance.device.nama}',
                    pesan  = (
                        f'Maintenance {maintenance.maintenance_type} pada '
                        f'{maintenance.device.nama} ({maintenance.device.lokasi}) '
                        f'tanggal {maintenance.date.strftime("%d %b %Y")} '
                        f'telah selesai dan memerlukan pengesahan.'
                    ),
                    level  = 'info',
                    url    = f'/maintenance/view/{maintenance.pk}/',
                    device = maintenance.device,
                )
            except Exception:
                pass
    return redirect('maintenance_list')


# ─────────────────────────────────────────────────────────────────────
# DETAIL
# ─────────────────────────────────────────────────────────────────────
@login_required
def maintenance_detail(request, pk):
    maintenance = get_object_or_404(Maintenance, pk=pk)
    device      = maintenance.device
    device_type = device.jenis.name.strip().upper() if device.jenis else ''

    plc_detail    = None
    router_detail = None
    radio_detail  = None
    voip_detail   = None
    mux_detail    = None
    rect_detail   = None
    tp_detail     = None
    genset_detail = None
    rtu_detail    = None
    sas_detail    = None
    roip_detail   = None
    ups_detail    = None

    if device_type == 'PLC':
        try:
            plc_detail = maintenance.maintenanceplc
        except MaintenancePLC.DoesNotExist:
            pass

    elif device_type in ('ROUTER', 'SWITCH'):
        try:
            router_detail = maintenance.maintenancerouter
        except MaintenanceRouter.DoesNotExist:
            pass

    elif device_type == 'RADIO':
        try:
            radio_detail = maintenance.maintenanceradio
        except MaintenanceRadio.DoesNotExist:
            pass

    elif device_type == 'VOIP':
        try:
            voip_detail = maintenance.maintenancevoip
        except MaintenanceVoIP.DoesNotExist:
            pass

    elif device_type == 'MULTIPLEXER':
        try:
            mux_detail = maintenance.maintenancemux
        except MaintenanceMux.DoesNotExist:
            pass

    elif device_type in ('RECTIFIER', 'CATU DAYA', 'CATUDAYA', 'RECTIFIER & BATTERY'):
        try:
            rect_detail = maintenance.maintenancerectifier
        except MaintenanceRectifier.DoesNotExist:
            pass

    elif device_type == 'TELEPROTEKSI':
        try:
            tp_detail = maintenance.maintenanceteleproteksi
        except MaintenanceTeleproteksi.DoesNotExist:
            pass

    elif device_type == 'GENSET':
        try:
            genset_detail = maintenance.maintenancegenset
        except MaintenanceGenset.DoesNotExist:
            pass

    elif device_type == 'RTU':
        try:
            rtu_detail = maintenance.maintenancertu
        except MaintenanceRTU.DoesNotExist:
            pass

    elif device_type in ('SAS', 'SERVER SCADA', 'GATEWAY SAS'):
        try:
            sas_detail = maintenance.maintenancesas
        except MaintenanceSAS.DoesNotExist:
            pass

    elif device_type in ('ROIP',):
        try:
            roip_detail = maintenance.maintenanceroip
        except MaintenanceRoIP.DoesNotExist:
            pass

    elif device_type == 'UPS':
        try:
            ups_detail = maintenance.maintenanceups
        except MaintenanceUPS.DoesNotExist:
            pass

    # Checklist peralatan terpasang untuk template radio
    radio_checklist = []
    if radio_detail:
        radio_checklist = [
            ('ada_radio',        'Radio',         radio_detail.ada_radio),
            ('ada_battery',      'Battery',       radio_detail.ada_battery),
            ('ada_power_supply', 'Power Supply',  radio_detail.ada_power_supply),
        ]

    # Checklist fisik untuk template router
    router_checklist = []
    if router_detail and device_type == 'ROUTER':
        router_checklist = [
            ('Kondisi Fisik Unit',       router_detail.kondisi_fisik),
            ('Indikator LED Link/Port',  router_detail.led_link),
            ('Kondisi Kabel & Konektor', router_detail.kondisi_kabel),
        ]

    # Checklist fisik untuk template switch
    switch_checklist = []
    if router_detail and device_type == 'SWITCH':
        switch_checklist = [
            ('Kondisi Fisik Unit',       router_detail.kondisi_fisik),
            ('Indikator LED Link/Port',  router_detail.led_link),
            ('Kondisi Kabel & Konektor', router_detail.kondisi_kabel),
            ('Status Switching / VLAN',  router_detail.status_routing),
        ]



    # Context untuk Rectifier detail
    rect_list = []
    bat_list  = []
    if rect_detail:
        d = rect_detail
        # rect_list: (rn, merk, tipe, kondisi, kapasitas, v_rect, v_bat, teg_pos, teg_neg, v_drop, a_rect, a_bat, a_load)
        for n in [1, 2]:
            rect_list.append((
                n,
                getattr(d, f'rect{n}_merk', ''),
                getattr(d, f'rect{n}_tipe', ''),
                getattr(d, f'rect{n}_kondisi', ''),
                getattr(d, f'rect{n}_kapasitas', ''),
                getattr(d, f'rect{n}_v_rectifier', None),
                getattr(d, f'rect{n}_v_battery', None),
                getattr(d, f'rect{n}_teg_pos_ground', None),
                getattr(d, f'rect{n}_teg_neg_ground', None),
                getattr(d, f'rect{n}_v_dropper', None),
                getattr(d, f'rect{n}_a_rectifier', None),
                getattr(d, f'rect{n}_a_battery', None),
                getattr(d, f'rect{n}_a_load', None),
            ))
        # bat_list: (bn, merk, tipe, kondisi, kapasitas, jumlah, kabel, mur, sel_rak, air, v_total, v_load, cells)
        for n in [1, 2]:
            bat_list.append((
                n,
                getattr(d, f'bat{n}_merk', ''),
                getattr(d, f'bat{n}_tipe', ''),
                getattr(d, f'bat{n}_kondisi', ''),
                getattr(d, f'bat{n}_kapasitas', ''),
                getattr(d, f'bat{n}_jumlah', None),
                getattr(d, f'bat{n}_kondisi_kabel', ''),
                getattr(d, f'bat{n}_kondisi_mur_baut', ''),
                getattr(d, f'bat{n}_kondisi_sel_rak', ''),
                getattr(d, f'bat{n}_air_battery', None),
                getattr(d, f'bat{n}_v_total', None),
                getattr(d, f'bat{n}_v_load', None),
                getattr(d, f'bat{n}_cells', []),
            ))

    # Context tambahan untuk MUX detail
    mux_slots = []
    mux_psu_list = []
    hs_list = []
    if mux_detail:
        for letter in 'ABCDEFGH':
            l = letter.lower()
            mux_slots.append((
                letter,
                getattr(mux_detail, f'slot_{l}_modul', ''),
                getattr(mux_detail, f'slot_{l}_isian', ''),
            ))
        mux_psu_list = [
            ('PSU 1', mux_detail.psu1_status, mux_detail.psu1_temp1, mux_detail.psu1_temp2, mux_detail.psu1_temp3),
            ('PSU 2', mux_detail.psu2_status, mux_detail.psu2_temp1, mux_detail.psu2_temp2, mux_detail.psu2_temp3),
            ('FAN',   mux_detail.fan_status,  None, None, None),
        ]
        hs_list = [('1', 'hs1'), ('2', 'hs2')]

    # Checklist untuk Mux (PSU & FAN status)
    mux_checklist = []
    if mux_detail:
        mux_checklist = [
            ('PSU 1', mux_detail.psu1_status),
            ('PSU 2', mux_detail.psu2_status),
            ('FAN',   mux_detail.fan_status),
        ]

    # Checklist untuk VoIP
    voip_checklist = []
    if voip_detail:
        voip_checklist = [
            ('Kondisi Fisik Perangkat', voip_detail.kondisi_fisik),
            ('NTP Server',             voip_detail.ntp_server),
            ('Web Config',             voip_detail.webconfig),
            ('Status Power Supply',    voip_detail.ps_status),
        ]

    return render(request, 'maintenance/maintenance_detail.html', {
        'maintenance':      maintenance,
        'device_type':      device_type,
        'plc_detail':       plc_detail,
        'router_detail':    router_detail,
        'radio_detail':     radio_detail,
        'voip_detail':      voip_detail,
        'voip_checklist':   voip_checklist,
        'mux_detail':       mux_detail,
        'mux_checklist':    mux_checklist,
        'mux_slots':        mux_slots,
        'mux_psu_list':     mux_psu_list,
        'hs_list':          hs_list,
        'rect_detail':      rect_detail,
        'rect_list':        rect_list,
        'bat_list':         bat_list,
        'rect_v_list':      [('Rectifier', rect_detail.rect1_v_rectifier if rect_detail else None, 'V'), ('Battery', rect_detail.rect1_v_battery if rect_detail else None, 'V'), ('V Load', rect_detail.rect1_v_load if rect_detail else None, 'V'), ('Teg(+) GND', rect_detail.rect1_teg_pos_ground if rect_detail else None, 'V'), ('Teg(-) GND', rect_detail.rect1_teg_neg_ground if rect_detail else None, 'V'), ('Dropper', rect_detail.rect1_v_dropper if rect_detail else None, 'V')] if rect_detail else [],
        'rect_a_list':      [('Rectifier', rect_detail.rect1_a_rectifier if rect_detail else None), ('Battery', rect_detail.rect1_a_battery if rect_detail else None), ('Load', rect_detail.rect1_a_load if rect_detail else None)] if rect_detail else [],
        'bat_kondisi_list': [('Kabel Battery', rect_detail.bat1_kondisi_kabel if rect_detail else ''), ('Mur & Baut', rect_detail.bat1_kondisi_mur_baut if rect_detail else ''), ('Sel & Rak', rect_detail.bat1_kondisi_sel_rak if rect_detail else '')] if rect_detail else [],
        'radio_checklist':  radio_checklist,
        'router_checklist': router_checklist,
        'switch_checklist': switch_checklist,
        'tp_detail':        tp_detail,
        'tp_akses_list':    [
            ('Akses TP',        tp_detail.akses_tp        if tp_detail else ''),
            ('Remote Akses TP', tp_detail.remote_akses_tp if tp_detail else ''),
        ] if tp_detail else [],
        'tp_skema_list': [
            {'n': n, 'command': getattr(tp_detail, f'skema_{n}_command', ''),
             'send_minus': getattr(tp_detail, f'skema_{n}_send_minus', None),
             'send_plus':  getattr(tp_detail, f'skema_{n}_send_plus',  None),
             'receive_minus': getattr(tp_detail, f'skema_{n}_receive_minus', None),
             'receive_plus':  getattr(tp_detail, f'skema_{n}_receive_plus',  None),
            } for n in range(1, 5)
        ] if tp_detail else [],
        'tp_uji_list': [
            ('Send Command 1',    tp_detail.skema_1_send_result    if tp_detail else ''),
            ('Receive Command 1', tp_detail.skema_1_receive_result if tp_detail else ''),
            ('Send Command 2',    tp_detail.skema_2_send_result    if tp_detail else ''),
            ('Receive Command 2', tp_detail.skema_2_receive_result if tp_detail else ''),
            ('Send Command 3',    tp_detail.skema_3_send_result    if tp_detail else ''),
            ('Receive Command 3', tp_detail.skema_3_receive_result if tp_detail else ''),
            ('Send Command 4',    tp_detail.skema_4_send_result    if tp_detail else ''),
            ('Receive Command 4', tp_detail.skema_4_receive_result if tp_detail else ''),
        ] if tp_detail else [],
        'genset_detail': genset_detail,
        'rtu_detail':    rtu_detail,
        'sas_detail':    sas_detail,
        # (label, val, ok_val, nok_val)
        'sas_kondisi_rows': [
            ('Kondisi Server/Gateway', sas_detail.kondisi_server,  'BERSIH', 'TIDAK BERSIH'),
            ('Kondisi Panel',          sas_detail.kondisi_panel,   'BERSIH', 'TIDAK BERSIH'),
            ('Exhaust Fan',            sas_detail.exhaust_fan,     'ADA, BERFUNGSI', 'ADA, TIDAK BERFUNGSI'),
            ('Indikasi Alarm/Error',   sas_detail.indikasi_alarm,  'TIDAK ADA', 'ADA'),
            ('Komm. Master Station',   sas_detail.komm_master,     'OK', 'ALARM'),
            ('Komm. IED',              sas_detail.komm_ied,        'OK', 'ALARM'),
            ('Time Synchronization',   sas_detail.time_sync,       'OK', 'NOK'),
        ] if sas_detail else [],
        'sas_peri_rows': [
            ('Ethernet Switch', sas_detail.peri_eth_switch),
            ('GPS',             sas_detail.peri_gps),
            ('Eth to Serial',   sas_detail.peri_eth_serial),
            ('Router',          sas_detail.peri_router),
        ] if sas_detail else [],
        'sas_perf_rows': [
            ('CPU Terpakai',     sas_detail.perf_cpu),
            ('RAM Terpakai',     sas_detail.perf_ram),
            ('Storage Terpakai', sas_detail.perf_storage),
        ] if sas_detail else [],
        'batere_list': [
            ('Air Accu',       genset_detail.air_accu        if genset_detail else None, 'mm'),
            ('Teg. Batere',    genset_detail.tegangan_batere if genset_detail else None, 'VDC'),
            ('Arus Pengisian', genset_detail.arus_pengisian  if genset_detail else None, 'A'),
        ] if genset_detail else [],
        'charger_list': [
            ('Teg. Charger', genset_detail.tegangan_charger   if genset_detail else None, 'VDC'),
            ('Arus Beban',   genset_detail.arus_beban_charger if genset_detail else None, 'A'),
        ] if genset_detail else [],
        'mdf_list': [
            ('Oil Pressure',      genset_detail.oil_pressure       if genset_detail else None, 'Kpa'),
            ('Engine Temp',       genset_detail.engine_temperature  if genset_detail else None, '°C'),
            ('Batere Condition',  genset_detail.batere_condition    if genset_detail else None, 'VDC'),
            ('RPM',               genset_detail.rpm                 if genset_detail else None, 'rpm'),
        ] if genset_detail else [],
        'roip_detail': roip_detail,
        'ups_detail':  ups_detail,
        'ups_ac_list': [
            ('V Input R-N',   ups_detail.v_input_r  if ups_detail else None, 'V'),
            ('V Input S-N',   ups_detail.v_input_s  if ups_detail else None, 'V'),
            ('V Input T-N',   ups_detail.v_input_t  if ups_detail else None, 'V'),
            ('Frek. Input',   ups_detail.f_input    if ups_detail else None, 'Hz'),
            ('V Output R-N',  ups_detail.v_output_r if ups_detail else None, 'V'),
            ('V Output S-N',  ups_detail.v_output_s if ups_detail else None, 'V'),
            ('V Output T-N',  ups_detail.v_output_t if ups_detail else None, 'V'),
            ('Frek. Output',  ups_detail.f_output   if ups_detail else None, 'Hz'),
            ('Arus Beban',    ups_detail.a_load      if ups_detail else None, 'A'),
            ('% Beban',       ups_detail.percent_load if ups_detail else None, '%'),
        ] if ups_detail else [],
        'ups_cells': [
            {
                'num': str(c.get('cell', '')).zfill(2),
                'vf':  c.get('v_float'),
                'vd0': c.get('vd_0'),
                'vd1': c.get('vd_1'),
                'vd2': c.get('vd_2'),
                'vd3': c.get('vd_3'),
            }
            for c in (ups_detail.bat_cells or [])
            if isinstance(c.get('cell'), int)
        ] if ups_detail else [],
    })




# ─────────────────────────────────────────────────────────────────────
# EDIT  (hanya status Open)
# ─────────────────────────────────────────────────────────────────────
@login_required
def maintenance_edit(request, pk):
    maintenance = get_object_or_404(Maintenance, pk=pk)
    # Corrective punya form & flow sendiri — jangan campur dengan preventive
    if maintenance.maintenance_type == 'Corrective':
        return redirect('corrective_edit', pk=pk)
    device      = maintenance.device
    detail_form_class, template = _get_detail_form_config(device)

    # Ambil detail object yang sudah ada (jika ada)
    detail_instance = None
    if detail_form_class:
        try:
            if detail_form_class.__name__ == 'MaintenancePLCForm':
                detail_instance = maintenance.maintenanceplc
            elif detail_form_class.__name__ == 'MaintenanceRouterForm':
                detail_instance = maintenance.maintenancerouter
            elif detail_form_class.__name__ == 'MaintenanceRadioForm':
                detail_instance = maintenance.maintenanceradio
            elif detail_form_class.__name__ == 'MaintenanceVoIPForm':
                detail_instance = maintenance.maintenancevoip
            elif detail_form_class.__name__ == 'MaintenanceMuxForm':
                detail_instance = maintenance.maintenancemux
            elif detail_form_class.__name__ == 'MaintenanceRectifierForm':
                detail_instance = maintenance.maintenancerectifier
            elif detail_form_class.__name__ == 'MaintenanceTeleproteksiForm':
                detail_instance = maintenance.maintenanceteleproteksi
            elif detail_form_class.__name__ == 'MaintenanceGensetForm':
                detail_instance = maintenance.maintenancegenset
            elif detail_form_class.__name__ == 'MaintenanceRTUForm':
                detail_instance = maintenance.maintenancertu
            elif detail_form_class.__name__ == 'MaintenanceSASForm':
                detail_instance = maintenance.maintenancesas
            elif detail_form_class.__name__ == 'MaintenanceRoIPForm':
                detail_instance = maintenance.maintenanceroip
            elif detail_form_class.__name__ == 'MaintenanceUPSForm':
                detail_instance = maintenance.maintenanceups
        except Exception:
            pass

    # Gunakan template edit yang sama dengan create
    edit_template = template  # reuse template yang sama

    if request.method == 'POST':
        mform = MaintenanceForm(request.POST, request.FILES, instance=maintenance)
        dform = detail_form_class(request.POST, instance=detail_instance) if detail_form_class else None

        if mform.is_valid() and (dform is None or dform.is_valid()):
            m = mform.save(commit=False)
            m.maintenance_type = 'Preventive'  # FIX: pastikan tidak hilang
            names_raw = request.POST.get('pelaksana_names', '[]')
            try:
                m.pelaksana_names = json.loads(names_raw)
            except (json.JSONDecodeError, ValueError):
                m.pelaksana_names = []
            m.save()
            if dform:
                detail = dform.save(commit=False)
                detail.maintenance = maintenance
                detail.save()
            return redirect('maintenance_view', pk=pk)
    else:
        mform = MaintenanceForm(instance=maintenance)
        dform = detail_form_class(instance=detail_instance) if detail_form_class else None

    slot_fields_edit = []
    if dform and dform.__class__.__name__ == 'MaintenanceMuxForm':
        for letter in 'ABCDEFGH':
            l = letter.lower()
            slot_fields_edit.append((letter, dform['slot_' + l + '_modul'], dform['slot_' + l + '_isian']))

    sas_ctx = _build_sas_context(dform)

    return render(request, edit_template, {
        'maintenance_form': mform,
        'detail_form':      dform,
        'device':           device,
        'is_edit':          True,
        'maintenance':      maintenance,
        'slot_fields':      slot_fields_edit,
        'pelaksana_names_json': json.dumps(maintenance.pelaksana_names or []),
        **sas_ctx,
    })

# ─────────────────────────────────────────────────────────────────────
# DELETE
# ─────────────────────────────────────────────────────────────────────
@login_required
@require_can_delete
def maintenance_delete(request, pk):
    maintenance = get_object_or_404(Maintenance, pk=pk)
    maintenance.delete()
    return redirect('maintenance_list')


# ─────────────────────────────────────────────────────────────────────
# LAPORAN BULANAN
# ─────────────────────────────────────────────────────────────────────
@login_required
def maintenance_report(request):
    today          = date.today()
    mode           = request.GET.get('mode', 'monthly')  # 'monthly' atau 'ytd'
    selected_month = int(request.GET.get('month') or today.month)
    selected_year  = int(request.GET.get('year')  or today.year)

    month_names = [
        'Januari','Februari','Maret','April','Mei','Juni',
        'Juli','Agustus','September','Oktober','November','Desember'
    ]

    if mode == 'ytd':
        # Bulan berjalan: Januari s/d bulan sekarang di tahun selected_year
        current_month = today.month if selected_year == today.year else 12
        maintenances = (
            Maintenance.objects
            .filter(date__year=selected_year, date__month__lte=current_month)
            .select_related('device', 'device__jenis', 'signed_by')
            .order_by('date')
        )
        period_label = f"Januari — {month_names[current_month-1]} {selected_year}"
        # Ringkasan per bulan untuk YTD
        monthly_summary = []
        for m in range(1, current_month + 1):
            m_qs = maintenances.filter(date__month=m)
            monthly_summary.append({
                'month_name': month_names[m-1],
                'total': m_qs.count(),
                'done':  m_qs.filter(status='Done').count(),
                'open':  m_qs.filter(status='Open').count(),
            })
    else:
        maintenances = (
            Maintenance.objects
            .filter(date__year=selected_year, date__month=selected_month)
            .select_related('device', 'device__jenis', 'signed_by')
            .order_by('date')
        )
        period_label   = f"{month_names[selected_month-1]} {selected_year}"
        monthly_summary = []

    total      = maintenances.count()
    done       = maintenances.filter(status='Done').count()
    open_count = maintenances.filter(status='Open').count()
    preventive = maintenances.filter(maintenance_type='Preventive').count()

    by_type_qs = (
        maintenances.values('device__jenis__name')
        .annotate(total=Count('id'))
        .order_by('device__jenis__name')
    )
    by_type = []
    for row in by_type_qs:
        jenis_name = row['device__jenis__name']
        done_c = maintenances.filter(device__jenis__name=jenis_name, status='Done').count()
        open_c = maintenances.filter(device__jenis__name=jenis_name, status='Open').count()
        by_type.append({**row, 'done': done_c, 'open': open_c})

    month_choices = [{'value': i+1, 'label': n} for i, n in enumerate(month_names)]
    first_year = (
        Maintenance.objects.order_by('date').first().date.year
        if Maintenance.objects.exists() else today.year
    )
    year_choices = list(range(first_year, today.year + 1))

    return render(request, 'maintenance/maintenance_report.html', {
        'maintenances':    maintenances,
        'summary':         {'total': total, 'done': done, 'open': open_count, 'preventive': preventive},
        'by_type':         by_type,
        'selected_month':  selected_month,
        'selected_year':   selected_year,
        'month_choices':   month_choices,
        'year_choices':    year_choices,
        'period_label':    period_label,
        'mode':            mode,
        'monthly_summary': monthly_summary,
    })


# ─────────────────────────────────────────────────────────────────────
# EXPORT EXCEL
# ─────────────────────────────────────────────────────────────────────
@login_required
def export_maintenance_excel(request):
    status    = request.GET.get('status') or ''
    lokasi    = request.GET.get('lokasi') or ''
    jenis_id  = request.GET.get('jenis') or ''
    date_from = request.GET.get('date_from') or ''
    date_to   = request.GET.get('date_to') or ''
    year      = int(request.GET.get('year') or date.today().year)
    month     = request.GET.get('month') or ''
    mode      = request.GET.get('mode', 'monthly')

    month_names = [
        'Januari','Februari','Maret','April','Mei','Juni',
        'Juli','Agustus','September','Oktober','November','Desember'
    ]

    qs = Maintenance.objects.select_related(
        'device','device__jenis','signed_by'
    ).prefetch_related('technicians').order_by('date')

    if mode == 'ytd':
        current_month = date.today().month if year == date.today().year else 12
        qs = qs.filter(date__year=year, date__month__lte=current_month)
        filename = f"pemeliharaan_YTD_{year}_Jan-{month_names[current_month-1]}.xlsx"
        title    = f"LAPORAN PEMELIHARAAN {year} (Januari — {month_names[current_month-1]})"
    else:
        if status:    qs = qs.filter(status=status)
        if lokasi:    qs = qs.filter(device__lokasi__iexact=lokasi)
        if jenis_id:  qs = qs.filter(device__jenis_id=jenis_id)
        if date_from: qs = qs.filter(date__gte=date_from)
        if date_to:   qs = qs.filter(date__lte=date_to)
        if year and month:
            qs = qs.filter(date__year=int(year), date__month=int(month))
            filename = f"pemeliharaan_{year}_{month_names[int(month)-1]}.xlsx"
            title    = f"DATA PEMELIHARAAN {month_names[int(month)-1]} {year}"
        else:
            filename = "pemeliharaan_fasop.xlsx"
            title    = "DATA PEMELIHARAAN PERALATAN FASOP UP2B"

    # ── Style helpers ────────────────────────────────────────────────
    hdr_fill  = PatternFill("solid", fgColor="0F172A")
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c_align   = Alignment(horizontal="center", vertical="center")
    l_align   = Alignment(vertical="center", wrap_text=True)
    thin      = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'),  bottom=Side(style='thin'))
    done_fill = PatternFill("solid", fgColor="D1FAE5")
    open_fill = PatternFill("solid", fgColor="FEF3C7")
    alt_fill  = PatternFill("solid", fgColor="F8FAFC")
    norm_fill = PatternFill("solid", fgColor="DCFCE7")
    abnorm_fill = PatternFill("solid", fgColor="FEE2E2")

    def write_header(ws, title_text, headers, col_widths, color="0F172A"):
        ncols = len(headers)
        last  = get_column_letter(ncols)
        ws.merge_cells(f"A1:{last}1")
        ws["A1"].value     = title_text
        ws["A1"].font      = Font(bold=True, size=12)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].fill      = PatternFill("solid", fgColor="EFF6FF")
        ws.row_dimensions[1].height = 28
        ws.merge_cells(f"A2:{last}2")
        ws["A2"].value     = f"Dicetak: {date.today().strftime('%d %B %Y')}"
        ws["A2"].alignment = Alignment(horizontal="center")
        ws["A2"].font      = Font(size=10, italic=True, color="64748B")
        ws.row_dimensions[3].height = 6
        hf = PatternFill("solid", fgColor=color)
        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=4, column=ci, value=h)
            cell.font      = hdr_font
            cell.fill      = hf
            cell.alignment = hdr_align
            cell.border    = thin
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[4].height = 22
        ws.freeze_panes = "A5"

    def write_rows_maintenance(ws, rows_qs, start_row=5):
        """Sheet mode maintenance biasa."""
        for ri, m in enumerate(rows_qs, 1):
            wr = ri + start_row - 1
            techs = ", ".join(
                t.get_full_name() or t.username for t in m.technicians.all()
            ) or (", ".join(m.pelaksana_names) if m.pelaksana_names else "-")
            row_data = [
                ri,
                m.date.strftime("%d/%m/%Y"),
                m.date.strftime("%B"),
                str(m.device),
                m.device.lokasi or "-",
                m.maintenance_type,
                techs,
                m.description or "-",
                m.status,
            ]
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=wr, column=ci, value=val)
                cell.border    = thin
                cell.alignment = c_align if ci in [1,2,3,6,9] else l_align
                if ci == 9:
                    cell.fill = done_fill if val == "Done" else open_fill
                    cell.font = Font(bold=True, color="065F46" if val == "Done" else "92400E")
                elif ri % 2 == 0:
                    cell.fill = alt_fill
            ws.row_dimensions[wr].height = 18

    def write_rows_assessment(ws, devices_qs, maint_map, start_row=5):
        """Sheet per jenis — format asesmen."""
        for ri, dev in enumerate(devices_qs, 1):
            wr = ri + start_row - 1
            # Ambil maintenance terakhir di periode YTD
            last_maint = maint_map.get(dev.pk)
            tgl_assesmen = last_maint.date.strftime("%d/%m/%Y") if last_maint else "—"
            # Hasil asesmen dari status_operasi device
            status_op = (dev.status_operasi or "").lower()
            if status_op == "operasi":
                hasil   = "Normal"
                h_fill  = norm_fill
                h_color = "166534"
            else:
                hasil   = "Tidak Normal"
                h_fill  = abnorm_fill
                h_color = "991B1B"
            row_data = [
                ri,
                "UP2B MAKASSAR",
                dev.lokasi or "—",
                dev.merk or "—",
                dev.type or "—",
                hasil,
                tgl_assesmen,
                dev.keterangan or "—",
            ]
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=wr, column=ci, value=val)
                cell.border    = thin
                cell.alignment = c_align if ci in [1,2,7] else l_align
                if ci == 6:
                    cell.fill = h_fill
                    cell.font = Font(bold=True, color=h_color)
                elif ri % 2 == 0:
                    cell.fill = alt_fill
            ws.row_dimensions[wr].height = 18

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if mode == "ytd":
        # ── Sheet Ringkasan YTD ──────────────────────────────────────
        ws_sum = wb.create_sheet("Ringkasan")
        ws_sum.sheet_properties.tabColor = "0F172A"
        ws_sum.merge_cells("A1:G1")
        ws_sum["A1"].value     = title
        ws_sum["A1"].font      = Font(bold=True, size=12)
        ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_sum["A1"].fill      = PatternFill("solid", fgColor="EFF6FF")
        ws_sum.row_dimensions[1].height = 28

        sum_headers = ["Bulan", "Total", "Preventive", "Corrective", "Done", "Open"]
        sum_widths  = [18, 10, 14, 14, 10, 10]
        for ci, (h, w) in enumerate(zip(sum_headers, sum_widths), 1):
            cell = ws_sum.cell(row=3, column=ci, value=h)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = hdr_align; cell.border = thin
            ws_sum.column_dimensions[get_column_letter(ci)].width = w
        ws_sum.row_dimensions[3].height = 20

        for ri, m_name in enumerate(month_names[:current_month], 1):
            m_qs = qs.filter(date__month=ri)
            row_data = [
                m_name, m_qs.count(),
                m_qs.filter(maintenance_type="Preventive").count(),
                m_qs.filter(maintenance_type="Corrective").count(),
                m_qs.filter(status="Done").count(),
                m_qs.filter(status="Open").count(),
            ]
            for ci, val in enumerate(row_data, 1):
                cell = ws_sum.cell(row=ri+3, column=ci, value=val)
                cell.border    = thin
                cell.alignment = c_align if ci > 1 else Alignment(vertical="center")
                if ri % 2 == 0: cell.fill = alt_fill
            ws_sum.row_dimensions[ri+3].height = 18

        # Total row
        tr = current_month + 4
        total_row = ["TOTAL", qs.count(),
                     qs.filter(maintenance_type="Preventive").count(),
                     qs.filter(maintenance_type="Corrective").count(),
                     qs.filter(status="Done").count(),
                     qs.filter(status="Open").count()]
        for ci, val in enumerate(total_row, 1):
            cell = ws_sum.cell(row=tr, column=ci, value=val)
            cell.font   = Font(bold=True)
            cell.fill   = PatternFill("solid", fgColor="DBEAFE")
            cell.border = thin
            cell.alignment = c_align if ci > 1 else Alignment(vertical="center")

        # ── Buat maint_map: device_pk → maintenance terakhir di YTD ─
        from django.db.models import Max
        # Ambil pk maintenance terbaru per device dalam periode YTD
        latest_ids = (
            qs.values("device_id")
              .annotate(latest=Max("date"))
        )
        maint_map = {}
        for row in latest_ids:
            m = qs.filter(device_id=row["device_id"], date=row["latest"]).first()
            if m:
                maint_map[row["device_id"]] = m

        # ── Sheet per jenis perangkat ────────────────────────────────
        JENIS_COLORS = {
            "Router": "2563EB", "Switch": "7C3AED", "PLC": "0891B2",
            "Radio": "D97706", "VoIP": "9333EA", "Multiplexer": "0D9488",
            "Catu Daya": "EA580C", "Rectifier": "DC2626",
            "Teleproteksi": "4F46E5", "Genset": "16A34A",
        }
        from devices.models import DeviceType
        asmt_headers = ["No", "UP2B", "Lokasi", "Merk", "Type",
                        "Hasil Asesmen (Status Aset)", "Tanggal Asesmen", "Keterangan"]
        asmt_widths  = [5, 18, 22, 20, 20, 28, 18, 30]

        all_jenis = DeviceType.objects.order_by("name")
        for jenis in all_jenis:
            devs = Device.objects.filter(
                jenis=jenis, is_deleted=False
            ).order_by("lokasi", "nama")
            if not devs.exists():
                continue
            sheet_name = jenis.name[:31]
            color = JENIS_COLORS.get(jenis.name, "334155")
            ws_j  = wb.create_sheet(sheet_name)
            ws_j.sheet_properties.tabColor = color
            write_header(ws_j, f"{jenis.name.upper()} — {title}", asmt_headers, asmt_widths, color)
            write_rows_assessment(ws_j, devs, maint_map)

        # ── Sheet semua data maintenance ─────────────────────────────
        ws_all = wb.create_sheet("Semua Data Maintenance")
        ws_all.sheet_properties.tabColor = "64748B"
        maint_headers = ["No","Tanggal","Bulan","Perangkat","Lokasi","Jenis Maint.","Pelaksana","Deskripsi","Status"]
        maint_widths  = [5, 12, 12, 25, 20, 14, 20, 35, 10]
        write_header(ws_all, title, maint_headers, maint_widths)
        write_rows_maintenance(ws_all, qs)

    else:
        ws = wb.create_sheet("Data Pemeliharaan")
        headers    = ["No","Tanggal","Bulan","Perangkat","Lokasi","Jenis Maint.","Pelaksana","Deskripsi","Status"]
        col_widths = [5, 12, 12, 25, 20, 14, 20, 35, 10]
        write_header(ws, title, headers, col_widths)
        write_rows_maintenance(ws, qs)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

# ─────────────────────────────────────────────────────────────────────
# BERITA ACARA ASESMEN PERALATAN
# ─────────────────────────────────────────────────────────────────────

BULAN_ID = ['Januari','Februari','Maret','April','Mei','Juni',
            'Juli','Agustus','September','Oktober','November','Desember']

BA_GRUP_MAP = {
    'TELEKOMUNIKASI': [
        'Router','Switch','Radio','VoIP','Multiplexer','PLC','Teleproteksi','RoIP',
        'Server Telkom','Master Clock','Catu Daya',
    ],
    'SCADA': [
        'RTU','SAS','Server SCADA','UPS','IED BCU','Clock Server',
        'Serial Server','Router SAS','Switch SAS','Inverter','GENSET',
    ],
    'PROSIS': [
        'RELE DEFENSE SCHEME','Master Trip','UFLS','DFR','Server Prosis',
    ],
}

def _ba_hitung(jenis_list, year, bulan, core_label=None):
    """Hitung rekap satu grup peralatan untuk Berita Acara.
    core_label: jika diisi, Workstation PC dengan spesifikasi.core=core_label
    akan ikut dihitung dalam grup ini.
    """
    devs = Device.objects.filter(is_deleted=False, jenis__name__in=jenis_list)
    if core_label:
        ws_qs = Device.objects.filter(
            is_deleted=False,
            jenis__name__iexact='Workstation PC',
            spesifikasi__core=core_label,
        )
        devs = (devs | ws_qs).distinct()
    total   = devs.count()
    normal  = devs.filter(status_operasi='operasi').count()
    tdk_nrm = devs.exclude(status_operasi='operasi').exclude(
                   status_operasi__isnull=True).exclude(status_operasi='').count()
    sudah_ids = set(Maintenance.objects.filter(
        device__in=devs,
        date__year=year,
        date__month__lte=bulan,
    ).values_list('device_id', flat=True))
    sudah = len(sudah_ids)
    belum = total - sudah
    return {'total': total, 'normal': normal, 'tidak_normal': tdk_nrm,
            'sudah': sudah, 'belum': belum}


def _ba_params(request):
    """Baca parameter filter dari GET: year, bulan, telkom[], scada[], prosis[]."""
    today = date.today()
    year  = int(request.GET.get('year',  today.year))
    bulan = int(request.GET.get('bulan', today.month))
    submitted = 'year' in request.GET or 'telkom' in request.GET
    if submitted:
        sel_telkom = request.GET.getlist('telkom') or []
        sel_scada  = request.GET.getlist('scada')  or []
        sel_prosis = request.GET.getlist('prosis') or []
    else:
        sel_telkom = BA_GRUP_MAP['TELEKOMUNIKASI']
        sel_scada  = BA_GRUP_MAP['SCADA']
        sel_prosis = BA_GRUP_MAP['PROSIS']
    return year, bulan, sel_telkom, sel_scada, sel_prosis


# Map nama grup → nilai spesifikasi.core Workstation PC yang ikut dihitung
_BA_CORE_LABEL = {
    'TELEKOMUNIKASI': 'Telekomunikasi',
    'SCADA':          'SCADA',
    'PROSIS':         'PROSIS',
}


def _ba_rekap(year, bulan, sel_telkom, sel_scada, sel_prosis):
    """Bangun list rekap [{grup, total, normal, tidak_normal, sudah, belum}]."""
    rekap = []
    for grup_nama, jenis_list in [
        ('TELEKOMUNIKASI', sel_telkom),
        ('SCADA',          sel_scada),
        ('PROSIS',         sel_prosis),
    ]:
        if jenis_list:
            core_label = _BA_CORE_LABEL.get(grup_nama)
            data = _ba_hitung(jenis_list, year, bulan, core_label=core_label)
            rekap.append({'grup': grup_nama, **data})
    return rekap


@login_required
def berita_acara_config(request):
    """Halaman konfigurasi Berita Acara — pilih peralatan, periode, pratinjau rekap."""
    today = date.today()
    year, bulan, sel_telkom, sel_scada, sel_prosis = _ba_params(request)
    rekap = _ba_rekap(year, bulan, sel_telkom, sel_scada, sel_prosis)

    first_year = (Maintenance.objects.order_by('date').first().date.year
                  if Maintenance.objects.exists() else today.year)
    year_choices  = list(range(first_year, today.year + 1))
    bulan_choices = list(enumerate(BULAN_ID, start=1))   # [(1,'Januari'), ...]

    groups = [
        {'key': 'telkom', 'nama': 'TELEKOMUNIKASI', 'icon': 'broadcast',        'border': '#bfdbfe', 'bg': '#eff6ff', 'color': '#1d4ed8',
         'jenis_list': BA_GRUP_MAP['TELEKOMUNIKASI'], 'selected': sel_telkom},
        {'key': 'scada',  'nama': 'SCADA',          'icon': 'cpu',              'border': '#bbf7d0', 'bg': '#f0fdf4', 'color': '#065f46',
         'jenis_list': BA_GRUP_MAP['SCADA'],          'selected': sel_scada},
        {'key': 'prosis', 'nama': 'PROSIS',         'icon': 'lightning-charge', 'border': '#fde68a', 'bg': '#fffbeb', 'color': '#92400e',
         'jenis_list': BA_GRUP_MAP['PROSIS'],         'selected': sel_prosis},
    ]

    return render(request, 'maintenance/berita_acara_config.html', {
        'rekap':         rekap,
        'groups':        groups,
        'selected_year': year,
        'selected_bulan': bulan,
        'year_choices':  year_choices,
        'bulan_choices': bulan_choices,
        'bulan_str':     BULAN_ID[bulan - 1],
        'tahun_str':     str(year),
    })


@login_required
def berita_acara_excel(request):
    """Generate Berita Acara Asesmen — landscape A4, siap print 1 halaman."""
    import os
    from django.conf import settings
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    year, bulan, sel_telkom, sel_scada, sel_prosis = _ba_params(request)
    bulan_str = BULAN_ID[bulan - 1]
    tahun_str = str(year)
    rekap = _ba_rekap(year, bulan, sel_telkom, sel_scada, sel_prosis)

    # ── Workbook ─────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Berita Acara'

    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins.left   = 0.4
    ws.page_margins.right  = 0.4
    ws.page_margins.top    = 0.4
    ws.page_margins.bottom = 0.4
    ws.page_margins.header = 0
    ws.page_margins.footer = 0
    ws.sheet_view.showGridLines = False

    # A=margin, B=label, C=GRUP, D=TOTAL, E=SUDAH, F=TIDAK NORMAL, G=BELUM, H=margin
    col_cfg = [
        ('A', 1.5),
        ('B', 6),
        ('C', 38),
        ('D', 14),
        ('E', 18),
        ('F', 16),
        ('G', 18),
        ('H', 1.5),
    ]
    for col_letter, width in col_cfg:
        ws.column_dimensions[col_letter].width = width

    for r in range(1, 45):
        ws.row_dimensions[r].height = 14

    thin   = Side(style='thin')
    TNR    = 'Times New Roman'

    def brd(l=None, r=None, t=None, b=None, all_thin=False):
        s = thin if all_thin else None
        return Border(
            left=l or s or Side(style=None),
            right=r or s or Side(style=None),
            top=t or s or Side(style=None),
            bottom=b or s or Side(style=None),
        )
    box = brd(all_thin=True)

    def C(row, col, val='', bold=False, sz=10, color='000000',
          ha='left', va='center', wrap=False, bg=None, border=None, italic=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name=TNR, bold=bold, size=sz, color=color, italic=italic)
        c.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
        if bg:     c.fill   = PatternFill('solid', fgColor=bg)
        if border: c.border = border
        return c

    def MG(r1, c1, r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    cB, cC, cD, cE, cF, cG = 2, 3, 4, 5, 6, 7

    # ── Logo PLN ─────────────────────────────────────────────────────
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'pln_logo_conv.png')
    if os.path.exists(logo_path):
        img        = XLImage(logo_path)
        img.width  = 55
        img.height = 55
        img.anchor = 'D1'
        ws.add_image(img)
        ws.row_dimensions[1].height = 42

    # ── Judul ─────────────────────────────────────────────────────────
    r = 2
    ws.row_dimensions[r].height = 10

    r = 3
    ws.row_dimensions[r].height = 22
    MG(r, cB, r, cG)
    C(r, cB, 'BERITA ACARA', bold=True, sz=14, ha='center', va='center')

    r = 4
    ws.row_dimensions[r].height = 18
    MG(r, cB, r, cG)
    C(r, cB, 'Penyampaian Data Hasil Asesmen Peralatan Fasilitas Operasi',
      bold=True, sz=12, ha='center', va='center')

    r = 5
    ws.row_dimensions[r].height = 16
    MG(r, cB, r, cG)
    C(r, cB, f'Bulan Januari s.d. {bulan_str} {tahun_str}',
      sz=11, ha='center', va='center')

    r = 6
    ws.row_dimensions[r].height = 8

    # ── Teks pembuka ──────────────────────────────────────────────────
    r = 7
    ws.row_dimensions[r].height = 28
    MG(r, cB, r, cG)
    C(r, cB,
      'Berikut disampaikan tabel rekap hasil asesmen peralatan Fasilitas Operasi '
      'UP2B Sistem Makassar sebagai berikut :',
      sz=10, va='center', wrap=True)

    r = 8
    ws.row_dimensions[r].height = 4

    # ── Tabel header ──────────────────────────────────────────────────
    HBG = 'BFBFBF'
    r = 9
    ws.row_dimensions[r].height   = 14
    ws.row_dimensions[r+1].height = 20

    for col in [cC, cD, cE, cF, cG]:
        MG(r, col, r+1, col)
    labels = {
        cC: 'GRUP',
        cD: 'TOTAL\nPERALATAN',
        cE: 'SUDAH\nMAINTENANCE',
        cF: 'TIDAK\nNORMAL',
        cG: 'BELUM\nMAINTENANCE',
    }
    for col, lbl in labels.items():
        C(r, col, lbl, bold=True, sz=10, ha='center', va='center',
          bg=HBG, border=box, wrap=True)
        ws.cell(r+1, col).border = box

    # ── Data rekap ────────────────────────────────────────────────────
    r = 11
    for i, row in enumerate(rekap):
        ws.row_dimensions[r+i].height = 18
        alt = 'F2F2F2' if i % 2 == 1 else None
        C(r+i, cC, row['grup'],        sz=10, ha='center', va='center', bg=alt, border=box)
        C(r+i, cD, row['total'],       sz=10, ha='center', va='center', bg=alt, border=box)
        cs = C(r+i, cE, row['sudah'], sz=10, ha='center', va='center', bg=alt, border=box)
        if row['sudah'] > 0:
            cs.font = Font(name=TNR, bold=True, size=10, color='375623')
        cn = C(r+i, cF, row['tidak_normal'], sz=10, ha='center', va='center', bg=alt, border=box)
        if row['tidak_normal'] > 0:
            cn.font = Font(name=TNR, bold=True, size=10, color='C00000')
        cb = C(r+i, cG, row['belum'], sz=10, ha='center', va='center', bg=alt, border=box)
        if row['belum'] > 0:
            cb.font = Font(name=TNR, bold=True, size=10, color='C55A11')

    r = r + len(rekap) + 1

    # ── Teks penutup ──────────────────────────────────────────────────
    ws.row_dimensions[r].height = 26
    MG(r, cB, r, cG)
    C(r, cB,
      'Berdasarkan rekap data di atas, dilampirkan detail informasi aset untuk hasil ABNORMAL.',
      sz=10, va='center', wrap=True)

    r += 2
    # ── Tanggal ───────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 16
    MG(r, cE, r, cG)
    C(r, cE, f'Makassar,     {bulan_str} {tahun_str}', sz=10, ha='center')

    r += 2
    # ── Tanda Tangan ─────────────────────────────────────────────────
    ws.row_dimensions[r].height = 16
    MG(r, cB, r, cC)
    C(r, cB, 'Disahkan Oleh :', sz=10, ha='center',
      border=brd(l=thin, r=thin, t=thin))
    MG(r, cE, r, cG)
    C(r, cE, 'Disusun Oleh :', sz=10, ha='center',
      border=brd(l=thin, r=thin, t=thin))

    r += 1
    ws.row_dimensions[r].height = 16
    MG(r, cB, r, cC)
    C(r, cB, 'MUP2B SISTEM MAKASSAR', bold=True, sz=10, ha='center',
      border=brd(l=thin, r=thin))
    MG(r, cE, r, cG)
    C(r, cE, 'ASMAN FASOP UP2B SISTEM MAKASSAR', bold=True, sz=10, ha='center',
      border=brd(l=thin, r=thin))

    for ttd_r in range(r+1, r+5):
        ws.row_dimensions[ttd_r].height = 18
        MG(ttd_r, cB, ttd_r, cC)
        ws.cell(ttd_r, cB).border = brd(l=thin, r=thin)
        MG(ttd_r, cE, ttd_r, cG)
        ws.cell(ttd_r, cE).border = brd(l=thin, r=thin)

    name_r = r + 5
    ws.row_dimensions[name_r].height = 16
    MG(name_r, cB, name_r, cC)
    C(name_r, cB, '(                                             )',
      sz=10, ha='center', border=brd(l=thin, r=thin, b=thin))
    MG(name_r, cE, name_r, cG)
    C(name_r, cE, '(                                             )',
      sz=10, ha='center', border=brd(l=thin, r=thin, b=thin))

    ws.print_area = f'A1:{get_column_letter(8)}{name_r + 1}'

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fname = f'Berita_Acara_Asesmen_{bulan_str}_{tahun_str}.xlsx'
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    wb.save(resp)
    return resp


@login_required
def berita_acara_pdf(request):
    """Generate Berita Acara Asesmen dalam format PDF via WeasyPrint."""
    import os
    from django.conf import settings
    from django.template.loader import render_to_string

    year, bulan, sel_telkom, sel_scada, sel_prosis = _ba_params(request)
    bulan_str = BULAN_ID[bulan - 1]
    tahun_str = str(year)
    rekap = _ba_rekap(year, bulan, sel_telkom, sel_scada, sel_prosis)

    logo_b64 = ''
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'pln_logo_conv.png')
    if os.path.exists(logo_path):
        import base64
        with open(logo_path, 'rb') as f:
            logo_b64 = base64.b64encode(f.read()).decode()

    ctx = {
        'rekap':     rekap,
        'bulan_str': bulan_str,
        'tahun_str': tahun_str,
        'logo_b64':  logo_b64,
    }

    try:
        import weasyprint
        html_string = render_to_string('maintenance/pdf/berita_acara.html', ctx)
        html = weasyprint.HTML(string=html_string)
        pdf_bytes = html.write_pdf()
        resp = HttpResponse(pdf_bytes, content_type='application/pdf')
        fname = f'Berita_Acara_Asesmen_{bulan_str}_{tahun_str}.pdf'
        resp['Content-Disposition'] = f'attachment; filename="{fname}"'
        return resp
    except ImportError:
        return HttpResponse('WeasyPrint tidak tersedia di server ini.', status=500)


# ─────────────────────────────────────────────────────────────────────
# EXPORT PDF LAPORAN PEMELIHARAAN
# ─────────────────────────────────────────────────────────────────────
@login_required
def export_maintenance_pdf(request, pk):
    # Coba WeasyPrint dulu, fallback ke ReportLab jika tidak tersedia
    try:
        from .pdf_weasy import build_pdf_weasy as build_pdf
    except (ImportError, OSError):
        from .pdf_export import build_pdf

    maintenance = get_object_or_404(Maintenance, pk=pk)
    device      = maintenance.device
    device_kind = device.jenis.name.strip().upper() if device.jenis else 'GENERIC'

    # ── Ambil detail sesuai jenis ──────────────────────────────────
    router_detail = plc_detail = radio_detail = None
    voip_detail = mux_detail = rect_detail = tp_detail = genset_detail = rtu_detail = sas_detail = roip_detail = ups_detail = None

    def _try(fn):
        try: return fn()
        except Exception: return None

    if device_kind in ('ROUTER', 'SWITCH'):
        router_detail = _try(lambda: maintenance.maintenancerouter)
    elif device_kind == 'PLC':
        plc_detail = _try(lambda: maintenance.maintenanceplc)
    elif device_kind == 'RADIO':
        radio_detail = _try(lambda: maintenance.maintenanceradio)
    elif device_kind == 'VOIP':
        voip_detail = _try(lambda: maintenance.maintenancevoip)
    elif device_kind == 'MULTIPLEXER':
        mux_detail = _try(lambda: maintenance.maintenancemux)
    elif device_kind in ('RECTIFIER', 'CATU DAYA', 'CATUDAYA', 'RECTIFIER & BATTERY'):
        rect_detail = _try(lambda: maintenance.maintenancerectifier)
    elif device_kind == 'TELEPROTEKSI':
        tp_detail = _try(lambda: maintenance.maintenanceteleproteksi)
    elif device_kind == 'GENSET':
        genset_detail = _try(lambda: maintenance.maintenancegenset)
    elif device_kind == 'RTU':
        rtu_detail = _try(lambda: maintenance.maintenancertu)
    elif device_kind in ('SAS', 'SERVER SCADA', 'GATEWAY SAS'):
        sas_detail = _try(lambda: maintenance.maintenancesas)
    elif device_kind in ('ROIP',):
        roip_detail = _try(lambda: maintenance.maintenanceroip)
    elif device_kind == 'UPS':
        ups_detail = _try(lambda: maintenance.maintenanceups)

    # Corrective detail
    corrective_detail = None
    if maintenance.maintenance_type == 'Corrective':
        corrective_detail = _try(lambda: maintenance.corrective_detail)

    # SFP JSON
    sfp_ports = []
    if router_detail and router_detail.sfp_port_data:
        try: sfp_ports = json.loads(router_detail.sfp_port_data)
        except Exception: pass

    # ── Susun dict data ────────────────────────────────────────────
    def _g(obj, attr, default=None):
        val = getattr(obj, attr, default)
        return val if val not in (None, '') else default

    # Signature dari asisten manager
    sigs = {}
    if maintenance.signed_by:
        try:
            sig_path = maintenance.signed_by.profile.signature.path
            if sig_path:
                sigs['asisten_manager'] = sig_path
        except Exception:
            pass
    # Signature pelaksana (technician pertama yang punya signature)
    for tech in maintenance.technicians.all():
        try:
            sp = tech.profile.signature.path
            if sp:
                sigs['operator'] = sp
                break
        except Exception:
            pass

    # Nama pelaksana: prioritaskan pelaksana_names (input manual), fallback ke User M2M
    if maintenance.pelaksana_names:
        techs_str = ', '.join(n for n in maintenance.pelaksana_names if n)
    else:
        techs_str = ', '.join(
            t.get_full_name() or t.username for t in maintenance.technicians.all()
        ) or '-'

    data = {
        'print_date':  dj_timezone.localtime(dj_timezone.now()).strftime('%d %B %Y  %H:%M'),
        'print_by':    request.user.get_full_name() or request.user.username,
        'device_kind': device_kind,
        'signatures':  sigs,

        'info': {
            'device_name':      device.nama,
            'device_type':      str(device.jenis) if device.jenis else '-',
            'lokasi':           _g(device, 'lokasi', '-'),
            'ip_address':       _g(device, 'ip_address', '-'),
            'serial_number':    _g(device, 'serial_number', '-'),
            'merk':            _g(device, 'merk', '-'),
            'type':             _g(device, 'type', '-'),
            'date':             dj_timezone.localtime(maintenance.date).strftime('%d %B %Y  %H:%M'),
            'maintenance_type': maintenance.maintenance_type,
            'technician':       techs_str,
            'status':           maintenance.status,
            'description':      maintenance.description or '',
            'catatan_am':       maintenance.catatan_am or '',
            'signed_by':        maintenance.signed_by.profile.get_display_name() if maintenance.signed_by and hasattr(maintenance.signed_by, 'profile') else (maintenance.signed_by.get_full_name() or maintenance.signed_by.username if maintenance.signed_by else ''),
        },

        'fisik': {
            'kondisi_fisik': _g(router_detail, 'kondisi_fisik', ''),
            'led_link':      _g(router_detail, 'led_link', ''),
            'kondisi_kabel': _g(router_detail, 'kondisi_kabel', ''),
        } if router_detail else {},
        'pengukuran': {
            'tegangan_input': _g(router_detail, 'tegangan_input'),
            'suhu_perangkat': _g(router_detail, 'suhu_perangkat'),
            'cpu_load':       _g(router_detail, 'cpu_load'),
            'memory_usage':   _g(router_detail, 'memory_usage'),
        } if router_detail else {},
        'port': {
            'jumlah_port_aktif': _g(router_detail, 'jumlah_port_aktif'),
            'jumlah_port_total': _g(router_detail, 'jumlah_port_total'),
            'status_routing':    _g(router_detail, 'status_routing', ''),
            'detail_port':       _g(router_detail, 'detail_port', ''),
        } if router_detail else {},
        'sfp_ports':        sfp_ports,
        'catatan_tambahan': (_g(router_detail, 'catatan_tambahan', '') if router_detail else ''),

        'plc': {
            'akses_plc':         _g(plc_detail, 'akses_plc', ''),
            'remote_akses_plc':  _g(plc_detail, 'remote_akses_plc', ''),
            'time_sync':         _g(plc_detail, 'time_sync', ''),
            'wave_trap':         _g(plc_detail, 'wave_trap', ''),
            'imu':               _g(plc_detail, 'imu', ''),
            'kabel_coaxial':     _g(plc_detail, 'kabel_coaxial', ''),
            'transmission_line': _g(plc_detail, 'transmission_line'),
            'rx_pilot_level':    _g(plc_detail, 'rx_pilot_level'),
            'freq_tx':           _g(plc_detail, 'freq_tx'),
            'bandwidth_tx':      _g(plc_detail, 'bandwidth_tx'),
            'freq_rx':           _g(plc_detail, 'freq_rx'),
            'bandwidth_rx':      _g(plc_detail, 'bandwidth_rx'),
            'modul_terpasang':   _g(plc_detail, 'modul_terpasang', []),
        } if plc_detail else {},

        'radio': {
            'suhu_ruangan':     _g(radio_detail, 'suhu_ruangan'),
            'kebersihan':       _g(radio_detail, 'kebersihan', ''),
            'lampu_penerangan': _g(radio_detail, 'lampu_penerangan', ''),
            'ada_radio':        _g(radio_detail, 'ada_radio', ''),
            'ada_battery':      _g(radio_detail, 'ada_battery', ''),
            'merk_battery':     _g(radio_detail, 'merk_battery', ''),
            'ada_power_supply': _g(radio_detail, 'ada_power_supply', ''),
            'merk_power_supply':_g(radio_detail, 'merk_power_supply', ''),
            'jenis_antena':     _g(radio_detail, 'jenis_antena', ''),
            'swr':              _g(radio_detail, 'swr', ''),
            'power_tx':         _g(radio_detail, 'power_tx'),
            'tegangan_battery': _g(radio_detail, 'tegangan_battery'),
            'tegangan_psu':     _g(radio_detail, 'tegangan_psu'),
            'frekuensi_tx':     _g(radio_detail, 'frekuensi_tx'),
            'frekuensi_rx':     _g(radio_detail, 'frekuensi_rx'),
            'catatan':          _g(radio_detail, 'catatan', ''),
        } if radio_detail else {},

        'voip': {
            'ip_address':        _g(voip_detail, 'ip_address', ''),
            'extension_number':  _g(voip_detail, 'extension_number', ''),
            'sip_server_1':      _g(voip_detail, 'sip_server_1', ''),
            'sip_server_2':      _g(voip_detail, 'sip_server_2', ''),
            'suhu_ruangan':      _g(voip_detail, 'suhu_ruangan'),
            'kondisi_fisik':     _g(voip_detail, 'kondisi_fisik', ''),
            'ntp_server':        _g(voip_detail, 'ntp_server', ''),
            'webconfig':         _g(voip_detail, 'webconfig', ''),
            'ps_merk':           _g(voip_detail, 'ps_merk', ''),
            'ps_tegangan_input': _g(voip_detail, 'ps_tegangan_input'),
            'ps_status':         _g(voip_detail, 'ps_status', ''),
            'catatan':           _g(voip_detail, 'catatan', ''),
        } if voip_detail else {},

        'mux': {
            'brand':         _g(mux_detail, 'brand', ''),
            'firmware':      _g(mux_detail, 'firmware', ''),
            'sync_source_1': _g(mux_detail, 'sync_source_1', ''),
            'sync_source_2': _g(mux_detail, 'sync_source_2', ''),
            'suhu_ruangan':  _g(mux_detail, 'suhu_ruangan'),
            'kebersihan':    _g(mux_detail, 'kebersihan', ''),
            'hs1_merk':      _g(mux_detail, 'hs1_merk', ''),
            'hs1_tx_bias':   _g(mux_detail, 'hs1_tx_bias'),
            'hs1_jarak':     _g(mux_detail, 'hs1_jarak'),
            'hs1_tx':        _g(mux_detail, 'hs1_tx'),
            'hs1_lambda':    _g(mux_detail, 'hs1_lambda'),
            'hs1_suhu':      _g(mux_detail, 'hs1_suhu'),
            'hs1_rx':        _g(mux_detail, 'hs1_rx'),
            'hs1_bandwidth': _g(mux_detail, 'hs1_bandwidth', ''),
            'hs2_merk':      _g(mux_detail, 'hs2_merk', ''),
            'hs2_tx_bias':   _g(mux_detail, 'hs2_tx_bias'),
            'hs2_jarak':     _g(mux_detail, 'hs2_jarak'),
            'hs2_tx':        _g(mux_detail, 'hs2_tx'),
            'hs2_lambda':    _g(mux_detail, 'hs2_lambda'),
            'hs2_suhu':      _g(mux_detail, 'hs2_suhu'),
            'hs2_rx':        _g(mux_detail, 'hs2_rx'),
            'hs2_bandwidth': _g(mux_detail, 'hs2_bandwidth', ''),
            'psu1_status':   _g(mux_detail, 'psu1_status', ''),
            'psu1_temp1':    _g(mux_detail, 'psu1_temp1'),
            'psu1_temp2':    _g(mux_detail, 'psu1_temp2'),
            'psu1_temp3':    _g(mux_detail, 'psu1_temp3'),
            'psu2_status':   _g(mux_detail, 'psu2_status', ''),
            'psu2_temp1':    _g(mux_detail, 'psu2_temp1'),
            'psu2_temp2':    _g(mux_detail, 'psu2_temp2'),
            'psu2_temp3':    _g(mux_detail, 'psu2_temp3'),
            'fan_status':    _g(mux_detail, 'fan_status', ''),
            'catatan':       _g(mux_detail, 'catatan', ''),
            **{f'slot_{l.lower()}_modul': _g(mux_detail, f'slot_{l.lower()}_modul', '')
               for l in 'ABCDEFGH'},
            **{f'slot_{l.lower()}_isian': _g(mux_detail, f'slot_{l.lower()}_isian', '')
               for l in 'ABCDEFGH'},
        } if mux_detail else {},

        'rectifier': {
            'suhu_ruangan':        _g(rect_detail, 'suhu_ruangan'),
            'exhaust_fan':         _g(rect_detail, 'exhaust_fan', ''),
            'kebersihan':          _g(rect_detail, 'kebersihan', ''),
            'lampu_penerangan':    _g(rect_detail, 'lampu_penerangan', ''),
            'rect1_merk':          _g(rect_detail, 'rect1_merk', ''),
            'rect1_tipe':          _g(rect_detail, 'rect1_tipe', ''),
            'rect1_kondisi':       _g(rect_detail, 'rect1_kondisi', ''),
            'rect1_kapasitas':     _g(rect_detail, 'rect1_kapasitas', ''),
            'rect1_v_rectifier':   _g(rect_detail, 'rect1_v_rectifier'),
            'rect1_v_battery':     _g(rect_detail, 'rect1_v_battery'),
            'rect1_teg_pos_ground':_g(rect_detail, 'rect1_teg_pos_ground'),
            'rect1_teg_neg_ground':_g(rect_detail, 'rect1_teg_neg_ground'),
            'rect1_v_dropper':     _g(rect_detail, 'rect1_v_dropper'),
            'rect1_v_load':        _g(rect_detail, 'rect1_v_load'),
            'rect1_a_rectifier':   _g(rect_detail, 'rect1_a_rectifier'),
            'rect1_a_battery':     _g(rect_detail, 'rect1_a_battery'),
            'rect1_a_load':        _g(rect_detail, 'rect1_a_load'),
            'bat1_merk':           _g(rect_detail, 'bat1_merk', ''),
            'bat1_tipe':           _g(rect_detail, 'bat1_tipe', ''),
            'bat1_kondisi':        _g(rect_detail, 'bat1_kondisi', ''),
            'bat1_kapasitas':      _g(rect_detail, 'bat1_kapasitas', ''),
            'bat1_jumlah':         _g(rect_detail, 'bat1_jumlah'),
            'bat1_kondisi_kabel':  _g(rect_detail, 'bat1_kondisi_kabel', ''),
            'bat1_kondisi_mur_baut':_g(rect_detail,'bat1_kondisi_mur_baut',''),
            'bat1_kondisi_sel_rak': _g(rect_detail,'bat1_kondisi_sel_rak', ''),
            'bat1_air_battery':    _g(rect_detail, 'bat1_air_battery'),
            'bat1_v_total':        _g(rect_detail, 'bat1_v_total'),
            'bat1_v_load':         _g(rect_detail, 'bat1_v_load'),
            'bat1_cells':          _g(rect_detail, 'bat1_cells', []),
            'catatan':             _g(rect_detail, 'catatan', ''),
        } if rect_detail else {},

        'tp': {
            'suhu_ruangan':       _g(tp_detail, 'suhu_ruangan'),
            'kebersihan_perangkat': _g(tp_detail, 'kebersihan_perangkat', ''),
            'kebersihan_panel':   _g(tp_detail, 'kebersihan_panel', ''),
            'lampu':              _g(tp_detail, 'lampu', ''),
            'link':               _g(tp_detail, 'link', ''),
            'tipe_tp':            _g(tp_detail, 'tipe_tp', ''),
            'versi_program':      _g(tp_detail, 'versi_program', ''),
            'address_tp':         _g(tp_detail, 'address_tp', ''),
            'port_comm':          _g(tp_detail, 'port_comm', ''),
            'akses_tp':           _g(tp_detail, 'akses_tp', ''),
            'remote_akses_tp':    _g(tp_detail, 'remote_akses_tp', ''),
            'jumlah_skema':       _g(tp_detail, 'jumlah_skema'),
            'skema_1_command':    _g(tp_detail, 'skema_1_command', ''),
            'skema_1_send_minus': _g(tp_detail, 'skema_1_send_minus'),
            'skema_1_send_plus':  _g(tp_detail, 'skema_1_send_plus'),
            'skema_1_receive_minus': _g(tp_detail, 'skema_1_receive_minus'),
            'skema_1_receive_plus':  _g(tp_detail, 'skema_1_receive_plus'),
            'skema_2_command':    _g(tp_detail, 'skema_2_command', ''),
            'skema_2_send_minus': _g(tp_detail, 'skema_2_send_minus'),
            'skema_2_send_plus':  _g(tp_detail, 'skema_2_send_plus'),
            'skema_2_receive_minus': _g(tp_detail, 'skema_2_receive_minus'),
            'skema_2_receive_plus':  _g(tp_detail, 'skema_2_receive_plus'),
            'skema_3_command':    _g(tp_detail, 'skema_3_command', ''),
            'skema_3_send_minus': _g(tp_detail, 'skema_3_send_minus'),
            'skema_3_send_plus':  _g(tp_detail, 'skema_3_send_plus'),
            'skema_3_receive_minus': _g(tp_detail, 'skema_3_receive_minus'),
            'skema_3_receive_plus':  _g(tp_detail, 'skema_3_receive_plus'),
            'skema_4_command':    _g(tp_detail, 'skema_4_command', ''),
            'skema_4_send_minus': _g(tp_detail, 'skema_4_send_minus'),
            'skema_4_send_plus':  _g(tp_detail, 'skema_4_send_plus'),
            'skema_4_receive_minus': _g(tp_detail, 'skema_4_receive_minus'),
            'skema_4_receive_plus':  _g(tp_detail, 'skema_4_receive_plus'),
            'skema_1_send_result':    _g(tp_detail, 'skema_1_send_result', ''),
            'skema_1_receive_result': _g(tp_detail, 'skema_1_receive_result', ''),
            'skema_2_send_result':    _g(tp_detail, 'skema_2_send_result', ''),
            'skema_2_receive_result': _g(tp_detail, 'skema_2_receive_result', ''),
            'skema_3_send_result':    _g(tp_detail, 'skema_3_send_result', ''),
            'skema_3_receive_result': _g(tp_detail, 'skema_3_receive_result', ''),
            'skema_4_send_result':    _g(tp_detail, 'skema_4_send_result', ''),
            'skema_4_receive_result': _g(tp_detail, 'skema_4_receive_result', ''),
            'time_sync':          _g(tp_detail, 'time_sync', ''),
            'loop_test':          _g(tp_detail, 'loop_test'),
            'catatan':            _g(tp_detail, 'catatan', ''),
        } if tp_detail else {},

        'genset': {
            'air_accu':           _g(genset_detail, 'air_accu'),
            'tegangan_batere':    _g(genset_detail, 'tegangan_batere'),
            'arus_pengisian':     _g(genset_detail, 'arus_pengisian'),
            'tegangan_charger':   _g(genset_detail, 'tegangan_charger'),
            'arus_beban_charger': _g(genset_detail, 'arus_beban_charger'),
            'radiator':           _g(genset_detail, 'radiator'),
            'kapasitas_tangki':   _g(genset_detail, 'kapasitas_tangki'),
            'tangki_bbm_sebelum': _g(genset_detail, 'tangki_bbm_sebelum'),
            'tangki_bbm_sesudah': _g(genset_detail, 'tangki_bbm_sesudah'),
            'bbm_terpakai':       genset_detail.bbm_terpakai if genset_detail else None,
            'persen_bbm_sesudah': genset_detail.persen_bbm_sesudah if genset_detail else None,
            'mcb':                _g(genset_detail, 'mcb', ''),
            'pelumas':            _g(genset_detail, 'pelumas', ''),
            'waktu_transisi':     _g(genset_detail, 'waktu_transisi'),
            'pln_f_r':  _g(genset_detail,'pln_f_r'),  'pln_f_s':  _g(genset_detail,'pln_f_s'),  'pln_f_t':  _g(genset_detail,'pln_f_t'),
            'pln_v_rn': _g(genset_detail,'pln_v_rn'), 'pln_v_sn': _g(genset_detail,'pln_v_sn'), 'pln_v_tn': _g(genset_detail,'pln_v_tn'),
            'pln_v_rs': _g(genset_detail,'pln_v_rs'), 'pln_v_st': _g(genset_detail,'pln_v_st'), 'pln_v_tr': _g(genset_detail,'pln_v_tr'),
            'pln_i_r':  _g(genset_detail,'pln_i_r'),  'pln_i_s':  _g(genset_detail,'pln_i_s'),  'pln_i_t':  _g(genset_detail,'pln_i_t'),
            'gen_f_r':  _g(genset_detail,'gen_f_r'),  'gen_f_s':  _g(genset_detail,'gen_f_s'),  'gen_f_t':  _g(genset_detail,'gen_f_t'),
            'gen_v_rn': _g(genset_detail,'gen_v_rn'), 'gen_v_sn': _g(genset_detail,'gen_v_sn'), 'gen_v_tn': _g(genset_detail,'gen_v_tn'),
            'gen_v_rs': _g(genset_detail,'gen_v_rs'), 'gen_v_st': _g(genset_detail,'gen_v_st'), 'gen_v_tr': _g(genset_detail,'gen_v_tr'),
            'gen_i_r':  _g(genset_detail,'gen_i_r'),  'gen_i_s':  _g(genset_detail,'gen_i_s'),  'gen_i_t':  _g(genset_detail,'gen_i_t'),
            'oil_pressure':       _g(genset_detail, 'oil_pressure'),
            'engine_temperature': _g(genset_detail, 'engine_temperature'),
            'batere_condition':   _g(genset_detail, 'batere_condition'),
            'rpm':                _g(genset_detail, 'rpm'),
            'counter_sebelum':    _g(genset_detail, 'counter_sebelum'),
            'counter_sesudah':    _g(genset_detail, 'counter_sesudah'),
            'selisih_counter':    genset_detail.selisih_counter if genset_detail else None,
            'waktu_start':        str(genset_detail.waktu_start) if genset_detail and genset_detail.waktu_start else '-',
            'waktu_stop':         str(genset_detail.waktu_stop)  if genset_detail and genset_detail.waktu_stop  else '-',
            'durasi_menit':       genset_detail.durasi_menit if genset_detail else None,
            'catatan':            _g(genset_detail, 'catatan', ''),
        } if genset_detail else {},

        'rtu': {
            'cp2016_jumlah': _g(rtu_detail, 'cp2016_jumlah'),
            'cp2016_data':   _g(rtu_detail, 'cp2016_data', {}),
            'cp2019_jumlah': _g(rtu_detail, 'cp2019_jumlah'),
            'cp2019_data':   _g(rtu_detail, 'cp2019_data', {}),
            'di2112_jumlah': _g(rtu_detail, 'di2112_jumlah'),
            'di2112_data':   _g(rtu_detail, 'di2112_data', {}),
            'do2210_jumlah': _g(rtu_detail, 'do2210_jumlah'),
            'do2210_data':   _g(rtu_detail, 'do2210_data', {}),
            'ai2300_data':   _g(rtu_detail, 'ai2300_data', {}),
            'ied_data':      _g(rtu_detail, 'ied_data', {}),
            'ps48_teg_beban':   _g(rtu_detail, 'ps48_teg_beban'),
            'ps48_arus_beban':  _g(rtu_detail, 'ps48_arus_beban'),
            'ps48_teg_supply':  _g(rtu_detail, 'ps48_teg_supply'),
            'ps48_arus_supply': _g(rtu_detail, 'ps48_arus_supply'),
            'ps110_teg_beban':  _g(rtu_detail, 'ps110_teg_beban'),
            'ps110_arus_beban': _g(rtu_detail, 'ps110_arus_beban'),
            'ps110_teg_supply': _g(rtu_detail, 'ps110_teg_supply'),
            'ps110_arus_supply':_g(rtu_detail, 'ps110_arus_supply'),
        } if rtu_detail else {},

        'sas': {
            'spek_merk':       _g(sas_detail, 'spek_merk', ''),
            'spek_type':       _g(sas_detail, 'spek_type', ''),
            'spek_cpu':        _g(sas_detail, 'spek_cpu', ''),
            'spek_ram':        _g(sas_detail, 'spek_ram', ''),
            'spek_gpu':        _g(sas_detail, 'spek_gpu', ''),
            'spek_storage':    _g(sas_detail, 'spek_storage', ''),
            'spek_firmware':   _g(sas_detail, 'spek_firmware', ''),
            'spek_config_ver': _g(sas_detail, 'spek_config_ver', ''),
            'spek_ip':         _g(sas_detail, 'spek_ip', ''),
            'modul_io':        _g(sas_detail, 'modul_io', ''),
            'kondisi_server':  _g(sas_detail, 'kondisi_server', ''),
            'kondisi_panel':   _g(sas_detail, 'kondisi_panel', ''),
            'temp_ruangan':    _g(sas_detail, 'temp_ruangan'),
            'temp_peralatan':  _g(sas_detail, 'temp_peralatan'),
            'exhaust_fan':     _g(sas_detail, 'exhaust_fan', ''),
            'peri_eth_switch': _g(sas_detail, 'peri_eth_switch', ''),
            'peri_gps':        _g(sas_detail, 'peri_gps', ''),
            'peri_eth_serial': _g(sas_detail, 'peri_eth_serial', ''),
            'peri_router':     _g(sas_detail, 'peri_router', ''),
            'jumlah_bay':      _g(sas_detail, 'jumlah_bay'),
            'peri_keterangan': _g(sas_detail, 'peri_keterangan', ''),
            'perf_cpu':        _g(sas_detail, 'perf_cpu', ''),
            'perf_ram':        _g(sas_detail, 'perf_ram', ''),
            'perf_storage':    _g(sas_detail, 'perf_storage', ''),
            'indikasi_alarm':  _g(sas_detail, 'indikasi_alarm', ''),
            'komm_master':     _g(sas_detail, 'komm_master', ''),
            'komm_ied':        _g(sas_detail, 'komm_ied', ''),
            'time_sync':       _g(sas_detail, 'time_sync', ''),
            'inv_kondisi':     _g(sas_detail, 'inv_kondisi', ''),
            'inv_teg_input':   _g(sas_detail, 'inv_teg_input'),
            'inv_arus_input':  _g(sas_detail, 'inv_arus_input'),
            'inv_teg_output':  _g(sas_detail, 'inv_teg_output'),
            'inv_arus_output': _g(sas_detail, 'inv_arus_output'),
            'ps_teg_input':    _g(sas_detail, 'ps_teg_input'),
            'ps_arus_input':   _g(sas_detail, 'ps_arus_input'),
            'ps_teg_output':   _g(sas_detail, 'ps_teg_output'),
            'ps_arus_output':  _g(sas_detail, 'ps_arus_output'),
        } if sas_detail else {},

        'roip': {
            'kondisi_fisik':    _g(roip_detail, 'kondisi_fisik', ''),
            'ntp_server':       _g(roip_detail, 'ntp_server', ''),
            'power_supply':     _g(roip_detail, 'power_supply', ''),
            'memory_usage':     _g(roip_detail, 'memory_usage'),
            'tx_volume_offset':        _g(roip_detail, 'tx_volume_offset'),
            'rx_volume_offset':        _g(roip_detail, 'rx_volume_offset'),
            'bridge_conn_source':      _g(roip_detail, 'bridge_conn_source', ''),
            'bridge_conn_destination': _g(roip_detail, 'bridge_conn_destination', ''),
            'dest_port_number':        _g(roip_detail, 'dest_port_number', ''),
            'source_port_number':      _g(roip_detail, 'source_port_number', ''),
            'ptt_attack_time':         _g(roip_detail, 'ptt_attack_time'),
            'ptt_release_time':  _g(roip_detail, 'ptt_release_time'),
            'ptt_voice_delay':   _g(roip_detail, 'ptt_voice_delay'),
            'ptt_vox_threshold': _g(roip_detail, 'ptt_vox_threshold'),
            'rx_attack_time':   _g(roip_detail, 'rx_attack_time'),
            'rx_release_time':  _g(roip_detail, 'rx_release_time'),
            'rx_voice_delay':   _g(roip_detail, 'rx_voice_delay'),
            'rx_vox_threshold': _g(roip_detail, 'rx_vox_threshold'),
            'test_radio_master': _g(roip_detail, 'test_radio_master', ''),
            'test_ping_master':  _g(roip_detail, 'test_ping_master'),
            'catatan':           _g(roip_detail, 'catatan', ''),
        } if roip_detail else {},

        'ups': {
            'ups_merk':          _g(ups_detail, 'ups_merk', ''),
            'ups_model':         _g(ups_detail, 'ups_model', ''),
            'ups_kapasitas':     _g(ups_detail, 'ups_kapasitas', ''),
            'ups_kondisi':       _g(ups_detail, 'ups_kondisi', ''),
            'v_input_r':         _g(ups_detail, 'v_input_r'),
            'v_input_s':         _g(ups_detail, 'v_input_s'),
            'v_input_t':         _g(ups_detail, 'v_input_t'),
            'f_input':           _g(ups_detail, 'f_input'),
            'v_output_r':        _g(ups_detail, 'v_output_r'),
            'v_output_s':        _g(ups_detail, 'v_output_s'),
            'v_output_t':        _g(ups_detail, 'v_output_t'),
            'f_output':          _g(ups_detail, 'f_output'),
            'a_load':            _g(ups_detail, 'a_load'),
            'percent_load':      _g(ups_detail, 'percent_load'),
            'bat_merk':          _g(ups_detail, 'bat_merk', ''),
            'bat_tipe':          _g(ups_detail, 'bat_tipe', ''),
            'bat_kapasitas':     _g(ups_detail, 'bat_kapasitas', ''),
            'bat_jumlah_cell':   _g(ups_detail, 'bat_jumlah_cell'),
            'bat_kondisi':       _g(ups_detail, 'bat_kondisi', ''),
            'bat_kondisi_kabel': _g(ups_detail, 'bat_kondisi_kabel', ''),
            'bat_v_total':       _g(ups_detail, 'bat_v_total'),
            'bat_cells':         _g(ups_detail, 'bat_cells', []),
            'catatan':           _g(ups_detail, 'catatan', ''),
        } if ups_detail else {},
    }

    # ── Corrective detail dict ─────────────────────────────────────
    data['maintenance_type'] = maintenance.maintenance_type
    if corrective_detail:
        data['corrective'] = {
            'jenis_kerusakan':         corrective_detail.get_jenis_kerusakan_display() if corrective_detail.jenis_kerusakan else '-',
            'deskripsi_masalah':       _g(corrective_detail, 'deskripsi_masalah', '-'),
            'tindakan':                _g(corrective_detail, 'tindakan', '-'),
            'komponen_diganti':        corrective_detail.komponen_diganti,
            'nama_komponen':           _g(corrective_detail, 'nama_komponen', ''),
            'komponen_terkait':        str(corrective_detail.komponen_terkait) if corrective_detail.komponen_terkait else '',
            'kondisi_sebelum':         _g(corrective_detail, 'kondisi_sebelum', ''),
            'kondisi_sesudah':         _g(corrective_detail, 'kondisi_sesudah', ''),
            'durasi_display':          corrective_detail.durasi_display,
            'status_perbaikan':        corrective_detail.get_status_perbaikan_display(),
            'gangguan_nomor':          corrective_detail.gangguan.nomor_gangguan if corrective_detail.gangguan else '',
            'gangguan_summary':        corrective_detail.gangguan.executive_summary if corrective_detail.gangguan else '',
            'foto_sebelum_path':       corrective_detail.foto_sebelum.path if corrective_detail.foto_sebelum else '',
            'foto_sesudah_path':       corrective_detail.foto_sesudah.path if corrective_detail.foto_sesudah else '',
        }

    # ── Generate & stream ──────────────────────────────────────────
    buffer = BytesIO()
    build_pdf(data, buffer)
    buffer.seek(0)

    clean_date = dj_timezone.localtime(maintenance.date).strftime('%d-%m-%Y_%H.%M')
    filename = f"LAPORAN_PEMELIHARAAN_{device.nama}_{clean_date}.pdf".replace(' ', '_')
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────────────────────────────────
# TANDA TANGAN (Asisten Manager)
# ─────────────────────────────────────────────────────────────────────
@login_required
def maintenance_sign(request, pk):
    """Asisten Manager menandatangani laporan pemeliharaan."""
    from django.utils import timezone
    maintenance = get_object_or_404(Maintenance, pk=pk)

    # Hanya asisten manager boleh sign
    try:
        is_am = request.user.profile.is_asisten_manager
    except Exception:
        is_am = False

    if not is_am:
        from django.contrib import messages
        messages.error(request, 'Hanya Asisten Manager yang dapat menandatangani laporan.')
        return redirect('maintenance_view', pk=pk)

    if request.method == 'POST':
        maintenance.signed_by = request.user
        maintenance.signed_at = timezone.now()
        maintenance.catatan_am = request.POST.get('catatan_am', '').strip()
        maintenance.save(update_fields=['signed_by', 'signed_at', 'catatan_am'])
        from django.contrib import messages
        messages.success(request, 'Laporan berhasil ditandatangani.')

    return redirect('maintenance_view', pk=pk)


@login_required
def maintenance_catatan_am_edit(request, pk):
    """Asisten Manager mengedit / menambah catatan setelah TTD."""
    maintenance = get_object_or_404(Maintenance, pk=pk)

    try:
        is_am = request.user.profile.is_asisten_manager
    except Exception:
        is_am = False

    if not is_am:
        from django.contrib import messages
        messages.error(request, 'Hanya Asisten Manager yang dapat mengubah catatan.')
        return redirect('maintenance_view', pk=pk)

    if request.method == 'POST':
        maintenance.catatan_am = request.POST.get('catatan_am', '').strip()
        maintenance.save(update_fields=['catatan_am'])
        from django.contrib import messages
        messages.success(request, 'Catatan berhasil diperbarui.')

    return redirect('maintenance_view', pk=pk)


# ─────────────────────────────────────────────────────────────────────
# PROFILE — upload tanda tangan
# ─────────────────────────────────────────────────────────────────────
@login_required
def profile_view(request):
    from devices.models import UserProfile, UserLoginLog
    from django.contrib import messages
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    if request.method == 'POST':
        if 'signature' in request.FILES:
            profile.signature = request.FILES['signature']
            profile.save()
            messages.success(request, 'Tanda tangan berhasil disimpan.')
        elif 'save_display_name' in request.POST:
            nama_lengkap = request.POST.get('display_name', '').strip()
            profile.display_name = nama_lengkap
            profile.save(update_fields=['display_name'])
            # Sync ke User.first_name / last_name agar get_full_name() ikut terupdate
            parts = nama_lengkap.split(' ', 1)
            request.user.first_name = parts[0]
            request.user.last_name  = parts[1] if len(parts) > 1 else ''
            request.user.save(update_fields=['first_name', 'last_name'])
            messages.success(request, f'Nama lengkap berhasil disimpan sebagai "{nama_lengkap}".')
        return redirect('profile_view')

    ctx = {'profile': profile}

    # ── Data eksklusif superuser ──────────────────────────────────
    if request.user.is_superuser:
        from django.contrib.auth.models import User as AuthUser
        from django.contrib.sessions.models import Session
        from django.utils import timezone

        # Pengguna yang sedang aktif (session valid + active_session_key cocok)
        now = timezone.now()
        valid_sessions = Session.objects.filter(expire_date__gt=now)
        active_user_ids = set()
        for s in valid_sessions:
            data = s.get_decoded()
            uid  = data.get('_auth_user_id')
            if uid:
                active_user_ids.add(int(uid))

        active_profiles = (
            UserProfile.objects
            .filter(user_id__in=active_user_ids)
            .exclude(active_session_key='')
            .select_related('user')
            .order_by('user__username')
        )

        # Log login/logout terbaru (100 record)
        login_logs = (
            UserLoginLog.objects
            .select_related('user')
            .order_by('-timestamp')[:100]
        )

        ctx['active_profiles'] = active_profiles
        ctx['login_logs']      = login_logs

    return render(request, 'maintenance/profile.html', ctx)


# ─────────────────────────────────────────────────────────────
# APPROVAL VIEW
# ─────────────────────────────────────────────────────────────

@login_required
def maintenance_approval(request):
    from django.contrib import messages as dj_messages
    try:
        is_am = request.user.profile.is_asisten_manager
    except Exception:
        is_am = False

    if not (request.user.is_superuser or is_am):
        dj_messages.error(request, 'Anda tidak memiliki akses ke halaman Approval.')
        return redirect('maintenance_list')

    tab = request.GET.get('tab', 'pending')

    # ── Filter terkunci per peralatan (jenis) & lokasi ─────────
    # Reset semua filter
    if 'reset_filter' in request.GET:
        request.session.pop('approval_jenis_id', None)
        request.session.pop('approval_lokasi', None)
        params = request.GET.copy()
        params.pop('reset_filter', None)
        qs = params.urlencode()
        return redirect(f"{request.path}?{qs}" if qs else request.path)

    # Simpan filter jenis ke session
    if 'jenis_id' in request.GET:
        val = request.GET.get('jenis_id', '').strip()
        if val:
            request.session['approval_jenis_id'] = val
        else:
            request.session.pop('approval_jenis_id', None)

    # Simpan filter lokasi ke session
    if 'lokasi' in request.GET:
        val = request.GET.get('lokasi', '').strip()
        if val:
            request.session['approval_lokasi'] = val
        else:
            request.session.pop('approval_lokasi', None)

    selected_jenis_id = request.session.get('approval_jenis_id', '')
    selected_lokasi   = request.session.get('approval_lokasi', '')

    jenis_list = DeviceType.objects.all().order_by('name')

    lokasi_list = (
        Device.objects
        .filter(is_deleted=False)
        .exclude(lokasi__isnull=True)
        .exclude(lokasi='')
        .annotate(lokasi_clean=Trim('lokasi'))
        .values_list('lokasi_clean', flat=True)
        .distinct().order_by('lokasi_clean')
    )

    pending_qs = (
        Maintenance.objects
        .filter(status='Done', signed_by__isnull=True)
        .select_related('device', 'device__jenis')
        .order_by('-date')
    )

    approved_qs = (
        Maintenance.objects
        .filter(signed_by__isnull=False)
        .select_related('device', 'device__jenis', 'signed_by')
        .order_by('-signed_at')
    )

    if selected_jenis_id:
        pending_qs  = pending_qs.filter(device__jenis_id=selected_jenis_id)
        approved_qs = approved_qs.filter(device__jenis_id=selected_jenis_id)

    if selected_lokasi:
        pending_qs  = pending_qs.filter(device__lokasi__iexact=selected_lokasi)
        approved_qs = approved_qs.filter(device__lokasi__iexact=selected_lokasi)

    selected_jenis_obj = None
    if selected_jenis_id:
        try:
            selected_jenis_obj = DeviceType.objects.get(pk=selected_jenis_id)
        except DeviceType.DoesNotExist:
            request.session.pop('approval_jenis_id', None)
            selected_jenis_id = ''

    return render(request, 'maintenance/approval.html', {
        'pending_list':       pending_qs,
        'approved_list':      approved_qs,
        'tab':                tab,
        'jenis_list':         jenis_list,
        'lokasi_list':        lokasi_list,
        'selected_jenis_id':  selected_jenis_id,
        'selected_jenis_obj': selected_jenis_obj,
        'selected_lokasi':    selected_lokasi,
    })


# ─────────────────────────────────────────────────────────────
# COVERAGE VIEW
# ─────────────────────────────────────────────────────────────

@login_required
def maintenance_coverage(request):
    from django.utils import timezone as tz

    year = request.GET.get('tahun', '')
    try:
        year = int(year)
    except (ValueError, TypeError):
        year = tz.now().year

    selected_lokasi = request.GET.get('lokasi', '')

    from jadwal.models import JADWAL_EXCLUDED_JENIS
    devices_qs = (
        Device.objects
        .filter(is_deleted=False, lokasi__isnull=False, host__isnull=True)
        .exclude(lokasi='')
        .exclude(lokasi__iexact='none')
        .exclude(jenis__name__in=JADWAL_EXCLUDED_JENIS)
        .select_related('jenis')
        .order_by('lokasi', 'nama')
    )

    # IDs of devices that have maintenance this year
    maintained_ids = set(
        Maintenance.objects
        .filter(device__in=devices_qs, date__year=year)
        .values_list('device_id', flat=True)
        .distinct()
    )

    # Last maintenance date per device
    from django.db.models import Max
    last_maint_qs = (
        Maintenance.objects
        .filter(device__in=devices_qs, date__year=year)
        .values('device_id')
        .annotate(last_date=Max('date'))
    )
    last_maint_map = {row['device_id']: row['last_date'] for row in last_maint_qs}

    # Last maintenance ID per device (for detail link)
    last_id_map = {}
    for m in (Maintenance.objects
              .filter(device__in=devices_qs, date__year=year)
              .order_by('device_id', '-date')
              .values('device_id', 'id')):
        if m['device_id'] not in last_id_map:
            last_id_map[m['device_id']] = m['id']

    if selected_lokasi:
        # Drill-down: detail per lokasi
        lokasi_devices = devices_qs.filter(lokasi__iexact=selected_lokasi)
        device_rows = []
        for d in lokasi_devices:
            device_rows.append({
                'device':           d,
                'has_maintenance':  d.id in maintained_ids,
                'last_date':        last_maint_map.get(d.id),
                'last_maint_id':    last_id_map.get(d.id),
            })
        return render(request, 'maintenance/coverage.html', {
            'device_rows':      device_rows,
            'selected_lokasi':  selected_lokasi,
            'year':             year,
            'detail_view':      True,
        })

    # Summary per lokasi
    lokasi_map = {}
    for d in devices_qs:
        lok = d.lokasi.strip()
        if lok not in lokasi_map:
            lokasi_map[lok] = {'lokasi': lok, 'total': 0, 'sudah': 0, 'devices': []}
        lokasi_map[lok]['total'] += 1
        has_maint = d.id in maintained_ids
        if has_maint:
            lokasi_map[lok]['sudah'] += 1
        lokasi_map[lok]['devices'].append({
            'device':        d,
            'has_maintenance': has_maint,
            'last_date':     last_maint_map.get(d.id),
        })

    lokasi_list = []
    for v in sorted(lokasi_map.values(), key=lambda x: x['lokasi']):
        v['belum'] = v['total'] - v['sudah']
        v['persen'] = int(v['sudah'] / v['total'] * 100) if v['total'] else 0
        lokasi_list.append(v)

    return render(request, 'maintenance/coverage.html', {
        'lokasi_list':     lokasi_list,
        'year':            year,
        'detail_view':     False,
        'selected_lokasi': '',
    })


# ─────────────────────────────────────────────────────────────
# CORRECTIVE MAINTENANCE VIEWS
# ─────────────────────────────────────────────────────────────

@login_required
@require_can_edit
def corrective_add(request, device_id=None, gangguan_id=None):
    """
    Form corrective maintenance ringkas.
    Bisa diakses dari:
      - Menu maintenance umum (GET /maintenance/corrective/add/)
      - Device detail (GET /maintenance/corrective/device/<id>/)
      - Gangguan detail (GET /maintenance/corrective/gangguan/<id>/)
    """
    from maintenance.models import MaintenanceCorrective
    from devices.models import Device
    from gangguan.models import Gangguan

    # Pre-fill device & gangguan jika dari context tertentu
    device_init   = Device.objects.filter(pk=device_id, is_deleted=False).first() if device_id else None
    gangguan_init = Gangguan.objects.filter(pk=gangguan_id).first() if gangguan_id else None
    # Kalau dari gangguan, device dari gangguan.peralatan
    if gangguan_init and not device_init and gangguan_init.peralatan:
        device_init = gangguan_init.peralatan

    if request.method == 'POST':
        # ── Ambil data form ──
        device_pk         = request.POST.get('device_id')
        tanggal           = request.POST.get('tanggal', '')
        # Pelaksana dari tag-input JS (JSON array)
        import json as _json
        _raw_pel = request.POST.get('pelaksana_names_input', '[]')
        try:
            pelaksana_list = _json.loads(_raw_pel)
        except Exception:
            pelaksana_list = [n.strip() for n in request.POST.get('pelaksana_names', '').split(',') if n.strip()]
        jenis_kerusakan   = request.POST.get('jenis_kerusakan', '')
        deskripsi_masalah = request.POST.get('deskripsi_masalah', '').strip()
        tindakan          = request.POST.get('tindakan', '').strip()
        komponen_diganti  = request.POST.get('komponen_diganti') == 'on'
        nama_komponen     = request.POST.get('nama_komponen', '').strip()
        komponen_terkait_pk = request.POST.get('komponen_terkait', '') or None
        kondisi_sebelum   = request.POST.get('kondisi_sebelum', '').strip()
        kondisi_sesudah   = request.POST.get('kondisi_sesudah', '').strip()
        durasi_jam        = request.POST.get('durasi_jam', '') or None
        durasi_menit      = request.POST.get('durasi_menit', '') or None
        status_perbaikan  = request.POST.get('status_perbaikan', 'selesai')
        gangguan_pk       = request.POST.get('gangguan_id', '') or None
        update_gangguan   = request.POST.get('update_gangguan', '')
        status_gangguan   = request.POST.get('status_gangguan', '')
        foto_sebelum      = request.FILES.get('foto_sebelum')
        foto_sesudah      = request.FILES.get('foto_sesudah')

        device = Device.objects.filter(pk=device_pk, is_deleted=False).first()
        if not device:
            device = device_init

        # Resolve komponen_terkait dari DeviceComponent
        from devices.models_komponen import DeviceComponent
        komponen_terkait_obj = None
        if komponen_terkait_pk:
            komponen_terkait_obj = DeviceComponent.objects.filter(pk=komponen_terkait_pk).first()

        if device and tanggal and deskripsi_masalah and tindakan:
            from datetime import datetime
            # Buat Maintenance record
            m_status = 'Done' if status_perbaikan == 'selesai' else 'Open'
            maint = Maintenance.objects.create(
                device           = device,
                maintenance_type = 'Corrective',
                date             = tanggal,
                description      = deskripsi_masalah,
                status           = m_status,
                pelaksana_names  = pelaksana_list,
            )

            # Buat detail corrective
            gangguan_obj = Gangguan.objects.filter(pk=gangguan_pk).first() if gangguan_pk else gangguan_init
            corr = MaintenanceCorrective(
                maintenance       = maint,
                gangguan          = gangguan_obj,
                jenis_kerusakan   = jenis_kerusakan,
                deskripsi_masalah = deskripsi_masalah,
                tindakan          = tindakan,
                komponen_diganti  = komponen_diganti,
                nama_komponen     = nama_komponen,
                komponen_terkait  = komponen_terkait_obj,
                kondisi_sebelum   = kondisi_sebelum,
                kondisi_sesudah   = kondisi_sesudah,
                durasi_jam        = int(durasi_jam) if durasi_jam else None,
                durasi_menit      = int(durasi_menit) if durasi_menit else None,
                status_perbaikan  = status_perbaikan,
            )
            if foto_sebelum: corr.foto_sebelum = foto_sebelum
            if foto_sesudah: corr.foto_sesudah = foto_sesudah
            corr.save()

            # Auto DeviceEvent kalau ada komponen diganti
            if komponen_diganti and nama_komponen:
                from devices.models import DeviceEvent
                from datetime import date as _date
                DeviceEvent.objects.create(
                    device         = device,
                    tipe           = 'penggantian',
                    tanggal        = tanggal[:10] if tanggal else _date.today(),
                    komponen       = nama_komponen,
                    nilai_lama     = kondisi_sebelum,
                    nilai_baru     = kondisi_sesudah,
                    catatan        = f'Dari corrective maintenance — {deskripsi_masalah[:100]}',
                    dilakukan_oleh = request.user,
                    gangguan       = gangguan_obj,
                )

                # Auto-update status DeviceComponent jika komponen_terkait dipilih
                if komponen_terkait_obj:
                    from django.utils import timezone as _tz
                    komponen_terkait_obj.status = 'diganti'
                    komponen_terkait_obj.tanggal_ganti = _tz.localdate()
                    komponen_terkait_obj.save(update_fields=['status', 'tanggal_ganti', 'updated_at'])

            # Update status gangguan jika diminta
            if gangguan_obj and update_gangguan and status_gangguan:
                gangguan_obj.status = status_gangguan
                if status_gangguan in ('resolved', 'closed') and not gangguan_obj.tanggal_resolved:
                    from django.utils import timezone
                    gangguan_obj.tanggal_resolved = timezone.now()
                gangguan_obj.save()

            # Notif ke AM — corrective selesai
            if status_perbaikan == 'selesai':
                try:
                    from notifikasi.views import notif_ke_am
                    g_info = f' (Tiket {gangguan_obj.nomor_gangguan})' if gangguan_obj else ''
                    notif_ke_am(
                        tipe   = 'corrective_selesai',
                        judul  = f'Corrective selesai — {device.nama}',
                        pesan  = (
                            f'Tindakan perbaikan pada {device.nama} ({device.lokasi})'
                            f'{g_info} telah selesai. Tindakan: {tindakan[:100]}'
                        ),
                        level  = 'success',
                        url    = f'/maintenance/view/{maint.pk}/',
                        device = device,
                    )
                except Exception:
                    pass

            # Redirect sesuai konteks
            if gangguan_obj and update_gangguan:
                return redirect('gangguan_detail', pk=gangguan_obj.pk)
            if gangguan_id:
                return redirect('gangguan_detail', pk=gangguan_id)
            if device_id:
                return redirect('device_view', pk=device_id)
            return redirect('maintenance_list')

    # Daftar gangguan aktif untuk dropdown
    from gangguan.models import Gangguan
    gangguan_aktif = Gangguan.objects.filter(
        status__in=['open', 'in_progress']
    ).order_by('-tanggal_gangguan')

    devices = Device.objects.filter(is_deleted=False).select_related('jenis').order_by('lokasi', 'nama')

    from datetime import date as _date
    return render(request, 'maintenance/corrective_form.html', {
        'device_init':    device_init,
        'gangguan_init':  gangguan_init,
        'gangguan_aktif': gangguan_aktif,
        'devices':        devices,
        'today_date':     _date.today().strftime('%Y-%m-%dT%H:%M'),
        'from_gangguan':  gangguan_id is not None,
        'from_device':    device_id is not None,
    })


# ─────────────────────────────────────────────────────────────────────
# DASHBOARD CATU DAYA — Discharge prediction + Inspection history
# ─────────────────────────────────────────────────────────────────────
@login_required
def catu_daya_dashboard(request):
    THRESHOLD = 38.0

    devices = Device.objects.filter(
        is_deleted=False,
        jenis__isnull=False,
    ).filter(
        Q(jenis__name__icontains='catu') |
        Q(jenis__name__icontains='rectifier')
    ).select_related('jenis').order_by('lokasi', 'nama')

    def _sum_field(cells, field):
        # Hanya jumlahkan sel INTEGER (bukan summary row vtotal/vload)
        vals = [float(c[field]) for c in cells
                if isinstance(c.get('cell'), int) and c.get(field) not in (None, '', 0)]
        return round(sum(vals), 3) if vals else None

    def _vtotal_row_val(cells, field):
        """Ambil nilai dari baris summary vtotal di bat1_cells."""
        row = next((c for c in cells if str(c.get('cell', '')) == 'vtotal'), None)
        if row:
            v = row.get(field)
            return round(float(v), 3) if v not in (None, '', 0) else None
        return None

    def _linreg(xs, ys):
        n = len(xs)
        sx = sum(xs); sy = sum(ys)
        sxy = sum(x * y for x, y in zip(xs, ys))
        sx2 = sum(x * x for x in xs)
        d = n * sx2 - sx * sx
        if d == 0:
            return None, None
        slope = (n * sxy - sx * sy) / d
        intercept = (sy - slope * sx) / n
        return slope, intercept

    def _fv(v):
        """Float value or None."""
        return round(float(v), 3) if v is not None else None

    devices_data = []
    for dev in devices:
        # Semua record inspection untuk device ini (urut terlama → terbaru)
        all_records = (
            MaintenanceRectifier.objects
            .filter(maintenance__device=dev)
            .select_related('maintenance')
            .order_by('maintenance__date')
        )
        if not all_records.exists():
            continue

        latest = all_records.last()

        # ── Discharge curve dari inspeksi terbaru ───────────────────
        cells = latest.bat1_cells or []

        # Prioritas: ambil dari baris summary "vtotal" di bat1_cells
        # (diisi user saat input form — nilai total bank tegangan per waktu discharge)
        # Fallback: jumlah sel individual (hanya sel integer, bukan summary row)
        def _vd(field):
            v = _vtotal_row_val(cells, field)
            return v if v is not None else _sum_field(cells, field)

        vd_0    = _vd('vd_0')
        vd_half = _vd('vd_half')
        vd_1    = _vd('vd_1')
        vd_2    = _vd('vd_2')

        # V Float sebagai titik referensi sebelum discharge (t = -0.5, tampil terpisah)
        v_float_discharge = _vtotal_row_val(cells, 'v_float') or _sum_field(cells, 'v_float') or latest.bat1_v_total

        actual = []
        if v_float_discharge:
            actual.append({'t': -0.5, 'v': round(float(v_float_discharge), 3), 'label': 'V Float'})
        for t, v in [(0, vd_0), (0.5, vd_half), (1, vd_1), (2, vd_2)]:
            if v is not None:
                actual.append({'t': t, 'v': v})

        prediction = []
        threshold_time = None
        slope = None
        # Regresi hanya dari titik VD (t >= 0), bukan V Float
        discharge_pts = [p for p in actual if p['t'] >= 0]
        if len(discharge_pts) >= 2:
            xs = [p['t'] for p in discharge_pts]
            ys = [p['v'] for p in discharge_pts]
            sl, intercept = _linreg(xs, ys)
            if sl is not None:
                slope = sl
                for t in [x / 2 for x in range(5, 21)]:
                    prediction.append({'t': t, 'v': round(sl * t + intercept, 3)})
                if sl < 0:
                    t_thr = (THRESHOLD - intercept) / sl
                    if t_thr > 0:
                        threshold_time = round(t_thr, 2)

        latest_v = discharge_pts[-1]['v'] if discharge_pts else (latest.bat1_v_total or 0)
        if latest_v <= THRESHOLD:
            status = 'danger'
        elif threshold_time is not None and threshold_time <= 10:
            status = 'warning'
        else:
            status = 'ok'

        # ── History semua inspeksi: V Batt, V Load, A Load, A Batt ──
        history = []
        for rec in all_records:
            history.append({
                'date':        str(rec.maintenance.date),
                'v_battery':   _fv(rec.rect1_v_battery),
                'v_load':      _fv(rec.bat1_v_load),
                'a_load':      _fv(rec.rect1_a_load),
                'a_battery':   _fv(rec.rect1_a_battery),
                'v_rectifier': _fv(rec.rect1_v_rectifier),
                'v_total':     _fv(rec.bat1_v_total),
            })

        # ── Data inservice inspection terbaru ───────────────────────
        latest_insp = (
            InspectionCatuDaya.objects
            .filter(inspection__device=dev)
            .select_related('inspection')
            .order_by('-inspection__tanggal')
            .first()
        )
        insp_data = {}
        if latest_insp:
            insp_data = {
                'insp_tanggal':         str(latest_insp.inspection.tanggal.date()),
                'insp_teg_baterai':     _fv(latest_insp.tegangan_baterai_dc),
                'insp_teg_load':        _fv(latest_insp.tegangan_load_dc),
                'insp_arus_load':       _fv(latest_insp.arus_load_dc),
                'insp_arus_baterai':    _fv(latest_insp.arus_baterai_dc),
                'insp_teg_input_ac':    _fv(latest_insp.tegangan_input_ac),
                'insp_arus_input_ac':   _fv(latest_insp.arus_input_ac),
                'insp_kondisi_rectifier': latest_insp.kondisi_rectifier or '',
                'insp_alarm_ground':    latest_insp.alarm_ground_fault or '',
                'insp_alarm_min_ac':    latest_insp.alarm_min_ac_fault or '',
                'insp_alarm_recti':     latest_insp.alarm_recti_fault or '',
                'insp_level_air':       latest_insp.level_air_bank or '',
                'insp_kondisi_keseluruhan': latest_insp.kondisi_keseluruhan or '',
                'insp_exhaust_fan':     latest_insp.exhaust_fan or '',
            }

        # ── Discharge curves dari inspeksi-inspeksi sebelumnya ──────────
        discharge_history = []
        for rec in list(all_records)[:-1]:   # semua kecuali latest, urut lama→baru
            h_cells = rec.bat1_cells or []
            h_vfloat = _vtotal_row_val(h_cells, 'v_float') or _sum_field(h_cells, 'v_float') or rec.bat1_v_total
            h_actual = []
            if h_vfloat:
                h_actual.append({'t': -0.5, 'v': round(float(h_vfloat), 3)})
            for _t, _field in [(0, 'vd_0'), (0.5, 'vd_half'), (1, 'vd_1'), (2, 'vd_2')]:
                _v = _vtotal_row_val(h_cells, _field) or _sum_field(h_cells, _field)
                if _v is not None:
                    h_actual.append({'t': _t, 'v': _v})
            if any(p['t'] >= 0 for p in h_actual):
                discharge_history.append({
                    'date': str(rec.maintenance.date)[:10],
                    'actual': h_actual,
                })

        devices_data.append({
            'device_id':      dev.id,
            'device_name':    dev.nama,
            'lokasi':         dev.lokasi or '-',
            'tanggal':        str(latest.maintenance.date),
            'v_float_total':  _vtotal_row_val(cells, 'v_float') or _sum_field(cells, 'v_float') or latest.bat1_v_total,
            'v_total_field':  latest.bat1_v_total,
            # latest pemeliharaan metrics
            'v_battery':      latest.rect1_v_battery,
            'v_load':         latest.bat1_v_load,
            'a_load':         latest.rect1_a_load,
            'a_battery':      latest.rect1_a_battery,
            'v_rectifier':    latest.rect1_v_rectifier,
            # discharge
            'actual':         actual,
            'prediction':     prediction,
            'threshold_time': threshold_time,
            'status':         status,
            'slope':          round(slope, 4) if slope else None,
            'bat_merk':       latest.bat1_merk or '-',
            'bat_tipe':       latest.bat1_tipe or '-',
            'bat_kapasitas':  latest.bat1_kapasitas or '-',
            # historical series
            'history':           history,
            'discharge_history': discharge_history,
            # inservice inspection
            **insp_data,
        })

    # ── UPS Dashboard Data ────────────────────────────────────────────
    UPS_VAC_THRESHOLD = 170.0  # VAC minimum — threshold prediksi backup

    ups_devices = Device.objects.filter(
        is_deleted=False,
        jenis__isnull=False,
    ).filter(
        Q(jenis__name__icontains='ups')
    ).select_related('jenis').order_by('lokasi', 'nama')

    def _avg3(a, b, c):
        vs = [float(x) for x in (a, b, c) if x is not None]
        return round(sum(vs) / len(vs), 2) if vs else None

    ups_data = []
    for udev in ups_devices:
        u_recs = (
            MaintenanceUPS.objects
            .filter(maintenance__device=udev)
            .select_related('maintenance')
            .order_by('maintenance__date')
        )
        if not u_recs.exists():
            continue

        u_latest = u_recs.last()
        u_cells  = u_latest.bat_cells or []

        def _u_vd(field):
            row = next((c for c in u_cells if str(c.get('cell', '')) == 'vtotal'), None)
            if row:
                v = row.get(field)
                if v not in (None, '', 0):
                    try:
                        return round(float(v), 3)
                    except (TypeError, ValueError):
                        pass
            vals = [float(c[field]) for c in u_cells
                    if isinstance(c.get('cell'), int) and c.get(field) not in (None, '', 0)]
            return round(sum(vals), 3) if vals else None

        u_vfloat = _u_vd('v_float')
        u_actual = []
        if u_vfloat:
            u_actual.append({'t': -0.5, 'v': u_vfloat, 'label': 'V Float'})
        for _t, _f in [(0, 'vd_0'), (1, 'vd_1'), (2, 'vd_2'), (3, 'vd_3')]:
            _v = _u_vd(_f)
            if _v is not None:
                u_actual.append({'t': _t, 'v': _v})

        # Threshold DC ketika output AC ≈ 170 VAC (1.75 V/cell end-of-discharge)
        n_cell = u_latest.bat_jumlah_cell or 0
        bat_thr = round(n_cell * 1.75, 1) if n_cell else None

        u_prediction = []
        backup_time  = None
        u_bat_slope  = None
        u_dpts = [p for p in u_actual if p['t'] >= 0]
        if len(u_dpts) >= 2:
            xs_u = [p['t'] for p in u_dpts]
            ys_u = [p['v'] for p in u_dpts]
            sl_u, ic_u = _linreg(xs_u, ys_u)
            if sl_u is not None and sl_u < 0:
                u_bat_slope = sl_u
                for _t2 in [x / 2 for x in range(7, 25)]:
                    u_prediction.append({'t': _t2, 'v': round(sl_u * _t2 + ic_u, 3)})
                if bat_thr:
                    _t_thr = (bat_thr - ic_u) / sl_u
                    if _t_thr > 0:
                        backup_time = round(_t_thr, 2)

        # History per maintenance record
        u_history = []
        for u_rec in u_recs:
            u_history.append({
                'date':          str(u_rec.maintenance.date),
                'v_input':       _avg3(u_rec.v_input_r, u_rec.v_input_s, u_rec.v_input_t),
                'v_input_r':     _fv(u_rec.v_input_r),
                'v_input_s':     _fv(u_rec.v_input_s),
                'v_input_t':     _fv(u_rec.v_input_t),
                'v_output':      _avg3(u_rec.v_output_r, u_rec.v_output_s, u_rec.v_output_t),
                'v_output_r':    _fv(u_rec.v_output_r),
                'v_output_s':    _fv(u_rec.v_output_s),
                'v_output_t':    _fv(u_rec.v_output_t),
                'f_input':       _fv(u_rec.f_input),
                'f_output':      _fv(u_rec.f_output),
                'a_load':        _fv(u_rec.a_load),
                'percent_load':  _fv(u_rec.percent_load),
                'bat_v_total':   _fv(u_rec.bat_v_total),
            })

        # Prediksi kapan V Input PLN turun ke 170 VAC (tren historis)
        vin_pred = None
        vin_pts = [(h['date'], h['v_input']) for h in u_history if h['v_input'] is not None]
        if len(vin_pts) >= 2:
            from datetime import date as _ddate
            _d0 = _ddate.fromisoformat(vin_pts[0][0])
            _xs = [(_ddate.fromisoformat(d) - _d0).days for d, _ in vin_pts]
            _ys = [v for _, v in vin_pts]
            _sl, _ic = _linreg(_xs, _ys)
            if _sl is not None and _sl < 0 and _ic > UPS_VAC_THRESHOLD:
                _x_thr = (UPS_VAC_THRESHOLD - _ic) / _sl
                _days_rem = int(_x_thr - _xs[-1])
                if _days_rem > 0:
                    vin_pred = {
                        'days':      _days_rem,
                        'rate_mo':   round(_sl * 30, 2),
                    }

        v_in_lat  = _avg3(u_latest.v_input_r, u_latest.v_input_s, u_latest.v_input_t)
        v_out_lat = _avg3(u_latest.v_output_r, u_latest.v_output_s, u_latest.v_output_t)

        if (v_in_lat and v_in_lat < 180) or (v_out_lat and v_out_lat < 180):
            u_status = 'danger'
        elif (v_in_lat and v_in_lat < 200) or (backup_time and backup_time < 3):
            u_status = 'warning'
        else:
            u_status = 'ok'

        # Discharge history dari record sebelumnya
        u_disc_hist = []
        for u_rec in list(u_recs)[:-1]:
            h_cells  = u_rec.bat_cells or []
            h_vfloat = next((round(float(c['v_float']), 3)
                             for c in h_cells if str(c.get('cell', '')) == 'vtotal'
                             and c.get('v_float') not in (None, '', 0)), None)
            h_actual = []
            if h_vfloat:
                h_actual.append({'t': -0.5, 'v': h_vfloat})
            for _ht, _hf in [(0, 'vd_0'), (1, 'vd_1'), (2, 'vd_2'), (3, 'vd_3')]:
                _row = next((c for c in h_cells if str(c.get('cell', '')) == 'vtotal'), None)
                _hv  = _row.get(_hf) if _row else None
                if _hv not in (None, '', 0):
                    try:
                        h_actual.append({'t': _ht, 'v': round(float(_hv), 3)})
                    except (TypeError, ValueError):
                        pass
            if any(p['t'] >= 0 for p in h_actual):
                u_disc_hist.append({
                    'date':   str(u_rec.maintenance.date)[:10],
                    'actual': h_actual,
                })

        ups_data.append({
            'device_id':      udev.id,
            'device_name':    udev.nama,
            'lokasi':         udev.lokasi or '-',
            'tanggal':        str(u_latest.maintenance.date),
            'v_input':        v_in_lat,
            'v_output':       v_out_lat,
            'v_input_r':      _fv(u_latest.v_input_r),
            'v_input_s':      _fv(u_latest.v_input_s),
            'v_input_t':      _fv(u_latest.v_input_t),
            'v_output_r':     _fv(u_latest.v_output_r),
            'v_output_s':     _fv(u_latest.v_output_s),
            'v_output_t':     _fv(u_latest.v_output_t),
            'f_input':        _fv(u_latest.f_input),
            'f_output':       _fv(u_latest.f_output),
            'a_load':         _fv(u_latest.a_load),
            'percent_load':   _fv(u_latest.percent_load),
            'ups_merk':       u_latest.ups_merk or '-',
            'ups_model':      u_latest.ups_model or '-',
            'ups_kapasitas':  u_latest.ups_kapasitas or '-',
            'ups_kondisi':    u_latest.ups_kondisi or '',
            'bat_merk':       u_latest.bat_merk or '-',
            'bat_tipe':       u_latest.bat_tipe or '-',
            'bat_kapasitas':  u_latest.bat_kapasitas or '-',
            'bat_jumlah_cell': n_cell,
            'bat_v_total':    _fv(u_latest.bat_v_total),
            'bat_kondisi':    u_latest.bat_kondisi or '',
            'bat_threshold':  bat_thr,
            'actual':         u_actual,
            'prediction':     u_prediction,
            'backup_time':    backup_time,
            'bat_slope':      round(u_bat_slope, 4) if u_bat_slope else None,
            'history':        u_history,
            'discharge_history': u_disc_hist,
            'vin_pred':       vin_pred,
            'status':         u_status,
        })

    selected_id = request.GET.get('device')

    # ── Genset Dashboard Data ─────────────────────────────────────────
    from datetime import date as _date_type
    _today = _date_type.today()

    genset_devices = Device.objects.filter(
        is_deleted=False,
        jenis__isnull=False,
        jenis__name__iexact='GENSET',
    ).select_related('jenis').order_by('lokasi', 'nama')

    def _avg(*vals):
        vs = [float(v) for v in vals if v is not None]
        return round(sum(vs) / len(vs), 2) if vs else None

    genset_data = []
    for gdev in genset_devices:
        g_recs = (
            MaintenanceGenset.objects
            .filter(maintenance__device=gdev)
            .select_related('maintenance')
            .order_by('maintenance__date')
        )
        if not g_recs.exists():
            genset_data.append({
                'device_id': gdev.id, 'device_name': gdev.nama,
                'lokasi': gdev.lokasi or '-',
                'last_date': None, 'days_since': None,
                'weekly_status': 'no_data',
                'bbm_before': None, 'bbm_after': None,
                'rpm': None, 'engine_temp': None,
                'oil_pressure': None, 'bat_volt': None,
                'radiator': None, 'counter': None,
                'waktu_transisi': None,
                'pln_v_avg': None, 'gen_v_avg': None,
                'pln_f_avg': None, 'gen_f_avg': None,
                'history': [],
            })
            continue

        g_latest = g_recs.last()
        last_date = g_latest.maintenance.date.date() if hasattr(g_latest.maintenance.date, 'date') else g_latest.maintenance.date
        days_since = (_today - last_date).days
        weekly_status = 'ok' if days_since <= 7 else 'late'

        history = []
        for r in g_recs.order_by('-maintenance__date')[:12][::-1]:
            history.append({
                'date':        str(r.maintenance.date)[:10],
                'bbm_before':  _fv(r.tangki_bbm_sebelum),
                'bbm_after':   _fv(r.tangki_bbm_sesudah),
                'rpm':         _fv(r.rpm),
                'engine_temp': _fv(r.engine_temperature),
                'oil_pressure':_fv(r.oil_pressure),
                'bat_volt':    _fv(r.tegangan_batere),
                'bat_cond':    _fv(r.batere_condition),
                'counter':     _fv(r.counter_sesudah),
                'radiator':    _fv(r.radiator),
                'transisi':    _fv(r.waktu_transisi),
                'pln_v_avg':   _avg(r.pln_v_rn, r.pln_v_sn, r.pln_v_tn),
                'gen_v_avg':   _avg(r.gen_v_rn, r.gen_v_sn, r.gen_v_tn),
                'pln_f_avg':   _avg(r.pln_f_r,  r.pln_f_s,  r.pln_f_t),
                'gen_f_avg':   _avg(r.gen_f_r,  r.gen_f_s,  r.gen_f_t),
                'durasi':      round(r.durasi_menit, 1) if r.durasi_menit else None,
            })

        genset_data.append({
            'device_id':     gdev.id,
            'device_name':   gdev.nama,
            'lokasi':        gdev.lokasi or '-',
            'last_date':     str(last_date),
            'days_since':    days_since,
            'weekly_status': weekly_status,
            'bbm_before':    _fv(g_latest.tangki_bbm_sebelum),
            'bbm_after':     _fv(g_latest.tangki_bbm_sesudah),
            'rpm':           _fv(g_latest.rpm),
            'engine_temp':   _fv(g_latest.engine_temperature),
            'oil_pressure':  _fv(g_latest.oil_pressure),
            'bat_volt':      _fv(g_latest.tegangan_batere),
            'bat_cond':      _fv(g_latest.batere_condition),
            'radiator':      _fv(g_latest.radiator),
            'counter':       _fv(g_latest.counter_sesudah),
            'waktu_transisi':_fv(g_latest.waktu_transisi),
            'pln_v_avg':     _avg(g_latest.pln_v_rn, g_latest.pln_v_sn, g_latest.pln_v_tn),
            'gen_v_avg':     _avg(g_latest.gen_v_rn, g_latest.gen_v_sn, g_latest.gen_v_tn),
            'pln_f_avg':     _avg(g_latest.pln_f_r,  g_latest.pln_f_s,  g_latest.pln_f_t),
            'gen_f_avg':     _avg(g_latest.gen_f_r,  g_latest.gen_f_s,  g_latest.gen_f_t),
            'history':       history,
        })
    if selected_id:
        try:
            selected_id = int(selected_id)
        except ValueError:
            selected_id = None

    if not selected_id and devices_data:
        selected_id = devices_data[0]['device_id']

    ups_selected_id = ups_data[0]['device_id'] if ups_data else None

    return render(request, 'maintenance/catu_daya_dashboard.html', {
        'devices_data_json':   json.dumps(devices_data),
        'devices_data':        devices_data,
        'selected_id':         selected_id,
        'threshold':           THRESHOLD,
        'ups_data_json':       json.dumps(ups_data),
        'ups_data':            ups_data,
        'ups_vac_threshold':   UPS_VAC_THRESHOLD,
        'ups_selected_id':     ups_selected_id,
        'genset_data':         genset_data,
        'genset_data_json':    json.dumps(genset_data),
    })


# ─────────────────────────────────────────────────────────────────────
# CORRECTIVE MAINTENANCE — EDIT
# ─────────────────────────────────────────────────────────────────────
@login_required
@require_can_edit
def corrective_edit(request, pk):
    """
    Edit corrective maintenance. pk = Maintenance.pk
    (maintenance_edit() meredirect ke sini untuk tipe Corrective).
    """
    from maintenance.models import MaintenanceCorrective
    from datetime import date as _date

    maint = get_object_or_404(Maintenance, pk=pk)
    corr  = get_object_or_404(MaintenanceCorrective, maintenance=maint)
    device = maint.device

    if request.method == 'POST':
        tanggal           = request.POST.get('tanggal', '').strip()
        pelaksana_raw     = request.POST.get('pelaksana_names_input') or request.POST.get('pelaksana_names', '')
        jenis_kerusakan   = request.POST.get('jenis_kerusakan', '')
        deskripsi_masalah = request.POST.get('deskripsi_masalah', '').strip()
        tindakan          = request.POST.get('tindakan', '').strip()
        komponen_diganti  = request.POST.get('komponen_diganti') == 'on'
        nama_komponen     = request.POST.get('nama_komponen', '').strip()
        kondisi_sebelum   = request.POST.get('kondisi_sebelum', '').strip()
        kondisi_sesudah   = request.POST.get('kondisi_sesudah', '').strip()
        durasi_jam        = request.POST.get('durasi_jam', '') or None
        durasi_menit      = request.POST.get('durasi_menit', '') or None
        status_perbaikan  = request.POST.get('status_perbaikan', 'selesai')
        foto_sebelum      = request.FILES.get('foto_sebelum')
        foto_sesudah      = request.FILES.get('foto_sesudah')

        if tanggal and deskripsi_masalah and tindakan:
            # Parse pelaksana (JSON array atau comma-separated)
            try:
                pelaksana_list = json.loads(pelaksana_raw) if pelaksana_raw.startswith('[') else \
                                 [n.strip() for n in pelaksana_raw.split(',') if n.strip()]
            except (json.JSONDecodeError, ValueError):
                pelaksana_list = [n.strip() for n in pelaksana_raw.split(',') if n.strip()]

            maint.date             = tanggal
            maint.description      = deskripsi_masalah
            maint.status           = 'Done' if status_perbaikan == 'selesai' else 'Open'
            maint.pelaksana_names  = pelaksana_list
            maint.save()

            corr.jenis_kerusakan   = jenis_kerusakan
            corr.deskripsi_masalah = deskripsi_masalah
            corr.tindakan          = tindakan
            corr.komponen_diganti  = komponen_diganti
            corr.nama_komponen     = nama_komponen
            corr.kondisi_sebelum   = kondisi_sebelum
            corr.kondisi_sesudah   = kondisi_sesudah
            corr.durasi_jam        = int(durasi_jam)   if durasi_jam   else None
            corr.durasi_menit      = int(durasi_menit) if durasi_menit else None
            corr.status_perbaikan  = status_perbaikan
            if foto_sebelum: corr.foto_sebelum = foto_sebelum
            if foto_sesudah: corr.foto_sesudah = foto_sesudah
            corr.save()

            return redirect('maintenance_view', pk=maint.pk)

    # Pre-fill context
    try:
        tanggal_str = maint.date.strftime('%Y-%m-%dT%H:%M')
    except Exception:
        tanggal_str = str(maint.date) if maint.date else ''

    pelaksana_json = json.dumps(maint.pelaksana_names or [])

    gangguan_aktif = Gangguan.objects.filter(
        status__in=['open', 'in_progress']
    ).order_by('-tanggal_gangguan')

    devices = Device.objects.filter(is_deleted=False).select_related('jenis').order_by('lokasi', 'nama')

    from datetime import date as _date
    return render(request, 'maintenance/corrective_form.html', {
        'device_init':          device,
        'gangguan_init':        corr.gangguan,
        'gangguan_aktif':       gangguan_aktif,
        'devices':              devices,
        'corr':                 corr,
        'is_edit':              True,
        'from_gangguan':        corr.gangguan is not None,
        'from_device':          True,
        'today_date':           tanggal_str,
        'pelaksana_init_json':  pelaksana_json,
    })


# ─────────────────────────────────────────────────────────────────────
# OFFLINE FORM — Download template & Upload hasil
# ─────────────────────────────────────────────────────────────────────
@login_required
def offline_form_download(request):
    """Placeholder — download template form offline (Excel/PDF)."""
    from django.http import HttpResponse
    return HttpResponse("Fitur offline form download belum tersedia.", status=501)


@login_required
def offline_form_upload(request):
    """Placeholder — upload hasil form offline."""
    from django.http import HttpResponse
    return HttpResponse("Fitur offline form upload belum tersedia.", status=501)


# ─────────────────────────────────────────────────────────────────────
# EKSPOR DATA — BERITA ACARA (Pemasangan / Pembongkaran / Penggantian)
# ─────────────────────────────────────────────────────────────────────

_BULAN_ID_FULL = [
    'Januari','Februari','Maret','April','Mei','Juni',
    'Juli','Agustus','September','Oktober','November','Desember',
]


def _load_ba_sig_b64(user):
    """Return base64-encoded signature image for user, or empty string."""
    if not user:
        return ''
    try:
        import os as _os, base64 as _b
        path = user.profile.signature.path
        if path and _os.path.exists(path):
            with open(path, 'rb') as f:
                return _b.b64encode(f.read()).decode()
    except Exception:
        pass
    return ''


def _ba_ttd_ctx(record):
    """Build TTD context dict for BA PDF templates."""
    def _display(u):
        if not u:
            return ''
        try:
            return u.profile.get_display_name()
        except Exception:
            return u.get_full_name() or u.username

    return {
        'ttd_engineer_name':   _display(record.ttd_engineer),
        'ttd_engineer_sig_b64': _load_ba_sig_b64(record.ttd_engineer),
        'ttd_am_name':         _display(record.ttd_am),
        'ttd_am_sig_b64':      _load_ba_sig_b64(record.ttd_am),
    }

def _load_logo_b64():
    import os, base64
    from django.conf import settings
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'pln_logo_conv.png')
    if os.path.exists(logo_path):
        with open(logo_path, 'rb') as f:
            return base64.b64encode(f.read()).decode()
    return ''


def _render_ba_pdf(template_name, ctx, filename):
    from django.template.loader import render_to_string
    try:
        import weasyprint
        html = render_to_string(template_name, ctx)
        pdf  = weasyprint.HTML(string=html).write_pdf()
        resp = HttpResponse(pdf, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp
    except ImportError:
        return HttpResponse('WeasyPrint tidak tersedia di server ini.', status=500)


def _format_tanggal(tanggal_str):
    """Convert '2026-04-15' → '15 April 2026' (locale-independent)."""
    try:
        from datetime import datetime
        d = datetime.strptime(tanggal_str, '%Y-%m-%d')
        return f'{d.day} {_BULAN_ID_FULL[d.month - 1]} {d.year}'
    except (ValueError, IndexError):
        return tanggal_str


def _ba_device_context():
    """Shared GET context: full device list + filter options."""
    devices = Device.objects.filter(
        is_deleted=False
    ).select_related('jenis').order_by('jenis__name', 'nama')

    jenis_list = list(
        DeviceType.objects.values_list('name', flat=True).order_by('name')
    )
    lokasi_list = list(
        Device.objects.filter(is_deleted=False)
        .exclude(lokasi='').values_list('lokasi', flat=True)
        .distinct().order_by('lokasi')
    )
    return devices, jenis_list, lokasi_list


def _ba_extra_ctx(tanggal, nomor_ba):
    """Hitung hari, bulan-tahun, tahun, dan nama file dari tanggal & nomor BA."""
    import re as _re
    from datetime import datetime as _dt
    HARI_ID = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
    try:
        d = _dt.strptime(tanggal, '%Y-%m-%d')
        tahun              = str(d.year)
        hari_display       = HARI_ID[d.weekday()]
        bulan_tahun_display = f'{_BULAN_ID_FULL[d.month - 1]} {d.year}'
    except (ValueError, IndexError):
        tahun = tanggal[:4] if tanggal else ''
        hari_display        = ''
        bulan_tahun_display = tanggal
    nomor_clean = _re.sub(r'[^\w]', '', nomor_ba) if nomor_ba else 'export'
    fname_base  = f'{nomor_clean}.BAPFASOPUP2BS-MKS{tahun}'
    return tahun, hari_display, bulan_tahun_display, fname_base


def _save_ba_record(jenis, nomor_ba, tanggal_str, pelaksana, nip, jabatan, catatan, rows, eviden_files, eviden_captions, user):
    """Simpan record BA + eviden ke database."""
    from datetime import datetime as _dt
    try:
        tanggal_date = _dt.strptime(tanggal_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        tanggal_date = None
    try:
        rec = BeritaAcaraRecord.objects.create(
            jenis=jenis,
            nomor_ba=nomor_ba,
            tanggal=tanggal_date,
            pelaksana=pelaksana,
            nip=nip,
            jabatan=jabatan,
            catatan=catatan,
            rows_data=rows,
            created_by=user,
        )
        for i, f in enumerate(eviden_files):
            try:
                f.seek(0)
            except Exception:
                pass
            catatan_ev = eviden_captions[i] if i < len(eviden_captions) else ''
            BeritaAcaraEviden.objects.create(ba=rec, gambar=f, catatan=catatan_ev, urutan=i)
    except Exception:
        pass  # jangan gagalkan export PDF jika penyimpanan error


@login_required
def ba_list(request):
    from django.contrib.auth import get_user_model as _get_user_model
    records = (
        BeritaAcaraRecord.objects
        .select_related('created_by', 'ttd_req_to', 'ttd_engineer', 'ttd_am')
        .prefetch_related('evidens')
        .order_by('-created_at')
    )
    _User = _get_user_model()
    engineers = (
        _User.objects
        .filter(profile__role='technician')
        .select_related('profile')
        .order_by('profile__display_name', 'first_name')
    )
    return render(request, 'maintenance/ba_list.html', {
        'records':   records,
        'engineers': engineers,
    })


@login_required
def ba_request_sign(request, pk):
    from django.contrib import messages as dj_messages
    from django.contrib.auth import get_user_model as _gum
    record = get_object_or_404(BeritaAcaraRecord, pk=pk)
    if not (request.user.is_superuser or record.created_by == request.user):
        dj_messages.error(request, 'Tidak memiliki izin.')
        return redirect('ba_list')
    if request.method == 'POST':
        eng_id = request.POST.get('engineer_id')
        try:
            eng = _gum().objects.select_related('profile').get(pk=eng_id, profile__role='technician')
            record.ttd_req_to = eng
            record.ttd_status = 'menunggu_engineer'
            record.save(update_fields=['ttd_req_to', 'ttd_status'])
            dj_messages.success(request, f'Permintaan TTD dikirim ke {eng.profile.get_display_name()}.')
        except Exception:
            dj_messages.error(request, 'Engineer tidak valid.')
    return redirect('ba_list')


@login_required
def ba_sign_engineer(request, pk):
    from django.contrib import messages as dj_messages
    record = get_object_or_404(BeritaAcaraRecord, pk=pk)
    if record.ttd_status != 'menunggu_engineer' or record.ttd_req_to_id != request.user.pk:
        dj_messages.error(request, 'Anda tidak dapat menandatangani BA ini.')
        return redirect('ba_list')
    if request.method == 'POST':
        record.ttd_engineer    = request.user
        record.ttd_engineer_at = dj_timezone.now()
        record.ttd_status      = 'signed_engineer'
        record.save(update_fields=['ttd_engineer', 'ttd_engineer_at', 'ttd_status'])
        dj_messages.success(request, 'Berita Acara berhasil Anda tandatangani.')
    return redirect('ba_list')


@login_required
def ba_sign_am(request, pk):
    from django.contrib import messages as dj_messages
    record = get_object_or_404(BeritaAcaraRecord, pk=pk)
    try:
        is_am = request.user.profile.is_asisten_manager
    except Exception:
        is_am = False
    if not (request.user.is_superuser or is_am):
        dj_messages.error(request, 'Hanya Asisten Manager yang dapat menandatangani sebagai AM.')
        return redirect('ba_list')
    if record.ttd_status != 'signed_engineer':
        dj_messages.error(request, 'BA belum ditandatangani oleh engineer.')
        return redirect('ba_list')
    if request.method == 'POST':
        record.ttd_am    = request.user
        record.ttd_am_at = dj_timezone.now()
        record.ttd_status = 'signed_am'
        record.save(update_fields=['ttd_am', 'ttd_am_at', 'ttd_status'])
        dj_messages.success(request, 'Berita Acara berhasil ditandatangani sebagai AM.')
    return redirect('ba_list')


@login_required
def ba_export(request, pk):
    import re as _re2
    record = get_object_or_404(BeritaAcaraRecord, pk=pk)

    import base64 as _b64
    eviden_list = []
    for ev in record.evidens.all():
        if ev.gambar:
            try:
                with open(ev.gambar.path, 'rb') as f:
                    data = f.read()
                name = ev.gambar.name.lower()
                if name.endswith('.png'):
                    mime = 'image/png'
                elif name.endswith('.gif'):
                    mime = 'image/gif'
                else:
                    mime = 'image/jpeg'
                eviden_list.append({'b64': _b64.b64encode(data).decode(), 'mime': mime, 'catatan': ev.catatan})
            except Exception:
                pass

    HARI_ID = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
    d = record.tanggal
    hari_display        = HARI_ID[d.weekday()]
    tanggal_display     = f'{d.day} {_BULAN_ID_FULL[d.month - 1]} {d.year}'
    bulan_tahun_display = f'{_BULAN_ID_FULL[d.month - 1]} {d.year}'

    ctx = {
        'logo_b64':             _load_logo_b64(),
        'nomor_ba':             record.nomor_ba,
        'tanggal_display':      tanggal_display,
        'hari_display':         hari_display,
        'bulan_tahun_display':  bulan_tahun_display,
        'pelaksana':            record.pelaksana,
        'nip':                  record.nip,
        'jabatan':              record.jabatan,
        'catatan':              record.catatan,
        'rows':                 record.rows_data,
        'eviden_list':          eviden_list,
        **_ba_ttd_ctx(record),
    }
    template_map = {
        'pemasangan':   'maintenance/pdf/ba_pemasangan.html',
        'pembongkaran': 'maintenance/pdf/ba_pembongkaran.html',
        'penggantian':  'maintenance/pdf/ba_penggantian.html',
    }
    template = template_map.get(record.jenis, 'maintenance/pdf/ba_pemasangan.html')
    nomor_clean = _re2.sub(r'[^\w]', '', record.nomor_ba) if record.nomor_ba else 'export'
    return _render_ba_pdf(template, ctx, f'{nomor_clean}.pdf')


@login_required
def ba_delete(request, pk):
    from django.contrib import messages as dj_messages
    record = get_object_or_404(BeritaAcaraRecord, pk=pk)
    if not (request.user.is_superuser or record.created_by == request.user):
        dj_messages.error(request, 'Anda tidak memiliki izin untuk menghapus BA ini.')
        return redirect('ba_list')
    if request.method == 'POST':
        for ev in record.evidens.all():
            if ev.gambar:
                try:
                    import os as _os
                    if _os.path.exists(ev.gambar.path):
                        _os.remove(ev.gambar.path)
                except Exception:
                    pass
        record.delete()
        dj_messages.success(request, 'Data BA berhasil dihapus.')
    return redirect('ba_list')


@login_required
def ba_preview(request, pk):
    import re as _re2
    record = get_object_or_404(BeritaAcaraRecord, pk=pk)

    import base64 as _b64
    eviden_list = []
    for ev in record.evidens.all():
        if ev.gambar:
            try:
                with open(ev.gambar.path, 'rb') as f:
                    data = f.read()
                name = ev.gambar.name.lower()
                if name.endswith('.png'):
                    mime = 'image/png'
                elif name.endswith('.gif'):
                    mime = 'image/gif'
                else:
                    mime = 'image/jpeg'
                eviden_list.append({'b64': _b64.b64encode(data).decode(), 'mime': mime, 'catatan': ev.catatan})
            except Exception:
                pass

    HARI_ID = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
    d = record.tanggal
    hari_display        = HARI_ID[d.weekday()]
    tanggal_display     = f'{d.day} {_BULAN_ID_FULL[d.month - 1]} {d.year}'
    bulan_tahun_display = f'{_BULAN_ID_FULL[d.month - 1]} {d.year}'

    ctx = {
        'logo_b64':             _load_logo_b64(),
        'nomor_ba':             record.nomor_ba,
        'tanggal_display':      tanggal_display,
        'hari_display':         hari_display,
        'bulan_tahun_display':  bulan_tahun_display,
        'pelaksana':            record.pelaksana,
        'nip':                  record.nip,
        'jabatan':              record.jabatan,
        'catatan':              record.catatan,
        'rows':                 record.rows_data,
        'eviden_list':          eviden_list,
        **_ba_ttd_ctx(record),
    }
    template_map = {
        'pemasangan':   'maintenance/pdf/ba_pemasangan.html',
        'pembongkaran': 'maintenance/pdf/ba_pembongkaran.html',
        'penggantian':  'maintenance/pdf/ba_penggantian.html',
    }
    template = template_map.get(record.jenis, 'maintenance/pdf/ba_pemasangan.html')
    from django.template.loader import render_to_string
    try:
        import weasyprint
        html = render_to_string(template, ctx)
        pdf  = weasyprint.HTML(string=html).write_pdf()
        nomor_clean = _re2.sub(r'[^\w]', '', record.nomor_ba) if record.nomor_ba else 'preview'
        resp = HttpResponse(pdf, content_type='application/pdf')
        resp['Content-Disposition'] = f'inline; filename="{nomor_clean}.pdf"'
        return resp
    except ImportError:
        return HttpResponse('WeasyPrint tidak tersedia di server ini.', status=500)


@login_required
def ba_pemasangan(request):
    devices, jenis_list, lokasi_list = _ba_device_context()

    if request.method == 'POST':
        nomor_input     = request.POST.get('nomor_ba', '').strip()
        tanggal         = request.POST.get('tanggal', '').strip()
        pelaksana       = request.POST.get('pelaksana', '').strip()
        nip             = request.POST.get('nip', '').strip()
        jabatan         = request.POST.get('jabatan', '').strip()
        catatan         = request.POST.get('catatan', '').strip()
        device_ids      = request.POST.getlist('device_ids[]')
        lokasi_tujuan   = request.POST.getlist('lokasi_tujuan[]')
        keterangan_list = request.POST.getlist('keterangan[]')

        dev_map = {
            str(d.pk): d
            for d in Device.objects.select_related('jenis').filter(pk__in=device_ids)
        }
        rows = []
        for i, did in enumerate(device_ids):
            dev = dev_map.get(did)
            if not dev:
                continue
            rows.append({
                'no':            i + 1,
                'nama':          dev.nama,
                'jenis':         dev.jenis.name if dev.jenis else '-',
                'serial_number': dev.serial_number or '-',
                'lokasi_tujuan': lokasi_tujuan[i] if i < len(lokasi_tujuan) else (dev.lokasi or '-'),
                'keterangan':    keterangan_list[i] if i < len(keterangan_list) else '',
            })

        tahun, hari, bulan_tahun, fname_base = _ba_extra_ctx(tanggal, nomor_input)
        nomor_ba = f'{nomor_input}.BA/FASOP/UP2BS-MKS/{tahun}' if nomor_input else ''
        import base64 as _b64
        eviden_files    = request.FILES.getlist('eviden')
        eviden_captions = request.POST.getlist('eviden_catatan[]')
        eviden_list = [
            {
                'b64':     _b64.b64encode(f.read()).decode(),
                'mime':    f.content_type or 'image/jpeg',
                'catatan': eviden_captions[i] if i < len(eviden_captions) else '',
            }
            for i, f in enumerate(eviden_files)
        ]
        ctx = {
            'logo_b64':          _load_logo_b64(),
            'nomor_ba':          nomor_ba,
            'tanggal_display':   _format_tanggal(tanggal),
            'hari_display':      hari,
            'bulan_tahun_display': bulan_tahun,
            'pelaksana':         pelaksana,
            'nip':               nip,
            'jabatan':           jabatan,
            'catatan':           catatan,
            'rows':              rows,
            'eviden_list':       eviden_list,
        }
        _save_ba_record('pemasangan', nomor_ba, tanggal, pelaksana, nip, jabatan, catatan, rows, eviden_files, eviden_captions, request.user)
        return _render_ba_pdf('maintenance/pdf/ba_pemasangan.html', ctx, f'{fname_base}.pdf')

    return render(request, 'maintenance/ba_pemasangan.html', {
        'devices':     devices,
        'jenis_list':  jenis_list,
        'lokasi_list': lokasi_list,
        'today':       date.today().isoformat(),
    })


@login_required
def ba_pembongkaran(request):
    devices, jenis_list, lokasi_list = _ba_device_context()

    if request.method == 'POST':
        nomor_input     = request.POST.get('nomor_ba', '').strip()
        tanggal         = request.POST.get('tanggal', '').strip()
        pelaksana       = request.POST.get('pelaksana', '').strip()
        nip             = request.POST.get('nip', '').strip()
        jabatan         = request.POST.get('jabatan', '').strip()
        catatan         = request.POST.get('catatan', '').strip()
        device_ids      = request.POST.getlist('device_ids[]')
        keterangan_list = request.POST.getlist('keterangan[]')

        dev_map = {
            str(d.pk): d
            for d in Device.objects.select_related('jenis').filter(pk__in=device_ids)
        }
        rows = []
        for i, did in enumerate(device_ids):
            dev = dev_map.get(did)
            if not dev:
                continue
            rows.append({
                'no':            i + 1,
                'nama':          dev.nama,
                'jenis':         dev.jenis.name if dev.jenis else '-',
                'serial_number': dev.serial_number or '-',
                'lokasi_asal':   dev.lokasi or '-',
                'keterangan':    keterangan_list[i] if i < len(keterangan_list) else '',
            })

        tahun, hari, bulan_tahun, fname_base = _ba_extra_ctx(tanggal, nomor_input)
        nomor_ba = f'{nomor_input}.BA/FASOP/UP2BS-MKS/{tahun}' if nomor_input else ''
        import base64 as _b64
        eviden_files    = request.FILES.getlist('eviden')
        eviden_captions = request.POST.getlist('eviden_catatan[]')
        eviden_list = [
            {
                'b64':     _b64.b64encode(f.read()).decode(),
                'mime':    f.content_type or 'image/jpeg',
                'catatan': eviden_captions[i] if i < len(eviden_captions) else '',
            }
            for i, f in enumerate(eviden_files)
        ]
        ctx = {
            'logo_b64':          _load_logo_b64(),
            'nomor_ba':          nomor_ba,
            'tanggal_display':   _format_tanggal(tanggal),
            'hari_display':      hari,
            'bulan_tahun_display': bulan_tahun,
            'pelaksana':         pelaksana,
            'nip':               nip,
            'jabatan':           jabatan,
            'catatan':           catatan,
            'rows':              rows,
            'eviden_list':       eviden_list,
        }
        _save_ba_record('pembongkaran', nomor_ba, tanggal, pelaksana, nip, jabatan, catatan, rows, eviden_files, eviden_captions, request.user)
        return _render_ba_pdf('maintenance/pdf/ba_pembongkaran.html', ctx, f'{fname_base}.pdf')

    return render(request, 'maintenance/ba_pembongkaran.html', {
        'devices':     devices,
        'jenis_list':  jenis_list,
        'lokasi_list': lokasi_list,
        'today':       date.today().isoformat(),
    })


@login_required
def ba_penggantian(request):
    devices, jenis_list, lokasi_list = _ba_device_context()

    if request.method == 'POST':
        nomor_input     = request.POST.get('nomor_ba', '').strip()
        tanggal         = request.POST.get('tanggal', '').strip()
        pelaksana       = request.POST.get('pelaksana', '').strip()
        nip             = request.POST.get('nip', '').strip()
        jabatan         = request.POST.get('jabatan', '').strip()
        catatan         = request.POST.get('catatan', '').strip()
        device_ids      = request.POST.getlist('device_ids[]')
        komponen_lama   = request.POST.getlist('komponen_lama[]')
        komponen_baru   = request.POST.getlist('komponen_baru[]')
        keterangan_list = request.POST.getlist('keterangan[]')

        dev_map = {
            str(d.pk): d
            for d in Device.objects.select_related('jenis').filter(pk__in=device_ids)
        }
        rows = []
        for i, did in enumerate(device_ids):
            dev = dev_map.get(did)
            if not dev:
                continue
            rows.append({
                'no':            i + 1,
                'nama':          dev.nama,
                'jenis':         dev.jenis.name if dev.jenis else '-',
                'komponen_lama': komponen_lama[i] if i < len(komponen_lama) else '',
                'komponen_baru': komponen_baru[i] if i < len(komponen_baru) else '',
                'keterangan':    keterangan_list[i] if i < len(keterangan_list) else '',
            })

        tahun, hari, bulan_tahun, fname_base = _ba_extra_ctx(tanggal, nomor_input)
        nomor_ba = f'{nomor_input}.BA/FASOP/UP2BS-MKS/{tahun}' if nomor_input else ''
        import base64 as _b64
        eviden_files    = request.FILES.getlist('eviden')
        eviden_captions = request.POST.getlist('eviden_catatan[]')
        eviden_list = [
            {
                'b64':     _b64.b64encode(f.read()).decode(),
                'mime':    f.content_type or 'image/jpeg',
                'catatan': eviden_captions[i] if i < len(eviden_captions) else '',
            }
            for i, f in enumerate(eviden_files)
        ]
        ctx = {
            'logo_b64':          _load_logo_b64(),
            'nomor_ba':          nomor_ba,
            'tanggal_display':   _format_tanggal(tanggal),
            'hari_display':      hari,
            'bulan_tahun_display': bulan_tahun,
            'pelaksana':         pelaksana,
            'nip':               nip,
            'jabatan':           jabatan,
            'catatan':           catatan,
            'rows':              rows,
            'eviden_list':       eviden_list,
        }
        _save_ba_record('penggantian', nomor_ba, tanggal, pelaksana, nip, jabatan, catatan, rows, eviden_files, eviden_captions, request.user)
        return _render_ba_pdf('maintenance/pdf/ba_penggantian.html', ctx, f'{fname_base}.pdf')

    return render(request, 'maintenance/ba_penggantian.html', {
        'devices':     devices,
        'jenis_list':  jenis_list,
        'lokasi_list': lokasi_list,
        'today':       date.today().isoformat(),
    })
