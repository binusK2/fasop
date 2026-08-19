import datetime
import json
import logging
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponseForbidden
from django.utils import timezone
from django.db.models import Q
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from fasop.hashids_helper import encode
from .models import RTU, RTULog, ZabbixHost, ZabbixEventLog, ZabbixWebhookLog
from .notifications import notif_zabbix_transisi

logger = logging.getLogger(__name__)


def _boundaries():
    """
    Kembalikan (now, today_start, month_start) semuanya timezone-aware UTC.
    Batas dihitung dari waktu LOKAL agar benar untuk zona Asia/Makassar.
    """
    tz_local   = timezone.get_current_timezone()
    now        = timezone.now()                         # UTC-aware
    now_local  = now.astimezone(tz_local)

    today_start = now_local.replace(
        hour=0, minute=0, second=0, microsecond=0
    )                                                   # masih aware (local tz)

    month_start = now_local.replace(
        day=1, hour=0, minute=0, second=0, microsecond=0
    )

    return now, today_start, month_start


def _calc_avail(logs_down, since, now, total_menit):
    """
    Hitung menit DOWN dari kumpulan log yang sudah di-filter,
    lalu kembalikan availability %.
    logs_down: queryset atau list RTULog (state=DOWN) yang overlap dengan [since, now].
    """
    down_menit = 0
    for log in logs_down:
        start = max(log.mulai, since)
        end   = min(log.selesai, now) if log.selesai else now
        down_menit += max(0, int((end - start).total_seconds() / 60))

    up_menit = max(0, total_menit - down_menit)
    return round(up_menit / total_menit * 100, 2)


@login_required
def dashboard(request):
    return render(request, 'device_mon/dashboard.html')


@login_required
def api_status(request):
    """
    JSON: status semua RTU + statistik availability.
    Dipanggil oleh dashboard setiap 60 detik.

    Optimasi: ambil semua log DOWN dalam satu query per periode,
    kemudian group by rtu_id di Python — tidak ada N+1.
    """
    now, today_start, month_start = _boundaries()

    rtus = list(RTU.objects.filter(aktif=True))
    if not rtus:
        return JsonResponse({
            'total_up': 0, 'total_down': 0, 'total_rtu': 0,
            'avail_hari': None, 'avail_bulan': None,
            'rtus': [], 'gangguan_terkini': [],
        })

    rtu_ids = [r.pk for r in rtus]

    # ── Ambil semua log DOWN relevan dalam SATU query per periode ──────
    # Hari ini
    down_today = list(
        RTULog.objects.filter(
            rtu_id__in=rtu_ids,
            state='DOWN',
            mulai__lt=now,
        ).filter(Q(selesai__gt=today_start) | Q(selesai__isnull=True))
    )

    # Bulan ini
    down_month = list(
        RTULog.objects.filter(
            rtu_id__in=rtu_ids,
            state='DOWN',
            mulai__lt=now,
        ).filter(Q(selesai__gt=month_start) | Q(selesai__isnull=True))
    )

    # Group log by rtu_id di Python (tanpa loop N query)
    def group_by_rtu(logs):
        d = {}
        for log in logs:
            d.setdefault(log.rtu_id, []).append(log)
        return d

    today_by_rtu = group_by_rtu(down_today)
    month_by_rtu = group_by_rtu(down_month)

    total_menit_today = max(1, int((now - today_start).total_seconds() / 60))
    total_menit_month = max(1, int((now - month_start).total_seconds() / 60))

    # ── Hitung per RTU ─────────────────────────────────────────────────
    rtu_data   = []
    total_up   = 0
    total_down = 0
    avail_hari_list = []

    for rtu in rtus:
        if rtu.state == 'UP':
            total_up  += 1
        elif rtu.state == 'DOWN':
            total_down += 1

        avail_hari  = _calc_avail(
            today_by_rtu.get(rtu.pk, []), today_start, now, total_menit_today
        )
        avail_bulan = _calc_avail(
            month_by_rtu.get(rtu.pk, []), month_start, now, total_menit_month
        )
        avail_hari_list.append(avail_hari)

        # Menit DOWN hari ini (hanya yang sudah selesai + yang masih jalan)
        down_menit_hari = 0
        for log in today_by_rtu.get(rtu.pk, []):
            start = max(log.mulai, today_start)
            end   = min(log.selesai, now) if log.selesai else now
            down_menit_hari += max(0, int((end - start).total_seconds() / 60))

        rtu_data.append({
            'id':              rtu.pk,
            'nama':            rtu.nama,
            'lokasi':          rtu.lokasi,
            'state':           rtu.state,
            'state_sejak':     rtu.state_sejak.isoformat() if rtu.state_sejak else None,
            'durasi_menit':    rtu.durasi_menit,
            'avail_hari':      avail_hari,
            'avail_bulan':     avail_bulan,
            'down_menit_hari': down_menit_hari,
        })

    avg_avail_hari  = round(sum(avail_hari_list) / len(avail_hari_list), 2) if avail_hari_list else None
    avg_avail_bulan = _calc_avail(down_month, month_start, now, total_menit_month * len(rtus)) if rtus else None

    # ── 10 log gangguan terkini ────────────────────────────────────────
    gangguan = RTULog.objects.filter(state='DOWN').select_related('rtu').order_by('-mulai')[:10]
    gangguan_data = [
        {
            'rtu':          g.rtu.nama,
            'state':        g.state,
            'mulai':        g.mulai.isoformat(),
            'selesai':      g.selesai.isoformat() if g.selesai else None,
            'durasi_menit': g.durasi_menit,
        }
        for g in gangguan
    ]

    return JsonResponse({
        'total_up':         total_up,
        'total_down':       total_down,
        'total_rtu':        len(rtu_data),
        'avail_hari':       avg_avail_hari,
        'avail_bulan':      avg_avail_bulan,
        'rtus':             rtu_data,
        'gangguan_terkini': gangguan_data,
    })


@login_required
def rtu_detail(request, pk):
    """Halaman detail satu RTU — info state, availability, histori log."""
    rtu = get_object_or_404(RTU, pk=pk)
    return render(request, 'device_mon/rtu_detail.html', {'rtu': rtu})


@login_required
def api_rtu_logs(request, pk):
    """
    JSON: log state RTU untuk chart timeline + tabel.
    ?hari=N  (default 7, max 30)
    """
    rtu = get_object_or_404(RTU, pk=pk)
    try:
        hari = min(max(int(request.GET.get('hari', 7)), 1), 30)
    except (ValueError, TypeError):
        hari = 7

    now, today_start, month_start = _boundaries()
    tz_local = timezone.get_current_timezone()
    since    = now - datetime.timedelta(days=hari)

    logs = (RTULog.objects
            .filter(rtu=rtu, mulai__gte=since)
            .order_by('mulai'))

    total_menit = max(1, int((now - since).total_seconds() / 60))
    down_logs   = [l for l in logs if l.state == 'DOWN']
    avail       = _calc_avail(down_logs, since, now, total_menit)

    # Availability hari ini & bulan ini
    down_today = [l for l in RTULog.objects.filter(
        rtu=rtu, state='DOWN', mulai__lt=now
    ).filter(Q(selesai__gt=today_start) | Q(selesai__isnull=True))]
    down_month = [l for l in RTULog.objects.filter(
        rtu=rtu, state='DOWN', mulai__lt=now
    ).filter(Q(selesai__gt=month_start) | Q(selesai__isnull=True))]

    avail_hari  = _calc_avail(down_today, today_start, now,
                              max(1, int((now - today_start).total_seconds() / 60)))
    avail_bulan = _calc_avail(down_month, month_start, now,
                              max(1, int((now - month_start).total_seconds() / 60)))

    # Total menit DOWN dalam periode
    down_menit_total = sum(
        max(0, int((
            (min(l.selesai, now) if l.selesai else now) -
            max(l.mulai, since)
        ).total_seconds() / 60))
        for l in down_logs
    )

    # Format log untuk tabel & chart
    log_data = []
    for l in logs:
        mulai_local   = l.mulai.astimezone(tz_local)
        selesai_local = l.selesai.astimezone(tz_local) if l.selesai else None
        # Durasi aktual (mungkin masih berjalan)
        dur = l.durasi_menit
        if dur is None and not l.selesai:
            dur = max(0, int((now - l.mulai).total_seconds() / 60))
        log_data.append({
            'state':        l.state,
            'mulai':        mulai_local.strftime('%d/%m %H:%M'),
            'selesai':      selesai_local.strftime('%d/%m %H:%M') if selesai_local else None,
            'durasi_menit': dur,
        })

    return JsonResponse({
        'nama':             rtu.nama,
        'lokasi':           rtu.lokasi,
        'state':            rtu.state,
        'state_sejak':      rtu.state_sejak.astimezone(tz_local).strftime('%d/%m/%Y %H:%M') if rtu.state_sejak else None,
        'durasi_menit':     rtu.durasi_menit,
        'avail_hari':       avail_hari,
        'avail_bulan':      avail_bulan,
        'avail_periode':    avail,
        'down_menit_total': down_menit_total,
        'hari':             hari,
        'logs':             log_data,
    })


@login_required
def gangguan_list(request):
    """Halaman histori gangguan."""
    logs = (RTULog.objects
            .filter(state='DOWN')
            .select_related('rtu')
            .order_by('-mulai')[:200])
    return render(request, 'device_mon/gangguan.html', {'logs': logs})


def _bulan_boundaries(tahun, bulan):
    """Batas awal & akhir (eksklusif) sebuah bulan, timezone-aware (local tz)."""
    tz_local = timezone.get_current_timezone()
    month_start = timezone.make_aware(datetime.datetime(tahun, bulan, 1), tz_local)
    if bulan == 12:
        next_tahun, next_bulan = tahun + 1, 1
    else:
        next_tahun, next_bulan = tahun, bulan + 1
    month_end = timezone.make_aware(datetime.datetime(next_tahun, next_bulan, 1), tz_local)
    return month_start, month_end


NAMA_BULAN = ['', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
              'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']


@login_required
def export_availability(request):
    """
    Download rekap availability RTU per bulan ke Excel.
    ?bulan=YYYY-MM  (default: bulan berjalan)

    Sheet "Detail Gangguan": tiap log DOWN RTU dalam bulan tsb, dengan kolom
    Eliminasi (dropdown Ya/Tidak) dan Durasi Efektif (formula Excel).
    Sheet "Rekap Availability": AV% per RTU dihitung via formula SUMIFS yang
    merujuk ke sheet Detail — jadi saat kolom Eliminasi diubah di Excel,
    nilai Av langsung ter-update otomatis tanpa perlu download ulang.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    tz_local = timezone.get_current_timezone()
    now = timezone.now()

    bulan_str = request.GET.get('bulan', '')
    try:
        tahun_str, bulan_num_str = bulan_str.split('-')
        tahun, bulan = int(tahun_str), int(bulan_num_str)
        if not (1 <= bulan <= 12):
            raise ValueError
    except (ValueError, TypeError):
        now_local = now.astimezone(tz_local)
        tahun, bulan = now_local.year, now_local.month

    month_start, month_end = _bulan_boundaries(tahun, bulan)
    effective_end = min(now, month_end)
    total_menit_bulan = max(1, int((effective_end - month_start).total_seconds() / 60))

    rtus = list(RTU.objects.filter(aktif=True))
    rtu_ids = [r.pk for r in rtus]

    down_logs = list(
        RTULog.objects.filter(
            rtu_id__in=rtu_ids,
            state='DOWN',
            mulai__lt=effective_end,
        ).filter(Q(selesai__gt=month_start) | Q(selesai__isnull=True))
        .select_related('rtu')
        .order_by('rtu__urutan', 'rtu__nama', 'mulai')
    )

    # ── Workbook & style ────────────────────────────────────────────
    wb = openpyxl.Workbook()

    hdr_fill = PatternFill('solid', fgColor='0F172A')
    hdr_font = Font(bold=True, color='60A5FA', size=10)
    thin     = Side(style='thin', color='1E293B')
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    center   = Alignment(horizontal='center', vertical='center')

    label_periode = f'{NAMA_BULAN[bulan]} {tahun}'

    # ── Sheet 1: Detail Gangguan ────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Detail Gangguan'

    headers1 = ['No', 'RTU', 'Lokasi', 'Turun (Down)', 'Naik (Up) Kembali',
                'Durasi Down (menit)', 'Eliminasi', 'Durasi Efektif (menit)']
    ws1.append(headers1)
    for col in range(1, len(headers1) + 1):
        cell = ws1.cell(1, col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = border
        cell.alignment = center

    for i, w in enumerate([6, 16, 18, 18, 18, 16, 12, 18], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    row_idx = 1
    for log in down_logs:
        row_idx += 1
        mulai_clip   = max(log.mulai, month_start).astimezone(tz_local)
        selesai_clip = (min(log.selesai, effective_end) if log.selesai else effective_end).astimezone(tz_local)
        durasi = max(0, int((selesai_clip - mulai_clip).total_seconds() / 60))

        ws1.append([
            row_idx - 1,
            log.rtu.nama,
            log.rtu.lokasi,
            mulai_clip.strftime('%d/%m/%Y %H:%M'),
            selesai_clip.strftime('%d/%m/%Y %H:%M') if log.selesai else 'Masih Down',
            durasi,
            'Tidak',
            f'=IF(G{row_idx}="Ya",0,F{row_idx})',
        ])
        for col in range(1, len(headers1) + 1):
            ws1.cell(row_idx, col).border = border
            ws1.cell(row_idx, col).alignment = center

    if row_idx == 1:
        # Tidak ada gangguan sama sekali pada bulan ini
        row_idx = 2
        ws1.append(['-', '-', '-', '-', '-', 0, 'Tidak', 0])
        for col in range(1, len(headers1) + 1):
            ws1.cell(row_idx, col).border = border
            ws1.cell(row_idx, col).alignment = center

    last_detail_row = row_idx

    dv = DataValidation(type='list', formula1='"Tidak,Ya"', allow_blank=False)
    dv.add(f'G2:G{last_detail_row}')
    ws1.add_data_validation(dv)

    ws1.freeze_panes = 'A2'

    # ── Sheet 2: Rekap Availability ─────────────────────────────────
    ws2 = wb.create_sheet('Rekap Availability')
    ws2.append([f'Periode: {label_periode}'])
    ws2.cell(1, 1).font = Font(bold=True, color='CCE0FF', size=11)
    ws2.append([])

    headers2 = ['No', 'RTU', 'Lokasi', 'Total Menit Bulan',
                'Total Down Efektif (menit)', 'Availability (%)']
    header_row2 = 3
    ws2.append(headers2)
    for col in range(1, len(headers2) + 1):
        cell = ws2.cell(header_row2, col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = border
        cell.alignment = center

    for i, w in enumerate([6, 16, 18, 18, 22, 16], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    row_idx2 = header_row2
    for rtu in rtus:
        row_idx2 += 1
        ws2.append([
            row_idx2 - header_row2,
            rtu.nama,
            rtu.lokasi,
            total_menit_bulan,
            f"=SUMIFS('Detail Gangguan'!$H$2:$H${last_detail_row},"
            f"'Detail Gangguan'!$B$2:$B${last_detail_row},B{row_idx2})",
            f'=ROUND((D{row_idx2}-E{row_idx2})/D{row_idx2}*100,2)',
        ])
        for col in range(1, len(headers2) + 1):
            ws2.cell(row_idx2, col).border = border
            ws2.cell(row_idx2, col).alignment = center

    first_data_row2 = header_row2 + 1
    last_rekap_row  = row_idx2
    ws2.append([])
    ws2.append(['', '', '', '', 'Rata-rata Availability',
                f'=ROUND(AVERAGE(F{first_data_row2}:F{last_rekap_row}),2)'])
    summary_row2 = last_rekap_row + 2
    ws2.cell(summary_row2, 5).font = Font(bold=True, color='34D399', size=10)
    ws2.cell(summary_row2, 6).font = Font(bold=True, color='34D399', size=10)

    ws2.freeze_panes = f'A{first_data_row2}'

    # ── Response ──────────────────────────────────────────────────
    filename = f'Availability_RTU_{tahun}-{bulan:02d}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def availability_report(request):
    """Halaman laporan availability per RTU — bulan ini."""
    now, _, month_start = _boundaries()

    rtu_ids = list(RTU.objects.filter(aktif=True).values_list('pk', flat=True))
    down_month = list(
        RTULog.objects.filter(
            rtu_id__in=rtu_ids,
            state='DOWN',
            mulai__lt=now,
        ).filter(Q(selesai__gt=month_start) | Q(selesai__isnull=True))
    )
    month_by_rtu   = {}
    for log in down_month:
        month_by_rtu.setdefault(log.rtu_id, []).append(log)

    total_menit = max(1, int((now - month_start).total_seconds() / 60))

    rtus = RTU.objects.filter(aktif=True)
    rows = []
    for rtu in rtus:
        avail = _calc_avail(month_by_rtu.get(rtu.pk, []), month_start, now, total_menit)
        rows.append({'rtu': rtu, 'avail': avail})

    tz_local   = timezone.get_current_timezone()
    month_start_local = month_start.astimezone(tz_local)

    return render(request, 'device_mon/availability.html', {
        'rows':   rows,
        'period': month_start_local.strftime('%B %Y'),
    })


# ═══════════════════════════════════════════════════════════════════════════
#  Zabbix — status host dipantau lewat Zabbix API (pull) + webhook (push).
#  Tampilan terpisah dari RTU di atas (sumber data beda: MSSQL vs Zabbix),
#  tapi tetap satu app Device Monitor supaya semua status peralatan
#  realtime ketemu di satu tempat.
# ═══════════════════════════════════════════════════════════════════════════
@login_required
def zbx_dashboard(request):
    """Ringkasan Zabbix: total status + breakdown per grup. Detail per host ada di zbx_group_detail."""
    return render(request, 'device_mon/zbx_dashboard.html')


@login_required
def zbx_api_summary(request):
    """JSON: ringkasan status Zabbix — total + breakdown per Host Group + problem terkini. Dipoll dashboard."""
    now, today_start, month_start = _boundaries()

    hosts = list(ZabbixHost.objects.filter(aktif=True))
    if not hosts:
        return JsonResponse({
            'total_ok': 0, 'total_problem': 0, 'total_unknown': 0, 'total_host': 0,
            'avail_hari': None, 'avail_bulan': None, 'groups': [], 'gangguan_terkini': [],
        })

    host_ids = [h.pk for h in hosts]
    down_today = list(
        ZabbixEventLog.objects.filter(
            host_id__in=host_ids, state='PROBLEM', mulai__lt=now,
        ).filter(Q(selesai__gt=today_start) | Q(selesai__isnull=True))
    )
    today_by_host = {}
    for log in down_today:
        today_by_host.setdefault(log.host_id, []).append(log)

    total_menit_today = max(1, int((now - today_start).total_seconds() / 60))

    total_ok = total_problem = total_unknown = 0
    avail_hari_list = []
    group_stats = {}   # nama grup -> {total, ok, problem, unknown, avail_list}

    for h in hosts:
        avail_hari = _calc_avail(today_by_host.get(h.pk, []), today_start, now, total_menit_today)
        avail_hari_list.append(avail_hari)

        if h.state == 'OK':
            total_ok += 1
        elif h.state == 'PROBLEM':
            total_problem += 1
        else:
            total_unknown += 1

        for gname in (h.group_list or ['(Tanpa Grup)']):
            g = group_stats.setdefault(gname, {'total': 0, 'ok': 0, 'problem': 0, 'unknown': 0, 'avail_list': []})
            g['total'] += 1
            g['avail_list'].append(avail_hari)
            if h.state == 'OK':
                g['ok'] += 1
            elif h.state == 'PROBLEM':
                g['problem'] += 1
            else:
                g['unknown'] += 1

    avg_avail_hari = round(sum(avail_hari_list) / len(avail_hari_list), 2) if avail_hari_list else None

    down_month = list(
        ZabbixEventLog.objects.filter(
            host_id__in=host_ids, state='PROBLEM', mulai__lt=now,
        ).filter(Q(selesai__gt=month_start) | Q(selesai__isnull=True))
    )
    total_menit_month = max(1, int((now - month_start).total_seconds() / 60))
    avg_avail_bulan = _calc_avail(down_month, month_start, now, total_menit_month * len(hosts))

    groups_data = [
        {
            'nama': gname,
            'total': g['total'],
            'ok': g['ok'],
            'problem': g['problem'],
            'unknown': g['unknown'],
            'avail_hari': round(sum(g['avail_list']) / len(g['avail_list']), 2) if g['avail_list'] else None,
        }
        for gname, g in sorted(group_stats.items())
    ]

    gangguan = (ZabbixEventLog.objects.filter(state='PROBLEM')
                .select_related('host').order_by('-mulai')[:10])
    gangguan_data = [{
        'host': g.host.nama,
        'severity': g.severity,
        'problem_name': g.problem_name,
        'mulai': g.mulai.isoformat(),
        'selesai': g.selesai.isoformat() if g.selesai else None,
        'durasi_menit': g.durasi_menit,
    } for g in gangguan]

    return JsonResponse({
        'total_ok': total_ok,
        'total_problem': total_problem,
        'total_unknown': total_unknown,
        'total_host': len(hosts),
        'avail_hari': avg_avail_hari,
        'avail_bulan': avg_avail_bulan,
        'groups': groups_data,
        'gangguan_terkini': gangguan_data,
    })


@login_required
def zbx_group_detail(request, group):
    """Tampilan detail satu Host Group Zabbix — grid host + chart availability, difilter ke grup ini."""
    return render(request, 'device_mon/zbx_group_detail.html', {'group': group})


@login_required
def zbx_api_status(request):
    """
    JSON: status host Zabbix + availability hari ini/bulan ini. Dipoll
    zbx_group_detail. ?group=<nama> memfilter ke satu Host Group saja
    (dipakai zbx_group_detail) — tanpa ?group, kembalikan semua host aktif.
    """
    now, today_start, month_start = _boundaries()

    hosts = list(ZabbixHost.objects.filter(aktif=True).select_related('device'))
    group_filter = request.GET.get('group', '').strip()
    if group_filter:
        hosts = [h for h in hosts if group_filter in h.group_list]

    if not hosts:
        return JsonResponse({
            'total_ok': 0, 'total_problem': 0, 'total_unknown': 0, 'total_host': 0,
            'avail_hari': None, 'avail_bulan': None, 'hosts': [], 'gangguan_terkini': [],
        })

    host_ids = [h.pk for h in hosts]
    down_today = list(
        ZabbixEventLog.objects.filter(
            host_id__in=host_ids, state='PROBLEM', mulai__lt=now,
        ).filter(Q(selesai__gt=today_start) | Q(selesai__isnull=True))
    )
    down_month = list(
        ZabbixEventLog.objects.filter(
            host_id__in=host_ids, state='PROBLEM', mulai__lt=now,
        ).filter(Q(selesai__gt=month_start) | Q(selesai__isnull=True))
    )

    def group_by_host(logs):
        d = {}
        for log in logs:
            d.setdefault(log.host_id, []).append(log)
        return d

    today_by_host = group_by_host(down_today)
    month_by_host = group_by_host(down_month)

    total_menit_today = max(1, int((now - today_start).total_seconds() / 60))
    total_menit_month = max(1, int((now - month_start).total_seconds() / 60))

    host_data = []
    total_ok = total_problem = total_unknown = 0
    avail_hari_list = []

    for h in hosts:
        if h.state == 'OK':
            total_ok += 1
        elif h.state == 'PROBLEM':
            total_problem += 1
        else:
            total_unknown += 1

        avail_hari = _calc_avail(today_by_host.get(h.pk, []), today_start, now, total_menit_today)
        avail_bulan = _calc_avail(month_by_host.get(h.pk, []), month_start, now, total_menit_month)
        avail_hari_list.append(avail_hari)

        host_data.append({
            'id': encode(h.pk),
            'nama': h.nama,
            'lokasi': h.lokasi,
            'device': h.device.nama if h.device_id else None,
            'state': h.state,
            'severity': h.severity,
            'problem_name': h.problem_name,
            'state_sejak': h.state_sejak.isoformat() if h.state_sejak else None,
            'durasi_menit': h.durasi_menit,
            'avail_hari': avail_hari,
            'avail_bulan': avail_bulan,
            'last_synced_at': h.last_synced_at.isoformat() if h.last_synced_at else None,
        })

    avg_avail_hari = round(sum(avail_hari_list) / len(avail_hari_list), 2) if avail_hari_list else None
    avg_avail_bulan = _calc_avail(down_month, month_start, now, total_menit_month * len(hosts)) if hosts else None

    gangguan = (ZabbixEventLog.objects.filter(state='PROBLEM', host_id__in=host_ids)
                .select_related('host').order_by('-mulai')[:10])
    gangguan_data = [{
        'host': g.host.nama,
        'severity': g.severity,
        'problem_name': g.problem_name,
        'mulai': g.mulai.isoformat(),
        'selesai': g.selesai.isoformat() if g.selesai else None,
        'durasi_menit': g.durasi_menit,
    } for g in gangguan]

    return JsonResponse({
        'total_ok': total_ok,
        'total_problem': total_problem,
        'total_unknown': total_unknown,
        'total_host': len(host_data),
        'avail_hari': avg_avail_hari,
        'avail_bulan': avg_avail_bulan,
        'hosts': host_data,
        'gangguan_terkini': gangguan_data,
    })


@login_required
def zbx_host_detail(request, pk):
    host = get_object_or_404(ZabbixHost, pk=pk)
    return render(request, 'device_mon/zbx_host_detail.html', {'host': host})


@login_required
def zbx_api_host_logs(request, pk):
    host = get_object_or_404(ZabbixHost, pk=pk)

    try:
        hari = min(max(int(request.GET.get('hari', 7)), 1), 30)
    except (ValueError, TypeError):
        hari = 7

    now, today_start, _ = _boundaries()
    tz_local = timezone.get_current_timezone()
    since = now - datetime.timedelta(days=hari)

    logs = ZabbixEventLog.objects.filter(host=host, mulai__gte=since).order_by('mulai')
    total_menit = max(1, int((now - since).total_seconds() / 60))
    problem_logs = [l for l in logs if l.state == 'PROBLEM']
    avail_periode = _calc_avail(problem_logs, since, now, total_menit)

    down_today = [l for l in ZabbixEventLog.objects.filter(
        host=host, state='PROBLEM', mulai__lt=now,
    ).filter(Q(selesai__gt=today_start) | Q(selesai__isnull=True))]
    avail_hari = _calc_avail(down_today, today_start, now,
                             max(1, int((now - today_start).total_seconds() / 60)))

    log_data = []
    for l in logs:
        mulai_local = l.mulai.astimezone(tz_local)
        selesai_local = l.selesai.astimezone(tz_local) if l.selesai else None
        dur = l.durasi_menit
        if dur is None and not l.selesai:
            dur = max(0, int((now - l.mulai).total_seconds() / 60))
        log_data.append({
            'state': l.state,
            'severity': l.severity,
            'problem_name': l.problem_name,
            'source': l.source,
            'mulai': mulai_local.strftime('%d/%m %H:%M'),
            'selesai': selesai_local.strftime('%d/%m %H:%M') if selesai_local else None,
            'durasi_menit': dur,
        })

    return JsonResponse({
        'nama': host.nama,
        'lokasi': host.lokasi,
        'state': host.state,
        'severity': host.severity,
        'problem_name': host.problem_name,
        'state_sejak': host.state_sejak.astimezone(tz_local).strftime('%d/%m/%Y %H:%M') if host.state_sejak else None,
        'durasi_menit': host.durasi_menit,
        'avail_hari': avail_hari,
        'avail_periode': avail_periode,
        'hari': hari,
        'logs': log_data,
    })


@login_required
def zbx_gangguan_list(request):
    logs = (ZabbixEventLog.objects.filter(state='PROBLEM')
            .select_related('host').order_by('-mulai')[:200])
    return render(request, 'device_mon/zbx_gangguan.html', {'logs': logs})


def _check_zbx_webhook_token(request):
    from django.conf import settings
    expected = getattr(settings, 'ZABBIX_WEBHOOK_TOKEN', '')
    if not expected:
        return False, 'ZABBIX_WEBHOOK_TOKEN belum dikonfigurasi di server.'
    got = request.headers.get('X-Zabbix-Webhook-Token') or request.GET.get('token', '')
    import hmac
    if not got or not hmac.compare_digest(str(got), str(expected)):
        return False, 'Token webhook tidak valid.'
    return True, ''


@csrf_exempt
@require_http_methods(["POST"])
def zbx_webhook_receiver(request):
    """
    Terima push realtime dari Zabbix Action (media type Webhook) saat trigger
    naik (PROBLEM) atau pulih (OK/RESOLVED). Lihat deploy/ZABBIX_INTEGRATION.md
    untuk skrip webhook Zabbix-nya dan daftar macro yang harus diisikan.

    Auth: header X-Zabbix-Webhook-Token: <ZABBIX_WEBHOOK_TOKEN>
          (atau ?token=... kalau Zabbix versi lama tidak bisa set header custom)

    Body JSON:
      {
        "event_status": "PROBLEM" | "OK" | "RESOLVED",
        "eventid": "12345",
        "hostid": "10084",
        "host": "nama-teknis",
        "host_visible_name": "Nama Tampilan",
        "severity": "High",
        "problem_name": "Free disk space is low",
        "event_time": "2026-08-18T10:00:00+08:00"   # opsional, ISO8601
      }
    """
    ok_token, msg = _check_zbx_webhook_token(request)
    raw_body = request.body.decode('utf-8', errors='replace')[:4000]

    if not ok_token:
        ZabbixWebhookLog.objects.create(ok=False, keterangan=msg, payload=raw_body)
        return HttpResponseForbidden(msg)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        ZabbixWebhookLog.objects.create(ok=False, keterangan='Body bukan JSON valid.', payload=raw_body)
        return JsonResponse({'status': 'error', 'message': 'Body bukan JSON valid.'}, status=400)

    hostid = str(data.get('hostid', '')).strip()
    if not hostid:
        ZabbixWebhookLog.objects.create(ok=False, keterangan='Field hostid kosong.', payload=raw_body)
        return JsonResponse({'status': 'error', 'message': "Field 'hostid' wajib diisi."}, status=400)

    raw_status = str(data.get('event_status', '')).strip().upper()
    state = 'OK' if raw_status in ('OK', 'RESOLVED') else 'PROBLEM' if raw_status == 'PROBLEM' else None
    if state is None:
        ZabbixWebhookLog.objects.create(
            ok=False, keterangan=f'event_status tidak dikenali: {raw_status!r}', payload=raw_body,
        )
        return JsonResponse({'status': 'error', 'message': 'event_status harus PROBLEM/OK/RESOLVED.'}, status=400)

    host_name = (data.get('host_visible_name') or data.get('host') or hostid).strip()
    eventid = str(data.get('eventid', '')).strip()

    now = timezone.now()
    event_time_raw = data.get('event_time')
    event_time = now
    if event_time_raw:
        try:
            event_time = datetime.datetime.fromisoformat(str(event_time_raw).replace('Z', '+00:00'))
            if timezone.is_naive(event_time):
                event_time = timezone.make_aware(event_time)
        except ValueError:
            event_time = now

    host, created = ZabbixHost.objects.get_or_create(
        zabbix_hostid=hostid,
        defaults={'zabbix_host': data.get('host', '') or host_name, 'nama': host_name},
    )

    # Idempotensi: kalau eventid ini sudah pernah dicatat untuk state yang sama, jangan dobel.
    if eventid and ZabbixEventLog.objects.filter(
        host=host, zabbix_eventid=eventid, state=state,
    ).exists():
        ZabbixWebhookLog.objects.create(
            ok=True, host=host, keterangan='Duplikat eventid, diabaikan.', payload=raw_body,
        )
        return JsonResponse({'status': 'ok', 'message': 'Duplikat, diabaikan.'})

    prev_state = host.state
    dur_tutup = None
    if prev_state != state or created:
        open_log = ZabbixEventLog.objects.filter(host=host, selesai__isnull=True).first()
        if open_log:
            selesai = event_time if state == 'OK' else now
            dur = max(0, int((selesai - open_log.mulai).total_seconds() / 60))
            open_log.selesai = selesai
            open_log.durasi_menit = dur
            open_log.save(update_fields=['selesai', 'durasi_menit'])
            dur_tutup = dur

        ZabbixEventLog.objects.create(
            host=host, state=state,
            severity=data.get('severity', '') or '',
            problem_name=data.get('problem_name', '') or '',
            zabbix_eventid=eventid, source='webhook',
            mulai=event_time if state == 'PROBLEM' else now,
        )

        host.state = state
        host.severity = data.get('severity', '') or '' if state == 'PROBLEM' else ''
        host.problem_name = data.get('problem_name', '') or '' if state == 'PROBLEM' else ''
        host.state_sejak = event_time if state == 'PROBLEM' else now
        host.last_synced_at = now
        host.save(update_fields=['state', 'severity', 'problem_name', 'state_sejak', 'last_synced_at'])

        if not created and host.aktif:
            notif_zabbix_transisi(host, state, dur_tutup_menit=dur_tutup)
    else:
        host.last_synced_at = now
        host.save(update_fields=['last_synced_at'])

    ZabbixWebhookLog.objects.create(ok=True, host=host, keterangan=f'{prev_state} -> {state}', payload=raw_body)
    return JsonResponse({'status': 'ok', 'host': host.nama, 'state': host.state})
