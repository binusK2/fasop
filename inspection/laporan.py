"""Skema kolom laporan inspeksi + pembangun workbook Excel harian.

Ini satu-satunya tempat yang mendefinisikan **kolom apa yang berlaku untuk
jenis peralatan apa**. Dipakai bersama oleh:

* halaman "Hasil Inspeksi Harian" (tab per jenis peralatan),
* export Excel harian — tombol di halaman itu maupun cron jam 12.00
  (`manage.py export_inspeksi_harian`),
* export rekap bulanan per ULTG.

Kalau kolomnya hidup di masing-masing view, Excel dan layar cepat berbeda —
itu yang dulu membuat sheet DFR dan Server ADS keluar semua strip: header-nya
memakai kolom rele (kebersihan_panel/kondisi_relay/sumber_dc) yang memang tidak
ada di kedua model itu, jadi setiap baris jatuh ke blok except.
"""
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from devices.models import Device


# ── Jenis perangkat yang bisa diinspeksi ────────────────────────────
# Key HARUS persis sama dengan DeviceType.name di database (case-sensitive)
INSPECTABLE_JENIS = {
    'Catu Daya':           'catu_daya',
    'RELE DEFENSE SCHEME': 'defense_scheme',
    'MASTER TRIP':         'master_trip',
    'UFLS':                'ufls',
    'DFR':                 'dfr',
    'SERVER PROSIS':       'server_ads',
    'Radio':               'telecom',
    'VoIP':                'telecom',
}

# Jenis khusus Dispatcher
TELECOM_JENIS = {'Radio', 'VoIP'}


def perangkat_operasi(qs):
    """Buang perangkat berstatus 'Tidak Operasi' dari queryset Device.

    Perangkat yang sudah tidak beroperasi tidak boleh muncul di halaman
    inspeksi operator, dan tidak ikut jadi pembagi progres inspeksi
    (kalau ikut dihitung, ia selamanya berstatus "belum diinspeksi").
    """
    return qs.exclude(status_operasi='tidak_operasi')


def is_alarm_inspection(insp):
    """True bila hasil satu inspeksi menunjukkan kondisi alarm/tidak normal,
    berdasarkan field yang benar-benar diisi lewat form.html untuk masing-masing jenis."""
    try:
        if insp.jenis == 'catu_daya':
            d = insp.detail_catu_daya
            return (d.kondisi_rectifier == 'alarm'
                    or d.alarm_ground_fault == 'ada'
                    or d.alarm_min_ac_fault == 'ada'
                    or d.alarm_recti_fault == 'ada'
                    or d.level_air_bank in ('bawah_level', 'atas_level')
                    or d.exhaust_fan == 'mati'
                    or d.kondisi_baterai == 'kotor'
                    or d.kondisi_keseluruhan == 'kotor')
        if insp.jenis == 'defense_scheme':
            d = insp.detail_defense_scheme
            return d.kondisi_relay == 'alarm' or d.status_indikator == 'tidak_normal'
        if insp.jenis in ('master_trip', 'ufls'):
            d = insp.detail_master_trip if insp.jenis == 'master_trip' else insp.detail_ufls
            return d.kondisi_relay == 'alarm' or d.indikator_led == 'tidak_normal'
        if insp.jenis == 'dfr':
            d = insp.detail_dfr
            return (d.kondisi_dfr == 'faulty' or d.healthy_status in ('faulty', 'alarm')
                    or d.indikasi_led_alarm == 'ada' or d.status_indikator == 'tidak_normal')
        if insp.jenis == 'server_ads':
            d = insp.detail_server_ads
            return (d.peralatan_server_ads == 'tidak_normal' or d.tampilan_hmi == 'tidak_normal'
                    or d.peralatan_gateway_ic3 == 'tidak_normal' or d.kondisi_switch_lan == 'mati'
                    or d.peralatan_power_supply == 'tidak_normal' or d.fan_panel == 'mati')
        if insp.jenis == 'telecom':
            return insp.detail_telecom.hasil_komunikasi == 'tidak_normal'
    except Exception:
        pass
    return False


# ─────────────────────────────────────────────────────────────────────
# SKEMA KOLOM PER JENIS PERALATAN
# ─────────────────────────────────────────────────────────────────────
@dataclass(frozen=True)
class Kolom:
    """Satu kolom hasil inspeksi.

    field  — nama field di model detail (InspectionCatuDaya, InspectionDFR, …)
    ok     — nilai yang dianggap normal; kosong = kolom informatif saja
    alarm  — bila nilainya bukan `ok`: True → merah (ikut hitungan alarm),
             False → kuning (perlu perhatian, bukan alarm)
    satuan — ditempel di belakang angka (V, A, °C, %)
    """
    field:  str
    label:  str
    lebar:  int = 16
    ok:     str = ''
    alarm:  bool = False
    satuan: str = ''


_K = Kolom

KOLOM_JENIS = {
    'catu_daya': (
        _K('kondisi_rectifier',       'Kondisi Rectifier',      17, ok='normal',     alarm=True),
        _K('mode_recti',              'Mode Rectifier',         14),
        _K('alarm_ground_fault',      'Alarm Ground Fault',     17, ok='tidak_ada',  alarm=True),
        _K('alarm_min_ac_fault',      'Alarm Min AC Fault',     17, ok='tidak_ada',  alarm=True),
        _K('alarm_recti_fault',       'Alarm Recti Fault',      17, ok='tidak_ada',  alarm=True),
        _K('kebersihan_ruangan',      'Kebersihan Ruangan',     17, ok='bersih'),
        _K('kondisi_baterai',         'Kondisi Baterai',        15, ok='bersih',     alarm=True),
        _K('level_air_bank',          'Level Air Bank',         16, ok='normal',     alarm=True),
        _K('kebersihan_ruangan_bank', 'Kebersihan Ruang Bank',  19, ok='bersih'),
        _K('kebersihan_bank',         'Kebersihan Bank',        15, ok='bersih'),
        _K('exhaust_fan',             'Exhaust Fan',            13, ok='nyala',      alarm=True),
        _K('tegangan_input_ac',       'Teg. Input AC',          14, satuan='V'),
        _K('arus_input_ac',           'Arus Input AC',          14, satuan='A'),
        _K('tegangan_load_dc',        'Teg. Load DC',           14, satuan='V'),
        _K('arus_load_dc',            'Arus Load DC',           14, satuan='A'),
        _K('tegangan_baterai_dc',     'Teg. Baterai DC',        15, satuan='V'),
        _K('arus_baterai_dc',         'Arus Baterai DC',        15, satuan='A'),
        _K('kondisi_keseluruhan',     'Kondisi Keseluruhan',    18, ok='bersih',     alarm=True),
        _K('catatan_rectifier',       'Catatan Rectifier',      28),
        _K('catatan_baterai',         'Catatan Baterai',        28),
    ),
    'defense_scheme': (
        _K('suhu_ruangan',      'Suhu Ruangan',       14, satuan='°C'),
        _K('kelembapan',        'Kelembapan',         13, satuan='%'),
        _K('kebersihan_panel',  'Kebersihan Panel',   16, ok='bersih'),
        _K('lampu_panel',       'Lampu Panel',        13, ok='nyala',  alarm=True),
        _K('kondisi_relay',     'Kondisi Rele',       14, ok='normal', alarm=True),
        _K('status_indikator',  'Status Indikator',   16, ok='normal', alarm=True),
        _K('selektor_blok_skema', 'Selektor Blok Skema', 18, ok='on'),
        _K('posisi_selektor',   'Posisi Selektor',    16, ok='on'),
        _K('kondisi_kabel_lan', 'Kondisi Kabel LAN',  17, ok='normal', alarm=True),
        _K('sumber_dc',         'Sumber DC',          13, satuan='V'),
        _K('catatan_relay',     'Catatan Rele',       28),
    ),
    'master_trip': (
        _K('suhu_ruangan',      'Suhu Ruangan',       14, satuan='°C'),
        _K('kelembapan',        'Kelembapan',         13, satuan='%'),
        _K('kebersihan_panel',  'Kebersihan Panel',   16, ok='bersih'),
        _K('lampu_panel',       'Lampu Panel',        13, ok='nyala',  alarm=True),
        _K('kondisi_relay',     'Kondisi Rele',       14, ok='normal', alarm=True),
        _K('indikator_led',     'Status Indikator',   16, ok='normal', alarm=True),
        _K('posisi_selektor',   'Posisi Selektor',    16, ok='on'),
        _K('kondisi_kabel_lan', 'Kondisi Kabel LAN',  17, ok='normal', alarm=True),
        _K('sumber_dc',         'Sumber DC',          13, satuan='V'),
        _K('catatan_relay',     'Catatan Rele',       28),
    ),
    # UFLS memakai pilihan yang berbeda untuk selektor & kabel LAN
    # (on_aktif / terpasang), jadi nilai `ok`-nya tidak boleh disamakan
    # dengan Master Trip meski labelnya mirip.
    'ufls': (
        _K('suhu_ruangan',      'Suhu Ruangan',       14, satuan='°C'),
        _K('kelembapan',        'Kelembapan',         13, satuan='%'),
        _K('kebersihan_panel',  'Kebersihan Panel',   16, ok='bersih'),
        _K('lampu_panel',       'Lampu Panel',        13, ok='nyala',     alarm=True),
        _K('kondisi_relay',     'Kondisi Rele',       14, ok='normal',    alarm=True),
        _K('indikator_led',     'Status Indikator',   16, ok='normal',    alarm=True),
        _K('posisi_selektor',   'Posisi Selektor',    16, ok='on_aktif'),
        _K('kondisi_kabel_lan', 'Kondisi Kabel LAN',  17, ok='terpasang', alarm=True),
        _K('sumber_dc',         'Sumber DC',          13, satuan='V'),
        _K('catatan_relay',     'Catatan Rele',       28),
    ),
    'dfr': (
        _K('suhu_ruangan',       'Suhu Ruangan',       14, satuan='°C'),
        _K('kelembapan',         'Kelembapan',         13, satuan='%'),
        _K('kebersihan_ruangan', 'Kebersihan',         14, ok='bersih'),
        _K('lampu_penerangan',   'Lampu Penerangan',   17, ok='baik',       alarm=True),
        _K('kondisi_dfr',        'Kondisi DFR',        14, ok='normal',     alarm=True),
        _K('healthy_status',     'Healthy Status',     15, ok='healthy',    alarm=True),
        _K('indikasi_led_alarm', 'Indikasi LED Alarm', 18, ok='tidak_ada',  alarm=True),
        _K('status_indikator',   'Status Indikator',   16, ok='normal',     alarm=True),
        _K('kondisi_kabel_lan',  'Kondisi Kabel LAN',  17, ok='normal',     alarm=True),
    ),
    'server_ads': (
        _K('suhu_ruangan',           'Suhu Ruangan',        14, satuan='°C'),
        _K('kelembapan',             'Kelembapan',          13, satuan='%'),
        _K('kebersihan_ruangan',     'Kebersihan',          14, ok='bersih'),
        _K('lampu_penerangan',       'Lampu Penerangan',    17, ok='baik',   alarm=True),
        _K('peralatan_server_ads',   'Peralatan Server ADS', 19, ok='normal', alarm=True),
        _K('tampilan_hmi',           'Tampilan HMI',        15, ok='normal', alarm=True),
        _K('peralatan_gateway_ic3',  'Gateway IC3 ADS',     17, ok='normal', alarm=True),
        _K('kondisi_switch_lan',     'Switch Kabel LAN',    17, ok='normal', alarm=True),
        _K('peralatan_power_supply', 'Power Supply',        15, ok='normal', alarm=True),
        _K('fan_panel',              'Fan Panel',           13, ok='nyala',  alarm=True),
    ),
    'telecom': (
        _K('hasil_komunikasi',  'Hasil Komunikasi', 17, ok='normal', alarm=True),
        _K('kualitas_suara',    'Kualitas Suara',   15, ok='baik'),
        _K('catatan_pengujian', 'Catatan Pengujian', 30),
    ),
}

# Nama relasi detail per jenis
DETAIL_ATTR = {
    'catu_daya':      'detail_catu_daya',
    'defense_scheme': 'detail_defense_scheme',
    'master_trip':    'detail_master_trip',
    'ufls':           'detail_ufls',
    'dfr':            'detail_dfr',
    'server_ads':     'detail_server_ads',
    'telecom':        'detail_telecom',
}

# Urutan tab di halaman harian = urutan sheet di Excel
JENIS_URUT = ('catu_daya', 'defense_scheme', 'master_trip', 'ufls',
              'dfr', 'server_ads', 'telecom')

JENIS_LABEL = {
    'catu_daya':      'Catu Daya',
    'defense_scheme': 'Rele Defense Scheme',
    'master_trip':    'Master Trip',
    'ufls':           'UFLS',
    'dfr':            'DFR',
    'server_ads':     'Server ADS',
    'telecom':        'Telekomunikasi',
}

JENIS_IKON = {
    'catu_daya':      'bi-battery-charging',
    'defense_scheme': 'bi-shield-shaded',
    'master_trip':    'bi-lightning-charge',
    'ufls':           'bi-graph-down-arrow',
    'dfr':            'bi-activity',
    'server_ads':     'bi-hdd-rack',
    'telecom':        'bi-broadcast',
}

JENIS_WARNA = {
    'catu_daya':      'EA580C',
    'defense_scheme': '7C3AED',
    'master_trip':    '2563EB',
    'ufls':           '059669',
    'dfr':            'DB2777',
    'server_ads':     '0891B2',
    'telecom':        '65A30D',
}


def device_types(jenis_key):
    """Nama DeviceType yang termasuk satu jenis inspeksi (telecom = Radio + VoIP)."""
    return [nama for nama, key in INSPECTABLE_JENIS.items() if key == jenis_key]


def kolom_jenis(jenis_key):
    return KOLOM_JENIS.get(jenis_key, ())


# ─────────────────────────────────────────────────────────────────────
# PENGISIAN NILAI
# ─────────────────────────────────────────────────────────────────────
def nilai_kolom(detail, kol):
    """Nilai satu kolom, siap dipakai template maupun openpyxl.

    status: 'kosong' | 'data' | 'normal' | 'perhatian' | 'alarm'
    raw   — angka asli untuk sel Excel (biar bisa dihitung), teks untuk pilihan.
    """
    raw = getattr(detail, kol.field, None) if detail is not None else None
    if raw is None or raw == '':
        return {'label': kol.label, 'raw': None, 'display': '—', 'status': 'kosong'}

    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        teks = f'{raw:g}' + (f' {kol.satuan}' if kol.satuan else '')
        return {'label': kol.label, 'raw': raw, 'display': teks, 'status': 'data'}

    getter = getattr(detail, f'get_{kol.field}_display', None)
    teks = getter() if callable(getter) else str(raw)
    if not kol.ok:
        status = 'data'
    elif raw == kol.ok:
        status = 'normal'
    else:
        status = 'alarm' if kol.alarm else 'perhatian'
    return {'label': kol.label, 'raw': teks, 'display': teks, 'status': status}


def _nama_operator(user):
    if not user:
        return '—'
    return user.get_full_name() or user.username


def baris_perangkat(dev, insp, jenis_key):
    """Satu baris hasil inspeksi satu perangkat (insp boleh None = belum diinspeksi)."""
    kolom = kolom_jenis(jenis_key)
    detail = getattr(insp, DETAIL_ATTR.get(jenis_key, ''), None) if insp else None
    sel = [nilai_kolom(detail, k) for k in kolom]

    if insp is None:
        status, status_label = 'belum', 'Belum Diinspeksi'
    elif is_alarm_inspection(insp):
        status, status_label = 'alarm', 'Alarm'
    elif insp.is_flagged:
        status, status_label = 'flag', 'Diflag'
    else:
        status, status_label = 'normal', 'Normal'

    return {
        'device':       dev,
        'nama':         dev.nama if dev else (insp.device.nama if insp else '—'),
        'lokasi':       (dev.lokasi if dev else insp.device.lokasi) or '—',
        'insp':         insp,
        'jam':          insp.tanggal.strftime('%H:%M') if insp else '—',
        'waktu':        insp.tanggal.strftime('%d/%m/%Y %H:%M') if insp else '—',
        'operator':     _nama_operator(insp.operator) if insp else '—',
        'catatan':      (insp.catatan if insp else '') or '',
        'sel':          sel,
        'status':       status,
        'status_label': status_label,
    }


def baris_telecom_pengujian(item, pengujian):
    """Baris dari batch Pengujian Telekomunikasi (dispatcher) — sumber kedua
    hasil telekomunikasi selain Inspection(jenis='telecom')."""
    tidak_normal = item.hasil == 'tidak_normal'
    sel = [
        {'label': 'Hasil Komunikasi', 'raw': item.get_hasil_display(),
         'display': item.get_hasil_display(),
         'status': 'alarm' if tidak_normal else 'normal'},
        {'label': 'Kualitas Suara', 'raw': None, 'display': '—', 'status': 'kosong'},
        {'label': 'Catatan Pengujian', 'raw': item.catatan or None,
         'display': item.catatan or '—',
         'status': 'data' if item.catatan else 'kosong'},
    ]
    return {
        'device':       item.device,
        'nama':         item.device.nama,
        'lokasi':       item.device.lokasi or '—',
        'insp':         None,
        'jam':          pengujian.created_at.strftime('%H:%M') if pengujian.created_at else '—',
        'waktu':        pengujian.tanggal.strftime('%d/%m/%Y') if pengujian.tanggal else '—',
        'operator':     _nama_operator(pengujian.dibuat_oleh),
        'catatan':      pengujian.catatan or '',
        'sel':          sel,
        'status':       'alarm' if tidak_normal else 'normal',
        'status_label': 'Alarm' if tidak_normal else 'Normal',
        'sumber':       'Pengujian Telekomunikasi',
    }


# ─────────────────────────────────────────────────────────────────────
# PENGAMBILAN DATA HARIAN
# ─────────────────────────────────────────────────────────────────────
def perangkat_jenis(jenis_key, lokasi_names=None):
    """Perangkat yang beroperasi untuk satu jenis inspeksi, urut lokasi lalu nama."""
    devs = (
        perangkat_operasi(Device.objects)
        .filter(is_deleted=False, jenis__name__in=device_types(jenis_key))
        .select_related('jenis')
        .order_by('lokasi', 'nama')
    )
    if lokasi_names is not None:
        devs = devs.filter(lokasi__in=lokasi_names)
    return devs


def baris_periode(jenis_key, filter_inspeksi, lokasi_names=None, sertakan_belum=True):
    """Baris hasil inspeksi satu jenis peralatan untuk rentang waktu apa pun.

    filter_inspeksi — kwargs untuk Inspection.objects.filter, mis.
    {'tanggal__date': tgl} (harian) atau {'tanggal__year': y, 'tanggal__month': m}.
    Kalau satu perangkat punya beberapa inspeksi di rentang itu, yang dipakai
    yang terbaru. Perangkat yang belum diinspeksi ikut dikembalikan (status
    'belum') supaya laporan menunjukkan cakupan, bukan cuma yang terisi.
    """
    from .models import Inspection  # hindari import melingkar

    devs = perangkat_jenis(jenis_key, lokasi_names)

    insp_map = {}
    for insp in (Inspection.objects
                 .filter(device__in=devs, jenis=jenis_key, **filter_inspeksi)
                 .select_related('device', 'operator')
                 .order_by('device_id', '-tanggal')):
        insp_map.setdefault(insp.device_id, insp)

    baris = []
    for dev in devs:
        insp = insp_map.get(dev.pk)
        if insp is None and not sertakan_belum:
            continue
        baris.append(baris_perangkat(dev, insp, jenis_key))
    return baris


def baris_bulanan(jenis_key, year, month, lokasi_names=None):
    return baris_periode(jenis_key, {'tanggal__year': year, 'tanggal__month': month},
                         lokasi_names)


def baris_harian(tanggal, jenis_key, lokasi_names=None, sertakan_belum=True):
    """Baris hasil inspeksi satu jenis peralatan pada satu tanggal.

    Khusus telekomunikasi, hasil batch Pengujian Telekomunikasi (dispatcher)
    ikut digabung — dua jalur itu mengisi tabel berbeda, dan laporan harian
    tidak boleh menampilkan Radio/VoIP sebagai "belum diinspeksi" hanya karena
    dispatcher memakai form batch.
    """
    from .models import PengujianTelecomItem  # hindari import melingkar

    baris = baris_periode(jenis_key, {'tanggal__date': tanggal},
                          lokasi_names, sertakan_belum)

    if jenis_key == 'telecom':
        devs = perangkat_jenis(jenis_key, lokasi_names)
        dev_ids = {d.pk for d in devs}
        extra = (
            PengujianTelecomItem.objects
            .filter(pengujian__tanggal=tanggal, device_id__in=dev_ids)
            .select_related('device', 'pengujian', 'pengujian__dibuat_oleh')
            .order_by('device__lokasi', 'device__nama')
        )
        sudah = {b['device'].pk for b in baris if b['insp'] is not None}
        tambahan = []
        for item in extra:
            if item.device_id in sudah:
                continue
            tambahan.append(baris_telecom_pengujian(item, item.pengujian))
            sudah.add(item.device_id)
        if tambahan:
            # Baris pengujian menggantikan baris "belum" perangkat yang sama
            id_tambahan = {b['device'].pk for b in tambahan}
            baris = [b for b in baris
                     if not (b['insp'] is None and b['device'].pk in id_tambahan)]
            baris.extend(tambahan)
            baris.sort(key=lambda b: (b['lokasi'], b['nama']))

    return baris


def ringkasan_baris(baris):
    """Hitungan status untuk satu kumpulan baris."""
    total = len(baris)
    belum = sum(1 for b in baris if b['status'] == 'belum')
    alarm = sum(1 for b in baris if b['status'] == 'alarm')
    flag  = sum(1 for b in baris if b['status'] == 'flag')
    sudah = total - belum
    return {
        'total':  total,
        'sudah':  sudah,
        'belum':  belum,
        'alarm':  alarm,
        'flag':   flag,
        'normal': sudah - alarm - flag,
        'pct':    round(sudah / total * 100) if total else 0,
    }


# ─────────────────────────────────────────────────────────────────────
# WORKBOOK HARIAN
# ─────────────────────────────────────────────────────────────────────
_THIN = Side(style='thin', color='CBD5E1')
_BRD  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HDR_FONT = Font(bold=True, color='FFFFFF', size=9)
_HDR_ALN  = Alignment(horizontal='center', vertical='center', wrap_text=True)
_C_ALN    = Alignment(horizontal='center', vertical='center')
_L_ALN    = Alignment(horizontal='left', vertical='center')

_FILL_STATUS = {
    'alarm':     ('FEE2E2', '991B1B'),
    'perhatian': ('FEF9C3', '854D0E'),
    'flag':      ('FEF3C7', '92400E'),
    'belum':     ('F1F5F9', '64748B'),
    'normal':    ('DCFCE7', '166534'),
}

_KOLOM_IDENTITAS = (
    ('No',             5),
    ('Nama Perangkat', 30),
    ('Lokasi / GI',    22),
    ('Jam',            8),
    ('Operator',       18),
)


def tulis_sheet_jenis(wb, jenis_key, baris, judul, pakai_tanggal=False):
    """Satu sheet Excel untuk satu jenis peralatan, kolomnya dari KOLOM_JENIS.

    pakai_tanggal — True untuk rekap bulanan (kolom waktu berisi tanggal+jam),
    False untuk laporan harian (cukup jam).
    """
    warna   = JENIS_WARNA.get(jenis_key, '334155')
    kolom   = kolom_jenis(jenis_key)
    ws      = wb.create_sheet(JENIS_LABEL[jenis_key][:31])
    ws.sheet_properties.tabColor = warna

    ident   = list(_KOLOM_IDENTITAS)
    if pakai_tanggal:
        ident[3] = ('Tgl Inspeksi', 17)
    headers = [h for h, _ in ident] + [k.label for k in kolom] + ['Catatan', 'Status']
    lebar   = [w for _, w in ident] + [k.lebar for k in kolom] + [30, 16]
    last_col = get_column_letter(len(headers) + 1)

    r = ringkasan_baris(baris)
    ws.merge_cells(f'B1:{last_col}1')
    ws['B1'].value     = judul
    ws['B1'].font      = Font(bold=True, size=12)
    ws['B1'].alignment = _C_ALN
    ws['B1'].fill      = PatternFill('solid', fgColor='EFF6FF')
    ws.row_dimensions[1].height = 24

    ws.merge_cells(f'B2:{last_col}2')
    ws['B2'].value = (f'{r["sudah"]} dari {r["total"]} perangkat terinspeksi '
                      f'({r["pct"]}%) — alarm {r["alarm"]}, diflag {r["flag"]}, '
                      f'belum {r["belum"]}')
    ws['B2'].font      = Font(size=9, italic=True, color='64748B')
    ws['B2'].alignment = _C_ALN

    ws.column_dimensions['A'].width = 2
    hdr_fill = PatternFill('solid', fgColor=warna)
    for ci, (h, w) in enumerate(zip(headers, lebar), 2):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font      = _HDR_FONT
        cell.fill      = hdr_fill
        cell.alignment = _HDR_ALN
        cell.border    = _BRD
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[4].height = 30
    ws.freeze_panes = 'D5'

    for i, b in enumerate(baris, 1):
        wr = i + 4
        waktu = b['waktu'] if pakai_tanggal else b['jam']
        nilai = ([i, b['nama'], b['lokasi'], waktu, b['operator']]
                 + [s['raw'] if s['raw'] is not None else '—' for s in b['sel']]
                 + [b['catatan'] or '—', b['status_label']])
        status_sel = [None] * len(ident) + [s['status'] for s in b['sel']] + [None, b['status']]

        for ci, (val, st) in enumerate(zip(nilai, status_sel), 2):
            cell = ws.cell(row=wr, column=ci, value=val)
            cell.border    = _BRD
            cell.alignment = _C_ALN if ci in (2, 5) else _L_ALN
            if st in _FILL_STATUS and st != 'normal':
                bg, fg = _FILL_STATUS[st]
                cell.fill = PatternFill('solid', fgColor=bg)
                cell.font = Font(bold=st in ('alarm', 'flag'), color=fg, size=9)
            elif st == 'normal' and ci == len(headers) + 1:
                bg, fg = _FILL_STATUS['normal']
                cell.fill = PatternFill('solid', fgColor=bg)
                cell.font = Font(bold=True, color=fg, size=9)
            elif i % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F8FAFC')
        ws.row_dimensions[wr].height = 17

    if not baris:
        ws.cell(row=5, column=2, value='Tidak ada perangkat jenis ini.').font = Font(
            italic=True, color='94A3B8', size=9)
    return ws


def _tulis_sheet_ringkasan(wb, data, tanggal, judul_lokasi):
    ws = wb.active
    ws.title = 'Ringkasan'
    ws.sheet_properties.tabColor = '0F172A'

    ws.merge_cells('B1:H1')
    ws['B1'].value     = 'LAPORAN HASIL INSPEKSI HARIAN'
    ws['B1'].font      = Font(bold=True, size=14, color='0F172A')
    ws['B1'].alignment = _C_ALN
    ws.row_dimensions[1].height = 26

    ws.merge_cells('B2:H2')
    ws['B2'].value     = f'{tanggal.strftime("%A, %d %B %Y")} — {judul_lokasi}'
    ws['B2'].font      = Font(size=10, italic=True, color='64748B')
    ws['B2'].alignment = _C_ALN

    headers = ['Jenis Peralatan', 'Total Perangkat', 'Terinspeksi', 'Belum',
               'Normal', 'Alarm', 'Diflag', '% Terinspeksi']
    lebar   = [26, 16, 13, 10, 10, 10, 10, 15]
    ws.column_dimensions['A'].width = 2
    for ci, (h, w) in enumerate(zip(headers, lebar), 2):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font      = _HDR_FONT
        cell.fill      = PatternFill('solid', fgColor='0F172A')
        cell.alignment = _HDR_ALN
        cell.border    = _BRD
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[4].height = 24

    wr = 5
    tot = {'total': 0, 'sudah': 0, 'belum': 0, 'normal': 0, 'alarm': 0, 'flag': 0}
    for jenis_key, baris in data:
        r = ringkasan_baris(baris)
        for k in tot:
            tot[k] += r[k]
        vals = [JENIS_LABEL[jenis_key], r['total'], r['sudah'], r['belum'],
                r['normal'], r['alarm'], r['flag'], f'{r["pct"]}%']
        for ci, val in enumerate(vals, 2):
            cell = ws.cell(row=wr, column=ci, value=val)
            cell.border    = _BRD
            cell.alignment = _L_ALN if ci == 2 else _C_ALN
            if ci == 7 and r['alarm']:
                cell.fill = PatternFill('solid', fgColor='FEE2E2')
                cell.font = Font(bold=True, color='991B1B', size=9)
        wr += 1

    pct_tot = round(tot['sudah'] / tot['total'] * 100) if tot['total'] else 0
    vals = ['TOTAL', tot['total'], tot['sudah'], tot['belum'],
            tot['normal'], tot['alarm'], tot['flag'], f'{pct_tot}%']
    for ci, val in enumerate(vals, 2):
        cell = ws.cell(row=wr, column=ci, value=val)
        cell.border    = _BRD
        cell.alignment = _L_ALN if ci == 2 else _C_ALN
        cell.font      = Font(bold=True, size=10)
        cell.fill      = PatternFill('solid', fgColor='E2E8F0')
    ws.freeze_panes = 'B5'
    return ws


def workbook_harian(tanggal, lokasi_names=None, judul_lokasi='Seluruh UIP3B Sulawesi'):
    """Workbook laporan harian: sheet Ringkasan + satu sheet per jenis peralatan.

    Kolom tiap sheet mengikuti KOLOM_JENIS, jadi sheet DFR berisi kolom DFR dan
    sheet Server ADS berisi kolom Server ADS — bukan kolom rele untuk semuanya.
    """
    data = [(jk, baris_harian(tanggal, jk, lokasi_names)) for jk in JENIS_URUT]
    wb = Workbook()
    _tulis_sheet_ringkasan(wb, data, tanggal, judul_lokasi)
    for jenis_key, baris in data:
        judul = f'{JENIS_LABEL[jenis_key].upper()} — {tanggal.strftime("%d %B %Y")}'
        tulis_sheet_jenis(wb, jenis_key, baris, judul)
    return wb


def nama_file_harian(tanggal):
    return f'Inspeksi_Harian_{tanggal.strftime("%Y-%m-%d")}.xlsx'
