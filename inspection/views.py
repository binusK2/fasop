from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Count, Q
from .models import (Inspection, InspectionCatuDaya, InspectionDefenseScheme,
                     InspectionMasterTrip, InspectionUFLS, InspectionDFR,
                     InspectionServerADS, InspectionTelecom,
                     PengujianTelecom, PengujianTelecomItem)
from devices.models import Device, DeviceType, SiteLocation
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Jenis perangkat yang bisa diinspeksi ────────────────────────────
# Key HARUS persis sama dengan DeviceType.name di database (case-sensitive)
INSPECTABLE_JENIS = {
    'Catu Daya':           'catu_daya',
    'RELE DEFENSE SCHEME': 'defense_scheme',
    'MASTER TRIP':         'master_trip',
    'UFLS':                'ufls',
    'DFR':                 'dfr',
    'SERVER PROSIS':       'server_ads',
    'Radio':               'telecom',
    'VoIP':                'telecom',
}

# Jenis khusus Dispatcher
TELECOM_JENIS = {'Radio', 'VoIP'}


def _is_alarm_inspection(insp):
    """True bila hasil satu inspeksi menunjukkan kondisi alarm/tidak normal,
    berdasarkan field yang benar-benar diisi lewat form.html untuk masing-masing jenis."""
    try:
        if insp.jenis == 'catu_daya':
            return insp.detail_catu_daya.kondisi_rectifier == 'alarm'
        if insp.jenis == 'defense_scheme':
            d = insp.detail_defense_scheme
            return d.kondisi_relay == 'alarm' or d.status_indikator == 'tidak_normal'
        if insp.jenis in ('master_trip', 'ufls'):
            d = insp.detail_master_trip if insp.jenis == 'master_trip' else insp.detail_ufls
            return d.kondisi_relay == 'alarm' or d.indikator_led == 'tidak_normal'
        if insp.jenis == 'dfr':
            d = insp.detail_dfr
            return (d.kondisi_dfr == 'faulty' or d.healthy_status in ('faulty', 'alarm')
                    or d.indikasi_led_alarm == 'ada' or d.status_indikator == 'tidak_normal')
        if insp.jenis == 'server_ads':
            d = insp.detail_server_ads
            return (d.peralatan_server_ads == 'tidak_normal' or d.tampilan_hmi == 'tidak_normal'
                    or d.peralatan_gateway_ic3 == 'tidak_normal' or d.kondisi_switch_lan == 'mati'
                    or d.peralatan_power_supply == 'tidak_normal' or d.fan_panel == 'mati')
        if insp.jenis == 'telecom':
            return insp.detail_telecom.hasil_komunikasi == 'tidak_normal'
    except Exception:
        pass
    return False


def require_operator(view_func):
    """Decorator: hanya operator, AM, superuser yang boleh akses."""
    from functools import wraps
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            from django.contrib.auth.views import redirect_to_login
            return redirect_to_login(request.get_full_path())
        if request.user.is_superuser:
            return view_func(request, *args, **kwargs)
        try:
            role = request.user.profile.role
        except Exception:
            role = ''
        if role in ('operator', 'technician', 'asisten_manager', 'dispatcher'):
            return view_func(request, *args, **kwargs)
        from django.shortcuts import render
        return render(request, '403.html',
                      {'message': 'Halaman ini hanya untuk Operator / Dispatcher.'}, status=403)
    return wrapper


# ─────────────────────────────────────────────────────────────────────
# PILIH LOKASI
# ─────────────────────────────────────────────────────────────────────
@login_required
@require_operator
def inspection_lokasi(request):
    """Halaman utama — pilih lokasi GI untuk inspeksi."""
    jenis_names = list(INSPECTABLE_JENIS.keys())

    # ── Filter lokasi berdasarkan ULTG user (khusus role operator) ──
    ultg          = None
    ultg_lokasi   = None   # list nama lokasi jika dibatasi ULTG
    try:
        profile = request.user.profile
        if profile.role == 'operator' and profile.ultg:
            ultg        = profile.ultg
            ultg_lokasi = ultg.get_lokasi_names()
    except Exception:
        pass

    lokasi_qs = (
        Device.objects
        .filter(is_deleted=False, jenis__name__in=jenis_names)
        .exclude(lokasi__isnull=True).exclude(lokasi='')
        .values_list('lokasi', flat=True)
        .distinct().order_by('lokasi')
    )

    # Terapkan filter ULTG jika ada
    if ultg_lokasi is not None:
        lokasi_qs = [l for l in lokasi_qs if l in ultg_lokasi]

    today = date.today()
    lokasi_stats = []
    for lok in lokasi_qs:
        insp_today = Inspection.objects.filter(
            device__lokasi=lok, tanggal__date=today,
        ).count()
        total_device = Device.objects.filter(
            is_deleted=False, jenis__name__in=jenis_names, lokasi=lok
        ).count()
        lokasi_stats.append({
            'lokasi': lok,
            'today':  insp_today,
            'total':  total_device,
        })

    return render(request, 'inspection/lokasi.html', {
        'lokasi_stats': lokasi_stats,
        'today':        today,
        'ultg':         ultg,
    })


# ─────────────────────────────────────────────────────────────────────
# DAFTAR PERANGKAT DI LOKASI
# ─────────────────────────────────────────────────────────────────────
@login_required
@require_operator
def inspection_device_list(request, lokasi):
    """Daftar perangkat inspectable di lokasi tertentu."""
    try:
        role = request.user.profile.role
    except Exception:
        role = ''

    # Dispatcher hanya lihat Radio & VoIP; role lain hanya lihat non-telecom
    if request.user.is_superuser or role == 'asisten_manager':
        jenis_names = list(INSPECTABLE_JENIS.keys())
    elif role == 'dispatcher':
        jenis_names = list(TELECOM_JENIS)
    else:
        jenis_names = [j for j in INSPECTABLE_JENIS if j not in TELECOM_JENIS]

    devices = (
        Device.objects
        .filter(is_deleted=False, jenis__name__in=jenis_names, lokasi=lokasi)
        .select_related('jenis')
        .order_by('jenis__name', 'nama')
    )

    # Build list of dict agar template tidak perlu filter gymnastics
    today = timezone.localdate()
    device_items = []
    for d in devices:
        last_insp = (
            Inspection.objects
            .filter(device=d)
            .select_related('operator')
            .order_by('-tanggal')
            .first()
        )
        # Inspeksi cukup 1x per hari — kalau inspeksi terakhir sudah hari ini,
        # tombol Inspeksi disembunyikan sampai berganti hari.
        sudah_hari_ini = bool(
            last_insp and timezone.localtime(last_insp.tanggal).date() == today
        )
        device_items.append({
            'device':         d,
            'last_insp':      last_insp,
            'sudah_hari_ini': sudah_hari_ini,
        })

    return render(request, 'inspection/device_list.html', {
        'lokasi':       lokasi,
        'devices':      devices,
        'device_items': device_items,
    })


# ─────────────────────────────────────────────────────────────────────
# FORM INSPEKSI
# ─────────────────────────────────────────────────────────────────────
@login_required
@require_operator
def inspection_form(request, device_pk):
    """Form input inspeksi inservice."""
    device = get_object_or_404(Device, pk=device_pk, is_deleted=False)
    jenis_name = device.jenis.name.strip() if device.jenis else ''
    jenis_key  = INSPECTABLE_JENIS.get(jenis_name)

    if not jenis_key:
        return render(request, '403.html',
                      {'message': f'Perangkat jenis "{jenis_name}" belum didukung untuk inspeksi.'})

    # Radio & VoIP hanya untuk Dispatcher (via pengujian telecom), bukan inservice inspection
    if jenis_name in TELECOM_JENIS:
        try:
            role = request.user.profile.role
        except Exception:
            role = ''
        if not (request.user.is_superuser or role in ('dispatcher', 'asisten_manager')):
            return render(request, '403.html',
                          {'message': 'Inspeksi Radio dan VoIP hanya dilakukan oleh Dispatcher melalui menu Pengujian Telekomunikasi.'}, status=403)

    # ── Batas 1x inspeksi per hari ────────────────────────────────────
    # Kalau perangkat sudah diinspeksi hari ini, form ditutup (GET & POST)
    # sampai berganti hari — mencegah pengisian ganda / bypass via URL.
    from django.contrib import messages
    if Inspection.objects.filter(device=device, tanggal__date=timezone.localdate()).exists():
        messages.info(request,
                      f'{device.nama} sudah diinspeksi hari ini. Inspeksi cukup satu kali per hari.')
        return redirect('inspection_device_list', lokasi=device.lokasi)

    if request.method == 'POST':
        if not request.FILES.get('foto'):
            messages.error(request, 'Foto dokumentasi wajib diisi sebelum menyimpan inspeksi.')
            return redirect('inspection_form', device_pk=device_pk)

        # ── Simpan Inspection induk ──────────────────────────────────
        insp = Inspection.objects.create(
            device   = device,
            jenis    = jenis_key,
            tanggal  = timezone.now(),
            operator = request.user,
            catatan  = request.POST.get('catatan', '').strip(),
            foto     = request.FILES['foto'],
        )

        # ── Simpan detail per jenis ──────────────────────────────────
        def g(key, default=''):
            return request.POST.get(key, default).strip()
        def gf(key):
            try:
                v = request.POST.get(key, '').strip()
                return float(v) if v else None
            except (ValueError, TypeError):
                return None

        if jenis_key == 'catu_daya':
            InspectionCatuDaya.objects.create(
                inspection              = insp,
                kondisi_rectifier       = g('kondisi_rectifier'),
                catatan_rectifier       = g('catatan_rectifier'),
                mode_recti              = g('mode_recti'),
                alarm_ground_fault      = g('alarm_ground_fault'),
                alarm_min_ac_fault      = g('alarm_min_ac_fault'),
                alarm_recti_fault       = g('alarm_recti_fault'),
                kebersihan_ruangan      = g('kebersihan_ruangan'),
                kondisi_baterai         = g('kondisi_baterai'),
                catatan_baterai         = g('catatan_baterai'),
                level_air_bank          = g('level_air_bank'),
                kebersihan_ruangan_bank = g('kebersihan_ruangan_bank'),
                kebersihan_bank         = g('kebersihan_bank'),
                exhaust_fan             = g('exhaust_fan'),
                tegangan_input_ac       = gf('tegangan_input_ac'),
                arus_input_ac           = gf('arus_input_ac'),
                tegangan_load_dc        = gf('tegangan_load_dc'),
                arus_load_dc            = gf('arus_load_dc'),
                tegangan_baterai_dc     = gf('tegangan_baterai_dc'),
                arus_baterai_dc         = gf('arus_baterai_dc'),
                kondisi_keseluruhan     = g('kondisi_keseluruhan'),
            )

        elif jenis_key == 'defense_scheme':
            InspectionDefenseScheme.objects.create(
                inspection          = insp,
                suhu_ruangan        = gf('suhu_ruangan'),
                kelembapan          = gf('kelembapan'),
                kebersihan_panel    = g('kebersihan_panel'),
                lampu_panel         = g('lampu_panel'),
                kondisi_relay       = g('kondisi_relay'),
                catatan_relay       = g('catatan_relay'),
                status_indikator    = g('status_indikator'),
                selektor_blok_skema = g('selektor_blok_skema'),
                posisi_selektor     = g('posisi_selektor'),
                kondisi_kabel_lan   = g('kondisi_kabel_lan'),
            )

        elif jenis_key == 'master_trip':
            InspectionMasterTrip.objects.create(
                inspection        = insp,
                suhu_ruangan      = gf('suhu_ruangan'),
                kelembapan        = gf('kelembapan'),
                kebersihan_panel  = g('kebersihan_panel'),
                lampu_panel       = g('lampu_panel'),
                kondisi_relay     = g('kondisi_relay'),
                indikator_led     = g('indikator_led'),
                catatan_relay     = g('catatan_relay'),
                posisi_selektor   = g('posisi_selektor'),
                kondisi_kabel_lan = g('kondisi_kabel_lan'),
                sumber_dc         = gf('sumber_dc'),
            )

        elif jenis_key == 'ufls':
            InspectionUFLS.objects.create(
                inspection        = insp,
                suhu_ruangan      = gf('suhu_ruangan'),
                kelembapan        = gf('kelembapan'),
                kebersihan_panel  = g('kebersihan_panel'),
                lampu_panel       = g('lampu_panel'),
                kondisi_relay     = g('kondisi_relay'),
                indikator_led     = g('indikator_led'),
                catatan_relay     = g('catatan_relay'),
                posisi_selektor   = g('posisi_selektor'),
                kondisi_kabel_lan = g('kondisi_kabel_lan'),
                sumber_dc         = gf('sumber_dc'),
            )

        elif jenis_key == 'dfr':
            InspectionDFR.objects.create(
                inspection         = insp,
                suhu_ruangan       = gf('suhu_ruangan'),
                kelembapan         = gf('kelembapan'),
                kebersihan_ruangan = g('kebersihan_ruangan'),
                lampu_penerangan   = g('lampu_penerangan'),
                kondisi_dfr        = g('kondisi_dfr'),
                healthy_status     = g('healthy_status'),
                indikasi_led_alarm = g('indikasi_led_alarm'),
                status_indikator   = g('status_indikator'),
                kondisi_kabel_lan  = g('kondisi_kabel_lan'),
            )

        elif jenis_key == 'server_ads':
            InspectionServerADS.objects.create(
                inspection            = insp,
                suhu_ruangan          = gf('suhu_ruangan'),
                kelembapan            = gf('kelembapan'),
                kebersihan_ruangan    = g('kebersihan_ruangan'),
                lampu_penerangan      = g('lampu_penerangan'),
                peralatan_server_ads  = g('peralatan_server_ads'),
                tampilan_hmi          = g('tampilan_hmi'),
                peralatan_gateway_ic3 = g('peralatan_gateway_ic3'),
                kondisi_switch_lan    = g('kondisi_switch_lan'),
                peralatan_power_supply= g('peralatan_power_supply'),
                fan_panel             = g('fan_panel'),
            )

        elif jenis_key == 'telecom':
            InspectionTelecom.objects.create(
                inspection        = insp,
                hasil_komunikasi  = g('hasil_komunikasi'),
                kualitas_suara    = g('kualitas_suara'),
                catatan_pengujian = g('catatan_pengujian'),
            )

        # ── Notifikasi ke AM jika ada kondisi Alarm/Tidak Normal ─────
        _kirim_notif_jika_perlu(insp, jenis_key, request.POST)

        return redirect('inspection_device_list', lokasi=device.lokasi)

    # GET — tampilkan form
    return render(request, 'inspection/form.html', {
        'device':      device,
        'jenis_key':   jenis_key,
        'jenis_label': dict(Inspection.JENIS_CHOICES).get(jenis_key, ''),
        'dev_merk':    device.merk or '—',
        'dev_type':    device.type or '—',
        'dev_ip':      str(device.ip_address) if device.ip_address else '—',
    })


def _kirim_notif_jika_perlu(insp, jenis_key, post_data):
    """Kirim notifikasi ke AM jika ada kondisi alarm/tidak normal."""
    perlu_notif = False
    pesan_detail = []

    if jenis_key == 'catu_daya':
        if post_data.get('kondisi_rectifier') == 'alarm':
            perlu_notif = True
            pesan_detail.append('Rectifier: ALARM')
        if post_data.get('kondisi_keseluruhan') == 'kotor':
            perlu_notif = True
            pesan_detail.append('Kondisi keseluruhan: KOTOR')

    elif jenis_key == 'defense_scheme':
        if post_data.get('kondisi_relay') == 'alarm':
            perlu_notif = True
            pesan_detail.append('Kondisi Relay Defense Scheme: ALARM')
        if post_data.get('status_indikator') == 'tidak_normal':
            perlu_notif = True
            pesan_detail.append('Status Indikator Defense Scheme: TIDAK NORMAL')

    elif jenis_key in ('master_trip', 'ufls'):
        if post_data.get('kondisi_relay') == 'alarm':
            perlu_notif = True
            pesan_detail.append(f'Kondisi Relay {jenis_key.upper()}: ALARM')
        if post_data.get('indikator_led') == 'tidak_normal':
            perlu_notif = True
            pesan_detail.append(f'Indikator LED {jenis_key.upper()}: TIDAK NORMAL')

    elif jenis_key == 'dfr':
        if post_data.get('kondisi_dfr') == 'faulty':
            perlu_notif = True
            pesan_detail.append('Kondisi DFR: FAULTY')
        if post_data.get('healthy_status') in ('faulty', 'alarm'):
            perlu_notif = True
            pesan_detail.append(f"Healthy Status DFR: {post_data.get('healthy_status').upper()}")
        if post_data.get('indikasi_led_alarm') == 'ada':
            perlu_notif = True
            pesan_detail.append('Indikasi LED Alarm DFR: ADA')
        if post_data.get('status_indikator') == 'tidak_normal':
            perlu_notif = True
            pesan_detail.append('Status Indikator DFR: TIDAK NORMAL')

    elif jenis_key == 'server_ads':
        for field, label in (
            ('peralatan_server_ads',   'Peralatan Server ADS'),
            ('tampilan_hmi',           'Tampilan HMI'),
            ('peralatan_gateway_ic3',  'Peralatan Gateway IC3 ADS'),
            ('peralatan_power_supply', 'Peralatan Power Supply'),
        ):
            if post_data.get(field) == 'tidak_normal':
                perlu_notif = True
                pesan_detail.append(f'{label}: TIDAK NORMAL')
        if post_data.get('kondisi_switch_lan') == 'mati':
            perlu_notif = True
            pesan_detail.append('Kondisi Switch Kabel LAN: MATI')
        if post_data.get('fan_panel') == 'mati':
            perlu_notif = True
            pesan_detail.append('Fan Panel: MATI')

    elif jenis_key == 'telecom':
        if post_data.get('hasil_komunikasi') == 'tidak_normal':
            perlu_notif = True
            pesan_detail.append('Pengujian Telekomunikasi: TIDAK NORMAL')

    if not perlu_notif:
        return

    try:
        from notifikasi.views import notif_ke_am
        notif_ke_am(
            tipe   = 'inspection_alert',
            judul  = f'Inspeksi Alert — {insp.device.nama}',
            pesan  = (
                f'Inspeksi inservice pada {insp.device.nama} ({insp.device.lokasi}) '
                f'tanggal {insp.tanggal.strftime("%d %b %Y %H:%M")} '
                f'mendeteksi kondisi: {", ".join(pesan_detail)}'
            ),
            level  = 'warning',
            url    = f'/inspection/riwayat/{insp.pk}/',
            device = insp.device,
        )
    except Exception:
        pass


# ─────────────────────────────────────────────────────────────────────
# RIWAYAT INSPEKSI
# ─────────────────────────────────────────────────────────────────────
@login_required
@require_operator
def inspection_riwayat(request, pk):
    """Detail riwayat satu inspeksi."""
    insp = get_object_or_404(Inspection, pk=pk)
    detail = None
    detail_rows = []
    rectifier_rows = []
    bank_rows = []
    env_rows = []
    relay_rows = []
    led_rows = []
    dfr_rows = []
    ads_rows = []

    def mk_row(label, val, ok_val, ok_display, nok_display=None, alarm=False):
        is_ok = (val == ok_val)
        display = ok_display if is_ok else (nok_display or val or '—')
        bg = '#fee2e2' if (not is_ok and alarm) else '#fef9c3'
        fg = '#991b1b' if (not is_ok and alarm) else '#854d0e'
        prefix = '⚠ ' if (not is_ok and alarm and val) else ''
        return {'label': label, 'val': val, 'ok': is_ok,
                'display': prefix + display if val else '—',
                'bg': bg, 'fg': fg}

    try:
        if insp.jenis == 'catu_daya':
            detail = insp.detail_catu_daya
            detail_rows = [
                ('Teg. Input AC',   detail.tegangan_input_ac,   'V'),
                ('Arus Input AC',   detail.arus_input_ac,       'A'),
                ('Teg. Load DC',    detail.tegangan_load_dc,    'V'),
                ('Arus Load DC',    detail.arus_load_dc,        'A'),
                ('Teg. Baterai DC', detail.tegangan_baterai_dc, 'V'),
                ('Arus Baterai DC', detail.arus_baterai_dc,     'A'),
            ]
            rectifier_rows = [
                mk_row('Kondisi Rectifier', detail.kondisi_rectifier,
                       'normal', 'Normal', 'Alarm', alarm=True),
                mk_row('Mode Rectifier', detail.mode_recti,
                       'float', detail.get_mode_recti_display() if detail.mode_recti else ''),
                mk_row('Alarm Ground Fault', detail.alarm_ground_fault,
                       'tidak_ada', 'Tidak Ada', 'Ada', alarm=True),
                mk_row('Alarm Min AC Fault', detail.alarm_min_ac_fault,
                       'tidak_ada', 'Tidak Ada', 'Ada', alarm=True),
                mk_row('Alarm Recti Fault', detail.alarm_recti_fault,
                       'tidak_ada', 'Tidak Ada', 'Ada', alarm=True),
                mk_row('Kebersihan Ruangan', detail.kebersihan_ruangan,
                       'bersih', 'Bersih', 'Kotor'),
            ]
            bank_rows = [
                mk_row('Level Air Bank', detail.level_air_bank,
                       'normal', 'Normal',
                       detail.get_level_air_bank_display() if detail.level_air_bank else '—'),
                mk_row('Kebersihan Ruangan Bank', detail.kebersihan_ruangan_bank,
                       'bersih', 'Bersih', 'Kotor'),
                mk_row('Kebersihan Bank', detail.kebersihan_bank,
                       'bersih', 'Bersih', 'Kotor'),
                mk_row('Exhaust Fan', detail.exhaust_fan,
                       'nyala', 'Nyala', 'Mati', alarm=True),
                mk_row('Kondisi Baterai', detail.kondisi_baterai,
                       'bersih', 'Bersih', 'Kotor'),
            ]

        elif insp.jenis == 'defense_scheme':
            detail = insp.detail_defense_scheme
            env_rows = [
                mk_row('Kebersihan', detail.kebersihan_panel, 'bersih', 'Bersih', 'Kotor'),
                mk_row('Lampu Penerangan', detail.lampu_panel, 'nyala', 'Nyala', 'Mati', alarm=True),
            ]
            relay_rows = [
                mk_row('Status Relay', detail.kondisi_relay, 'normal', 'Normal', 'Alarm', alarm=True),
            ]
            led_rows = [
                mk_row('Status Indikator', detail.status_indikator, 'normal', 'Normal', 'Tidak Normal', alarm=True),
                mk_row('Selector Blok Skema', detail.selektor_blok_skema, 'on', 'ON', 'OFF'),
                mk_row('Selector Target', detail.posisi_selektor, 'on', 'ON', 'Blok'),
                mk_row('Kondisi Kabel LAN', detail.kondisi_kabel_lan, 'normal', 'Normal', 'Terlepas', alarm=True),
            ]

        elif insp.jenis == 'master_trip':
            detail = insp.detail_master_trip
            env_rows = [
                mk_row('Kebersihan', detail.kebersihan_panel, 'bersih', 'Bersih', 'Kotor'),
                mk_row('Lampu Penerangan', detail.lampu_panel, 'nyala', 'Nyala', 'Mati', alarm=True),
            ]
            relay_rows = [
                mk_row('Status Relay', detail.kondisi_relay, 'normal', 'Normal', 'Alarm', alarm=True),
                mk_row('Selektor Target', detail.posisi_selektor, 'on', 'ON', 'Blok'),
                mk_row('Kondisi Kabel LAN', detail.kondisi_kabel_lan, 'normal', 'Normal', 'Terlepas', alarm=True),
            ]
            led_rows = [
                mk_row('Status Indikator', detail.indikator_led, 'normal', 'Normal', 'Tidak Normal', alarm=True),
            ]

        elif insp.jenis == 'ufls':
            detail = insp.detail_ufls
            env_rows = [
                mk_row('Kebersihan', detail.kebersihan_panel, 'bersih', 'Bersih', 'Kotor'),
                mk_row('Lampu Penerangan', detail.lampu_panel, 'nyala', 'Nyala', 'Mati', alarm=True),
            ]
            relay_rows = [
                mk_row('Status Relay', detail.kondisi_relay, 'normal', 'Normal', 'Alarm', alarm=True),
            ]
            led_rows = [
                mk_row('Status Indikator', detail.indikator_led, 'normal', 'Normal', 'Tidak Normal', alarm=True),
            ]
            if detail.sumber_dc is not None:
                detail_rows = [('Sumber DC', detail.sumber_dc, 'V')]

        elif insp.jenis == 'dfr':
            detail = insp.detail_dfr
            env_rows = [
                mk_row('Kebersihan', detail.kebersihan_ruangan, 'bersih', 'Bersih', 'Kotor'),
                mk_row('Lampu Penerangan', detail.lampu_penerangan, 'baik', 'Baik',
                       detail.get_lampu_penerangan_display() if detail.lampu_penerangan else '—', alarm=True),
            ]
            dfr_rows = [
                mk_row('Kondisi DFR', detail.kondisi_dfr, 'normal', 'Normal', 'Faulty', alarm=True),
                mk_row('Healthy Status', detail.healthy_status, 'healthy', 'Healthy',
                       detail.get_healthy_status_display() if detail.healthy_status else '—', alarm=True),
                mk_row('Indikasi LED Alarm', detail.indikasi_led_alarm, 'tidak_ada', 'Tidak Ada', 'Ada', alarm=True),
                mk_row('Status Indikator', detail.status_indikator, 'normal', 'Normal', 'Tidak Normal', alarm=True),
                mk_row('Kondisi Kabel LAN', detail.kondisi_kabel_lan, 'normal', 'Normal', 'Terlepas', alarm=True),
            ]

        elif insp.jenis == 'server_ads':
            detail = insp.detail_server_ads
            env_rows = [
                mk_row('Kebersihan', detail.kebersihan_ruangan, 'bersih', 'Bersih', 'Kotor'),
                mk_row('Lampu Penerangan', detail.lampu_penerangan, 'baik', 'Baik',
                       detail.get_lampu_penerangan_display() if detail.lampu_penerangan else '—', alarm=True),
            ]
            ads_rows = [
                mk_row('Peralatan Server ADS', detail.peralatan_server_ads, 'normal', 'Normal', 'Tidak Normal', alarm=True),
                mk_row('Tampilan HMI', detail.tampilan_hmi, 'normal', 'Normal',
                       detail.get_tampilan_hmi_display() if detail.tampilan_hmi else '—', alarm=True),
                mk_row('Peralatan Gateway IC3 ADS', detail.peralatan_gateway_ic3, 'normal', 'Normal', 'Tidak Normal', alarm=True),
                mk_row('Kondisi Switch Kabel LAN', detail.kondisi_switch_lan, 'normal', 'Normal', 'Mati', alarm=True),
                mk_row('Peralatan Power Supply', detail.peralatan_power_supply, 'normal', 'Normal', 'Tidak Normal', alarm=True),
                mk_row('Fan Panel', detail.fan_panel, 'nyala', 'Nyala', 'Mati', alarm=True),
            ]

        elif insp.jenis == 'telecom':
            detail = insp.detail_telecom

    except Exception:
        pass

    return render(request, 'inspection/riwayat_detail.html', {
        'insp':           insp,
        'detail':         detail,
        'detail_rows':    detail_rows,
        'rectifier_rows': rectifier_rows,
        'bank_rows':      bank_rows,
        'env_rows':       env_rows,
        'relay_rows':     relay_rows,
        'led_rows':       led_rows,
        'dfr_rows':       dfr_rows,
        'ads_rows':       ads_rows,
    })


@login_required
def inspection_delete(request, pk):
    """Hapus satu record inspeksi — hanya superuser."""
    if not request.user.is_superuser:
        return render(request, '403.html',
                      {'message': 'Hanya superuser yang dapat menghapus inspeksi.'}, status=403)
    insp = get_object_or_404(Inspection, pk=pk)
    if request.method == 'POST':
        lokasi = insp.device.lokasi
        insp.delete()
        from django.contrib import messages
        messages.success(request, 'Inspeksi berhasil dihapus.')
        return redirect('inspection_device_list', lokasi=lokasi)
    return render(request, '403.html', {'message': 'Method tidak diizinkan.'}, status=405)


@login_required
@require_operator
def inspection_riwayat_device(request, device_pk):
    """Riwayat semua inspeksi untuk satu device + trend kondisi."""
    import json as _json
    from datetime import timedelta
    device = get_object_or_404(Device, pk=device_pk, is_deleted=False)
    inspeksi_qs = Inspection.objects.filter(device=device).order_by('-tanggal')

    inspeksi_list = []
    for insp in inspeksi_qs:
        insp.status_kondisi = 'alarm' if _is_alarm_inspection(insp) else 'normal'
        inspeksi_list.append(insp)

    # ── Trend 30 hari ─────────────────────────────────────────────
    today = date.today()
    trend_labels  = []
    trend_kondisi = []
    trend_vload   = []
    trend_vbat    = []
    trend_sdc     = []

    for i in range(29, -1, -1):
        d = today - timedelta(days=i)
        trend_labels.append(d.strftime('%d/%m'))
        insp_day = next((x for x in inspeksi_list if x.tanggal.date() == d), None)
        if insp_day:
            trend_kondisi.append(0 if insp_day.status_kondisi == 'alarm' else 1)
            try:
                if insp_day.jenis == 'catu_daya':
                    det = insp_day.detail_catu_daya
                    trend_vload.append(float(det.tegangan_load_dc) if det.tegangan_load_dc else None)
                    trend_vbat.append(float(det.tegangan_baterai_dc) if det.tegangan_baterai_dc else None)
                    trend_sdc.append(None)
                else:
                    attr = 'detail_defense_scheme' if insp_day.jenis == 'defense_scheme' else f'detail_{insp_day.jenis}'
                    det = getattr(insp_day, attr)
                    trend_sdc.append(float(det.sumber_dc) if det.sumber_dc else None)
                    trend_vload.append(None); trend_vbat.append(None)
            except Exception:
                trend_vload.append(None); trend_vbat.append(None); trend_sdc.append(None)
        else:
            trend_kondisi.append(None)
            trend_vload.append(None); trend_vbat.append(None); trend_sdc.append(None)

    total   = len(inspeksi_list)
    alarm_c = sum(1 for x in inspeksi_list if x.status_kondisi == 'alarm')
    flag_c  = sum(1 for x in inspeksi_list if x.is_flagged)

    return render(request, 'inspection/riwayat_device.html', {
        'device':         device,
        'inspeksi_list':  inspeksi_list,
        'total':          total,
        'alarm_count':    alarm_c,
        'flag_count':     flag_c,
        'trend_labels':   _json.dumps(trend_labels),
        'trend_kondisi':  _json.dumps(trend_kondisi),
        'trend_vload':    _json.dumps(trend_vload),
        'trend_vbat':     _json.dumps(trend_vbat),
        'trend_sdc':      _json.dumps(trend_sdc),
        'jenis_device':   device.jenis.name if device.jenis else '',
    })


# ─────────────────────────────────────────────────────────────────────
# API — LAST INSPECTION (untuk form gangguan)
# ─────────────────────────────────────────────────────────────────────
@login_required
def inspection_api_last(request):
    """JSON: inspeksi terakhir sebuah device — dipakai form gangguan."""
    from django.http import JsonResponse
    device_id = request.GET.get('device')
    if not device_id:
        return JsonResponse({'has_alarm': False})

    insp = (
        Inspection.objects
        .filter(device_id=device_id)
        .order_by('-tanggal')
        .first()
    )
    if not insp:
        return JsonResponse({'has_alarm': False})

    detail    = f'Tanggal: {insp.tanggal.strftime("%d %b %Y %H:%M")} · Jenis: {insp.get_jenis_display()}'
    has_alarm = _is_alarm_inspection(insp)
    warning   = (f'Inspeksi terakhir ({insp.tanggal.strftime("%d %b %Y")}) mendeteksi kondisi tidak normal'
                 if has_alarm else '')

    if insp.is_flagged:
        has_alarm = True
        warning   = warning or f'Inspeksi terakhir ({insp.tanggal.strftime("%d %b %Y")}) DIFLAG'
        detail   += f' · Diflag: {insp.flag_catatan or "-"}'

    return JsonResponse({
        'has_alarm': has_alarm,
        'warning':   warning,
        'detail':    detail,
        'insp_pk':   insp.pk,
    })


# FLAG / UNFLAG INSPECTION
# ─────────────────────────────────────────────────────────────────────
@login_required
def inspection_flag(request, pk):
    """Flag inspection — hanya Engineer/AM/Superuser."""
    from devices.permissions import can_edit
    if not (request.user.is_superuser or can_edit(request.user)):
        from django.http import JsonResponse
        return JsonResponse({'error': 'Tidak punya akses'}, status=403)

    insp = get_object_or_404(Inspection, pk=pk)

    if request.method == 'POST':
        catatan = request.POST.get('catatan', '').strip()
        from django.utils import timezone as tz
        insp.is_flagged   = True
        insp.flag_catatan = catatan
        insp.flagged_by   = request.user
        insp.flagged_at   = tz.now()
        insp.save(update_fields=['is_flagged','flag_catatan','flagged_by','flagged_at'])

        # Notifikasi ke operator
        try:
            from notifikasi.views import notif_ke_am
            if insp.operator:
                from notifikasi.models import Notifikasi
                Notifikasi.objects.create(
                    user    = insp.operator,
                    tipe    = 'inspection_flag',
                    judul   = f'Inspeksi Diflag — {insp.device.nama}',
                    pesan   = (
                        f'Inspeksi {insp.device.nama} ({insp.device.lokasi}) '
                        f'tanggal {insp.tanggal.strftime("%d %b %Y %H:%M")} '
                        f'diflag oleh {request.user.get_full_name() or request.user.username}.'
                        + (f' Catatan: {catatan}' if catatan else '')
                    ),
                    level   = 'warning',
                    url     = f'/inspection/riwayat/{insp.pk}/',
                    device  = insp.device,
                )
        except Exception:
            pass

        from django.contrib import messages
        messages.warning(request, f'Inspeksi berhasil diflag.')
    return redirect('inspection_riwayat', pk=pk)


@login_required
def inspection_unflag(request, pk):
    """Hapus flag inspection."""
    from devices.permissions import can_edit
    if not (request.user.is_superuser or can_edit(request.user)):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden()

    insp = get_object_or_404(Inspection, pk=pk)
    if request.method == 'POST':
        insp.is_flagged   = False
        insp.flag_catatan = ''
        insp.flagged_by   = None
        insp.flagged_at   = None
        insp.save(update_fields=['is_flagged','flag_catatan','flagged_by','flagged_at'])
        from django.contrib import messages
        messages.success(request, 'Flag berhasil dihapus.')
    return redirect('inspection_riwayat', pk=pk)


# EXPORT LAPORAN PER ULTG
# ─────────────────────────────────────────────────────────────────────
@login_required
def inspection_export_ultg(request):
    """Export laporan inspeksi per ULTG — sheet terpisah per jenis perangkat."""
    from devices.models import ULTG
    from devices.permissions import can_edit

    if not (request.user.is_superuser or can_edit(request.user)):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden()

    ultg_id   = request.GET.get('ultg', '')
    month     = int(request.GET.get('month', date.today().month))
    year      = int(request.GET.get('year', date.today().year))

    BULAN_ID = ['Januari','Februari','Maret','April','Mei','Juni',
                'Juli','Agustus','September','Oktober','November','Desember']
    bulan_str = BULAN_ID[month - 1]

    # Ambil ULTG
    if ultg_id:
        ultg = get_object_or_404(ULTG, pk=ultg_id)
        lokasi_names = ultg.get_lokasi_names()
        ultg_label   = ultg.nama
    else:
        ultg         = None
        lokasi_names = None
        ultg_label   = 'Semua ULTG'

    INSPECTABLE = {
        'Catu Daya':           'catu_daya',
        'RELE DEFENSE SCHEME': 'defense_scheme',
        'MASTER TRIP':         'master_trip',
        'UFLS':                'ufls',
    }

    # ── Style helpers ────────────────────────────────────────────
    thin     = Side(style='thin')
    brd      = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font = Font(bold=True, color='FFFFFF', size=9)
    hdr_aln  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c_aln    = Alignment(horizontal='center', vertical='center')
    l_aln    = Alignment(vertical='center', wrap_text=True)
    alt_fill = PatternFill('solid', fgColor='F8FAFC')
    alarm_fill = PatternFill('solid', fgColor='FEE2E2')
    flag_fill  = PatternFill('solid', fgColor='FEF9C3')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    title_main = f'LAPORAN INSERVICE INSPECTION — {ultg_label.upper()} — {bulan_str} {year}'

    # ─────────────────────────────────────────────────────────────
    # SHEET RINGKASAN
    # ─────────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet('Ringkasan')
    ws_sum.sheet_properties.tabColor = '0F172A'
    ws_sum.column_dimensions['A'].width = 3
    ws_sum.column_dimensions['B'].width = 30
    ws_sum.column_dimensions['C'].width = 12
    ws_sum.column_dimensions['D'].width = 14
    ws_sum.column_dimensions['E'].width = 12
    ws_sum.column_dimensions['F'].width = 12
    ws_sum.column_dimensions['G'].width = 16

    ws_sum.merge_cells('B1:G1')
    ws_sum['B1'].value     = title_main
    ws_sum['B1'].font      = Font(bold=True, size=12)
    ws_sum['B1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_sum['B1'].fill      = PatternFill('solid', fgColor='EFF6FF')
    ws_sum.row_dimensions[1].height = 28

    ws_sum.merge_cells('B2:G2')
    ws_sum['B2'].value     = f"Dicetak: {date.today().strftime('%d %B %Y')}"
    ws_sum['B2'].font      = Font(size=10, italic=True, color='64748B')
    ws_sum['B2'].alignment = Alignment(horizontal='center')

    sum_heads = ['Jenis Perangkat','Total Terdaftar','Terinspeksi','Belum Inspeksi','Normal','Alarm/Flag']
    sum_widths_col = ['B','C','D','E','F','G']
    hdr_fill_dark = PatternFill('solid', fgColor='1E293B')
    for ci, (col, h) in enumerate(zip(sum_widths_col, sum_heads)):
        cell = ws_sum[f'{col}4']
        cell.value     = h
        cell.font      = hdr_font
        cell.fill      = hdr_fill_dark
        cell.alignment = hdr_aln
        cell.border    = brd
    ws_sum.row_dimensions[4].height = 22

    row_sum = 5
    JENIS_COLORS = {
        'Catu Daya':'EA580C', 'RELE DEFENSE SCHEME':'7C3AED',
        'MASTER TRIP':'2563EB', 'UFLS':'059669',
    }

    for jenis_name, jenis_key in INSPECTABLE.items():
        devs_qs = Device.objects.filter(is_deleted=False, jenis__name=jenis_name)
        if lokasi_names:
            devs_qs = devs_qs.filter(lokasi__in=lokasi_names)
        total = devs_qs.count()
        if total == 0:
            continue

        insp_qs = Inspection.objects.filter(
            device__in=devs_qs,
            tanggal__year=year, tanggal__month=month
        )
        insp_ids = set(insp_qs.values_list('device_id', flat=True).distinct())
        terinspeksi = len(insp_ids)
        belum       = total - terinspeksi

        # Hitung alarm
        alarm = sum(1 for insp in insp_qs.select_related('device') if _is_alarm_inspection(insp))
        flag_n = insp_qs.filter(is_flagged=True).count()
        abnormal = alarm + flag_n

        color = JENIS_COLORS.get(jenis_name, '334155')
        row_fill = PatternFill('solid', fgColor=color + '22' if len(color) == 6 else 'F8FAFC')
        vals = [jenis_name, total, terinspeksi, belum, terinspeksi - abnormal, abnormal]
        for ci, (col, val) in enumerate(zip(sum_widths_col, vals)):
            cell = ws_sum[f'{col}{row_sum}']
            cell.value     = val
            cell.border    = brd
            cell.alignment = c_aln if ci > 0 else Alignment(vertical='center')
            if ci == 5 and val > 0:
                cell.font = Font(bold=True, color='991B1B')
                cell.fill = PatternFill('solid', fgColor='FEE2E2')
        ws_sum.row_dimensions[row_sum].height = 18
        row_sum += 1

    ws_sum.freeze_panes = 'B5'

    # ─────────────────────────────────────────────────────────────
    # SHEET PER JENIS PERANGKAT
    # ─────────────────────────────────────────────────────────────
    for jenis_name, jenis_key in INSPECTABLE.items():
        devs_qs = Device.objects.filter(
            is_deleted=False, jenis__name=jenis_name
        ).select_related('jenis').order_by('lokasi', 'nama')
        if lokasi_names:
            devs_qs = devs_qs.filter(lokasi__in=lokasi_names)
        if not devs_qs.exists():
            continue

        # Inspeksi bulan ini per device (ambil yang terbaru)
        insp_map = {}
        for insp in Inspection.objects.filter(
            device__in=devs_qs,
            tanggal__year=year, tanggal__month=month,
            jenis=jenis_key
        ).select_related('device', 'operator').order_by('device_id', '-tanggal'):
            if insp.device_id not in insp_map:
                insp_map[insp.device_id] = insp

        color  = JENIS_COLORS.get(jenis_name, '334155')
        ws     = wb.create_sheet(jenis_name[:31])
        ws.sheet_properties.tabColor = color

        # Header
        if jenis_key == 'catu_daya':
            headers = ['No','Nama Perangkat','Lokasi','Tgl Inspeksi','Operator',
                       'Kond. Rectifier','Mode Recti','Alarm GF','Alarm AC','Alarm Recti',
                       'Kebersihan Ruangan','Level Air Bank','Exhaust Fan',
                       'Teg Load DC (V)','Teg Baterai DC (V)','Arus Load DC (A)','Status']
            col_w  = [4,28,22,14,16,16,14,12,12,12,18,16,12,16,18,16,12]
        else:
            headers = ['No','Nama Perangkat','Lokasi','Tgl Inspeksi','Operator',
                       'Suhu Ruangan (°C)','Kebersihan Panel','Lampu Panel',
                       'Kondisi Rele','Status Indikator',
                       'Posisi Selektor','Kabel LAN','Sumber DC (V)','Status']
            col_w  = [4,28,22,14,16,14,16,14,14,14,16,14,14,12]

        ncols = len(headers)
        last_col = get_column_letter(ncols + 1)

        # Title sheet
        ws.merge_cells(f'B1:{last_col}1')
        ws['B1'].value     = f'{jenis_name.upper()} — {title_main}'
        ws['B1'].font      = Font(bold=True, size=11)
        ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B1'].fill      = PatternFill('solid', fgColor='EFF6FF')
        ws.row_dimensions[1].height = 26

        ws.merge_cells(f'B2:{last_col}2')
        ws['B2'].value     = f"Dicetak: {date.today().strftime('%d %B %Y')} | {terinspeksi if (devs_qs.count() > 0) else '?'} dari {devs_qs.count()} perangkat terinspeksi"
        ws['B2'].font      = Font(size=10, italic=True, color='64748B')
        ws['B2'].alignment = Alignment(horizontal='center')

        ws.column_dimensions['A'].width = 2
        hdr_fill_j = PatternFill('solid', fgColor=color)
        for ci, (h, w) in enumerate(zip(headers, col_w), 2):
            cell = ws.cell(row=4, column=ci, value=h)
            cell.font      = hdr_font
            cell.fill      = hdr_fill_j
            cell.alignment = hdr_aln
            cell.border    = brd
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[4].height = 22
        ws.freeze_panes = f'B5'

        # Data rows
        for ri, dev in enumerate(devs_qs, 1):
            wr   = ri + 4
            insp = insp_map.get(dev.pk)

            if insp:
                tgl_str  = insp.tanggal.strftime('%d/%m/%Y %H:%M')
                op_str   = insp.operator.get_full_name() or insp.operator.username if insp.operator else '—'
                is_alarm = False
                is_flag  = insp.is_flagged

                if jenis_key == 'catu_daya':
                    try:
                        d = insp.detail_catu_daya
                        is_alarm = d.kondisi_rectifier == 'alarm'
                        row_vals = [
                            ri, dev.nama, dev.lokasi or '—', tgl_str, op_str,
                            d.get_kondisi_rectifier_display() if d.kondisi_rectifier else '—',
                            d.get_mode_recti_display() if d.mode_recti else '—',
                            d.get_alarm_ground_fault_display() if d.alarm_ground_fault else '—',
                            d.get_alarm_min_ac_fault_display() if d.alarm_min_ac_fault else '—',
                            d.get_alarm_recti_fault_display() if d.alarm_recti_fault else '—',
                            d.get_kebersihan_ruangan_display() if d.kebersihan_ruangan else '—',
                            d.get_level_air_bank_display() if d.level_air_bank else '—',
                            d.get_exhaust_fan_display() if d.exhaust_fan else '—',
                            d.tegangan_load_dc or '—',
                            d.tegangan_baterai_dc or '—',
                            d.arus_load_dc or '—',
                            '⚠ ALARM' if is_alarm else ('🚩 Flag' if is_flag else '✓ Normal'),
                        ]
                    except Exception:
                        row_vals = [ri, dev.nama, dev.lokasi or '—', tgl_str, op_str] + ['—'] * (len(headers) - 5)
                else:
                    try:
                        attr = 'detail_defense_scheme' if jenis_key == 'defense_scheme' else f'detail_{jenis_key}'
                        d = getattr(insp, attr)
                        is_alarm = _is_alarm_inspection(insp)
                        status_field = 'status_indikator' if jenis_key == 'defense_scheme' else 'indikator_led'
                        status_display_fn = f'get_{status_field}_display'
                        status_val = getattr(d, status_field)
                        row_vals = [
                            ri, dev.nama, dev.lokasi or '—', tgl_str, op_str,
                            d.suhu_ruangan if d.suhu_ruangan is not None else '—',
                            d.get_kebersihan_panel_display() if d.kebersihan_panel else '—',
                            d.get_lampu_panel_display() if d.lampu_panel else '—',
                            d.get_kondisi_relay_display() if d.kondisi_relay else '—',
                            getattr(d, status_display_fn)() if status_val else '—',
                            d.get_posisi_selektor_display() if d.posisi_selektor else '—',
                            d.get_kondisi_kabel_lan_display() if d.kondisi_kabel_lan else '—',
                            d.sumber_dc if d.sumber_dc is not None else '—',
                            '⚠ Faulty' if is_alarm else ('🚩 Flag' if is_flag else '✓ Normal'),
                        ]
                    except Exception:
                        row_vals = [ri, dev.nama, dev.lokasi or '—', tgl_str, op_str] + ['—'] * (len(headers) - 5)
            else:
                # Belum diinspeksi
                is_alarm = False
                is_flag  = False
                row_vals = [ri, dev.nama, dev.lokasi or '—', '—', '—'] + ['—'] * (len(headers) - 5)
                row_vals[-1] = 'Belum Diinspeksi'

            status_val = row_vals[-1]
            for ci, val in enumerate(row_vals, 2):
                cell = ws.cell(row=wr, column=ci, value=val)
                cell.border    = brd
                cell.alignment = c_aln if ci in [2, 4] else l_aln
                # Warna status
                if ci == len(headers) + 1:
                    if '⚠' in str(status_val) or 'Faulty' in str(status_val):
                        cell.fill = alarm_fill
                        cell.font = Font(bold=True, color='991B1B', size=9)
                    elif '🚩' in str(status_val):
                        cell.fill = flag_fill
                        cell.font = Font(bold=True, color='92400E', size=9)
                    elif 'Belum' in str(status_val):
                        cell.fill = PatternFill('solid', fgColor='F1F5F9')
                        cell.font = Font(color='64748B', size=9)
                    else:
                        cell.fill = PatternFill('solid', fgColor='DCFCE7')
                        cell.font = Font(bold=True, color='166534', size=9)
                elif ri % 2 == 0 and ci != len(headers) + 1:
                    cell.fill = alt_fill
            ws.row_dimensions[wr].height = 18

    # Response
    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fname = f'Inspeksi_{ultg_label.replace(" ","_")}_{bulan_str}_{year}.xlsx'
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    wb.save(resp)
    return resp


# DASHBOARD INSPECTION (Engineer / AM)
# ─────────────────────────────────────────────────────────────────────
@login_required
def inspection_dashboard(request):
    """Dashboard Inservice Inspection untuk Engineer dan AM — semua ULTG/GI."""
    from devices.permissions import can_edit
    # Hanya engineer, AM, superuser — bukan operator murni
    if not (request.user.is_superuser or can_edit(request.user)):
        return redirect('inspection_lokasi')

    from devices.models import ULTG
    from datetime import timedelta

    today = date.today()
    # Jenis perangkat yang dihitung di dashboard — semua jenis dengan form inservice
    # inspection sendiri (Radio/VoIP dikecualikan, dihitung lewat dashboard Pengujian Telekomunikasi).
    INSPECTABLE = ['Catu Daya', 'RELE DEFENSE SCHEME', 'MASTER TRIP', 'UFLS', 'DFR', 'SERVER PROSIS']

    # ── Stat global ──────────────────────────────────────────────
    insp_today      = Inspection.objects.filter(tanggal__date=today).count()
    insp_bulan      = Inspection.objects.filter(tanggal__year=today.year, tanggal__month=today.month).count()
    insp_total      = Inspection.objects.count()

    # Alarm total bulan ini
    alarm_bulan = sum(
        1 for insp in Inspection.objects.filter(
            tanggal__year=today.year, tanggal__month=today.month
        ).select_related('device')
        if _is_alarm_inspection(insp)
    )

    # Device total inspectable
    total_device_inspectable = Device.objects.filter(
        is_deleted=False, jenis__name__in=INSPECTABLE
    ).count()

    # Sudah diinspeksi bulan ini
    insp_ids_bulan = Inspection.objects.filter(
        tanggal__year=today.year, tanggal__month=today.month
    ).values_list('device_id', flat=True).distinct()
    sudah_bulan = len(set(insp_ids_bulan))
    pct_bulan   = round(sudah_bulan / total_device_inspectable * 100) if total_device_inspectable else 0

    # ── Progress per ULTG ────────────────────────────────────────
    ultg_stats = []
    for ultg in ULTG.objects.prefetch_related('lokasi').order_by('nama'):
        lokasi_names = ultg.get_lokasi_names()
        if not lokasi_names:
            continue
        devs = Device.objects.filter(
            is_deleted=False, jenis__name__in=INSPECTABLE, lokasi__in=lokasi_names
        )
        total = devs.count()
        if total == 0:
            continue
        inspected = Inspection.objects.filter(
            device__in=devs,
            tanggal__year=today.year, tanggal__month=today.month
        ).values_list('device_id', flat=True).distinct()
        sudah = len(set(inspected))
        pct   = round(sudah / total * 100) if total else 0
        insp_hari = Inspection.objects.filter(device__in=devs, tanggal__date=today).count()

        # Alarm bulan ini di ULTG ini
        alarm_ultg = sum(
            1 for insp in Inspection.objects.filter(
                device__in=devs, tanggal__year=today.year, tanggal__month=today.month
            )
            if _is_alarm_inspection(insp)
        )

        ultg_stats.append({
            'ultg':    ultg,
            'total':   total,
            'sudah':   sudah,
            'belum':   total - sudah,
            'pct':     pct,
            'hari':    insp_hari,
            'alarm':   alarm_ultg,
            'lokasi_count': len(lokasi_names),
        })

    # ── Progress per jenis perangkat ─────────────────────────────
    jenis_stats = []
    for jenis_name in INSPECTABLE:
        devs  = Device.objects.filter(is_deleted=False, jenis__name=jenis_name)
        total = devs.count()
        if total == 0:
            continue
        inspected = Inspection.objects.filter(
            device__in=devs, tanggal__year=today.year, tanggal__month=today.month
        ).values_list('device_id', flat=True).distinct()
        sudah = len(set(inspected))
        jenis_stats.append({
            'jenis': jenis_name, 'total': total,
            'sudah': sudah, 'belum': total - sudah,
            'pct': round(sudah / total * 100) if total else 0,
        })

    # ── Trend 30 hari ────────────────────────────────────────────
    trend_30 = []
    for i in range(29, -1, -1):
        d = today - timedelta(days=i)
        count = Inspection.objects.filter(tanggal__date=d).count()
        trend_30.append({'label': d.strftime('%d/%m'), 'count': count})

    # ── Alarm terbaru ─────────────────────────────────────────────
    alarm_list = []
    for insp in Inspection.objects.select_related('device','device__jenis','operator').order_by('-tanggal')[:100]:
        if _is_alarm_inspection(insp):
            alarm_list.append(insp)
            if len(alarm_list) >= 10:
                break

    # ── Inspeksi diflag ───────────────────────────────────────────
    flagged_list = Inspection.objects.filter(
        is_flagged=True
    ).select_related('device', 'flagged_by', 'operator').order_by('-flagged_at')[:20]

    # ── Inspeksi terbaru ─────────────────────────────────────────
    recent = Inspection.objects.select_related('device','device__jenis','operator').order_by('-tanggal')[:15]

    from devices.models import ULTG as ULTGModel
    all_ultg = ULTGModel.objects.order_by('nama')
    BULAN_ID = ['Januari','Februari','Maret','April','Mei','Juni',
                'Juli','Agustus','September','Oktober','November','Desember']
    month_choices = [{'val': i+1, 'label': n} for i,n in enumerate(BULAN_ID)]
    first_year = Inspection.objects.order_by('tanggal').values_list('tanggal__year', flat=True).first() or today.year
    year_choices = list(range(first_year, today.year + 1))

    return render(request, 'inspection/dashboard.html', {
        'today':                   today,
        'insp_today':              insp_today,
        'insp_bulan':              insp_bulan,
        'insp_total':              insp_total,
        'alarm_bulan':             alarm_bulan,
        'total_device_inspectable':total_device_inspectable,
        'sudah_bulan':             sudah_bulan,
        'pct_bulan':               pct_bulan,
        'ultg_stats':              ultg_stats,
        'jenis_stats':             jenis_stats,
        'trend_30':                trend_30,
        'alarm_list':              alarm_list,
        'flagged_list':            flagged_list,
        'recent':                  recent,
        'all_ultg':                all_ultg,
        'month_choices':           month_choices,
        'year_choices':            year_choices,
    })


# EXPORT LAPORAN EXCEL
# ─────────────────────────────────────────────────────────────────────
@login_required
@require_operator
def inspection_export(request):
    """Export laporan inspeksi ke Excel."""
    lokasi    = request.GET.get('lokasi', '')
    jenis     = request.GET.get('jenis', '')
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')

    qs = Inspection.objects.select_related(
        'device', 'device__jenis', 'operator'
    ).order_by('-tanggal')

    if lokasi:    qs = qs.filter(device__lokasi=lokasi)
    if jenis:     qs = qs.filter(jenis=jenis)
    if date_from: qs = qs.filter(tanggal__date__gte=date_from)
    if date_to:   qs = qs.filter(tanggal__date__lte=date_to)

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'Laporan Inspeksi'

    hdr_fill = PatternFill('solid', fgColor='0F172A')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    hdr_aln  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c_aln    = Alignment(horizontal='center', vertical='center')
    l_aln    = Alignment(vertical='center', wrap_text=True)
    thin     = Side(style='thin')
    brd      = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill('solid', fgColor='F8FAFC')

    # Title
    ws.merge_cells('A1:J1')
    ws['A1'].value = 'LAPORAN INSPEKSI INSERVICE — FASOP UP2B MAKASSAR'
    ws['A1'].font  = Font(bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill  = PatternFill('solid', fgColor='EFF6FF')
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:J2')
    ws['A2'].value = f"Dicetak: {date.today().strftime('%d %B %Y')}"
    ws['A2'].font  = Font(size=10, italic=True, color='64748B')
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['No','Tanggal','Perangkat','Lokasi','Jenis','Operator',
               'Kondisi Utama','Nilai Utama','Status','Catatan']
    widths  = [5, 14, 25, 20, 20, 18, 18, 18, 14, 30]

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = hdr_aln; cell.border = brd
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[4].height = 20
    ws.freeze_panes = 'A5'

    for ri, insp in enumerate(qs, 1):
        wr = ri + 4
        # Ambil kondisi dan nilai utama berdasarkan jenis
        kondisi_utama = nilai_utama = status = ''
        try:
            if insp.jenis == 'catu_daya':
                d = insp.detail_catu_daya
                kondisi_utama = d.get_kondisi_rectifier_display() if d.kondisi_rectifier else '—'
                nilai_utama   = f"Vload: {d.tegangan_load_dc or '—'} V"
                status        = 'ALARM' if d.kondisi_rectifier == 'alarm' else 'Normal'
            elif insp.jenis == 'defense_scheme':
                d = insp.detail_defense_scheme
                kondisi_utama = d.get_kondisi_relay_display() if d.kondisi_relay else '—'
                nilai_utama   = f"DC: {d.sumber_dc or '—'} V"
                status        = 'ALARM' if d.kondisi_relay == 'alarm' else \
                                ('Tidak Normal' if d.status_indikator == 'tidak_normal' else 'Normal')
            elif insp.jenis == 'master_trip':
                d = insp.detail_master_trip
                kondisi_utama = d.get_kondisi_relay_display() if d.kondisi_relay else '—'
                nilai_utama   = f"DC: {d.sumber_dc or '—'} V"
                status        = 'ALARM' if d.kondisi_relay == 'alarm' else \
                                ('Tidak Normal' if d.indikator_led == 'tidak_normal' else 'Normal')
            elif insp.jenis == 'ufls':
                d = insp.detail_ufls
                kondisi_utama = d.get_kondisi_relay_display() if d.kondisi_relay else '—'
                nilai_utama   = f"DC: {d.sumber_dc or '—'} V"
                status        = 'ALARM' if d.kondisi_relay == 'alarm' else \
                                ('Tidak Normal' if d.indikator_led == 'tidak_normal' else 'Normal')
            elif insp.jenis == 'dfr':
                d = insp.detail_dfr
                kondisi_utama = d.get_kondisi_dfr_display() if d.kondisi_dfr else '—'
                nilai_utama   = d.get_healthy_status_display() if d.healthy_status else '—'
                status        = 'ALARM' if _is_alarm_inspection(insp) else 'Normal'
            elif insp.jenis == 'server_ads':
                d = insp.detail_server_ads
                kondisi_utama = d.get_peralatan_server_ads_display() if d.peralatan_server_ads else '—'
                nilai_utama   = d.get_tampilan_hmi_display() if d.tampilan_hmi else '—'
                status        = 'ALARM' if _is_alarm_inspection(insp) else 'Normal'
            elif insp.jenis == 'telecom':
                d = insp.detail_telecom
                kondisi_utama = d.get_hasil_komunikasi_display() if d.hasil_komunikasi else '—'
                nilai_utama   = d.get_kualitas_suara_display() if d.kualitas_suara else '—'
                status        = 'ALARM' if _is_alarm_inspection(insp) else 'Normal'
        except Exception:
            kondisi_utama = nilai_utama = status = '—'

        op_name = insp.operator.get_full_name() or insp.operator.username if insp.operator else '—'
        row_data = [
            ri,
            insp.tanggal.strftime('%d/%m/%Y %H:%M'),
            str(insp.device),
            insp.device.lokasi or '—',
            insp.get_jenis_display(),
            op_name,
            kondisi_utama,
            nilai_utama,
            status,
            insp.catatan or '—',
        ]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=wr, column=ci, value=val)
            cell.border    = brd
            cell.alignment = c_aln if ci in [1,2,9] else l_aln
            if ci == 9 and status in ('ALARM', 'Tidak Normal'):
                cell.fill = PatternFill('solid', fgColor='FEE2E2')
                cell.font = Font(bold=True, color='C00000', size=10)
            elif ri % 2 == 0:
                cell.fill = alt_fill
        ws.row_dimensions[wr].height = 18

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="Laporan_Inspeksi_{date.today().strftime("%Y%m%d")}.xlsx"'
    wb.save(resp)
    return resp


# ─────────────────────────────────────────────────────────────────────
# PENGUJIAN TELEKOMUNIKASI
# ─────────────────────────────────────────────────────────────────────

def _require_dispatcher_or_above(view_func):
    """Dispatcher, AM, superuser."""
    from functools import wraps
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            from django.contrib.auth.views import redirect_to_login
            return redirect_to_login(request.get_full_path())
        if request.user.is_superuser:
            return view_func(request, *args, **kwargs)
        try:
            role = request.user.profile.role
        except Exception:
            role = ''
        if role in ('dispatcher', 'asisten_manager'):
            return view_func(request, *args, **kwargs)
        return render(request, '403.html',
                      {'message': 'Halaman ini hanya untuk Dispatcher / Asisten Manager.'}, status=403)
    return wrapper


def _get_branch_lokasi(user):
    """Return list of lokasi names scoped to dispatcher's branch; None = no filter."""
    try:
        profile = user.profile
        if profile.role == 'dispatcher' and profile.branch:
            return list(
                SiteLocation.objects
                .filter(branch=profile.branch)
                .values_list('nama', flat=True)
            )
    except Exception:
        pass
    return None  # no branch filter


def _filter_by_branch(qs, user):
    """Apply branch-based lokasi filter for dispatcher users."""
    lokasi_list = _get_branch_lokasi(user)
    if lokasi_list is not None:
        qs = qs.filter(lokasi__in=lokasi_list)
    return qs


def _pengujian_form_impl(request, jenis):
    """Shared implementation for pengujian form (radio / voip / all)."""
    radio_qs = (
        Device.objects
        .filter(is_deleted=False, jenis__name='Radio')
        .select_related('jenis')
        .order_by('lokasi', 'nama')
    )
    voip_qs = (
        Device.objects
        .filter(is_deleted=False, jenis__name='VoIP')
        .select_related('jenis')
        .order_by('lokasi', 'nama')
    )
    # Filter by dispatcher's branch
    radio_qs = _filter_by_branch(radio_qs, request.user)
    voip_qs  = _filter_by_branch(voip_qs, request.user)

    # Jenis-specific selection
    if jenis == 'radio':
        active_devices = list(radio_qs)
        radio_devices  = active_devices
        voip_devices   = []
    elif jenis == 'voip':
        active_devices = list(voip_qs)
        radio_devices  = []
        voip_devices   = active_devices
    else:  # 'all'
        radio_devices  = list(radio_qs)
        voip_devices   = list(voip_qs)
        active_devices = radio_devices + voip_devices

    if request.method == 'POST':
        tanggal = request.POST.get('tanggal') or date.today()
        catatan = request.POST.get('catatan', '')
        pengujian = PengujianTelecom.objects.create(
            jenis=jenis,
            tanggal=tanggal,
            lokasi='',
            dibuat_oleh=request.user,
            catatan=catatan,
        )
        for d in active_devices:
            hasil = request.POST.get(f'hasil_{d.pk}', 'normal')
            cat   = request.POST.get(f'catatan_{d.pk}', '')
            PengujianTelecomItem.objects.create(
                pengujian=pengujian,
                device=d,
                hasil=hasil,
                catatan=cat,
            )
        from django.contrib import messages
        messages.success(request, 'Pengujian berhasil disimpan.')
        return redirect('pengujian_telecom_detail', pk=pengujian.pk)

    # Branch info for display
    try:
        branch_nama = request.user.profile.branch.nama if request.user.profile.branch else None
    except Exception:
        branch_nama = None

    return render(request, 'inspection/pengujian_telecom_form.html', {
        'radio_devices': radio_devices,
        'voip_devices':  voip_devices,
        'jenis':         jenis,
        'branch_nama':   branch_nama,
        'today':         date.today().isoformat(),
    })


@login_required
@_require_dispatcher_or_above
def pengujian_telecom_form(request):
    return _pengujian_form_impl(request, jenis='all')


@login_required
@_require_dispatcher_or_above
def pengujian_radio_form(request):
    return _pengujian_form_impl(request, jenis='radio')


@login_required
@_require_dispatcher_or_above
def pengujian_voip_form(request):
    return _pengujian_form_impl(request, jenis='voip')


@login_required
def pengujian_telecom_dashboard(request):
    """Dashboard hasil pengujian telekomunikasi."""
    from django.db.models import Count, Q as Qlocal
    from django.utils import timezone as tz
    from datetime import timedelta

    today      = date.today()
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)

    base_qs = PengujianTelecom.objects.all()
    try:
        role = request.user.profile.role
    except Exception:
        role = ''
    if role == 'dispatcher':
        base_qs = base_qs.filter(dibuat_oleh=request.user)

    def _stats(qs):
        agg = qs.annotate(
            tn=Count('items', filter=Qlocal(items__hasil='tidak_normal'))
        ).aggregate(
            total_sesi=Count('id'),
            total_perangkat=Count('items'),
            total_tidak_normal=Count('items', filter=Qlocal(items__hasil='tidak_normal')),
        )
        return agg

    stats_hari_ini  = _stats(base_qs.filter(tanggal=today))
    stats_minggu    = _stats(base_qs.filter(tanggal__gte=week_start))
    stats_bulan     = _stats(base_qs.filter(tanggal__gte=month_start))
    stats_semua     = _stats(base_qs)

    # Recent 10 sessions with annotation
    recent = (
        base_qs
        .select_related('dibuat_oleh')
        .annotate(
            total_device=Count('items'),
            total_tidak_normal=Count('items', filter=Qlocal(items__hasil='tidak_normal')),
        )
        .order_by('-tanggal', '-created_at')[:10]
    )

    # Per-jenis breakdown (bulan ini)
    radio_bulan = _stats(base_qs.filter(tanggal__gte=month_start, jenis='radio'))
    voip_bulan  = _stats(base_qs.filter(tanggal__gte=month_start, jenis='voip'))
    all_bulan   = _stats(base_qs.filter(tanggal__gte=month_start, jenis='all'))

    return render(request, 'inspection/pengujian_telecom_dashboard.html', {
        'stats_hari_ini': stats_hari_ini,
        'stats_minggu':   stats_minggu,
        'stats_bulan':    stats_bulan,
        'stats_semua':    stats_semua,
        'radio_bulan':    radio_bulan,
        'voip_bulan':     voip_bulan,
        'all_bulan':      all_bulan,
        'recent':         recent,
        'today':          today,
    })


@login_required
def pengujian_telecom_list(request):
    """Daftar semua sesi pengujian, bisa difilter by jenis."""
    from django.db.models import Count, Q as Qlocal
    jenis_filter = request.GET.get('jenis', '')

    qs = (
        PengujianTelecom.objects
        .select_related('dibuat_oleh')
        .annotate(
            total_device=Count('items'),
            total_tidak_normal=Count('items', filter=Qlocal(items__hasil='tidak_normal')),
        )
        .all()
    )
    try:
        role = request.user.profile.role
    except Exception:
        role = ''
    if role == 'dispatcher':
        qs = qs.filter(dibuat_oleh=request.user)
    if jenis_filter in ('radio', 'voip', 'all'):
        qs = qs.filter(jenis=jenis_filter)

    return render(request, 'inspection/pengujian_telecom_list.html', {
        'pengujian_list': qs,
        'jenis_filter':   jenis_filter,
    })


@login_required
def pengujian_telecom_detail(request, pk):
    """Detail satu sesi pengujian."""
    pengujian = get_object_or_404(PengujianTelecom, pk=pk)
    items = pengujian.items.select_related('device', 'device__jenis').all()

    radio_items = [i for i in items if i.device.jenis.name == 'Radio']
    voip_items  = [i for i in items if i.device.jenis.name == 'VoIP']

    return render(request, 'inspection/pengujian_telecom_detail.html', {
        'pengujian':   pengujian,
        'radio_items': radio_items,
        'voip_items':  voip_items,
    })


@login_required
@_require_dispatcher_or_above
def pengujian_telecom_delete(request, pk):
    """Hapus satu sesi pengujian (POST only)."""
    pengujian = get_object_or_404(PengujianTelecom, pk=pk)
    if request.method == 'POST':
        pengujian.delete()
        from django.contrib import messages
        messages.success(request, 'Data pengujian berhasil dihapus.')
        return redirect('pengujian_telecom_list')
    return redirect('pengujian_telecom_detail', pk=pk)
