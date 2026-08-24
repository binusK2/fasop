"""Tes laporan inspeksi harian: kolom per jenis peralatan, penyaringan
perangkat Tidak Operasi, halaman /inspection/harian/, dan arsip Excel-nya."""
import io
import os
import tempfile
from datetime import timedelta

from django.contrib.auth.models import User
from django.core.management import call_command
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone

from devices.models import Device, DeviceType

from .laporan import (KOLOM_JENIS, JENIS_URUT, baris_harian, kolom_jenis,
                      ringkasan_baris, workbook_harian)
from .models import Inspection, InspectionCatuDaya, InspectionDFR


def _siapkan_profil(user, role='operator'):
    """Profil siap pakai: role diisi dan tidak dipaksa ganti password
    (ForcePasswordChangeMiddleware me-redirect user baru ke /ganti-password/)."""
    from devices.models import UserProfile
    profil, _ = UserProfile.objects.get_or_create(user=user)
    profil.role = role
    profil.force_password_change = False
    profil.save()
    return profil


class SkemaKolomTests(TestCase):
    """Kolom laporan harus benar-benar ada di model detail jenis itu."""

    MODEL = {
        'catu_daya':      InspectionCatuDaya,
        'dfr':            InspectionDFR,
    }

    def test_setiap_jenis_punya_kolom(self):
        for jenis_key in JENIS_URUT:
            self.assertTrue(kolom_jenis(jenis_key), f'{jenis_key} tanpa kolom')

    def test_field_kolom_ada_di_model(self):
        for jenis_key, model in self.MODEL.items():
            nama_field = {f.name for f in model._meta.get_fields()}
            for kol in KOLOM_JENIS[jenis_key]:
                self.assertIn(kol.field, nama_field,
                              f'{jenis_key}: {kol.field} tidak ada di {model.__name__}')

    def test_nilai_ok_adalah_pilihan_sah(self):
        for jenis_key, model in self.MODEL.items():
            for kol in KOLOM_JENIS[jenis_key]:
                if not kol.ok:
                    continue
                pilihan = [v for v, _ in model._meta.get_field(kol.field).choices or []]
                self.assertIn(kol.ok, pilihan,
                              f'{jenis_key}.{kol.field}: ok={kol.ok!r} bukan pilihan sah')


class LaporanHarianTests(TestCase):
    def setUp(self):
        self.hari_ini = timezone.localdate()
        self.operator = User.objects.create_user(username='operator', password='rahasia')
        _siapkan_profil(self.operator, role='operator')
        self.jenis_cd = DeviceType.objects.create(name='Catu Daya')

        self.dev = Device.objects.create(
            nama='CATU DAYA GI TELLO', jenis=self.jenis_cd, merk='X', lokasi='GI TELLO')
        self.dev_belum = Device.objects.create(
            nama='CATU DAYA GI DAYA', jenis=self.jenis_cd, merk='X', lokasi='GI DAYA')
        self.dev_mati = Device.objects.create(
            nama='CATU DAYA BEKAS', jenis=self.jenis_cd, merk='X', lokasi='GI TELLO',
            status_operasi='tidak_operasi')

        insp = Inspection.objects.create(
            device=self.dev, jenis='catu_daya', operator=self.operator)
        InspectionCatuDaya.objects.create(
            inspection=insp, kondisi_rectifier='alarm', kebersihan_ruangan='bersih',
            tegangan_load_dc=48.2)
        self.insp = insp

    def test_perangkat_tidak_operasi_tidak_masuk_laporan(self):
        nama = [b['nama'] for b in baris_harian(self.hari_ini, 'catu_daya')]
        self.assertIn(self.dev.nama, nama)
        self.assertIn(self.dev_belum.nama, nama)
        self.assertNotIn(self.dev_mati.nama, nama)

    def test_status_dan_ringkasan(self):
        baris = baris_harian(self.hari_ini, 'catu_daya')
        per_nama = {b['nama']: b for b in baris}
        self.assertEqual(per_nama[self.dev.nama]['status'], 'alarm')
        self.assertEqual(per_nama[self.dev_belum.nama]['status'], 'belum')

        r = ringkasan_baris(baris)
        self.assertEqual((r['total'], r['sudah'], r['belum'], r['alarm']), (2, 1, 1, 1))

    def test_nilai_kolom_terisi_sesuai_field_jenisnya(self):
        baris = baris_harian(self.hari_ini, 'catu_daya')
        b = next(x for x in baris if x['nama'] == self.dev.nama)
        nilai = {s['label']: s for s in b['sel']}
        self.assertEqual(nilai['Kondisi Rectifier']['display'], 'Alarm')
        self.assertEqual(nilai['Kondisi Rectifier']['status'], 'alarm')
        self.assertEqual(nilai['Kebersihan Ruangan']['status'], 'normal')
        self.assertEqual(nilai['Teg. Load DC']['display'], '48.2 V')

    def test_hari_lain_kosong(self):
        r = ringkasan_baris(baris_harian(self.hari_ini - timedelta(days=1), 'catu_daya'))
        self.assertEqual((r['sudah'], r['belum']), (0, 2))

    def test_halaman_harian_menampilkan_tab_dan_baris(self):
        self.client.force_login(self.operator)
        resp = self.client.get(reverse('inspection_harian'))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Catu Daya')
        self.assertContains(resp, self.dev.nama)
        self.assertNotContains(resp, self.dev_mati.nama)
        # tab jenis lain ikut dirender
        self.assertContains(resp, 'Server ADS')

    def test_halaman_harian_menghormati_parameter_tanggal(self):
        self.client.force_login(self.operator)
        kemarin = (self.hari_ini - timedelta(days=1)).isoformat()
        resp = self.client.get(reverse('inspection_harian'),
                               {'tanggal': kemarin, 'jenis': 'dfr'})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context['jenis_aktif'], 'dfr')
        self.assertEqual(resp.context['tanggal_str'], kemarin)

    def test_tanggal_ngawur_jatuh_ke_hari_ini(self):
        self.client.force_login(self.operator)
        resp = self.client.get(reverse('inspection_harian'), {'tanggal': 'bukan-tanggal'})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context['tanggal'], self.hari_ini)

    def test_export_excel_terunduh(self):
        self.client.force_login(self.operator)
        resp = self.client.get(reverse('inspection_harian_export'))
        self.assertEqual(resp.status_code, 200)
        self.assertIn('spreadsheetml', resp['Content-Type'])
        self.assertIn('Inspeksi_Harian_', resp['Content-Disposition'])

    def test_workbook_punya_sheet_setiap_jenis(self):
        wb = workbook_harian(self.hari_ini)
        self.assertEqual(wb.sheetnames[0], 'Ringkasan')
        self.assertEqual(len(wb.sheetnames), len(JENIS_URUT) + 1)

        ws = wb['Catu Daya']
        header = [c.value for c in ws[4] if c.value]
        self.assertEqual(header[:5], ['No', 'Nama Perangkat', 'Lokasi / GI', 'Jam', 'Operator'])
        self.assertIn('Kondisi Rectifier', header)
        # kolom milik jenis lain tidak boleh bocor ke sheet ini
        self.assertNotIn('Healthy Status', header)

        ws_dfr = wb['DFR']
        header_dfr = [c.value for c in ws_dfr[4] if c.value]
        self.assertIn('Healthy Status', header_dfr)
        self.assertNotIn('Kondisi Rectifier', header_dfr)


class ExportInspeksiHarianCommandTests(TestCase):
    def setUp(self):
        self.hari_ini = timezone.localdate()
        jenis = DeviceType.objects.create(name='Catu Daya')
        Device.objects.create(nama='CD-01', jenis=jenis, merk='X', lokasi='GI TELLO')

    def test_menulis_file_per_hari_ke_folder_tujuan(self):
        with tempfile.TemporaryDirectory() as folder:
            call_command('export_inspeksi_harian', '--dir', folder, '--days', '2',
                         stdout=io.StringIO())
            for i in range(2):
                tgl = self.hari_ini - timedelta(days=i)
                path = os.path.join(folder, tgl.strftime('%Y-%m'),
                                    f'Inspeksi_Harian_{tgl.isoformat()}.xlsx')
                self.assertTrue(os.path.exists(path), f'{path} tidak dibuat')
                self.assertGreater(os.path.getsize(path), 0)

    def test_dry_run_tidak_menulis_apa_apa(self):
        with tempfile.TemporaryDirectory() as folder:
            call_command('export_inspeksi_harian', '--dir', folder, '--dry-run',
                         stdout=io.StringIO())
            self.assertEqual(os.listdir(folder), [])

    def test_menjalankan_ulang_menimpa_file_yang_sama(self):
        with tempfile.TemporaryDirectory() as folder:
            call_command('export_inspeksi_harian', '--dir', folder, stdout=io.StringIO())
            call_command('export_inspeksi_harian', '--dir', folder, stdout=io.StringIO())
            bulan = os.path.join(folder, self.hari_ini.strftime('%Y-%m'))
            self.assertEqual(len(os.listdir(bulan)), 1)

class ExportRekapUltgTests(TestCase):
    """Rekap bulanan per ULTG memakai skema kolom yang sama dengan laporan harian."""

    def setUp(self):
        self.admin = User.objects.create_superuser(username='admin', password='rahasia')
        _siapkan_profil(self.admin, role='asisten_manager')
        jenis_cd  = DeviceType.objects.create(name='Catu Daya')
        jenis_dfr = DeviceType.objects.create(name='DFR')
        Device.objects.create(nama='CD-01', jenis=jenis_cd, merk='X', lokasi='GI TELLO')
        Device.objects.create(nama='DFR-01', jenis=jenis_dfr, merk='X', lokasi='GI TELLO')

    def test_sheet_per_jenis_dan_kolomnya(self):
        from openpyxl import load_workbook

        self.client.force_login(self.admin)
        resp = self.client.get(reverse('inspection_export_ultg'))
        self.assertEqual(resp.status_code, 200)

        wb = load_workbook(io.BytesIO(resp.content))
        # DFR dulu tidak pernah ikut diekspor
        self.assertIn('DFR', wb.sheetnames)
        self.assertIn('Catu Daya', wb.sheetnames)

        header = [c.value for c in wb['DFR'][4] if c.value]
        self.assertIn('Tgl Inspeksi', header)      # rekap bulanan pakai tanggal, bukan jam
        self.assertIn('Healthy Status', header)    # kolom milik DFR, bukan kolom rele
        self.assertNotIn('Kondisi Rele', header)
