import logging
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse

from .models import ScadaAvSession, ScadaAvFile
from .forms import ScadaAvUploadForm
from .calculator import run_full_calculation

logger = logging.getLogger(__name__)


# ── List ─────────────────────────────────────────────────────────────────────

@login_required
def scada_av_list(request):
    sessions = ScadaAvSession.objects.select_related('dibuat_oleh').prefetch_related('files')
    return render(request, 'scada_av/list.html', {'sessions': sessions})


# ── Upload & Hitung ───────────────────────────────────────────────────────────

@login_required
def scada_av_upload(request):
    if request.method == 'POST':
        form = ScadaAvUploadForm(request.POST, request.FILES)
        if form.is_valid():
            cd = form.cleaned_data

            session = ScadaAvSession.objects.create(
                nama         = cd['nama'],
                keterangan   = cd.get('keterangan', ''),
                periode_awal  = cd['periode_awal'],
                periode_akhir = cd['periode_akhir'],
                master        = cd['master'],
                input_type    = cd['input_type'],
                calc_type     = cd['calc_type'],
                dibuat_oleh   = request.user,
                status        = 'pending',
            )

            # Simpan semua file yang diupload
            uploaded_files = request.FILES.getlist('files')
            for f in uploaded_files:
                ScadaAvFile.objects.create(
                    session  = session,
                    file     = f,
                    filename = f.name,
                    ukuran   = f.size,
                )

            # Jalankan kalkulasi langsung (synchronous)
            try:
                run_full_calculation(session.pk)
                messages.success(request, f'Kalkulasi selesai dalam {session.durasi_hitung:.1f}s.')
            except Exception as exc:
                session.refresh_from_db()
                messages.error(request, f'Kalkulasi gagal: {session.error_message or str(exc)}')

            return redirect('scada_av_detail', pk=session.pk)
    else:
        form = ScadaAvUploadForm()

    return render(request, 'scada_av/upload.html', {'form': form})


# ── Detail ────────────────────────────────────────────────────────────────────

@login_required
def scada_av_detail(request, pk):
    session = get_object_or_404(
        ScadaAvSession.objects.prefetch_related('files', 'rtu_results', 'rcd_bay_results'),
        pk=pk,
    )

    rtu_results = session.rtu_results.all() if session.has_rtu else None
    rcd_summary = getattr(session, 'rcd_summary', None) if session.has_rcd else None
    rcd_bays    = session.rcd_bay_results.all() if session.has_rcd else None

    # Summary RTU
    rtu_avg_av = None
    if rtu_results and rtu_results.exists():
        vals = [r.rtu_availability for r in rtu_results]
        rtu_avg_av = round(sum(vals) / len(vals) * 100, 2)

    ctx = {
        'session':    session,
        'rtu_results': rtu_results,
        'rcd_summary': rcd_summary,
        'rcd_bays':    rcd_bays,
        'rtu_avg_av':  rtu_avg_av,
    }
    return render(request, 'scada_av/detail.html', ctx)


# ── Download Excel ────────────────────────────────────────────────────────────

@login_required
def scada_av_download(request, pk):
    """Buat ulang file Excel dari data yang sudah tersimpan di DB."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    session = get_object_or_404(ScadaAvSession, pk=pk)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # hapus sheet kosong

    header_fill  = PatternFill('solid', fgColor='1E3A5F')
    header_font  = Font(bold=True, color='FFFFFF', size=11)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin         = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'),
    )
    alt_fill = PatternFill('solid', fgColor='F0F9FF')

    def _apply_header(ws, headers, widths, row=1):
        for ci, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = header_font; c.fill = header_fill
            c.alignment = center_align; c.border = thin
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[row].height = 22

    def _title_row(ws, text, ncols):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = ws['A1']
        c.value = text
        c.font  = Font(bold=True, size=13)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill  = PatternFill('solid', fgColor='EFF6FF')
        ws.row_dimensions[1].height = 26

    # ── Sheet RTU ─────────────────────────────────────────────────────────────
    if session.has_rtu and session.rtu_results.exists():
        ws = wb.create_sheet('RTU Availability')
        _title_row(ws, f'RTU AVAILABILITY — {session.nama}  ({session.periode_awal} s/d {session.periode_akhir})', 9)
        headers = ['No','RTU','Nama Lengkap','Jumlah Downtime','Total Downtime',
                   'RTU AV (%)','Link AV (%)','Overall (%)','RTU Downtime']
        widths  = [5,  15,   25,             16,              16,
                   13,         13,            13,             16]
        _apply_header(ws, headers, widths, row=2)
        ws.freeze_panes = 'A3'
        for i, r in enumerate(session.rtu_results.all(), 1):
            row = i + 2
            vals = [i, r.rtu, r.long_name, r.downtime_occurences, r.total_downtime_hms,
                    r.rtu_av_pct, r.link_av_pct, r.overall_pct, _fmt_hms(r.rtu_downtime_s)]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.border = thin
                c.alignment = center_align if ci in (1,4,5,6,7,8,9) else Alignment(vertical='center')
                if i % 2 == 0:
                    c.fill = alt_fill
            ws.row_dimensions[row].height = 18

    # ── Sheet RCD Bay ─────────────────────────────────────────────────────────
    if session.has_rcd and session.rcd_bay_results.exists():
        ws2 = wb.create_sheet('RCD Bay')
        _title_row(ws2, f'RCD SUCCESS RATE (BAY) — {session.nama}', 12)
        h2 = ['No','GI (B1)','B2','Bay (B3)','Jumlah RC','Sukses','Gagal',
              'Success Rate (%)','Open Sukses','Open Gagal','Close Sukses','Close Gagal']
        w2 = [5,  20,       12,  25,          12,          10,      10,
              16,              12,            12,            13,           13]
        _apply_header(ws2, h2, w2, row=2)
        ws2.freeze_panes = 'A3'
        for i, b in enumerate(session.rcd_bay_results.all(), 1):
            row = i + 2
            vals = [i, b.station, b.bay_b2, b.bay_b3, b.occurences, b.success, b.failed,
                    b.success_pct, b.open_success, b.open_failed, b.close_success, b.close_failed]
            for ci, v in enumerate(vals, 1):
                c = ws2.cell(row=row, column=ci, value=v)
                c.border = thin
                c.alignment = center_align if ci in (1,5,6,7,8,9,10,11,12) else Alignment(vertical='center')
                if i % 2 == 0:
                    c.fill = alt_fill
            ws2.row_dimensions[row].height = 18

    # ── Sheet RCD Summary ─────────────────────────────────────────────────────
    try:
        rcd_sum = session.rcd_summary
        ws3 = wb.create_sheet('RCD Summary')
        _title_row(ws3, f'RCD SUMMARY — {session.nama}', 2)
        rows_data = [
            ('Total RC', rcd_sum.total_count),
            ('Valid RC', rcd_sum.total_valid),
            ('Sukses', rcd_sum.total_success),
            ('Gagal', rcd_sum.total_failed),
            ('Repetisi', rcd_sum.total_reps),
            ('Unused', rcd_sum.total_marked_unused),
            ('Success Rate', f'{rcd_sum.success_pct}%'),
            ('Close Success Rate', f'{rcd_sum.success_close_pct}%'),
            ('Open Success Rate', f'{rcd_sum.success_open_pct}%'),
        ]
        for i, (label, val) in enumerate(rows_data, 1):
            r = i + 1
            ws3.cell(row=r, column=1, value=label).border = thin
            ws3.cell(row=r, column=2, value=val).border = thin
            ws3.row_dimensions[r].height = 18
        ws3.column_dimensions['A'].width = 22
        ws3.column_dimensions['B'].width = 18
    except Exception:
        pass

    if not wb.sheetnames:
        wb.create_sheet('Kosong')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fname = f'scada_av_{session.pk}_{session.periode_awal}_{session.periode_akhir}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    wb.save(response)
    return response


def _fmt_hms(seconds: float) -> str:
    total = int(seconds)
    h, rem = divmod(total, 3600)
    m, s   = divmod(rem, 60)
    return f'{h:02d}:{m:02d}:{s:02d}'
