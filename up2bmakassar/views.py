from datetime import datetime, timedelta, time

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from . import ofdb
from .models import KinerjaAnalogHarian, KinerjaDigitalHarian, RemoteControl, SitePath1


def _parse_date(request, name, default):
    raw = request.GET.get(name)
    if raw:
        try:
            return datetime.strptime(raw, '%Y-%m-%d').date()
        except ValueError:
            pass
    return default


def _rentang(request):
    """
    Rentang tanggal halaman kinerja. Default: awal bulan berjalan s/d kemarin
    (hari berjalan belum lengkap datanya, sync jalan H+1).
    """
    kemarin = timezone.localdate() - timedelta(days=1)
    dari = _parse_date(request, 'tanggal_dari', kemarin.replace(day=1))
    sampai = _parse_date(request, 'tanggal_sampai', kemarin)
    if sampai < dari:
        sampai = dari
    return dari, sampai


def durasi(detik):
    """Format detik -> 'dd:hh:mm:ss', sama seperti kolom durasi di app up2bmakassar."""
    try:
        d = timedelta(seconds=float(detik or 0))
    except (TypeError, ValueError):
        return '00:00:00:00'
    return f'{d.days:02d}:{d.seconds // 3600:02d}:{(d.seconds // 60) % 60:02d}:{d.seconds % 60:02d}'


def _filter_tampilan(qs):
    """
    Sembunyikan site yang dinonaktifkan di admin (SitePath1.aktif=False).

    Ini filter TAMPILAN saja -- semua titik tetap dihitung & tersimpan oleh sync,
    jadi mengaktifkan kembali sebuah site langsung memunculkan datanya tanpa perlu
    sync ulang. (Dulu filternya ada di sisi sync, akibatnya site yang dimatikan
    tidak punya data sama sekali.)
    """
    nonaktif = {
        (v or '').strip()
        for v in SitePath1.objects.filter(aktif=False).values_list('path1', flat=True)
    } - {''}
    return qs.exclude(path1__in=list(nonaktif)) if nonaktif else qs


def _avail(uptime, alltime):
    """Availability = total uptime / total normal time * 100 (bukan rata-rata persen harian)."""
    return round(uptime / alltime * 100, 2) if alltime else 0


def _lengkapi(row):
    """Tambahkan downtime + kolom durasi terformat + availability ke satu baris agregat."""
    uptime = float(row.get('uptime') or 0)
    alltime = float(row.get('alltime') or 0)
    if 'path1' in row:
        row['b1'] = row.get('b1') or row['path1']
    row['uptime'] = uptime
    row['alltime'] = alltime
    row['downtime'] = max(alltime - uptime, 0)
    row['durasi_uptime'] = durasi(uptime)
    row['durasi_downtime'] = durasi(row['downtime'])
    row['durasi_normal'] = durasi(alltime)
    row['avail'] = _avail(uptime, alltime)
    return row


def _ringkasan(rows, jumlah_titik):
    total_uptime = sum(r['uptime'] for r in rows)
    total_alltime = sum(r['alltime'] for r in rows)
    return {
        'jumlah_titik': jumlah_titik,
        'jumlah_station': len(rows),
        'avail': _avail(total_uptime, total_alltime),
        'durasi_uptime': durasi(total_uptime),
        'durasi_downtime': durasi(max(total_alltime - total_uptime, 0)),
        'durasi_normal': durasi(total_alltime),
    }


@login_required
def dashboard(request):
    from django.db.models import Count, Sum

    tanggal = timezone.localdate() - timedelta(days=1)

    def _rekap(model, jenis):
        qs = _filter_tampilan(model.objects.filter(tanggal=tanggal, jenis=jenis))
        agg = qs.aggregate(
            titik=Count('point_number', distinct=True),
            uptime=Sum('uptime_detik'),
            alltime=Sum('alltime_detik'),
        )
        station = [
            _lengkapi(r) for r in qs.values('path1', 'b1').annotate(
                uptime=Sum('uptime_detik'), alltime=Sum('alltime_detik')
            )
        ]
        station.sort(key=lambda r: r['avail'])
        return {
            'jumlah_titik': agg['titik'] or 0,
            'avail': _avail(float(agg['uptime'] or 0), float(agg['alltime'] or 0)),
            'terbaik': station[-1] if station else None,
            'terburuk': station[0] if station else None,
        }

    return render(request, 'up2bmakassar/dashboard.html', {
        'tanggal': tanggal,
        'ringkasan_analog': _rekap(KinerjaAnalogHarian, ofdb.JENIS_TELEMETERING),
        'ringkasan_digital': _rekap(KinerjaDigitalHarian, ofdb.JENIS_TELESIGNAL),
    })


# ── Halaman kinerja Telemetering / Telesignal ────────────────────────────────────
#
# Bentuk tampilannya mengikuti app up2bmakassar: tabel REKAP per station (B1) —
# Station | Point | Normal Time | Uptime | Downtime | Avail (%) — dan kalau satu
# station dipilih, tabel DETAIL per titik (Point Number | B1 | B2 | B3 | Element |
# Up | Uptime | Downtime | Avail). Availability dihitung dari SUM(uptime)/SUM(alltime),
# sama seperti serializer rekap di up2bmakassar, bukan rata-rata kolom performance.

REKAP_HEADERS = ['Station', 'Jumlah Point', 'Normal Time (dd:hh:mm:ss)',
                 'Uptime (dd:hh:mm:ss)', 'Downtime (dd:hh:mm:ss)', 'Avail (%)']
REKAP_FIELDS = ['b1', 'poin', 'durasi_normal', 'durasi_uptime', 'durasi_downtime', 'avail']

DETAIL_HEADERS = ['Point Number', 'B1', 'B2', 'B3', 'Element', 'Up',
                  'Normal Time (dd:hh:mm:ss)', 'Uptime (dd:hh:mm:ss)',
                  'Downtime (dd:hh:mm:ss)', 'Avail (%)']
DETAIL_FIELDS = ['point_number', 'b1', 'b2', 'b3', 'elem', 'up',
                 'durasi_normal', 'durasi_uptime', 'durasi_downtime', 'avail']


def _kinerja_ctx(request, model, jenis):
    from django.db.models import Count, Q, Sum

    dari, sampai = _rentang(request)
    q = request.GET.get('q', '').strip()
    station = request.GET.get('station', '').strip()

    qs = model.objects.filter(jenis=jenis, tanggal__gte=dari, tanggal__lte=sampai)

    qs = _filter_tampilan(qs)

    if q:
        qs = qs.filter(
            Q(b1__icontains=q) | Q(b2__icontains=q) | Q(b3__icontains=q) | Q(elem__icontains=q)
            | Q(path1__icontains=q) | Q(path2__icontains=q) | Q(path3__icontains=q)
        )

    rekap = [
        _lengkapi(r) for r in qs.values('path1', 'b1').annotate(
            poin=Count('point_number', distinct=True),
            uptime=Sum('uptime_detik'),
            alltime=Sum('alltime_detik'),
        ).order_by('b1', 'path1')
    ]

    jumlah_titik = qs.values('point_number').distinct().count()

    detail = []
    if station:
        detail = [
            _lengkapi(r) for r in qs.filter(path1=station).values(
                'point_number', 'b1', 'b2', 'b3', 'elem'
            ).annotate(
                up=Sum('jumlah_up'),
                uptime=Sum('uptime_detik'),
                alltime=Sum('alltime_detik'),
            ).order_by('b2', 'b3', 'elem', 'point_number')
        ]

    return {
        'jenis': jenis,
        'jenis_pilihan': ofdb.JENIS_DIGITAL,
        'tanggal_dari': dari, 'tanggal_sampai': sampai,
        'q': q, 'station': station,
        'station_label': detail[0]['b1'] if detail else station,
        'rekap': rekap, 'detail': detail,
        'ringkasan': _ringkasan(rekap, jumlah_titik),
        'rekap_headers': REKAP_HEADERS, 'detail_headers': DETAIL_HEADERS,
    }


def _xlsx(sheets, filename):
    """
    Bangun file xlsx dari beberapa sheet dan kembalikan sebagai HttpResponse.
    `sheets` = list of (judul, headers, fields, rows-dict). Sheet tanpa baris dilewati.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    header_fill = PatternFill('solid', fgColor='1D4ED8')
    header_font = Font(bold=True, color='FFFFFF')
    pertama = True

    for judul, headers, fields, rows in sheets:
        if not rows and not pertama:
            continue
        ws = wb.active if pertama else wb.create_sheet()
        ws.title = judul
        pertama = False

        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal='center')
        for r_idx, row in enumerate(rows, start=2):
            for c_idx, f in enumerate(fields, start=1):
                ws.cell(row=r_idx, column=c_idx, value=row.get(f))
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


def _kinerja_export(ctx, jenis):
    """Export xlsx: sheet Rekap per station + sheet Detail (kalau ada station dipilih)."""
    return _xlsx(
        [('Rekap', REKAP_HEADERS, REKAP_FIELDS, ctx['rekap']),
         ('Detail', DETAIL_HEADERS, DETAIL_FIELDS, ctx['detail'])],
        f"kinerja_{jenis.lower()}_{ctx['tanggal_dari']}_sd_{ctx['tanggal_sampai']}.xlsx",
    )


def _kinerja_page(request, model, jenis, template, **extra):
    ctx = _kinerja_ctx(request, model, jenis)
    if request.GET.get('export') == '1':
        return _kinerja_export(ctx, jenis)
    ctx.update(extra)
    return render(request, template, ctx)


@login_required
def kinerja_analog(request):
    """Telemetering — titik ANALOG dengan induk point type TELEMETERING."""
    return _kinerja_page(
        request, KinerjaAnalogHarian, ofdb.JENIS_TELEMETERING,
        'up2bmakassar/kinerja_analog.html',
        judul='Kinerja Telemetering',
        subjudul='Titik ANALOG dengan induk point type TELEMETERING (OFDB scd_his_analog)',
        perintah_sync='sync_kinerja_analog',
        pilih_jenis=False,
    )


@login_required
def kinerja_digital(request):
    """
    Telesignal — titik DIGITAL dengan induk point type TELESIGNAL (default).
    ?jenis=RTU|MASTER|TELEKOMUNIKASI untuk jenis digital lain yang ikut disinkron.
    """
    jenis = (request.GET.get('jenis') or ofdb.JENIS_TELESIGNAL).upper()
    if jenis not in ofdb.JENIS_DIGITAL:
        jenis = ofdb.JENIS_TELESIGNAL
    return _kinerja_page(
        request, KinerjaDigitalHarian, jenis,
        'up2bmakassar/kinerja_digital.html',
        judul=f'Kinerja {jenis.title()}',
        subjudul=f'Titik DIGITAL dengan induk point type {jenis} (OFDB scd_his_digital)',
        perintah_sync=f'sync_kinerja_digital --jenis {jenis}',
        pilih_jenis=True,
    )


# ── Kinerja RC — agregat dari RemoteControl (sudah di-resolve oleh sync_rc) ────────
#
# Bentuknya mengikuti app up2bmakassar: rekap per bay dengan kunci B1/B2/B3/
# Element/Info (path1..path5 di scd_his_rc), kolom Jumlah RC | Sukses | Gagal |
# Performance. Klik satu bay -> daftar RC individualnya (waktu, operator, hasil).
#
# CATATAN soal rumus: serializer up2bmakassar menghitung
#     performance = sukses / gagal * 100
# yang keliru -- bay tanpa kegagalan menghasilkan ZeroDivisionError lalu jadi 0%,
# dan 10 sukses : 2 gagal jadi 500%. Di sini dipakai rumus yang benar,
#     performance = sukses / jumlah RC * 100,
# jadi angkanya memang tidak akan sama persis dengan halaman RC app lama.

RC_REKAP_HEADERS = ['B1', 'B2', 'B3', 'Element', 'Info',
                    'Jumlah RC', 'Sukses', 'Gagal', 'Performance (%)']
RC_REKAP_FIELDS = ['b1', 'b2', 'b3', 'elem', 'info',
                   'jumlah', 'sukses', 'gagal', 'performance']

RC_DETAIL_HEADERS = ['Waktu Perintah', 'B1', 'B2', 'B3', 'Element', 'Info',
                     'Operator', 'Status Perintah', 'Waktu Respon', 'Hasil']
RC_DETAIL_FIELDS = ['waktu', 'b1', 'b2', 'b3', 'elem', 'info',
                    'operator', 'status_eksekusi', 'waktu_respon', 'status_respon']

RC_DETAIL_MAX = 2000


def _rc_ctx(request):
    from django.db.models import Case, Count, IntegerField, Q, Sum, When
    from django.db.models.functions import Coalesce

    today = timezone.localdate()
    tanggal_dari = _parse_date(request, 'tanggal_dari', today.replace(day=1))
    tanggal_sampai = _parse_date(request, 'tanggal_sampai', today)
    if tanggal_sampai < tanggal_dari:
        tanggal_sampai = tanggal_dari

    q = request.GET.get('q', '').strip()
    # Bay yang dipilih untuk ditampilkan detailnya (semua kunci grouping).
    pilih = {k: request.GET.get(k, '') for k in ('b1', 'b2', 'b3', 'elem', 'info')}
    ada_pilihan = any(pilih.values())

    qs = RemoteControl.objects.filter(tanggal__gte=tanggal_dari, tanggal__lte=tanggal_sampai)
    if q:
        qs = qs.filter(
            Q(b1__icontains=q) | Q(b2__icontains=q) | Q(b3__icontains=q)
            | Q(elem__icontains=q) | Q(operator__icontains=q)
        )

    def _sukses_gagal():
        return dict(
            jumlah=Count('id'),
            sukses=Coalesce(Sum(Case(When(status_respon='BERHASIL', then=1),
                                     output_field=IntegerField())), 0),
            gagal=Coalesce(Sum(Case(When(status_respon='GAGAL', then=1),
                                    output_field=IntegerField())), 0),
        )

    rekap = list(
        qs.values('b1', 'b2', 'b3', 'elem', 'path5')
        .annotate(**_sukses_gagal())
        .order_by('b1', 'b2', 'b3', 'elem')
    )
    for r in rekap:
        r['info'] = r.pop('path5', '')
        r['performance'] = round(r['sukses'] / r['jumlah'] * 100, 2) if r['jumlah'] else 0

    detail = []
    if ada_pilihan:
        d_qs = qs.filter(b1=pilih['b1'], b2=pilih['b2'], b3=pilih['b3'],
                         elem=pilih['elem'], path5=pilih['info'])
        for rc in d_qs.order_by('-datum_eksekusi')[:RC_DETAIL_MAX]:
            detail.append({
                'waktu': timezone.localtime(rc.datum_eksekusi).strftime('%Y-%m-%d %H:%M:%S')
                         if rc.datum_eksekusi else '',
                'b1': rc.b1, 'b2': rc.b2, 'b3': rc.b3, 'elem': rc.elem, 'info': rc.path5,
                'operator': rc.operator,
                'status_eksekusi': rc.status_eksekusi,
                'waktu_respon': timezone.localtime(rc.datum_respon).strftime('%Y-%m-%d %H:%M:%S')
                                if rc.datum_respon else '',
                'status_respon': rc.status_respon,
            })

    total_jumlah = sum(r['jumlah'] for r in rekap)
    total_sukses = sum(r['sukses'] for r in rekap)
    ringkasan = {
        'jumlah_bay': len(rekap),
        'total_rc': total_jumlah,
        'total_sukses': total_sukses,
        'total_gagal': total_jumlah - total_sukses,
        'rata_rata': round(total_sukses / total_jumlah * 100, 2) if total_jumlah else 0,
    }

    return {
        'tanggal_dari': tanggal_dari, 'tanggal_sampai': tanggal_sampai, 'q': q,
        'pilih': pilih, 'ada_pilihan': ada_pilihan,
        'rekap': rekap, 'detail': detail, 'ringkasan': ringkasan,
        'rekap_headers': RC_REKAP_HEADERS, 'detail_headers': RC_DETAIL_HEADERS,
        'detail_dibatasi': len(detail) >= RC_DETAIL_MAX,
    }


@login_required
def kinerja_rc(request):
    ctx = _rc_ctx(request)
    if request.GET.get('export') == '1':
        return _xlsx([
            ('Rekap', RC_REKAP_HEADERS, RC_REKAP_FIELDS, ctx['rekap']),
            ('Detail', RC_DETAIL_HEADERS, RC_DETAIL_FIELDS, ctx['detail']),
        ], f"kinerja_rc_{ctx['tanggal_dari']}_sd_{ctx['tanggal_sampai']}.xlsx")
    return render(request, 'up2bmakassar/kinerja_rc.html', ctx)


# ── SOE Log — query on-demand read-only ke OFDB, tidak disimpan di PostgreSQL ──────

def _soe_range(request):
    """Parse tanggal_dari/tanggal_sampai (default: hari ini). Return (date_dari, date_sampai)."""
    today = timezone.localdate()

    def _p(name, default):
        raw = request.GET.get(name)
        if raw:
            try:
                return datetime.strptime(raw, '%Y-%m-%d').date()
            except ValueError:
                pass
        return default

    dari = _p('tanggal_dari', today)
    sampai = _p('tanggal_sampai', today)
    if sampai < dari:
        sampai = dari
    return dari, sampai


def _soe_fetch(request):
    """
    Ambil SOE dari OFDB untuk rentang & filter. Return dict siap render/export.
    Query HANYA dijalankan kalau form sudah disubmit (ada param `cari`) -- saat
    pertama buka halaman (belum submit) tabel dibiarkan kosong supaya tidak
    langsung menarik ribuan baris.
    """
    dari, sampai = _soe_range(request)
    filters = {k: request.GET.get(k, '').strip() for k in ofdb.SOE_FILTER_COLUMNS}
    submitted = 'cari' in request.GET

    ctx = {
        'tanggal_dari': dari, 'tanggal_sampai': sampai,
        'filters': filters, 'submitted': submitted,
        'headers': ofdb.SOE_HEADERS, 'rows': [], 'truncated': False,
        'error': None, 'max_rows': ofdb.SOE_MAX_ROWS,
    }

    if not submitted:
        return ctx

    dt_start = datetime.combine(dari, time.min)
    dt_end = datetime.combine(sampai, time.max)

    try:
        conn = ofdb.get_connection()
    except Exception as e:
        ctx['error'] = f'Gagal konek OFDB: {e}'
        return ctx
    try:
        cursor = conn.cursor()
        rows, truncated = ofdb.query_soe(cursor, dt_start, dt_end, filters=filters)
        ctx['rows'] = rows
        ctx['truncated'] = truncated
    except Exception as e:
        ctx['error'] = f'Gagal query SOE: {e}'
    finally:
        conn.close()
    return ctx


@login_required
def soe_log(request):
    if request.GET.get('export') == '1':
        return _soe_export(request)
    return render(request, 'up2bmakassar/soe_log.html', _soe_fetch(request))


def _soe_export(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    ctx = _soe_fetch(request)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'SOE Log'

    header_fill = PatternFill('solid', fgColor='1D4ED8')
    header_font = Font(bold=True, color='FFFFFF')
    for col, h in enumerate(ctx['headers'], start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center')

    for r_idx, row in enumerate(ctx['rows'], start=2):
        for c_idx, val in enumerate(row, start=1):
            # kolom pertama (time_stamp) -> string rapi
            if c_idx == 1 and val is not None:
                val = val.strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=r_idx, column=c_idx, value=val)

    for col in range(1, len(ctx['headers']) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    fname = f"soe_log_{ctx['tanggal_dari']}_sd_{ctx['tanggal_sampai']}.xlsx"
    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    wb.save(resp)
    return resp
